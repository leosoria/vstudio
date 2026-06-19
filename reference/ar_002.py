"""
AR_002 - Credit limit vs outstanding balance analysis.

This module replicates the VBA AR2 logic in Python.

VBA equivalent:
- Create_AR2_Output
- Populate_AR02_From_ZTFI098
- Enrich_AR02_With_Credit_Limits

Output:
- Adds/replaces sheet AR02 in:
    output/LBR_Results_AR_YYYYMMDD.xlsx

AR_002 is independent from the other AR controls:
- If the output workbook exists, AR_002 appends/replaces AR02.
- If the output workbook does not exist, AR_002 creates it.
- AR_002 does not require AR_001 to run before.
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.ar_common import (
    build_customer_erp_map,
    build_intercompany_exclusion_set,
    get_ar_output_file,
    load_ar_input_data,
    normalize_company_output,
    normalize_customer_key,
    normalize_erp_code,
    open_or_create_ar_output_workbook,
    recreate_ar_sheet,
    save_ar_output_workbook,
    to_datetime_value,
    to_number,
    validate_required_columns,
)


AR02_SHEET_NAME = "AR02"

AR02_HEADERS = [
    "Company",
    "ERP Customer Code",
    "Customer Name",
    "Credit Limit Main Currency",
    "Outstanding Balance",
    "Max DaysPastDue",
    "Credit Limit Currency",
    "Company Main Currency",
    "Differences",
]

ZTFI_REQUIRED_COLUMNS = [
    "Empresa",
    "Nº doc.",
    "Cliente",
    "Grp. Emp.",
    "Nome Cli/For",
    "Dt.base prazo pgto",
    "Val. da Transação",
]

CUSTOMER_REQUIRED_COLUMNS = [
    "Cliente",
    "Grupo",
]

SALESFORCE_CURRENT_REQUIRED_COLUMNS = [
    "Código ERP",
    "Limite Aprovado",
    "Status do Limite",
]


def run_ar_002(context):
    """
    Run AR_002 control.

    Expected context keys:
    - project_folder
    - config_path
    - input_folder
    - output_folder
    - module
    - control

    AR_002 loads common AR input data and writes AR02 to the final workbook.
    """

    ar_input_data = context.get("ar_input_data")

    if ar_input_data is None:
        ar_input_data = load_ar_input_data(context)

    ztfi_df = ar_input_data["ztfi"]
    customer_df = ar_input_data["customer"]
    sf_current_df = ar_input_data["salesforce_current"]

    validate_ar002_inputs(
        ztfi_df=ztfi_df,
        customer_df=customer_df,
        sf_current_df=sf_current_df,
    )

    ar02_rows = build_ar02_rows(
        ztfi_df=ztfi_df,
        customer_df=customer_df,
        sf_current_df=sf_current_df,
        module_config=context["module"],
    )

    output_file = get_ar_output_file(context)

    write_ar02_to_workbook(
        output_file=output_file,
        ar02_rows=ar02_rows,
    )

    print(f"AR_002 completed. Sheet {AR02_SHEET_NAME} added to {output_file}")


def validate_ar002_inputs(ztfi_df, customer_df, sf_current_df):
    """
    Validate required input columns before processing.
    """

    missing_ztfi_columns = validate_required_columns(
        ztfi_df,
        ZTFI_REQUIRED_COLUMNS,
    )

    if missing_ztfi_columns:
        raise ValueError(
            "AR_002 missing required columns in ZTFI098 data: "
            f"{missing_ztfi_columns}"
        )

    missing_customer_columns = validate_required_columns(
        customer_df,
        CUSTOMER_REQUIRED_COLUMNS,
    )

    if missing_customer_columns:
        raise ValueError(
            "AR_002 missing required columns in Customer data: "
            f"{missing_customer_columns}"
        )

    missing_sf_current_columns = validate_required_columns(
        sf_current_df,
        SALESFORCE_CURRENT_REQUIRED_COLUMNS,
    )

    if missing_sf_current_columns:
        raise ValueError(
            "AR_002 missing required columns in Salesforce current data: "
            f"{missing_sf_current_columns}"
        )


def build_ar02_rows(
    ztfi_df,
    customer_df,
    sf_current_df,
    module_config,
):
    """
    Build AR02 output rows.

    Replicates VBA:
    - Populate_AR02_From_ZTFI098
    - Enrich_AR02_With_Credit_Limits
    """

    cut_off_date = to_datetime_value(module_config.get("to", ""))

    if cut_off_date is None or str(cut_off_date) == "NaT":
        raise ValueError("Invalid module TO date. Could not calculate Max DaysPastDue.")

    cut_off_date = cut_off_date.normalize()

    customer_erp_map = build_customer_erp_map(customer_df)
    excluded_customers = build_intercompany_exclusion_set()
    salesforce_limit_map = build_ar02_salesforce_limit_map(sf_current_df)

    ar02_map = {}

    for _, row in ztfi_df.iterrows():
        doc_number = row.get("Nº doc.", "")

        if is_blank(doc_number):
            continue

        sap_customer_number = row.get("Cliente", "")
        sap_customer_key = normalize_customer_key(sap_customer_number)

        if sap_customer_key in excluded_customers:
            continue

        company_code = normalize_company_output(row.get("Empresa", ""))

        erp_customer_code_from_zt = normalize_erp_code(row.get("Grp. Emp.", ""))

        if sap_customer_key in customer_erp_map:
            customer_code = normalize_erp_code(customer_erp_map[sap_customer_key])
        else:
            customer_code = erp_customer_code_from_zt

        if customer_code == "":
            customer_code = sap_customer_key

        ar02_key = f"{company_code}|{customer_code}"

        outstanding_balance = to_number(
            row.get("Val. da Transação", 0),
            default=0.0,
        )

        due_date = to_datetime_value(row.get("Dt.base prazo pgto", ""))

        if due_date is None or str(due_date) == "NaT":
            days_past_due = ""
        else:
            due_date = due_date.normalize()
            days_past_due = int((cut_off_date - due_date).days)

        if ar02_key not in ar02_map:
            ar02_map[ar02_key] = {
                "Company": company_code,
                "ERP Customer Code": customer_code,
                "Customer Name": row.get("Nome Cli/For", ""),
                "Credit Limit Main Currency": "",
                "Outstanding Balance": 0.0,
                "Max DaysPastDue": days_past_due,
                "Credit Limit Currency": "",
                "Company Main Currency": "BRL",
                "Differences": "",
            }

        ar02_map[ar02_key]["Outstanding Balance"] += outstanding_balance

        current_max_days = ar02_map[ar02_key]["Max DaysPastDue"]

        if isinstance(days_past_due, int):
            if is_blank(current_max_days):
                ar02_map[ar02_key]["Max DaysPastDue"] = days_past_due
            elif days_past_due > int(current_max_days):
                ar02_map[ar02_key]["Max DaysPastDue"] = days_past_due

    enrich_rows_with_credit_limits(
        ar02_map=ar02_map,
        salesforce_limit_map=salesforce_limit_map,
    )

    return list(ar02_map.values())


def build_ar02_salesforce_limit_map(sf_current_df):
    """
    Build dictionary of valid current credit limits by ERP customer code.

    This replicates the VBA AR2 logic:
    - Exclude Status do Limite = Bloqueado
    - Exclude Status do Limite = Expirado
    - Sum Limite Aprovado by Código ERP
    - No ERP numeric range filter
    """

    limit_map = {}

    if sf_current_df.empty:
        return limit_map

    for _, row in sf_current_df.iterrows():
        customer_code = normalize_erp_code(row.get("Código ERP", ""))
        credit_limit = row.get("Limite Aprovado", 0)
        limit_status = str(row.get("Status do Limite", "")).strip()

        if customer_code == "":
            continue

        if limit_status.upper() in ["BLOQUEADO", "EXPIRADO"]:
            continue

        if customer_code not in limit_map:
            limit_map[customer_code] = 0.0

        limit_map[customer_code] += to_number(credit_limit, default=0.0)

    return limit_map


def enrich_rows_with_credit_limits(ar02_map, salesforce_limit_map):
    """
    Enrich AR02 rows with Salesforce credit limits.

    VBA logic:
    - If customer exists in valid limit dictionary:
        Credit Limit Main Currency = total limit
        Credit Limit Currency = BRL
        Differences = credit limit - outstanding balance
    - Else blank.
    """

    for row in ar02_map.values():
        customer_code = normalize_erp_code(row["ERP Customer Code"])
        outstanding_balance = to_number(row["Outstanding Balance"], default=0.0)

        if customer_code not in salesforce_limit_map:
            row["Credit Limit Main Currency"] = ""
            row["Credit Limit Currency"] = ""
            row["Differences"] = ""
            continue

        credit_limit_amount = to_number(
            salesforce_limit_map[customer_code],
            default=0.0,
        )

        row["Credit Limit Main Currency"] = credit_limit_amount
        row["Credit Limit Currency"] = "BRL"
        row["Differences"] = credit_limit_amount - outstanding_balance


def write_ar02_to_workbook(output_file, ar02_rows):
    """
    Add/replaces AR02 sheet in the final output workbook.

    If the workbook does not exist, create it.
    This allows AR_002 to run independently from all other AR controls.
    """

    workbook = open_or_create_ar_output_workbook(output_file)

    worksheet = recreate_ar_sheet(workbook, AR02_SHEET_NAME)

    write_headers(worksheet)
    write_rows(worksheet, ar02_rows)
    format_worksheet(worksheet)

    save_ar_output_workbook(workbook, output_file)


def write_headers(worksheet):
    """
    Write AR02 headers.
    """

    for column_index, header in enumerate(AR02_HEADERS, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


def write_rows(worksheet, ar02_rows):
    """
    Write AR02 detail rows.
    """

    for row_index, row_data in enumerate(ar02_rows, start=2):
        for column_index, header in enumerate(AR02_HEADERS, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=row_data.get(header, ""),
            )


def format_worksheet(worksheet):
    """
    Apply AR02 worksheet formatting.
    """

    money_columns = [
        "Credit Limit Main Currency",
        "Outstanding Balance",
        "Differences",
    ]

    integer_columns = [
        "Max DaysPastDue",
    ]

    text_columns = [
        "Company",
        "ERP Customer Code",
    ]

    for header in money_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "#,##0.00"

    for header in integer_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0"

    for header in text_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "@"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def get_header_column_index(worksheet, header_name):
    """
    Return one-based column index for a header in row 1.
    """

    expected_header = str(header_name).strip().upper()

    for column_index in range(1, worksheet.max_column + 1):
        current_header = worksheet.cell(row=1, column=column_index).value
        current_header = "" if current_header is None else str(current_header).strip().upper()

        if current_header == expected_header:
            return column_index

    raise ValueError(
        f"Column not found in sheet '{worksheet.title}': {header_name}"
    )


def is_blank(value):
    """
    Return True for None, empty string, or pandas-like NaN text.
    """

    if value is None:
        return True

    value_text = str(value).strip()

    if value_text == "":
        return True

    if value_text.lower() in ["nan", "nat", "none"]:
        return True

    return False
