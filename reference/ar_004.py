"""
AR_004 - Customers over 120 days past due with Salesforce credit limits.

This control creates/replaces only the AR04 sheet in the shared AR output file.

Logic replicated from VBA AR4:
- Load ZTFI098, Customer, and Salesforce current input data.
- Exclude intercompany customers.
- Build the SAP customer -> ERP customer map from Customer.
- Sum valid Salesforce current credit limits by ERP customer code, excluding
  limits with status Bloqueado or Expirado.
- Aggregate ZTFI098 balances by Company + ERP Customer Code.
- Convert BRL transaction balances to USD using the common AR BRL->USD FX rate.
- Calculate Max DaysPastDue using the module TO date as cutoff date.
- Write/recreate only the AR04 sheet in the shared output workbook.
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn

from core.ar_common import (
    build_customer_erp_map,
    build_intercompany_exclusion_set,
    get_ar_output_file,
    load_ar_input_data,
    normalize_company_output,
    normalize_customer_key,
    normalize_erp_code,
    open_or_create_ar_output_workbook,
    recreate_ar_sheet,
    save_ar_output_workbook,
    to_datetime_value,
    to_number,
    validate_required_columns,
)


AR04_SHEET_NAME = "AR04"

AR04_HEADERS = [
    "Company",
    "ERP Customer Code",
    "Customer Name",
    "Credit Limit Main Currency",
    "Outstanding Balance",
    "Max DaysPastDue",
    "Credit Limit Currency",
    "Company Main Currency",
]

ZTFI_REQUIRED_COLUMNS = [
    "Empresa",
    "Nº doc.",
    "Cliente",
    "Grp. Emp.",
    "Nome Cli/For",
    "Dt.base prazo pgto",
    "Moeda transação",
    "Val. da Transação",
]

CUSTOMER_REQUIRED_COLUMNS = [
    "Cliente",
    "Grupo",
]

SALESFORCE_CURRENT_REQUIRED_COLUMNS = [
    "Código ERP",
    "Limite Aprovado",
    "Status do Limite",
]

EXCLUDED_SALESFORCE_LIMIT_STATUSES = [
    "BLOQUEADO",
    "EXPIRADO",
]


def run_ar_004(context):
    """
    Execute AR_004.

    This function:
    - Loads common AR input data.
    - Builds AR04 output rows.
    - Opens or creates the shared AR output workbook.
    - Recreates only the AR04 sheet.
    - Saves the shared output workbook.
    """
    input_data = context.get("ar_input_data")

    if input_data is None:
        input_data = load_ar_input_data(context)

    output_rows = build_ar04_rows(input_data, context)

    output_file = get_ar_output_file(context)
    workbook = open_or_create_ar_output_workbook(output_file)
    worksheet = recreate_ar_sheet(workbook, AR04_SHEET_NAME)

    write_headers(worksheet)
    write_rows(worksheet, output_rows)
    format_worksheet(worksheet)

    save_ar_output_workbook(workbook, output_file)

    print(f"AR_004 completed. Rows written: {len(output_rows)}")
    print(f"Output file: {output_file}")


def build_ar04_rows(input_data, context):
    """
    Build AR04 output rows.

    Aggregation key:
        Company + ERP Customer Code

    Intercompany customers are excluded using build_intercompany_exclusion_set().
    """
    ztfi_df = input_data["ztfi"]
    customer_df = input_data["customer"]
    sf_current_df = input_data["salesforce_current"]

    validate_ar004_inputs(
        ztfi_df=ztfi_df,
        customer_df=customer_df,
        sf_current_df=sf_current_df,
    )

    cutoff_date = get_cutoff_date(input_data, context)

    fx_rate_brl_to_usd = to_number(
        input_data.get("fx_rate_brl_to_usd", 0),
        default=0.0,
    )

    customer_erp_map = build_customer_erp_map(customer_df)
    excluded_customers = build_intercompany_exclusion_set()
    salesforce_limit_map = build_ar04_salesforce_limit_map(sf_current_df)

    ar04_by_key = {}

    for _, row in ztfi_df.iterrows():
        doc_number = row.get("Nº doc.", "")

        if is_blank(doc_number):
            continue

        sap_customer_key = normalize_customer_key(row.get("Cliente", ""))

        if sap_customer_key in excluded_customers:
            continue

        company_output = normalize_company_output(row.get("Empresa", ""))
        customer_name = clean_text(row.get("Nome Cli/For", ""))
        erp_customer_code_from_ztfi = normalize_erp_code(row.get("Grp. Emp.", ""))

        if sap_customer_key in customer_erp_map:
            erp_customer_code = normalize_erp_code(customer_erp_map[sap_customer_key])
        else:
            erp_customer_code = erp_customer_code_from_ztfi

        if erp_customer_code == "":
            continue

        transaction_amount = to_number(
            row.get("Val. da Transação", 0),
            default=0.0,
        )

        document_currency = clean_text(row.get("Moeda transação", "")).upper()

        if document_currency == "BRL":
            applied_fx_rate = fx_rate_brl_to_usd
        else:
            applied_fx_rate = 1

        open_amount_usd = transaction_amount * applied_fx_rate

        due_date = to_datetime_value(row.get("Dt.base prazo pgto", ""))

        if due_date is not None and str(due_date) != "NaT":
            days_past_due = (cutoff_date.date() - due_date.date()).days
        else:
            days_past_due = 0

        ar04_key = f"{company_output}|{erp_customer_code}"

        if ar04_key not in ar04_by_key:
            ar04_by_key[ar04_key] = {
                "Company": company_output,
                "ERP Customer Code": erp_customer_code,
                "Customer Name": customer_name,
                "Credit Limit Main Currency": salesforce_limit_map.get(
                    erp_customer_code,
                    0,
                ),
                "Outstanding Balance": 0,
                "Max DaysPastDue": days_past_due,
                "Credit Limit Currency": "BRL",
                "Company Main Currency": "BRL",
            }

        ar04_row = ar04_by_key[ar04_key]

        ar04_row["Outstanding Balance"] = (
            ar04_row["Outstanding Balance"] + open_amount_usd
        )

        if days_past_due > int(ar04_row["Max DaysPastDue"]):
            ar04_row["Max DaysPastDue"] = days_past_due

        if ar04_row["Customer Name"] == "" and customer_name != "":
            ar04_row["Customer Name"] = customer_name

    return list(ar04_by_key.values())


def validate_ar004_inputs(ztfi_df, customer_df, sf_current_df):
    """
    Validate required input columns before processing.
    """
    missing_ztfi_columns = validate_required_columns(
        ztfi_df,
        ZTFI_REQUIRED_COLUMNS,
    )

    if missing_ztfi_columns:
        raise ValueError(
            "AR_004 missing required columns in ZTFI098 data: "
            f"{missing_ztfi_columns}"
        )

    missing_customer_columns = validate_required_columns(
        customer_df,
        CUSTOMER_REQUIRED_COLUMNS,
    )

    if missing_customer_columns:
        raise ValueError(
            "AR_004 missing required columns in Customer data: "
            f"{missing_customer_columns}"
        )

    missing_sf_current_columns = validate_required_columns(
        sf_current_df,
        SALESFORCE_CURRENT_REQUIRED_COLUMNS,
    )

    if missing_sf_current_columns:
        raise ValueError(
            "AR_004 missing required columns in Salesforce current data: "
            f"{missing_sf_current_columns}"
        )


def build_ar04_salesforce_limit_map(sf_current_df):
    """
    Build valid Salesforce current credit-limit map by ERP customer code.

    VBA equivalent:
    - Normalize Código ERP as numeric text without leading zeros when numeric.
    - Exclude Status do Limite = Bloqueado or Expirado.
    - Sum Limite Aprovado by Código ERP.
    """
    limit_map = {}

    for _, row in sf_current_df.iterrows():
        erp_customer_code = normalize_erp_code(row.get("Código ERP", ""))

        if erp_customer_code == "":
            continue

        limit_status = clean_text(row.get("Status do Limite", "")).upper()

        if limit_status in EXCLUDED_SALESFORCE_LIMIT_STATUSES:
            continue

        credit_limit = to_number(
            row.get("Limite Aprovado", 0),
            default=0.0,
        )

        if erp_customer_code not in limit_map:
            limit_map[erp_customer_code] = 0

        limit_map[erp_customer_code] = limit_map[erp_customer_code] + credit_limit

    return limit_map


def get_cutoff_date(input_data, context):
    """
    Return cutoff date used to calculate days past due.

    VBA source used:
        ThisWorkbook.Sheets("Parameters LBR").Range("B7")

    Python source priority:
    1. Date exposed by load_ar_input_data(context), if available.
    2. Module period end date from config.xlsx: context["module"]["to"].
    """
    possible_input_data_keys = [
        "cutoff_date",
        "cut_off_date",
        "period_end_date",
        "to_date",
        "date_to",
    ]

    for key in possible_input_data_keys:
        if key in input_data:
            cutoff_date = to_datetime_value(input_data[key])

            if cutoff_date is not None and str(cutoff_date) != "NaT":
                return cutoff_date

    module_config = context.get("module", {})

    possible_module_keys = [
        "to",
        "To",
        "date_to",
        "period_end_date",
        "cutoff_date",
        "cut_off_date",
    ]

    for key in possible_module_keys:
        if key in module_config:
            cutoff_date = to_datetime_value(module_config[key])

            if cutoff_date is not None and str(cutoff_date) != "NaT":
                return cutoff_date

    raise ValueError(
        "AR_004 requires a cutoff date to calculate Max DaysPastDue. "
        "No valid cutoff date was found in load_ar_input_data(context) or "
        "context['module']. Expected the module end date from config.xlsx, "
        "usually context['module']['to']."
    )


def write_headers(worksheet):
    """
    Write AR04 headers.
    """
    for column_index, header in enumerate(AR04_HEADERS, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


def write_rows(worksheet, ar04_rows):
    """
    Write AR04 rows.
    """
    for row_index, row_data in enumerate(ar04_rows, start=2):
        for column_index, header in enumerate(AR04_HEADERS, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=row_data.get(header, ""),
            )


def format_worksheet(worksheet):
    """
    Apply AR04 worksheet formatting.
    """
    money_columns = [
        "Credit Limit Main Currency",
        "Outstanding Balance",
    ]

    integer_columns = [
        "Max DaysPastDue",
    ]

    text_columns = [
        "Company",
        "ERP Customer Code",
    ]

    for header in money_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "#,##0.00"

    for header in integer_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0"

    for header in text_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "@"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    apply_max_days_past_due_filter(worksheet)


def apply_max_days_past_due_filter(worksheet):
    """
    Apply the AR04 visual filter for Max DaysPastDue > 120.

    Important OpenPyXL detail:
    - OpenPyXL can write the autofilter criteria to the file.
    - OpenPyXL does not calculate/apply the filter result like Excel does.
    - To make the workbook open already filtered, rows that do not match
      the criteria must also be hidden explicitly.
    """
    if worksheet.max_row <= 1:
        return

    max_days_column_index = get_header_column_index(worksheet, "Max DaysPastDue")
    filter_column = FilterColumn(colId=max_days_column_index - 1)
    filter_column.customFilters = CustomFilters(
        customFilter=[
            CustomFilter(operator="greaterThan", val="120"),
        ]
    )

    worksheet.auto_filter.filterColumn.append(filter_column)

    for row_index in range(2, worksheet.max_row + 1):
        max_days_value = worksheet.cell(
            row=row_index,
            column=max_days_column_index,
        ).value

        max_days_number = to_number(max_days_value, default=0.0)
        worksheet.row_dimensions[row_index].hidden = max_days_number <= 120


def get_header_column_index(worksheet, header_name):
    """
    Return one-based column index for a header in row 1.
    """
    expected_header = str(header_name).strip().upper()

    for column_index in range(1, worksheet.max_column + 1):
        current_header = worksheet.cell(row=1, column=column_index).value
        current_header = "" if current_header is None else str(current_header).strip().upper()

        if current_header == expected_header:
            return column_index

    raise ValueError(
        f"Column not found in sheet '{worksheet.title}': {header_name}"
    )


def clean_text(value):
    """
    Return stripped text, treating pandas-like blanks as empty string.
    """
    if is_blank(value):
        return ""

    return str(value).strip()


def is_blank(value):
    """
    Return True for None, empty string, or pandas-like NaN text.
    """
    if value is None:
        return True

    value_text = str(value).strip()

    if value_text == "":
        return True

    if value_text.lower() in ["nan", "nat", "none"]:
        return True

    return False
