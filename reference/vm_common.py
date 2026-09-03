"""Common infrastructure for LBR Vendor Management controls.

This module resolves, loads, normalizes and validates the shared SAP ECC
extracts used by VM controls. Control-specific analytics do not belong here.

Required input:
    LBR VM_VENDORS_YYYYMMDD.xlsx

Optional paired inputs:
    LBR VM_VPBSIK_YYYYMMDD.xlsx
    LBR VM_VPBSAK_YYYYMMDD.xlsx

Optional accounting-document header input:
    LBR_VM_BKPF_YYYYMMDD.xlsx

Optional bank-change inputs:
    LBR VM_BANK_CDHDR_YYYYMMDD.xlsx
    LBR VM_BANK_CDPOS_YYYYMMDD.xlsx

Optional employee input:
    LBR_VM_EMP_YYYYMMDD.xlsx

YYYYMMDD is derived strictly from the VM module TO date. Files from another
period, with another name or from another scope are never used as fallback.
"""

import re
import tempfile
import unicodedata
from collections.abc import Iterable, Iterator, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.intercompanies import INTERCOMPANIES

VM_INPUT_SHEET = "Hoja1"
VM_INPUT_SHEET_ALIASES = (
    "Sheet1",
    "Hoja1",
)
VM_HEADER_ROW = 1

VM_VENDOR_FILE_TEMPLATE = "LBR VM_VENDORS_{period}.xlsx"
VM_BSIK_FILE_TEMPLATE = "LBR VM_VPBSIK_{period}.xlsx"
VM_BSAK_FILE_TEMPLATE = "LBR VM_VPBSAK_{period}.xlsx"
VM_BKPF_FILE_TEMPLATE = "LBR_VM_BKPF_{period}.xlsx"
VM_EMPLOYEE_FILE_TEMPLATE = "LBR_VM_EMP_{period}.xlsx"
VM_BANK_CDHDR_FILE_TEMPLATE = "LBR VM_BANK_CDHDR_{period}.xlsx"
VM_BANK_CDPOS_FILE_TEMPLATE = "LBR VM_BANK_CDPOS_{period}.xlsx"

# SAP ECC document types equivalent to OPCH AP invoices.
VM_INVOICE_DOCUMENT_TYPES = frozenset(
    {
        "RE",
        "KR",
    }
)

VM_LAST_INVOICE_COLUMNS = (
    "Last Invoice Number",
    "Last Transaction Date",
    "Last Inv Amt Doc Currency",
    "Last Inv Amt Doc Currency Indicator",
)

VM_OUTPUT_FILE_TEMPLATE = "LBR_Results_VM_{period}.xlsx"

# VM01 is the official visual-format reference.
VM_OUTPUT_HEADER_FILL = "FFD9EAF7"

VM_DATE_NUMBER_FORMAT = "DD/MM/YYYY"
VM_AMOUNT_NUMBER_FORMAT = "General"
VM_INTEGER_NUMBER_FORMAT = "General"

INVALID_TEXT_VALUES = {
    "",
    "nan",
    "none",
    "nat",
    "<na>",
}

VENDOR_KEY_COLUMNS = (
    "Company",
    "Vendor Code",
)

BANK_KEY_COLUMNS = (
    "Bank Country",
    "Bank Code",
    "Bank Account",
)

VENDOR_BANK_KEY_COLUMNS = (
    *VENDOR_KEY_COLUMNS,
    *BANK_KEY_COLUMNS,
)

POSTING_KEY_COLUMNS = (
    "Company",
    "Fiscal Year",
    "Accounting Document",
    "Accounting Document Line",
)

POSTING_HEADER_KEY_COLUMNS = (
    "Company",
    "Fiscal Year",
    "Accounting Document",
)

POSTING_HEADER_COLUMNS = (
    "Company",
    "Fiscal Year",
    "Accounting Document",
    "Posting Date",
    "Posting User",
)

POSTING_HEADER_REQUIRED_COLUMNS = (
    "Company",
    "Fiscal Year",
    "Accounting Document",
    "Posting Date",
    "Posting User",
)

POSTING_HEADER_PHYSICAL_COLUMNS = (
    "Company Code",
    "Document Number",
    "Fiscal Year",
    "Posting Date",
    "User name",
)

POSTING_HEADER_ALIASES = {
    "Company": (
        "Company Code",
    ),
    "Fiscal Year": (
        "Fiscal Year",
    ),
    "Accounting Document": (
        "Document Number",
    ),
    "Posting Date": (
        "Posting Date",
    ),
    "Posting User": (
        "User name",
    ),
}

CHANGE_KEY_COLUMNS = (
    "Change Object Class",
    "Object Value",
    "Change Document",
)

EMPLOYEE_KEY_COLUMNS = (
    "Company",
    "Employee Code",
)

BANK_DETAIL_COLUMNS = (
    "Bank Country",
    "Bank Code",
    "Bank Account",
    "Account Holder Name",
    "Bank Valid From",
    "Bank Valid To",
)

VENDOR_COLUMNS = (
    "Company",
    "Company Name",
    "Vendor Code",
    "Vendor Name",
    "Country",
    "City",
    "District",
    "ZipCode",
    "State",
    "Street",
    "Phone1",
    "Tax Number 1",
    "Tax Number 2",
    "Tax Number 3",
    "Tax Number 4",
    "Tax Number 5",
    "VAT Registration Number",
    "Tax Jurisdiction",
    "Tax Type",
    "Tax Number Type",
    "Account Group",
    "Trading Partner",
    "One-Time Vendor",
    "Central Deletion Flag",
    "Central Posting Block",
    "Purchasing Block",
    "Central Payment Block",
    "Payment Block",
    "Company Deletion Flag",
    "Company Posting Block",
    "Payment Terms",
    "Payment Methods",
    "Vendor Created Date",
    "Vendor Created By",
    "Vendor Updated Date",
    "Company Extension Date",
    "Company Extension Created By",
    "Bank Country",
    "Bank Code",
    "Bank Account",
    "Account Holder Name",
    "Bank Valid From",
    "Bank Valid To",
)

VENDOR_REQUIRED_COLUMNS = (
    "Company",
    "Vendor Code",
)

VENDOR_COLUMN_ALIASES = {
    "Company": (
        "Company",
        "Company Code",
        "CoCd",
        "BUKRS",
        "LFB1-BUKRS",
    ),
    "Company Name": (
        "Company Name",
        "BUTXT",
        "T001-BUTXT",
    ),
    "Vendor Code": (
        "Vendor Code",
        "Vendor",
        "LIFNR",
        "LFA1-LIFNR",
    ),
    "Vendor Name": (
        "Vendor Name",
        "Name 1",
        "NAME1",
        "LFA1-NAME1",
    ),
    "Country": (
        "Country",
        "Cty",
        "LAND1",
        "LFA1-LAND1",
    ),
    "City": (
        "City",
        "ORT01",
        "LFA1-ORT01",
    ),
    "District": (
        "District",
        "ORT02",
        "LFA1-ORT02",
    ),
    "ZipCode": (
        "ZipCode",
        "Postal Code",
        "PostalCode",
        "PSTLZ",
        "LFA1-PSTLZ",
    ),
    "State": (
        "State",
        "Region",
        "Rg",
        "REGIO",
        "LFA1-REGIO",
    ),
    "Street": (
        "Street",
        "STRAS",
        "LFA1-STRAS",
    ),
    "Phone1": (
        "Phone1",
        "Telephone",
        "Telephone 1",
        "TELF1",
        "LFA1-TELF1",
    ),
    "Tax Number 1": (
        "Tax Number 1",
        "STCD1",
        "LFA1-STCD1",
    ),
    "Tax Number 2": (
        "Tax Number 2",
        "STCD2",
        "LFA1-STCD2",
    ),
    "Tax Number 3": (
        "Tax Number 3",
        "STCD3",
        "LFA1-STCD3",
    ),
    "Tax Number 4": (
        "Tax Number 4",
        "STCD4",
        "LFA1-STCD4",
    ),
    "Tax Number 5": (
        "Tax Number 5",
        "STCD5",
        "LFA1-STCD5",
    ),
    "VAT Registration Number": (
        "VAT Registration Number",
        "STCEG",
        "LFA1-STCEG",
    ),
    "Tax Jurisdiction": (
        "Tax Jurisdiction",
        "Tax Jur.",
        "TXJCD",
        "LFA1-TXJCD",
    ),
    "Tax Type": (
        "Tax Type",
        "FITYP",
        "LFA1-FITYP",
    ),
    "Tax Number Type": (
        "Tax Number Type",
        "STCDT",
        "LFA1-STCDT",
    ),
    "Account Group": (
        "Account Group",
        "Group",
        "KTOKK",
        "LFA1-KTOKK",
    ),
    "Trading Partner": (
        "Trading Partner",
        "VBUND",
        "LFA1-VBUND",
    ),
    "One-Time Vendor": (
        "One-Time Vendor",
        "XCPDK",
        "LFA1-XCPDK",
    ),
    "Central Deletion Flag": (
        "Central Deletion Flag",
        "LOEVM",
        "LFA1-LOEVM",
    ),
    "Central Posting Block": (
        "Central Posting Block",
        "SPERR",
        "LFA1-SPERR",
    ),
    "Purchasing Block": (
        "Purchasing Block",
        "SPERM",
        "LFA1-SPERM",
    ),
    "Central Payment Block": (
        "Central Payment Block",
        "SPERZ",
        "LFA1-SPERZ",
    ),
    "Payment Block": (
        "Payment Block",
        "ZAHLS",
        "LFB1-ZAHLS",
    ),
    "Company Deletion Flag": (
        "Company Deletion Flag",
        "LFB1_LOEVM",
        "LFB1-LOEVM",
    ),
    "Company Posting Block": (
        "Company Posting Block",
        "LFB1_SPERR",
        "LFB1-SPERR",
    ),
    "Payment Terms": (
        "Payment Terms",
        "ZTER",
        "ZTERM",
        "LFB1-ZTERM",
    ),
    "Payment Methods": (
        "Payment Methods",
        "ZWELS",
        "LFB1-ZWELS",
    ),
    "Vendor Created Date": (
        "Vendor Created Date",
        "Date",
        "LFA1-ERDAT",
    ),
    "Vendor Created By": (
        "Vendor Created By",
        "Created by",
        "LFA1-ERNAM",
    ),
    "Vendor Updated Date": (
        "Vendor Updated Date",
        "UPDAT",
        "LFA1-UPDAT",
    ),
    "Company Extension Date": (
        "Company Extension Date",
        "Date.1",
        "LFB1-ERDAT",
    ),
    "Company Extension Created By": (
        "Company Extension Created By",
        "Created by.1",
        "LFB1-ERNAM",
    ),
    "Bank Country": (
        "Bank Country",
        "Ctry",
        "BANKS",
        "LFBK-BANKS",
    ),
    "Bank Code": (
        "Bank Code",
        "Bank Key",
        "BANKL",
        "LFBK-BANKL",
    ),
    "Bank Account": (
        "Bank Account",
        "BANKN",
        "LFBK-BANKN",
    ),
    "Account Holder Name": (
        "Account Holder Name",
        "Acct holder",
        "KOINH",
        "LFBK-KOINH",
    ),
    "Bank Valid From": (
        "Bank Valid From",
        "KOVON",
        "LFBK-KOVON",
    ),
    "Bank Valid To": (
        "Bank Valid To",
        "KOBIS",
        "LFBK-KOBIS",
    ),
}



# Exact physical aliases present in the current LBR VM vendor extract.
#
# These assignments complete optional fields that are present in the workbook
# but are not resolved by the generic SAP aliases.
VENDOR_COLUMN_ALIASES.update(
    {
        "VAT Registration Number": (
            "VAT Registration No.",
        ),
        "Tax Type": (
            "Tax type",
        ),
        "Tax Number Type": (
            "Tax no.ty.",
        ),
        "Trading Partner": (
            "Tr.Prt",
        ),
        "One-Time Vendor": (
            "One-time",
        ),
        "Central Payment Block": (
            "SPERZ",
        ),
        "Payment Block": (
            "ZAHLS",
        ),
        "Company Posting Block": (
            "B",
        ),
        "Vendor Updated Date": (
            "Conf.date",
        ),
        "Bank Valid From": (
            "Valid from",
        ),
        "Bank Valid To": (
            "Eff.to",
        ),
    }
)



POSTING_COLUMNS = (
    "Company",
    "Vendor Code",
    "Fiscal Year",
    "Accounting Document",
    "Accounting Document Line",
    "Posting Date",
    "Document Type",
    "Debit/Credit Indicator",
    "Document Currency",
    "Amount in Document Currency",
    "Amount in Local Currency",
    "Clearing Date",
    "Clearing Document",
)

POSTING_REQUIRED_COLUMNS = (
    "Company",
    "Vendor Code",
    "Fiscal Year",
    "Accounting Document",
    "Accounting Document Line",
    "Posting Date",
)

EMPLOYEE_COLUMNS = (
    "Company",
    "Employee Code",
    "Employee Name",
    "Organizational Valid From",
    "Organizational Valid To",
    "Employment Status",
    "Status Valid From",
    "Status Valid To",
    "Personnel Area",
    "Personnel Subarea",
    "Employee Group",
    "Employee Subgroup",
    "Organizational Unit",
    "Position",
    "Cost Center",
)

EMPLOYEE_ALIASES = {
    "Company": (
        "Company",
        "Company Code",
        "BUKRS",
        "PA0001-BUKRS",
        "PAOOOI-BUKRS",
    ),
    "Employee Code": (
        "Employee Code",
        "Personnel Number",
        "PERNR",
        "PA0001-PERNR",
        "PAOOOI-PERNR",
    ),
    "Employee Name": (
        "Employee Name",
        "Formatted Name",
        "ENAME",
        "PA0001-ENAME",
        "PAOOOI-ENAME",
    ),
    "Organizational Valid From": (
        "Organizational Valid From",
        "PA0001-BEGDA",
        "PAOOOI-BEGDA",
    ),
    "Organizational Valid To": (
        "Organizational Valid To",
        "PA0001-ENDDA",
        "PAOOOI-ENDDA",
    ),
    "Employment Status": (
        "Employment Status",
        "STAT2",
        "PA0000-STAT2",
        "PAOOOO-STAT2",
    ),
    "Status Valid From": (
        "Status Valid From",
        "PA0000-BEGDA",
        "PAOOOO-BEGDA",
    ),
    "Status Valid To": (
        "Status Valid To",
        "PA0000-ENDDA",
        "PAOOOO-ENDDA",
    ),
    "Personnel Area": (
        "Personnel Area",
        "WERKS",
        "PA0001-WERKS",
        "PAOOOI-WERKS",
    ),
    "Personnel Subarea": (
        "Personnel Subarea",
        "BTRTL",
        "PA0001-BTRTL",
        "PAOOOI-BTRTL",
    ),
    "Employee Group": (
        "Employee Group",
        "PERSG",
        "PA0001-PERSG",
        "PAOOOI-PERSG",
    ),
    "Employee Subgroup": (
        "Employee Subgroup",
        "PERSK",
        "PA0001-PERSK",
        "PAOOOI-PERSK",
    ),
    "Organizational Unit": (
        "Organizational Unit",
        "ORGEH",
        "PA0001-ORGEH",
        "PAOOOI-ORGEH",
    ),
    "Position": (
        "Position",
        "PLANS",
        "PA0001-PLANS",
        "PAOOOI-PLANS",
    ),
    "Cost Center": (
        "Cost Center",
        "KOSTL",
        "PA0001-KOSTL",
        "PAOOOI-KOSTL",
    ),
}

# Exact physical headers used by the current LBR VM employee extract.
#
# Keep this block after the complete EMPLOYEE_ALIASES dictionary so these
# assignments override generic SAP aliases deterministically.
EMPLOYEE_ALIASES.update(
    {
        "Company": (
            "CoCd",
        ),
        "Employee Code": (
            "PersNo",
        ),
        "Employee Name": (
            "Name of employee or applicant",
        ),
        "Organizational Valid From": (
            "Start Date",
        ),
        "Organizational Valid To": (
            "End Date",
        ),
        "Employment Status": (
            "S",
        ),
        "Status Valid From": (
            "Start Date.1",
        ),
        "Status Valid To": (
            "End Date.1",
        ),
        "Personnel Area": (
            "PA",
        ),
        "Personnel Subarea": (
            "PSubarea",
        ),
        "Employee Group": (
            "EEGrp",
        ),
        "Employee Subgroup": (
            "ESgrp",
        ),
        "Organizational Unit": (
            "Org.unit",
        ),
        "Position": (
            "Position",
        ),
        "Cost Center": (
            "Cost Ctr",
        ),
    }
)

EMPLOYEE_REQUIRED_COLUMNS = (
    "Company",
    "Employee Code",
    "Employee Name",
    "Organizational Valid From",
    "Organizational Valid To",
    "Employment Status",
    "Status Valid From",
    "Status Valid To",
)

CDHDR_COLUMNS = (
    "Change Object Class",
    "Object Value",
    "Change Document",
    "Changed By",
    "Change Date",
    "Change Time",
    "Transaction Code",
)

CDHDR_ALIASES = {
    "Change Object Class": (
        "Change Object Class",
        "Change doc. object",
        "OBJECTCLAS",
        "CDHDR-OBJECTCLAS",
    ),
    "Object Value": (
        "Object Value",
        "OBJECTID",
        "CDHDR-OBJECTID",
    ),
    "Change Document": (
        "Change Document",
        "Document number",
        "CHANGENR",
        "CDHDR-CHANGENR",
    ),
    "Changed By": (
        "Changed By",
        "User",
        "User Name",
        "USERNAME",
        "CDHDR-USERNAME",
    ),
    "Change Date": (
        "Change Date",
        "Date",
        "UDATE",
        "CDHDR-UDATE",
    ),
    "Change Time": (
        "Change Time",
        "Time",
        "UTIME",
        "CDHDR-UTIME",
    ),
    "Transaction Code": (
        "Transaction Code",
        "TCODE",
        "CDHDR-TCODE",
    ),
}

CDPOS_COLUMNS = (
    "Change Object Class",
    "Object Value",
    "Change Document",
    "Changed Table",
    "Changed Record Key",
    "Changed Field",
    "Change Type",
    "New Value",
    "Old Value",
)

CDPOS_ALIASES = {
    "Change Object Class": (
        "Change Object Class",
        "Change doc. object",
        "OBJECTCLAS",
        "CDPOS-OBJECTCLAS",
    ),
    "Object Value": (
        "Object Value",
        "OBJECTID",
        "CDPOS-OBJECTID",
    ),
    "Change Document": (
        "Change Document",
        "Document number",
        "CHANGENR",
        "CDPOS-CHANGENR",
    ),
    "Changed Table": (
        "Changed Table",
        "Table Name",
        "TABNAME",
        "CDPOS-TABNAME",
    ),
    "Changed Record Key": (
        "Changed Record Key",
        "Table Key",
        "TABKEY",
        "CDPOS-TABKEY",
    ),
    "Changed Field": (
        "Changed Field",
        "Field Name",
        "FNAME",
        "CDPOS-FNAME",
    ),
    "Change Type": (
        "Change Type",
        "Change ID",
        "CHNGIND",
        "CDPOS-CHNGIND",
    ),
    "New Value": (
        "New Value",
        "VALUE_NEW",
        "CDPOS-VALUE_NEW",
    ),
    "Old Value": (
        "Old Value",
        "VALUE_OLD",
        "CDPOS-VALUE_OLD",
    ),
}


def safe_text(value: Any) -> str:
    """Return trimmed text without missing-value placeholders."""
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value).strip()

    if text.casefold() in INVALID_TEXT_VALUES:
        return ""

    return text


def is_blank(value: Any) -> bool:
    """Return whether a value is empty after safe conversion."""
    return safe_text(value) == ""


def remove_accents(value: Any) -> str:
    """Return text without Unicode accents."""
    normalized = unicodedata.normalize("NFKD", safe_text(value))

    return "".join(
        character
        for character in normalized
        if not unicodedata.combining(character)
    )


def normalize_upper_text(value: Any) -> str:
    """Return accent-insensitive uppercase text."""
    return " ".join(remove_accents(value).upper().split())


def normalize_header_lookup(value: Any) -> str:
    """Return punctuation-insensitive header key."""
    return re.sub(
        r"[^A-Z0-9]+",
        "",
        normalize_upper_text(value),
    )


def normalize_exact_key(value: Any) -> str:
    """Return uppercase alphanumeric comparison key."""
    return re.sub(
        r"[^A-Z0-9]+",
        "",
        normalize_upper_text(value),
    )


# Legal-entity suffixes removed from vendor names before VM01 fuzzy matching.
#
# The source name is already normalized to uppercase, without accents or
# punctuation. Compound legal forms such as "S.A.", "S.R.L." and
# "S. DE R.L. DE C.V." are therefore removed token by token from the end.
VENDOR_COMPANY_SUFFIX_TOKENS = frozenset(
    {
        # Generic / English
        "LTD",
        "LTDA",
        "LIMITED",
        "LIMITADA",
        "LDA",
        "INC",
        "INCORPORATED",
        "INCORPORATION",
        "CORP",
        "CORPORATION",
        "CO",
        "COMPANY",
        "COMPANIA",
        "CIA",
        "LLC",
        "LLP",
        "LLLP",
        "LP",
        "GP",
        "SP",
        "PLC",
        "PC",
        "PA",
        "PLLC",
        "PSC",
        "TRUST",
        "FUND",
        "FOUNDATION",
        "FOUNDATIONS",
        "HOLDING",
        "HOLDINGS",
        "GROUP",
        "GROUPE",
        "NGO",
        "NPO",
        "ASSOCIATION",
        "INSTITUTE",
        "INSTITUTES",
        "SOCIETY",
        "UNION",
        "SYNDICATE",
        "COOP",
        "COOPERATIVE",
        "COOPERATIVA",
        "POOL",
        "NATIONAL",
        "FEDERAL",
        "INDUSTRIES",
        "IND",
        "BANK",
        "BANKERS",
        "CLUB",
        "INTERNATIONAL",
        "INTL",
        "UNLTD",
        "ULTD",
        "NL",
        "NO",
        "LIABILITY",
        "PVT",
        "PTE",
        "PTY",
        "BK",
        "CC",
        "LC",
        "SMLLC",
        "CIC",
        "CIO",
        "CCC",
        "PRIVATE",
        "PROPRIETARY",
        "PROFESSIONAL",
        "REGISTERED",
        "PARTNERSHIP",
        "JOINT",
        "VENTURE",
        "VENTURES",
        "JV",
        "OF",
        "ON",
        "AT",

        # Spanish / Latin America
        "SA",
        "SAC",
        "SACI",
        "SAICF",
        "SACIF",
        "SAS",
        "SCA",
        "SRL",
        "SR",
        "EIRL",
        "SAPI",
        "SAB",
        "SAD",
        "SAL",
        "SGR",
        "SC",
        "SCP",
        "SCS",
        "SCCL",
        "SCOP",
        "SL",
        "SLL",
        "SLNE",
        "SCRA",
        "SOC",
        "SOCIEDAD",
        "SOCIEDADES",
        "ANONIMA",
        "LIMITADAS",
        "COLECTIVA",
        "COMANDITARIA",
        "COMANDITA",
        "RESPONSABILIDAD",
        "RECIPROCA",
        "GARANTIA",
        "ACCIONES",
        "SIMPLIFICADA",
        "UNIPERSONAL",
        "EMPRESA",
        "EMPRESAS",
        "SUCESORES",
        "SUC",
        "EU",
        "Y",

        # Portuguese
        "SGPS",
        "EIRELI",
        "ME",
        "EPP",
        "MEI",
        "ABERTA",
        "FECHADA",
        "SF",

        # Italian
        "SPA",
        "SAPA",
        "SNC",
        "SAA",

        # German
        "GMBH",
        "AG",
        "KG",
        "KGAA",
        "OHG",
        "GBR",
        "EG",
        "UG",
        "MBH",

        # French
        "SARL",
        "SARLU",
        "EURL",
        "SASU",
        "SCI",
        "SEM",
        "GIE",
        "SEP",
        "FCP",
        "SICAV",
        "SCE",
        "EEIG",

        # Dutch / Belgian / Nordic / other
        "BV",
        "NV",
        "CV",
        "VOF",
        "AB",
        "OY",
        "AS",
        "ASA",
        "APS",
        "SE",
        "SARF",
        "SDN",
        "BHD",
        "KK",
        "YK",
        "ZAO",
        "OAO",
        "OOO",
        "PAO",
        "TEO",
        "TEORANTA",
        "OYJ",

        # Individual tokens composing compound legal forms:
        # S A, C V, S DE R L DE C V, etc.
        "S",
        "A",
        "C",
        "V",
        "L",
        "R",
        "E",
        "U",
        "I",
        "P",
        "G",
        "F",
        "B",
        "T",
        "K",
        "N",
        "D",
        "M",
        "DE",
        "EN",
        "RL",
        "DEL",
        "LA",
        "EL",
        "DA",
    }
)


def clean_vendor_name(
    value: Any,
    suffix_tokens: Iterable[Any] | None = None,
) -> str:
    """
    Normalize one vendor name for VM01 fuzzy comparison.

    Rules:
    - missing values become an empty string;
    - accents and diacritics are removed;
    - text is converted to uppercase;
    - punctuation and separators become spaces;
    - repeated whitespace is collapsed;
    - trailing legal-entity suffixes are removed token by token;
    - at least one name token is retained.

    ``suffix_tokens`` is optional so controls may supply an explicitly approved
    suffix population without duplicating the normalization implementation.
    """
    text = normalize_upper_text(value)

    if text == "":
        return ""

    text = re.sub(
        r"[^A-Z0-9]+",
        " ",
        text,
    ).strip()

    if text == "":
        return ""

    tokens = text.split()

    if suffix_tokens is None:
        normalized_suffixes = VENDOR_COMPANY_SUFFIX_TOKENS
    else:
        normalized_suffixes = frozenset(
            normalize_upper_text(token)
            for token in suffix_tokens
            if not is_blank(token)
        )

    while (
        len(tokens) > 1
        and tokens[-1] in normalized_suffixes
    ):
        tokens.pop()

    return " ".join(tokens)



def normalize_phone(value: Any) -> str:
    """Return only digits from a telephone value."""
    return re.sub(r"\D+", "", safe_text(value))


def normalize_tax_id(value: Any) -> str:
    """Return uppercase alphanumeric Tax ID."""
    return normalize_exact_key(value)


def normalize_identifier(value: Any) -> str:
    """Normalize an identifier without producing values such as 12345.0."""
    text = safe_text(value)

    if re.fullmatch(r"[+-]?\d+\.0+", text):
        return text[:text.index(".")]

    return text


def normalize_company(value: Any) -> str:
    """Normalize numeric SAP company codes to four digits."""
    text = normalize_identifier(value)

    if text.isdigit():
        return text.zfill(4)

    return text.upper()


def normalize_vendor_code(value: Any) -> str:
    """Normalize numeric SAP vendor codes to ten digits."""
    text = normalize_identifier(value)

    if text.isdigit():
        return text.zfill(10)

    return text.upper()


def normalize_employee_code(value: Any) -> str:
    """Normalize numeric SAP personnel numbers to eight digits."""
    text = normalize_identifier(value)

    if text.isdigit():
        return text.zfill(8)

    return text.upper()


def normalize_document_number(value: Any) -> str:
    """Normalize numeric SAP accounting documents to ten digits."""
    text = normalize_identifier(value)

    if text.isdigit():
        return text.zfill(10)

    return text.upper()


def normalize_line_number(value: Any) -> str:
    """Normalize numeric SAP accounting lines to three digits."""
    text = normalize_identifier(value)

    if text.isdigit():
        return text.zfill(3)

    return text.upper()


def normalize_date_text(value: Any) -> str:
    """
    Return ISO date text when the value can be parsed.

    Native Excel/Python dates and ISO-formatted strings are parsed without
    dayfirst. Other textual dates use the LBR day-first convention.
    """
    if is_blank(value):
        return ""

    # pandas Timestamp and native Excel datetime/date values can be normalized
    # directly without format inference.
    if not isinstance(
        value,
        str,
    ):
        try:
            parsed = pd.Timestamp(
                value
            )
        except (
            TypeError,
            ValueError,
            OverflowError,
        ):
            return safe_text(
                value
            )

        if pd.isna(parsed):
            return safe_text(
                value
            )

        return parsed.strftime(
            "%Y-%m-%d"
        )

    text = safe_text(
        value
    )

    # ISO date or ISO datetime: year is unambiguous and dayfirst must not be
    # supplied, avoiding pandas warnings.
    if re.fullmatch(
        r"\d{4}-\d{2}-\d{2}",
        text,
    ):
        parsed = pd.to_datetime(
            text,
            format="%Y-%m-%d",
            errors="coerce",
        )

    elif re.fullmatch(
        r"\d{4}-\d{2}-\d{2}[ T].+",
        text,
    ):
        parsed = pd.to_datetime(
            text,
            errors="coerce",
        )

    else:
        parsed = pd.to_datetime(
            text,
            errors="coerce",
            dayfirst=True,
        )

    if pd.isna(parsed):
        return text

    return pd.Timestamp(
        parsed
    ).strftime(
        "%Y-%m-%d"
    )


def parse_config_date(
    value: Any,
    field_name: str,
) -> pd.Timestamp:
    """Parse one required VM configuration date."""
    text = safe_text(value)

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        parsed = pd.to_datetime(
            text,
            format="%Y-%m-%d",
            errors="coerce",
        )
    else:
        parsed = pd.to_datetime(
            value,
            errors="coerce",
            dayfirst=True,
        )

    if pd.isna(parsed):
        raise ValueError(
            f"VM CONFIG {field_name} is empty or invalid: {value!r}."
        )

    return pd.Timestamp(parsed).normalize()


def get_period_suffix(module_config: dict[str, Any]) -> str:
    """Return YYYYMMDD from the VM module TO date."""
    return parse_config_date(
        module_config.get("to", ""),
        "TO",
    ).strftime("%Y%m%d")


def get_vm_period(
    context: dict[str, Any],
) -> tuple[pd.Timestamp, pd.Timestamp]:
    """Return normalized VM FROM and TO dates."""
    module_config = context.get("module")

    if not isinstance(module_config, dict):
        raise ValueError(
            "VM context requires context['module'] configuration."
        )

    date_from = parse_config_date(
        module_config.get("from", ""),
        "FROM",
    )
    date_to = parse_config_date(
        module_config.get("to", ""),
        "TO",
    )

    if date_from > date_to:
        raise ValueError(
            "VM CONFIG FROM cannot be later than TO."
        )

    return date_from, date_to


def build_vendor_key(
    company: Any,
    vendor_code: Any,
) -> str:
    """Build mandatory Company + Vendor Code key."""
    company = normalize_company(company)
    vendor_code = normalize_vendor_code(vendor_code)

    if company == "" or vendor_code == "":
        return ""

    return f"{company}|{vendor_code}"


def build_employee_key(
    company: Any,
    employee_code: Any,
) -> str:
    """Build mandatory Company + Employee Code key."""
    company = normalize_company(company)
    employee_code = normalize_employee_code(employee_code)

    if company == "" or employee_code == "":
        return ""

    return f"{company}|{employee_code}"


def require_nonblank_normalized_key(
    value: Any,
    *,
    field_name: str = "key",
) -> str:
    """Normalize a key and reject empty normalized values."""
    normalized = normalize_exact_key(value)

    if normalized == "":
        raise ValueError(
            f"VM grouping failed because {field_name} is empty "
            "after normalization."
        )

    return normalized


def validate_group_has_distinct_vendors(
    dataframe: pd.DataFrame,
    *,
    vendor_key_column: str = "Vendor Key",
    minimum_vendors: int = 2,
) -> None:
    """Require a group to contain at least two distinct vendors."""
    if vendor_key_column not in dataframe.columns:
        raise ValueError(
            f"VM group requires column '{vendor_key_column}'."
        )

    vendor_keys = {
        safe_text(value)
        for value in dataframe[vendor_key_column]
        if not is_blank(value)
    }

    if len(vendor_keys) < minimum_vendors:
        raise ValueError(
            f"VM group requires {minimum_vendors} distinct vendors; "
            f"found {len(vendor_keys)}."
        )


def iter_company_key_groups(
    dataframe: pd.DataFrame,
    normalized_key_column: str,
) -> Iterator[tuple[tuple[str, str], pd.DataFrame]]:
    """Yield nonblank groups, always separated by Company."""
    required = {
        "Company",
        normalized_key_column,
    }
    missing = sorted(required.difference(dataframe.columns))

    if missing:
        raise ValueError(
            f"VM grouping requires missing columns: {missing}."
        )

    working = dataframe.copy()
    working["Company"] = working["Company"].map(normalize_company)
    working[normalized_key_column] = working[
        normalized_key_column
    ].map(normalize_exact_key)

    working = working.loc[
        working["Company"].ne("")
        & working[normalized_key_column].ne("")
    ]

    for group_key, group in working.groupby(
        ["Company", normalized_key_column],
        sort=True,
        dropna=False,
    ):
        yield group_key, group.copy()


def _expected_file(
    context: dict[str, Any],
    template: str,
) -> Path:
    input_folder = Path(context["input_folder"])

    if not input_folder.is_dir():
        raise FileNotFoundError(
            f"VM input folder was not found: {input_folder}"
        )

    module_config = context.get("module")

    if not isinstance(module_config, dict):
        raise ValueError(
            "VM input resolution requires context['module']."
        )

    expected_name = template.format(
        period=get_period_suffix(module_config)
    )

    return input_folder / expected_name


def find_vm_input_file(
    context: dict[str, Any],
    template: str = VM_VENDOR_FILE_TEMPLATE,
    *,
    required: bool = True,
) -> Path | None:
    """
    Resolve one exact period-specific VM input.

    Filename matching is case-insensitive only for filesystem casing, including
    .xlsx versus .XLSX. No fallback to another name, scope or period is allowed.
    """
    expected_file = _expected_file(
        context,
        template,
    )

    if expected_file.is_file():
        return expected_file

    input_folder = expected_file.parent
    expected_name = expected_file.name.casefold()

    case_insensitive_matches = [
        candidate
        for candidate in input_folder.iterdir()
        if (
            candidate.is_file()
            and candidate.name.casefold()
            == expected_name
        )
    ]

    if len(case_insensitive_matches) == 1:
        return case_insensitive_matches[0]

    if len(case_insensitive_matches) > 1:
        raise ValueError(
            "VM input filename is ambiguous under case-insensitive "
            f"matching. Expected: {expected_file.name}. "
            "Matches: "
            f"{[path.name for path in case_insensitive_matches]}"
        )

    if required:
        raise FileNotFoundError(
            "VM input was not found. "
            f"Expected exact file: {expected_file}. "
            "Filename casing may differ, but no fallback to another "
            "file or period is allowed."
        )

    return None



def resolve_vm_sheet_name(
    file_path: str | Path,
    expected_sheet: str = VM_INPUT_SHEET,
) -> str:
    """
    Resolve one VM worksheet using approved case-insensitive aliases.

    The default SAP export worksheet may be named Sheet1 or Hoja1 depending
    on the Excel/SAP language. No arbitrary first-sheet fallback is allowed.
    """
    file_path = Path(file_path)

    try:
        excel_file = pd.ExcelFile(
            file_path
        )
    except PermissionError as error:
        raise PermissionError(
            f"VM input is locked: {file_path}. "
            "Close Excel and retry."
        ) from error
    except Exception as error:
        raise ValueError(
            f"VM workbook is not readable: "
            f"{file_path}: {error}"
        ) from error

    try:
        if (
            normalize_header_lookup(expected_sheet)
            == normalize_header_lookup(VM_INPUT_SHEET)
        ):
            approved_names = {
                normalize_header_lookup(alias)
                for alias in VM_INPUT_SHEET_ALIASES
            }
        else:
            # A caller requesting a non-default worksheet still receives exact
            # case-insensitive matching, without fallback to Sheet1/Hoja1.
            approved_names = {
                normalize_header_lookup(
                    expected_sheet
                )
            }

        matches = [
            sheet_name
            for sheet_name in excel_file.sheet_names
            if normalize_header_lookup(
                sheet_name
            ) in approved_names
        ]

        available_sheets = list(
            excel_file.sheet_names
        )

    finally:
        excel_file.close()

    if not matches:
        raise ValueError(
            f"VM worksheet '{expected_sheet}' was not found in "
            f"{file_path}. Approved names: "
            f"{sorted(VM_INPUT_SHEET_ALIASES)}. "
            f"Available sheets: {available_sheets}"
        )

    if len(matches) > 1:
        raise ValueError(
            "VM input contains more than one approved input "
            f"worksheet in {file_path}: {matches}. "
            "Keep only one of Sheet1 or Hoja1."
        )

    return matches[0]


def read_vm_excel(
    file_path: str | Path,
    *,
    sheet_name: str = VM_INPUT_SHEET,
    usecols: Sequence[str] | None = None,
) -> pd.DataFrame:
    """
    Read one VM Excel export safely as object values.

    When usecols is supplied, only exact physical column names are selected.
    Exact matching allows loaders to distinguish source headers that differ
    only by letter casing.
    """
    resolved_sheet = resolve_vm_sheet_name(
        file_path,
        sheet_name,
    )

    selected_columns = None

    if usecols is not None:
        selected_columns = {
            safe_text(column)
            for column in usecols
            if not is_blank(column)
        }

        if not selected_columns:
            raise ValueError(
                "VM Excel usecols cannot be an empty collection."
            )

    try:
        dataframe = pd.read_excel(
            file_path,
            sheet_name=resolved_sheet,
            header=VM_HEADER_ROW - 1,
            dtype=object,
            usecols=(
                (
                    lambda column: safe_text(column)
                    in selected_columns
                )
                if selected_columns is not None
                else None
            ),
        )
    except PermissionError as error:
        raise PermissionError(
            f"VM input is locked: {file_path}. "
            "Close Excel and retry."
        ) from error
    except Exception as error:
        raise ValueError(
            f"VM worksheet '{resolved_sheet}' could not be read "
            f"from {file_path}: {error}"
        ) from error

    dataframe = dataframe.dropna(
        axis=0,
        how="all",
    ).dropna(
        axis=1,
        how="all",
    )

    dataframe.columns = [
        safe_text(column)
        for column in dataframe.columns
    ]

    if selected_columns is not None:
        missing_columns = sorted(
            selected_columns.difference(
                dataframe.columns
            )
        )

        if missing_columns:
            raise ValueError(
                f"VM worksheet '{resolved_sheet}' in {file_path} "
                "is missing requested physical columns: "
                f"{missing_columns}. Available selected headers: "
                f"{list(dataframe.columns)}"
            )

    if dataframe.empty:
        raise ValueError(
            f"VM worksheet '{resolved_sheet}' in {file_path} "
            "contains no data rows."
        )

    return dataframe


def resolve_columns(
    dataframe: pd.DataFrame,
    aliases: dict[str, Sequence[str]],
    required_columns: Sequence[str],
    *,
    source_name: str,
) -> dict[str, Any]:
    """Resolve explicit aliases and reject ambiguous or missing headers."""
    normalized_headers: dict[str, list[Any]] = {}

    for column in dataframe.columns:
        key = normalize_header_lookup(column)

        if key != "":
            normalized_headers.setdefault(
                key,
                [],
            ).append(column)

    duplicate_headers = {
        key: columns
        for key, columns in normalized_headers.items()
        if len(columns) > 1
    }

    if duplicate_headers:
        raise ValueError(
            f"{source_name} contains multiple headers with the same "
            f"normalized name: {duplicate_headers}."
        )

    mapping: dict[str, Any] = {}
    missing: list[str] = []
    ambiguous: list[str] = []
    physical_usage: dict[Any, list[str]] = {}

    for canonical_name, possible_names in aliases.items():
        matches: list[Any] = []

        for possible_name in possible_names:
            matches.extend(
                normalized_headers.get(
                    normalize_header_lookup(possible_name),
                    [],
                )
            )

        matches = list(dict.fromkeys(matches))

        if len(matches) == 1:
            physical_column = matches[0]
            mapping[canonical_name] = physical_column
            physical_usage.setdefault(
                physical_column,
                [],
            ).append(canonical_name)
        elif len(matches) > 1:
            ambiguous.append(
                f"{canonical_name}: {matches}"
            )
        elif canonical_name in required_columns:
            missing.append(canonical_name)

    reused_columns = {
        column: canonical_names
        for column, canonical_names in physical_usage.items()
        if len(canonical_names) > 1
    }

    if missing or ambiguous or reused_columns:
        messages = []

        if missing:
            messages.append(
                f"missing required headers: {sorted(missing)}"
            )

        if ambiguous:
            messages.append(
                f"ambiguous headers: {sorted(ambiguous)}"
            )

        if reused_columns:
            messages.append(
                "headers resolving to multiple canonical names: "
                f"{reused_columns}"
            )

        raise ValueError(
            f"{source_name} header validation failed "
            f"({'; '.join(messages)}). "
            f"Available headers: {list(dataframe.columns)}"
        )

    return mapping


def _select_and_complete(
    dataframe: pd.DataFrame,
    mapping: dict[str, Any],
    expected_columns: Sequence[str],
) -> pd.DataFrame:
    """Select resolved fields and add absent optional columns."""
    selected = dataframe[
        list(mapping.values())
    ].copy()

    selected = selected.rename(
        columns={
            physical_name: canonical_name
            for canonical_name, physical_name in mapping.items()
        }
    )

    for column in expected_columns:
        if column not in selected.columns:
            selected[column] = ""

    selected = selected.loc[
        :,
        list(expected_columns),
    ]

    for column in selected.columns:
        selected[column] = selected[column].map(safe_text)

    return selected


def _require_complete_keys(
    dataframe: pd.DataFrame,
    key_columns: Sequence[str],
    *,
    source_name: str,
) -> None:
    """Reject rows with incomplete mandatory logical keys."""
    invalid = dataframe.loc[
        dataframe[list(key_columns)].apply(
            lambda row: any(
                is_blank(value)
                for value in row
            ),
            axis=1,
        )
    ]

    if invalid.empty:
        return

    rows = [
        int(index) + VM_HEADER_ROW + 1
        if isinstance(index, int)
        else str(index)
        for index in invalid.index[:20]
    ]

    raise ValueError(
        f"{source_name} contains blank mandatory keys "
        f"at Excel rows: {rows}."
    )


def _nonblank_values(
    series: pd.Series,
) -> list[str]:
    """Return sorted distinct nonblank values."""
    return sorted({
        safe_text(value)
        for value in series
        if not is_blank(value)
    })


def _validate_consistency(
    dataframe: pd.DataFrame,
    *,
    key_columns: Sequence[str],
    value_columns: Sequence[str],
    source_name: str,
) -> None:
    """Reject contradictory values for the same logical key."""
    conflicts = []

    for group_key, group in dataframe.groupby(
        list(key_columns),
        sort=True,
        dropna=False,
    ):
        differences = {}

        for column in value_columns:
            values = _nonblank_values(group[column])

            if len(values) > 1:
                differences[column] = values

        if differences:
            conflicts.append(
                {
                    "key": group_key,
                    "values": differences,
                }
            )

    if conflicts:
        raise ValueError(
            f"{source_name} contains contradictory values for the "
            f"same logical key: {conflicts[:20]}"
        )


def _first_nonblank(
    series: pd.Series,
) -> str:
    """Return the first nonblank value from a validated group."""
    for value in series:
        text = safe_text(value)

        if text != "":
            return text

    return ""


def _collapse_groups(
    dataframe: pd.DataFrame,
    *,
    key_columns: Sequence[str],
    output_columns: Sequence[str],
) -> pd.DataFrame:
    """Collapse consistent groups deterministically."""
    rows = []

    for _, group in dataframe.groupby(
        list(key_columns),
        sort=True,
        dropna=False,
    ):
        rows.append({
            column: _first_nonblank(group[column])
            for column in output_columns
        })

    return pd.DataFrame(
        rows,
        columns=list(output_columns),
    )


def _posting_aliases(
    source: str,
) -> dict[str, tuple[str, ...]]:
    """
    Return the exact physical aliases used by LBR VPBSIK and VPBSAK.

    ``source`` is retained in the signature because callers identify each
    population as BSIK or BSAK, but both extracts currently use the same
    physical headers.
    """
    if source not in {
        "BSIK",
        "BSAK",
    }:
        raise ValueError(
            f"Unsupported VM posting source: {source!r}."
        )

    return {
        "Company": (
            "CoCd",
        ),
        "Vendor Code": (
            "Vendor",
        ),
        "Fiscal Year": (
            "Year",
        ),
        "Accounting Document": (
            "DocumentNo",
        ),
        "Accounting Document Line": (
            "Itm",
        ),
        "Posting Date": (
            "Pstng Date",
        ),
        "Document Type": (
            "Doc. Type",
        ),
        "Debit/Credit Indicator": (
            "D/C",
        ),
        "Document Currency": (
            "Crcy",
        ),
        "Amount in Document Currency": (
            "Amount",
        ),
        "Amount in Local Currency": (
            "Amount in LC",
        ),
        "Clearing Date": (
            "Clearing",
        ),
        "Clearing Document": (
            "Clrng doc.",
        ),
    }


def _validate_vendor_duplicate_company_columns(
    dataframe: pd.DataFrame,
) -> None:
    """
    Validate duplicate company columns produced by the SAP join/export.

    CoCd and CoCd.1 represent the company code from joined tables and must
    agree whenever both values are populated.
    """
    if (
        "CoCd" not in dataframe.columns
        or "CoCd.1" not in dataframe.columns
    ):
        return

    primary = dataframe[
        "CoCd"
    ].map(normalize_company)

    secondary = dataframe[
        "CoCd.1"
    ].map(normalize_company)

    conflict = (
        primary.ne("")
        & secondary.ne("")
        & primary.ne(secondary)
    )

    if not conflict.any():
        return

    examples = (
        dataframe.loc[
            conflict,
            [
                "CoCd",
                "CoCd.1",
            ],
        ]
        .head(20)
        .to_dict("records")
    )

    raise ValueError(
        "VM VENDORS contains conflicting company codes between "
        f"CoCd and CoCd.1. Examples: {examples}"
    )


def load_vm_vendors(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> pd.DataFrame:
    """Load required vendor extract without business exclusions."""
    input_file = find_vm_input_file(
        context,
        VM_VENDOR_FILE_TEMPLATE,
        required=True,
    )

    raw = read_vm_excel(
        input_file,
        sheet_name=sheet_name,
    )

    _validate_vendor_duplicate_company_columns(
        raw
    )

    mapping = resolve_columns(
        raw,
        VENDOR_COLUMN_ALIASES,
        VENDOR_REQUIRED_COLUMNS,
        source_name="VM VENDORS",
    )

    result = _select_and_complete(
        raw,
        mapping,
        VENDOR_COLUMNS,
    )

    result["Company"] = result["Company"].map(
        normalize_company
    )
    result["Vendor Code"] = result["Vendor Code"].map(
        normalize_vendor_code
    )

    identifier_columns = (
        "Phone1",
        "Tax Number 1",
        "Tax Number 2",
        "Tax Number 3",
        "Tax Number 4",
        "Tax Number 5",
        "VAT Registration Number",
        "Account Group",
        "Trading Partner",
        "Bank Country",
        "Bank Code",
        "Bank Account",
    )

    for column in identifier_columns:
        result[column] = result[column].map(
            normalize_identifier
        )

    date_columns = (
        "Vendor Created Date",
        "Vendor Updated Date",
        "Company Extension Date",
        "Bank Valid From",
        "Bank Valid To",
    )

    for column in date_columns:
        result[column] = result[column].map(
            normalize_date_text
        )

    _require_complete_keys(
        result,
        VENDOR_KEY_COLUMNS,
        source_name="VM VENDORS",
    )

    result["Vendor Key"] = [
        build_vendor_key(company, vendor_code)
        for company, vendor_code in zip(
            result["Company"],
            result["Vendor Code"],
        )
    ]

    return result.drop_duplicates(
        ignore_index=True
    )


def build_vendor_master_population(
    vendors: pd.DataFrame,
) -> pd.DataFrame:
    """Build one consistent row per Company + Vendor Code."""
    missing = sorted(
        set(VENDOR_COLUMNS).difference(vendors.columns)
    )

    if missing:
        raise ValueError(
            f"VM vendor master requires missing columns: {missing}."
        )

    master_columns = [
        column
        for column in VENDOR_COLUMNS
        if column not in BANK_DETAIL_COLUMNS
    ]

    value_columns = [
        column
        for column in master_columns
        if column not in VENDOR_KEY_COLUMNS
    ]

    _validate_consistency(
        vendors,
        key_columns=VENDOR_KEY_COLUMNS,
        value_columns=value_columns,
        source_name="VM vendor master",
    )

    result = _collapse_groups(
        vendors,
        key_columns=VENDOR_KEY_COLUMNS,
        output_columns=master_columns,
    )

    result["Vendor Key"] = [
        build_vendor_key(company, vendor_code)
        for company, vendor_code in zip(
            result["Company"],
            result["Vendor Code"],
        )
    ]

    return result.sort_values(
        list(VENDOR_KEY_COLUMNS),
        kind="mergesort",
    ).reset_index(drop=True)


def build_vendor_bank_population(
    vendors: pd.DataFrame,
) -> pd.DataFrame:
    """Build one row per distinct nonblank vendor bank combination."""
    required = {
        *VENDOR_KEY_COLUMNS,
        *BANK_DETAIL_COLUMNS,
    }
    missing = sorted(
        required.difference(vendors.columns)
    )

    if missing:
        raise ValueError(
            f"VM vendor banks require missing columns: {missing}."
        )

    has_bank = vendors[
        list(BANK_KEY_COLUMNS)
    ].apply(
        lambda row: any(
            not is_blank(value)
            for value in row
        ),
        axis=1,
    )

    working = vendors.loc[has_bank].copy()

    output_columns = [
        *VENDOR_KEY_COLUMNS,
        *BANK_DETAIL_COLUMNS,
    ]

    if working.empty:
        return pd.DataFrame(
            columns=[
                *output_columns,
                "Vendor Key",
            ]
        )

    _validate_consistency(
        working,
        key_columns=VENDOR_BANK_KEY_COLUMNS,
        value_columns=(
            "Account Holder Name",
            "Bank Valid From",
            "Bank Valid To",
        ),
        source_name="VM vendor banks",
    )

    result = _collapse_groups(
        working,
        key_columns=VENDOR_BANK_KEY_COLUMNS,
        output_columns=output_columns,
    )

    result["Vendor Key"] = [
        build_vendor_key(company, vendor_code)
        for company, vendor_code in zip(
            result["Company"],
            result["Vendor Code"],
        )
    ]

    return result.sort_values(
        list(VENDOR_BANK_KEY_COLUMNS),
        kind="mergesort",
    ).reset_index(drop=True)


def _load_posting_source(
    file_path: Path,
    source: str,
    *,
    sheet_name: str,
) -> pd.DataFrame:
    """Load and validate one BSIK or BSAK source."""
    raw = read_vm_excel(
        file_path,
        sheet_name=sheet_name,
    )

    mapping = resolve_columns(
        raw,
        _posting_aliases(source),
        POSTING_REQUIRED_COLUMNS,
        source_name=f"VM {source}",
    )

    result = _select_and_complete(
        raw,
        mapping,
        POSTING_COLUMNS,
    )

    result["Company"] = result["Company"].map(
        normalize_company
    )
    result["Vendor Code"] = result["Vendor Code"].map(
        normalize_vendor_code
    )
    result["Fiscal Year"] = result["Fiscal Year"].map(
        normalize_identifier
    )
    result["Accounting Document"] = result[
        "Accounting Document"
    ].map(normalize_document_number)
    result["Accounting Document Line"] = result[
        "Accounting Document Line"
    ].map(normalize_line_number)
    result["Clearing Document"] = result[
        "Clearing Document"
    ].map(normalize_document_number)

    for column in (
        "Posting Date",
        "Clearing Date",
    ):
        result[column] = result[column].map(
            normalize_date_text
        )

    result["Posting Source"] = source

    # SAP exports may include footer/total rows after the posting population.
    # Exclude only rows where every mandatory posting-key component is blank.
    # Partially populated keys remain audit errors.
    posting_key_values = result.loc[
        :,
        list(POSTING_KEY_COLUMNS),
    ].astype(
        "string"
    ).fillna(
        ""
    )

    posting_key_blank = pd.DataFrame(
        {
            column: posting_key_values[
                column
            ].str.strip().eq("")
            for column in POSTING_KEY_COLUMNS
        },
        index=result.index,
    )

    fully_blank_key = posting_key_blank.all(
        axis=1
    )

    residual_rows = int(
        fully_blank_key.sum()
    )

    if residual_rows:
        print(
            f"VM {source}: excluded {residual_rows} "
            "footer/total row(s) with fully blank posting keys."
        )

        result = result.loc[
            ~fully_blank_key
        ].copy()

    if result.empty:
        raise ValueError(
            f"VM {source} contains no posting rows after "
            "excluding fully blank footer/total rows."
        )

    _require_complete_keys(
        result,
        POSTING_KEY_COLUMNS,
        source_name=f"VM {source}",
    )

    result.attrs[
        "excluded_fully_blank_key_rows"
    ] = residual_rows
    

    value_columns = [
        column
        for column in result.columns
        if column not in (
            *POSTING_KEY_COLUMNS,
            "Posting Source",
        )
    ]

    _validate_consistency(
        result,
        key_columns=POSTING_KEY_COLUMNS,
        value_columns=value_columns,
        source_name=f"VM {source}",
    )

    return result.drop_duplicates(
        ignore_index=True
    )


def load_vm_vendor_postings(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    """Load optional paired BSIK/BSAK posting extracts."""
    bsik_file = find_vm_input_file(
        context,
        VM_BSIK_FILE_TEMPLATE,
        required=False,
    )
    bsak_file = find_vm_input_file(
        context,
        VM_BSAK_FILE_TEMPLATE,
        required=False,
    )

    metadata = {
        "available": False,
        "warnings": [],
    }

    if bsik_file is None and bsak_file is None:
        metadata["warnings"].append(
            "VM postings are unavailable because both "
            "BSIK and BSAK files are absent."
        )
        return None, metadata

    if bsik_file is None or bsak_file is None:
        missing_name = (
            VM_BSIK_FILE_TEMPLATE
            if bsik_file is None
            else VM_BSAK_FILE_TEMPLATE
        )

        raise FileNotFoundError(
            "VM postings population is incomplete. "
            "BSIK and BSAK must both be present for the "
            f"configured period. Missing template: {missing_name}."
        )

    bsik = _load_posting_source(
        bsik_file,
        "BSIK",
        sheet_name=sheet_name,
    )
    bsak = _load_posting_source(
        bsak_file,
        "BSAK",
        sheet_name=sheet_name,
    )

    cross_source_duplicates = bsik[
        list(POSTING_KEY_COLUMNS)
    ].merge(
        bsak[list(POSTING_KEY_COLUMNS)],
        how="inner",
        on=list(POSTING_KEY_COLUMNS),
    )

    if not cross_source_duplicates.empty:
        raise ValueError(
            "VM postings contain keys present in both BSIK "
            "and BSAK: "
            f"{cross_source_duplicates.head(20).to_dict('records')}"
        )

    postings = pd.concat(
        [
            bsik,
            bsak,
        ],
        ignore_index=True,
    )

    postings = postings.sort_values(
        [
            *POSTING_KEY_COLUMNS,
            "Posting Source",
        ],
        kind="mergesort",
    ).reset_index(drop=True)

    metadata.update(
        {
            "available": True,
            "bsik_rows": len(bsik),
            "bsak_rows": len(bsak),
            "posting_rows": len(postings),
            "warnings": [],
        }
    )

    return postings, metadata


def load_vm_posting_headers(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """
    Load and validate SAP ECC accounting-document headers.

    The source is reduced during the Excel read to the exact physical columns
    required for posting-user attribution. Posting User is preserved as a
    trimmed technical identifier without destructive normalization.
    """
    input_file = find_vm_input_file(
        context,
        VM_BKPF_FILE_TEMPLATE,
        required=True,
    )

    raw = read_vm_excel(
        input_file,
        sheet_name=sheet_name,
        usecols=POSTING_HEADER_PHYSICAL_COLUMNS,
    )

    source_rows = len(raw)

    mapping = resolve_columns(
        raw,
        POSTING_HEADER_ALIASES,
        POSTING_HEADER_REQUIRED_COLUMNS,
        source_name="VM BKPF",
    )

    headers = _select_and_complete(
        raw,
        mapping,
        POSTING_HEADER_COLUMNS,
    )

    headers["Company"] = headers[
        "Company"
    ].map(
        normalize_company
    )

    headers["Fiscal Year"] = headers[
        "Fiscal Year"
    ].map(
        normalize_identifier
    )

    headers["Accounting Document"] = headers[
        "Accounting Document"
    ].map(
        normalize_document_number
    )

    headers["Posting Date"] = headers[
        "Posting Date"
    ].map(
        normalize_date_text
    )

    headers["Posting User"] = headers[
        "Posting User"
    ].map(
        safe_text
    )

    _require_complete_keys(
        headers,
        POSTING_HEADER_KEY_COLUMNS,
        source_name="VM BKPF",
    )

    posting_date_text = (
        headers["Posting Date"]
        .astype("string")
        .fillna("")
        .str.strip()
    )

    posting_dates = pd.to_datetime(
        posting_date_text,
        format="%Y-%m-%d",
        errors="coerce",
    )

    blank_posting_date = posting_date_text.eq("")

    if blank_posting_date.any():
        examples = (
            headers.loc[
                blank_posting_date,
                list(POSTING_HEADER_KEY_COLUMNS),
            ]
            .head(20)
            .to_dict("records")
        )

        raise ValueError(
            "VM BKPF contains blank Posting Date values. "
            f"Examples: {examples}"
        )

    invalid_posting_date = (
        posting_date_text.ne("")
        & posting_dates.isna()
    )

    if invalid_posting_date.any():
        examples = (
            posting_date_text.loc[
                invalid_posting_date
            ]
            .drop_duplicates()
            .head(20)
            .tolist()
        )

        raise ValueError(
            "VM BKPF contains invalid Posting Date values. "
            f"Examples: {examples}"
        )

    date_from, date_to = get_vm_period(
        context
    )

    outside_period = (
        posting_dates.lt(date_from)
        | posting_dates.gt(date_to)
    )

    if outside_period.any():
        examples = (
            headers.loc[
                outside_period,
                [
                    *POSTING_HEADER_KEY_COLUMNS,
                    "Posting Date",
                ],
            ]
            .head(20)
            .to_dict("records")
        )

        raise ValueError(
            "VM BKPF contains Posting Date values outside "
            f"CONFIG FROM/TO ({date_from.date()} to "
            f"{date_to.date()}). Examples: {examples}"
        )

    key_columns = list(
        POSTING_HEADER_KEY_COLUMNS
    )

    posting_user_nonblank = headers.loc[
        headers["Posting User"].ne("")
    ]

    posting_user_counts = (
        posting_user_nonblank.groupby(
            key_columns,
            sort=False,
            observed=True,
            dropna=False,
        )["Posting User"]
        .nunique()
    )

    conflicting_user_keys = posting_user_counts.loc[
        posting_user_counts.gt(1)
    ]

    if not conflicting_user_keys.empty:
        conflict_examples = (
            conflicting_user_keys
            .head(20)
            .reset_index()
            .loc[
                :,
                key_columns,
            ]
            .to_dict("records")
        )

        raise ValueError(
            "VM BKPF contains conflicting Posting User values "
            "for the same Company/Fiscal Year/Accounting "
            f"Document key. Examples: {conflict_examples}"
        )

    posting_date_counts = (
        headers.groupby(
            key_columns,
            sort=False,
            observed=True,
            dropna=False,
        )["Posting Date"]
        .nunique()
    )

    conflicting_date_keys = posting_date_counts.loc[
        posting_date_counts.gt(1)
    ]

    if not conflicting_date_keys.empty:
        conflict_examples = (
            conflicting_date_keys
            .head(20)
            .reset_index()
            .loc[
                :,
                key_columns,
            ]
            .to_dict("records")
        )

        raise ValueError(
            "VM BKPF contains conflicting Posting Date values "
            "for the same Company/Fiscal Year/Accounting "
            f"Document key. Examples: {conflict_examples}"
        )

    headers["_Posting User Blank"] = headers[
        "Posting User"
    ].eq("")

    headers = headers.sort_values(
        [
            *key_columns,
            "_Posting User Blank",
            "Posting User",
        ],
        kind="mergesort",
    )

    duplicate_rows = int(
        headers.duplicated(
            subset=key_columns,
            keep="first",
        ).sum()
    )

    headers = (
        headers.drop_duplicates(
            subset=key_columns,
            keep="first",
        )
        .drop(
            columns="_Posting User Blank"
        )
        .loc[
            :,
            list(POSTING_HEADER_COLUMNS),
        ]
        .reset_index(drop=True)
    )

    blank_posting_users = int(
        headers["Posting User"].eq("").sum()
    )

    metadata = {
        "available": True,
        "source": "BKPF-USNAM",
        "source_rows": source_rows,
        "header_rows": len(headers),
        "duplicate_rows_collapsed": duplicate_rows,
        "posting_user_nonblank_rows": int(
            headers["Posting User"].ne("").sum()
        ),
        "posting_user_blank_rows": blank_posting_users,
        "posting_user_conflicts": 0,
        "posting_date_from": (
            posting_dates.min().strftime("%Y-%m-%d")
        ),
        "posting_date_to": (
            posting_dates.max().strftime("%Y-%m-%d")
        ),
        "warnings": [],
    }

    if blank_posting_users:
        metadata["warnings"].append(
            "VM BKPF contains accounting-document headers "
            "with blank Posting User."
        )

    return headers, metadata


def build_vm_last_invoice_population(
    postings: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build one latest AP-invoice row per Company + Vendor Code.

    LBR equivalents of OPCH invoices:
    - RE: invoice receipt / logistics invoice;
    - KR: FI vendor invoice.

    Selection order:
    1. Posting Date descending;
    2. Fiscal Year descending;
    3. Accounting Document descending;
    4. Accounting Document Line descending.

    No FX conversion is applied. The amount remains in document currency.
    """
    output_columns = [
        "Company",
        "Vendor Code",
        *VM_LAST_INVOICE_COLUMNS,
    ]

    if postings is None or postings.empty:
        return pd.DataFrame(
            columns=output_columns
        )

    required_columns = {
        "Company",
        "Vendor Code",
        "Fiscal Year",
        "Accounting Document",
        "Accounting Document Line",
        "Posting Date",
        "Document Type",
        "Document Currency",
        "Amount in Document Currency",
    }

    missing_columns = sorted(
        required_columns.difference(
            postings.columns
        )
    )

    if missing_columns:
        raise ValueError(
            "VM last-invoice population requires missing "
            f"posting columns: {missing_columns}."
        )

    working = postings.loc[
        :,
        list(required_columns),
    ].copy()

    working["Company"] = working[
        "Company"
    ].map(
        normalize_company
    )

    working["Vendor Code"] = working[
        "Vendor Code"
    ].map(
        normalize_vendor_code
    )

    working["Document Type"] = working[
        "Document Type"
    ].map(
        normalize_upper_text
    )

    working = working.loc[
        working["Document Type"].isin(
            VM_INVOICE_DOCUMENT_TYPES
        )
    ].copy()

    if working.empty:
        return pd.DataFrame(
            columns=output_columns
        )

    working["_Posting Date"] = pd.to_datetime(
        working["Posting Date"],
        format="%Y-%m-%d",
        errors="coerce",
    )

    invalid_date = (
        working["Posting Date"]
        .astype("string")
        .fillna("")
        .str.strip()
        .ne("")
        & working["_Posting Date"].isna()
    )

    if invalid_date.any():
        examples = (
            working.loc[
                invalid_date,
                "Posting Date",
            ]
            .astype("string")
            .drop_duplicates()
            .head(20)
            .tolist()
        )

        raise ValueError(
            "VM invoice postings contain invalid Posting Date "
            f"values. Examples: {examples}"
        )

    missing_date = working[
        "_Posting Date"
    ].isna()

    if missing_date.any():
        examples = (
            working.loc[
                missing_date,
                [
                    "Company",
                    "Vendor Code",
                    "Accounting Document",
                ],
            ]
            .head(20)
            .to_dict("records")
        )

        raise ValueError(
            "VM invoice postings contain blank Posting Date "
            f"values. Examples: {examples}"
        )

    amount_text = (
        working[
            "Amount in Document Currency"
        ]
        .astype("string")
        .fillna("")
        .str.strip()
    )

    working["_Invoice Amount"] = pd.to_numeric(
        amount_text,
        errors="coerce",
    )

    invalid_amount = (
        amount_text.ne("")
        & working["_Invoice Amount"].isna()
    )

    if invalid_amount.any():
        examples = (
            amount_text.loc[
                invalid_amount
            ]
            .drop_duplicates()
            .head(20)
            .tolist()
        )

        raise ValueError(
            "VM invoice postings contain invalid document-currency "
            f"amounts. Examples: {examples}"
        )

    # OPCH.DocTotal is presented as a positive invoice total.
    working["_Invoice Amount"] = working[
        "_Invoice Amount"
    ].abs()

    working["_Fiscal Year Sort"] = pd.to_numeric(
        working["Fiscal Year"],
        errors="coerce",
    ).fillna(-1)

    working["_Document Sort"] = working[
        "Accounting Document"
    ].map(
        normalize_document_number
    )

    working["_Line Sort"] = working[
        "Accounting Document Line"
    ].map(
        normalize_line_number
    )

    working = working.sort_values(
        [
            "Company",
            "Vendor Code",
            "_Posting Date",
            "_Fiscal Year Sort",
            "_Document Sort",
            "_Line Sort",
        ],
        ascending=[
            True,
            True,
            False,
            False,
            False,
            False,
        ],
        kind="mergesort",
    )

    latest = working.drop_duplicates(
        subset=[
            "Company",
            "Vendor Code",
        ],
        keep="first",
    ).copy()

    latest["Last Invoice Number"] = latest[
        "Accounting Document"
    ].map(
        normalize_document_number
    )

    latest["Last Transaction Date"] = latest[
        "_Posting Date"
    ].dt.strftime(
        "%Y-%m-%d"
    )

    latest[
        "Last Inv Amt Doc Currency"
    ] = latest[
        "_Invoice Amount"
    ]

    latest[
        "Last Inv Amt Doc Currency Indicator"
    ] = latest[
        "Document Currency"
    ].map(
        normalize_upper_text
    )

    result = latest.loc[
        :,
        output_columns,
    ].copy()

    duplicate_vendor = result.duplicated(
        subset=[
            "Company",
            "Vendor Code",
        ],
        keep=False,
    )

    if duplicate_vendor.any():
        examples = (
            result.loc[
                duplicate_vendor,
                [
                    "Company",
                    "Vendor Code",
                ],
            ]
            .head(20)
            .to_dict("records")
        )

        raise ValueError(
            "VM latest-invoice population is not unique by "
            f"Company/Vendor Code. Examples: {examples}"
        )

    return result.reset_index(
        drop=True
    )


def _employee_date_to_text(
    value: Any,
) -> str:
    """
    Normalize one employee validity date to YYYY-MM-DD text.

    Native Excel/Python dates are formatted directly, including SAP's
    9999-12-31 open-ended validity sentinel.
    """
    if is_blank(value):
        return ""

    strftime_method = getattr(
        value,
        "strftime",
        None,
    )

    if callable(strftime_method):
        try:
            return strftime_method(
                "%Y-%m-%d"
            )
        except (
            TypeError,
            ValueError,
            OverflowError,
        ):
            pass

    return normalize_date_text(
        value
    )


def _parse_employee_validity_column(
    series: pd.Series,
    *,
    column_name: str,
) -> tuple[pd.Series, pd.Series]:
    """
    Return comparison dates and display text for one EMP validity column.

    SAP 9999-12-31 means open-ended validity. For period comparisons it is
    represented internally by pandas' maximum normalized Timestamp, while the
    display series retains 9999-12-31.

    Every other nonblank unparseable value remains an audit error.
    """
    display_text = (
        series.map(
            _employee_date_to_text
        )
        .astype("string")
        .fillna("")
        .str.strip()
    )

    open_ended = display_text.eq(
        "9999-12-31"
    )

    parseable_text = display_text.mask(
        open_ended,
        "",
    )

    comparison_date = pd.to_datetime(
        parseable_text,
        format="%Y-%m-%d",
        errors="coerce",
    )

    comparison_date = pd.Series(
        comparison_date,
        index=series.index,
        dtype="datetime64[ns]",
    )

    comparison_date.loc[
        open_ended
    ] = pd.Timestamp.max.normalize()

    invalid = (
        display_text.ne("")
        & comparison_date.isna()
    )

    if invalid.any():
        examples = (
            display_text.loc[
                invalid
            ]
            .drop_duplicates()
            .head(20)
            .tolist()
        )

        rows = [
            int(index) + VM_HEADER_ROW + 1
            if isinstance(index, int)
            else str(index)
            for index in series.index[
                invalid
            ][:20]
        ]

        raise ValueError(
            f"VM EMP column '{column_name}' contains invalid "
            f"validity dates at Excel rows {rows}. "
            f"Examples: {examples}"
        )

    return comparison_date, display_text



def load_vm_employees(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    """
    Load optional employees and apply period-overlap validation.

    SAP's 9999-12-31 validity end date is treated as open-ended for date
    comparisons while remaining 9999-12-31 in the normalized output.
    """
    input_file = find_vm_input_file(
        context,
        VM_EMPLOYEE_FILE_TEMPLATE,
        required=False,
    )

    metadata = {
        "available": False,
        "warnings": [],
    }

    if input_file is None:
        metadata["warnings"].append(
            "VM employees are unavailable. Employee comparison "
            "controls cannot run."
        )

        return None, metadata

    raw = read_vm_excel(
        input_file,
        sheet_name=sheet_name,
    )

    mapping = resolve_columns(
        raw,
        EMPLOYEE_ALIASES,
        EMPLOYEE_REQUIRED_COLUMNS,
        source_name="VM EMP",
    )

    employees = _select_and_complete(
        raw,
        mapping,
        EMPLOYEE_COLUMNS,
    )

    employees["Company"] = employees[
        "Company"
    ].map(
        normalize_company
    )

    employees["Employee Code"] = employees[
        "Employee Code"
    ].map(
        normalize_employee_code
    )

    employees["Employment Status"] = employees[
        "Employment Status"
    ].map(
        normalize_identifier
    )

    _require_complete_keys(
        employees,
        EMPLOYEE_KEY_COLUMNS,
        source_name="VM EMP",
    )

    employees["Employee Key"] = [
        build_employee_key(
            company,
            employee_code,
        )
        for company, employee_code in zip(
            employees["Company"],
            employees["Employee Code"],
        )
    ]

    date_from, date_to = get_vm_period(
        context
    )

    date_columns = (
        "Organizational Valid From",
        "Organizational Valid To",
        "Status Valid From",
        "Status Valid To",
    )

    parsed_dates = {}
    display_dates = {}

    for column in date_columns:
        source_text = (
            employees[column]
            .astype("string")
            .fillna("")
            .str.strip()
        )

        # Excel returned native datetime values. _select_and_complete converted
        # those values to strings such as "2026-07-31 00:00:00". Extract the
        # unambiguous ISO date portion without pandas format inference.
        iso_date = (
            source_text.str.extract(
                r"^(\d{4}-\d{2}-\d{2})(?:\s+.*)?$",
                expand=False,
            )
            .astype("string")
            .fillna("")
        )

        # Also accept explicit day-first text if a future export returns dates
        # as DD/MM/YYYY or DD.MM.YYYY instead of native Excel dates.
        day_first_parts = source_text.str.extract(
            r"^(\d{2})[/.](\d{2})[/.](\d{4})(?:\s+.*)?$",
            expand=True,
        )

        day_first_available = (
            iso_date.eq("")
            & day_first_parts.notna().all(axis=1)
        )

        if day_first_available.any():
            iso_date.loc[
                day_first_available
            ] = (
                day_first_parts.loc[
                    day_first_available,
                    2,
                ]
                + "-"
                + day_first_parts.loc[
                    day_first_available,
                    1,
                ]
                + "-"
                + day_first_parts.loc[
                    day_first_available,
                    0,
                ]
            )

        open_ended = iso_date.eq(
            "9999-12-31"
        )

        parseable_date = iso_date.mask(
            open_ended,
            "",
        )

        parsed = pd.to_datetime(
            parseable_date,
            format="%Y-%m-%d",
            errors="coerce",
        )

        parsed = pd.Series(
            parsed,
            index=employees.index,
            dtype="datetime64[ns]",
        )

        # pandas cannot represent year 9999. Use its maximum normalized date
        # only for the internal period-overlap comparison.
        parsed.loc[
            open_ended
        ] = pd.Timestamp.max.normalize()

        invalid = (
            source_text.ne("")
            & (
                iso_date.eq("")
                | parsed.isna()
            )
        )

        if invalid.any():
            rows = [
                int(index) + VM_HEADER_ROW + 1
                if isinstance(index, int)
                else str(index)
                for index in employees.index[
                    invalid
                ][:20]
            ]

            examples = (
                source_text.loc[
                    invalid
                ]
                .drop_duplicates()
                .head(20)
                .tolist()
            )

            raise ValueError(
                f"VM EMP column '{column}' contains invalid "
                f"validity dates at Excel rows {rows}. "
                f"Examples: {examples}"
            )

        parsed_dates[column] = parsed
        display_dates[column] = iso_date

    overlap = (
        parsed_dates[
            "Organizational Valid From"
        ].le(date_to)
        & parsed_dates[
            "Organizational Valid To"
        ].ge(date_from)
        & parsed_dates[
            "Status Valid From"
        ].le(date_to)
        & parsed_dates[
            "Status Valid To"
        ].ge(date_from)
    )

    active = employees[
        "Employment Status"
    ].eq("3")

    valid = employees.loc[
        overlap & active
    ].copy()

    # Preserve the normalized SAP display dates. Open-ended records remain
    # 9999-12-31 rather than being exposed as pandas' internal maximum date.
    for column in date_columns:
        valid[column] = (
            display_dates[column]
            .loc[valid.index]
            .astype("string")
            .fillna("")
        )

    valid = valid.drop_duplicates(
        ignore_index=True
    )

    metadata.update(
        {
            "available": True,
            "source_rows": len(employees),
            "valid_period_rows": len(valid),
            "distinct_employees": valid[
                "Employee Key"
            ].nunique(),
            "organizational_open_ended_rows": int(
                display_dates[
                    "Organizational Valid To"
                ].eq("9999-12-31").sum()
            ),
            "status_open_ended_rows": int(
                display_dates[
                    "Status Valid To"
                ].eq("9999-12-31").sum()
            ),
        }
    )

    if valid.empty:
        metadata["warnings"].append(
            "The VM employee file exists but contains no active "
            "rows overlapping CONFIG FROM/TO."
        )

    return (
        valid.sort_values(
            [
                "Company",
                "Employee Code",
                "Organizational Valid From",
            ],
            kind="mergesort",
        ).reset_index(
            drop=True
        ),
        metadata,
    )



def load_vm_bank_changes(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    """Load optional paired CDHDR/CDPOS bank-change extracts."""
    header_file = find_vm_input_file(
        context,
        VM_BANK_CDHDR_FILE_TEMPLATE,
        required=False,
    )
    position_file = find_vm_input_file(
        context,
        VM_BANK_CDPOS_FILE_TEMPLATE,
        required=False,
    )

    metadata = {
        "available": False,
        "warnings": [],
    }

    if header_file is None and position_file is None:
        metadata["warnings"].append(
            "VM bank-change CDHDR and CDPOS files are absent."
        )
        return None, metadata

    if header_file is None or position_file is None:
        missing_source = (
            "CDHDR"
            if header_file is None
            else "CDPOS"
        )

        raise FileNotFoundError(
            "VM bank-change population is incomplete. "
            f"The {missing_source} file is missing."
        )

    raw_headers = read_vm_excel(
        header_file,
        sheet_name=sheet_name,
    )
    raw_positions = read_vm_excel(
        position_file,
        sheet_name=sheet_name,
    )

    header_mapping = resolve_columns(
        raw_headers,
        CDHDR_ALIASES,
        CHANGE_KEY_COLUMNS,
        source_name="VM BANK CDHDR",
    )
    position_mapping = resolve_columns(
        raw_positions,
        CDPOS_ALIASES,
        (
            *CHANGE_KEY_COLUMNS,
            "Changed Table",
            "Changed Field",
        ),
        source_name="VM BANK CDPOS",
    )

    headers = _select_and_complete(
        raw_headers,
        header_mapping,
        CDHDR_COLUMNS,
    )
    positions = _select_and_complete(
        raw_positions,
        position_mapping,
        CDPOS_COLUMNS,
    )

    for dataframe in (
        headers,
        positions,
    ):
        dataframe["Change Object Class"] = dataframe[
            "Change Object Class"
        ].map(normalize_upper_text)
        dataframe["Object Value"] = dataframe[
            "Object Value"
        ].map(normalize_vendor_code)
        dataframe["Change Document"] = dataframe[
            "Change Document"
        ].map(normalize_identifier)

    headers["Change Date"] = headers[
        "Change Date"
    ].map(normalize_date_text)

    _require_complete_keys(
        headers,
        CHANGE_KEY_COLUMNS,
        source_name="VM BANK CDHDR",
    )
    _require_complete_keys(
        positions,
        CHANGE_KEY_COLUMNS,
        source_name="VM BANK CDPOS",
    )

    if headers["Change Object Class"].ne("KRED").any():
        raise ValueError(
            "VM BANK CDHDR contains OBJECTCLAS values "
            "other than KRED."
        )

    if positions["Change Object Class"].ne("KRED").any():
        raise ValueError(
            "VM BANK CDPOS contains OBJECTCLAS values "
            "other than KRED."
        )

    invalid_table = positions[
        "Changed Table"
    ].map(normalize_upper_text).ne("LFBK")

    if invalid_table.any():
        raise ValueError(
            "VM BANK CDPOS contains TABNAME values "
            "other than LFBK."
        )

    headers = headers.drop_duplicates(
        ignore_index=True
    )

    header_value_columns = [
        column
        for column in CDHDR_COLUMNS
        if column not in CHANGE_KEY_COLUMNS
    ]

    _validate_consistency(
        headers,
        key_columns=CHANGE_KEY_COLUMNS,
        value_columns=header_value_columns,
        source_name="VM BANK CDHDR",
    )

    headers = _collapse_groups(
        headers,
        key_columns=CHANGE_KEY_COLUMNS,
        output_columns=CDHDR_COLUMNS,
    )

    positions = positions.drop_duplicates(
        ignore_index=True
    )

    changes = positions.merge(
        headers,
        how="left",
        on=list(CHANGE_KEY_COLUMNS),
        validate="many_to_one",
        indicator=True,
    )

    unmatched = changes["_merge"].ne("both")

    if unmatched.any():
        keys = changes.loc[
            unmatched,
            list(CHANGE_KEY_COLUMNS),
        ].head(20).to_dict("records")

        raise ValueError(
            "VM BANK CDPOS rows have no matching CDHDR "
            f"header: {keys}"
        )

    changes = changes.drop(
        columns="_merge"
    )

    changes["Vendor Code"] = changes[
        "Object Value"
    ].map(normalize_vendor_code)

    changes["Change Event Key"] = (
        changes["Change Object Class"]
        + "|"
        + changes["Object Value"]
        + "|"
        + changes["Change Document"]
    )

    changes = changes.sort_values(
        [
            *CHANGE_KEY_COLUMNS,
            "Changed Record Key",
            "Changed Field",
        ],
        kind="mergesort",
    ).reset_index(drop=True)

    metadata.update(
        {
            "available": True,
            "header_rows": len(headers),
            "position_rows": len(positions),
            "change_events": changes[
                "Change Event Key"
            ].nunique(),
        }
    )

    return changes, metadata


def get_intercompany_vendor_codes() -> set[str]:
    """
    Return normalized SAP identifiers from core.intercompanies.

    The shared intercompany file names its identifier field ``customer``
    because it was originally created for AR. VM compares the same configured
    identifiers against normalized Vendor Code values.
    """
    codes = set()

    for item in INTERCOMPANIES:
        if not isinstance(
            item,
            dict,
        ):
            raise ValueError(
                "core.intercompanies.INTERCOMPANIES must contain "
                "dictionaries."
            )

        raw_code = item.get(
            "customer",
            "",
        )

        normalized_code = normalize_vendor_code(
            raw_code
        )

        if normalized_code != "":
            codes.add(
                normalized_code
            )

    if not codes:
        raise ValueError(
            "core.intercompanies.INTERCOMPANIES contains no "
            "valid identifiers for VM exclusions."
        )

    return codes


def get_valid_vendor_population(
    vendor_master: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """
    Apply the approved VM valid-population rules.

    Exclusions:
    - central deletion flag;
    - company deletion flag;
    - vendor prefixes E and T;
    - employee/functionary Account Group ZFUN;
    - configured intercompany identifiers from core.intercompanies.
    """
    required = {
        "Company",
        "Vendor Code",
        "Central Deletion Flag",
        "Company Deletion Flag",
        "Account Group",
        "Trading Partner",
    }

    missing = sorted(
        required.difference(
            vendor_master.columns
        )
    )

    if missing:
        raise ValueError(
            f"VM valid population requires missing columns: "
            f"{missing}."
        )

    result = vendor_master.copy()

    central_deleted = result[
        "Central Deletion Flag"
    ].map(
        lambda value: not is_blank(value)
    )

    company_deleted = result[
        "Company Deletion Flag"
    ].map(
        lambda value: not is_blank(value)
    )

    prefix_excluded = result[
        "Vendor Code"
    ].map(
        normalize_vendor_code
    ).str.startswith(
        (
            "E",
            "T",
        )
    )

    employee_account_group = result[
        "Account Group"
    ].map(
        normalize_upper_text
    ).eq(
        "ZFUN"
    )

    intercompany_vendor_codes = (
        get_intercompany_vendor_codes()
    )

    intercompany_vendor = result[
        "Vendor Code"
    ].map(
        normalize_vendor_code
    ).isin(
        intercompany_vendor_codes
    )

    metadata = {
        "input_rows": len(result),
        "excluded_central_deletion_flag": int(
            central_deleted.sum()
        ),
        "excluded_company_deletion_flag": int(
            company_deleted.sum()
        ),
        "excluded_vendor_prefix_e_or_t": int(
            prefix_excluded.sum()
        ),
        "excluded_employee_account_group_zfun": int(
            employee_account_group.sum()
        ),
        "configured_intercompany_vendor_codes": len(
            intercompany_vendor_codes
        ),
        "excluded_intercompany_vendor_code": int(
            intercompany_vendor.sum()
        ),
        "trading_partner_nonblank_rows": int(
            result["Trading Partner"].map(
                lambda value: not is_blank(value)
            ).sum()
        ),
        "warnings": [],
    }

    excluded = (
        central_deleted
        | company_deleted
        | prefix_excluded
        | employee_account_group
        | intercompany_vendor
    )

    result = result.loc[
        ~excluded
    ].copy()

    metadata["output_rows"] = len(
        result
    )

    return (
        result.sort_values(
            list(
                VENDOR_KEY_COLUMNS
            ),
            kind="mergesort",
        ).reset_index(
            drop=True
        ),
        metadata,
    )


def resolve_tax_business_number(
    dataframe: pd.DataFrame,
    priority_columns: Sequence[str],
    *,
    output_column: str = "Tax/Business Number",
    normalized_output_column: str | None = None,
) -> pd.DataFrame:
    """Resolve Tax/Business Number using explicit functional priority."""
    if not priority_columns:
        raise ValueError(
            "Tax/Business Number requires an explicit ordered "
            "priority of columns."
        )

    missing = [
        column
        for column in priority_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise ValueError(
            "Tax/Business Number priority contains unavailable "
            f"columns: {missing}."
        )

    result = dataframe.copy()

    resolved = pd.Series(
        "",
        index=result.index,
        dtype=object,
    )

    for column in priority_columns:
        candidate = result[column].map(safe_text)
        use_candidate = (
            resolved.eq("")
            & candidate.ne("")
        )
        resolved.loc[use_candidate] = candidate.loc[
            use_candidate
        ]

    result[output_column] = resolved

    if normalized_output_column is not None:
        result[normalized_output_column] = result[
            output_column
        ].map(normalize_tax_id)

    return result


def load_vm_populations(
    context: dict[str, Any],
    *,
    sheet_name: str = VM_INPUT_SHEET,
) -> dict[str, Any]:
    """Load required vendors and every available optional population."""
    vendor_source = load_vm_vendors(
        context,
        sheet_name=sheet_name,
    )

    vendor_master = build_vendor_master_population(
        vendor_source
    )
    vendor_banks = build_vendor_bank_population(
        vendor_source
    )

    vendor_postings, posting_metadata = (
        load_vm_vendor_postings(
            context,
            sheet_name=sheet_name,
        )
    )

    employees, employee_metadata = load_vm_employees(
        context,
        sheet_name=sheet_name,
    )

    bank_changes, bank_change_metadata = (
        load_vm_bank_changes(
            context,
            sheet_name=sheet_name,
        )
    )

    warnings = [
        *posting_metadata["warnings"],
        *employee_metadata["warnings"],
        *bank_change_metadata["warnings"],
    ]

    return {
        "vendor_source": vendor_source,
        "vendor_master": vendor_master,
        "vendor_banks": vendor_banks,
        "vendor_postings": vendor_postings,
        "employees": employees,
        "bank_changes": bank_changes,
        "metadata": {
            "period": get_period_suffix(
                context["module"]
            ),
            "vendor_source_rows": len(vendor_source),
            "vendor_master_rows": len(vendor_master),
            "vendor_bank_rows": len(vendor_banks),
            "vendors_without_bank": int(
                (
                    ~vendor_master["Vendor Key"].isin(
                        vendor_banks["Vendor Key"]
                    )
                ).sum()
            ),
            "postings": posting_metadata,
            "employees": employee_metadata,
            "bank_changes": bank_change_metadata,
            "warnings": warnings,
        },
    }


def get_vm_output_file(
    context: dict[str, Any],
) -> Path:
    """
    Return the period-specific VM result workbook.

    Output:
        output/LBR_Results_VM_YYYYMMDD.xlsx
    """
    output_folder = Path(
        context["output_folder"]
    )

    module_config = context.get("module")

    if not isinstance(module_config, dict):
        raise ValueError(
            "VM output resolution requires context['module']."
        )

    period = get_period_suffix(module_config)

    output_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    return output_folder / VM_OUTPUT_FILE_TEMPLATE.format(
        period=period
    )


def _vm_excel_safe_value(value: Any) -> Any:
    """Convert pandas and numpy values into openpyxl-compatible values."""
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if isinstance(
        value,
        (
            datetime,
            date,
            str,
            int,
            float,
            bool,
        ),
    ):
        return value

    # Handle numpy scalar values without importing numpy directly.
    item_method = getattr(
        value,
        "item",
        None,
    )

    if callable(item_method):
        try:
            scalar = item_method()

            if isinstance(
                scalar,
                (
                    datetime,
                    date,
                    str,
                    int,
                    float,
                    bool,
                ),
            ):
                return scalar
        except (TypeError, ValueError):
            pass

    return str(value)


def _validate_vm_sheet_name(sheet_name: Any) -> str:
    """Validate an Excel-compatible result sheet name."""
    normalized = safe_text(sheet_name)

    if normalized == "":
        raise ValueError(
            "VM output sheet name cannot be empty."
        )

    if len(normalized) > 31:
        raise ValueError(
            f"VM output sheet name exceeds 31 characters: "
            f"{normalized!r}."
        )

    if re.search(r"[\[\]:*?/\\]", normalized):
        raise ValueError(
            f"VM output sheet name contains invalid Excel "
            f"characters: {normalized!r}."
        )

    return normalized


def _normalize_output_column_set(
    columns: Iterable[Any] | None,
) -> set[str]:
    """Normalize an optional output-formatting column collection."""
    if columns is None:
        return set()

    return {
        safe_text(column)
        for column in columns
        if not is_blank(column)
    }


def _format_vm_control_sheet(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    date_columns: Iterable[Any] | None = None,
    amount_columns: Iterable[Any] | None = None,
    integer_columns: Iterable[Any] | None = None,
) -> None:
    """
    Apply standard LBR VM formatting without modifying analytical values.

    Formatting performs no workbook read and no additional DataFrame copy.
    """
    if worksheet.max_row < 1:
        return

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=VM_OUTPUT_HEADER_FILL,
    )
    header_font = Font(
        bold=True,
    )

    header_positions = {}

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

        header_name = safe_text(cell.value)

        if header_name != "":
            header_positions[header_name] = cell.column

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    date_column_set = _normalize_output_column_set(
        date_columns
    )
    amount_column_set = _normalize_output_column_set(
        amount_columns
    )
    integer_column_set = _normalize_output_column_set(
        integer_columns
    )

    formats = {}

    for column_name in date_column_set:
        formats[column_name] = VM_DATE_NUMBER_FORMAT

    for column_name in amount_column_set:
        formats[column_name] = VM_AMOUNT_NUMBER_FORMAT

    for column_name in integer_column_set:
        formats[column_name] = VM_INTEGER_NUMBER_FORMAT

    for column_name, number_format in formats.items():
        column_index = header_positions.get(
            column_name
        )

        if column_index is None:
            continue

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = number_format

    # Width estimation is bounded to the first 2,000 data rows so a large
    # exception population does not trigger another full workbook traversal.
    width_sample_end = min(
        worksheet.max_row,
        2001,
    )

    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        maximum_length = len(
            safe_text(column_name)
        )

        for row_index in range(
            2,
            width_sample_end + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(
            maximum_length + 2,
            45,
        )


def _save_vm_workbook_atomic(
    workbook,
    output_file: Path,
) -> None:
    """
    Save the VM workbook through a temporary file and replace atomically.

    This reduces the possibility of leaving a damaged result workbook if the
    save operation fails midway.
    """
    output_file = Path(output_file)
    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary_file = None

    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_file.stem}_",
            suffix=".xlsx",
            dir=output_file.parent,
            delete=False,
        ) as handle:
            temporary_file = Path(handle.name)

        workbook.save(temporary_file)

        temporary_file.replace(output_file)
        temporary_file = None

    except PermissionError as error:
        raise PermissionError(
            f"VM output is open or locked: {output_file}. "
            "Close Excel and retry."
        ) from error

    except Exception:
        if (
            temporary_file is not None
            and temporary_file.exists()
        ):
            try:
                temporary_file.unlink()
            except OSError:
                pass

        raise

    finally:
        close_method = getattr(
            workbook,
            "close",
            None,
        )

        if callable(close_method):
            close_method()


def write_vm_control_sheet(
    context: dict[str, Any],
    sheet_name: str,
    dataframe: pd.DataFrame,
    date_columns: Iterable[Any] | None = None,
    amount_columns: Iterable[Any] | None = None,
    integer_columns: Iterable[Any] | None = None,
) -> Path:
    """
    Write or replace one VM result sheet while preserving all other VM sheets.

    The workbook is opened and saved once per invocation. The input workbook is
    never reopened by this function.

    Parameters
    ----------
    context:
        Runner context containing output_folder and module configuration.
    sheet_name:
        VM result sheet, for example VM01 or VM02.
    dataframe:
        Final control output with columns already in required eaorder.
    date_columns:
        Columns formatted as DD/MM/YYYY.
    amount_columns:
        Columns formatted as #,##0.00.
    integer_columns:
        Columns formatted as #,##0.

    Returns
    -------
    pathlib.Path
        Path of LBR_Results_VM_YYYYMMDD.xlsx.
    """
    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(
            "write_vm_control_sheet requires a pandas DataFrame."
        )

    normalized_sheet_name = _validate_vm_sheet_name(
        sheet_name
    )
    output_file = get_vm_output_file(
        context
    )

    if output_file.exists():
        try:
            workbook = load_workbook(
                output_file
            )
        except PermissionError as error:
            raise PermissionError(
                f"VM output is open or locked: {output_file}. "
                "Close Excel and retry."
            ) from error
        except Exception as error:
            raise ValueError(
                f"VM output workbook is not readable: "
                f"{output_file}: {error}"
            ) from error
    else:
        workbook = Workbook()

        default_sheet = workbook.active

        if default_sheet is not None:
            workbook.remove(default_sheet)

    if normalized_sheet_name in workbook.sheetnames:
        workbook.remove(
            workbook[normalized_sheet_name]
        )

    worksheet = workbook.create_sheet(
        normalized_sheet_name
    )

    # Header row.
    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=safe_text(column_name),
        )

    # Data rows. itertuples avoids DataFrame.iterrows and apply(axis=1).
    for row_index, row_values in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        ),
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=_vm_excel_safe_value(value),
            )

    _format_vm_control_sheet(
        worksheet=worksheet,
        dataframe=dataframe,
        date_columns=date_columns,
        amount_columns=amount_columns,
        integer_columns=integer_columns,
    )

    _save_vm_workbook_atomic(
        workbook=workbook,
        output_file=output_file,
    )

    print(f"VM output file: {output_file}")
    print(f"VM output sheet: {normalized_sheet_name}")
    print(f"VM output rows: {len(dataframe)}")

    return output_file
