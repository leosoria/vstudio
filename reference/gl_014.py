"""GL_014 - Analysis Of General Ledger Accounts By General Journal Type.

For LBR, the functional journal-type dimension is ``Transaction Code``
(``BKPF-TCODE``). It is intentionally kept distinct from Document Type
(``BKPF-BLART``): the LHA output matrix requests Transaction Code and the LBR
GL controls expose both fields separately.

The control reads only the required BSIS/BSAS columns, filters each source
before concatenation, removes cross-source technical duplicates, counts unique
Company/Fiscal Year/Document journals, and replaces only sheet GL14.
"""

import importlib.util
import re
import shutil
import tempfile
from pathlib import Path
from time import perf_counter

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAME = "GL14"
REPORT_CURRENCY = "USD"
HEADER_FILL = "D9EAF7"
AMOUNT_FORMAT = '#,##0.00;[Red]-#,##0.00'
INTEGER_FORMAT = '#,##0'
ALLOWED_EXTENSIONS = {
    ".xlsx",
    ".xls",
    ".csv",
    ".txt",
}

BSIS_KEYWORDS = [
    "LBR GL_JE_BSIS",
    "LBR_GL_JE_BSIS",
]
BSAS_KEYWORDS = [
    "LBR GL_JE_BSAS",
    "LBR_GL_JE_BSAS",
]
MASTER_KEYWORDS = [
    "LBR GL_MD",
    "LBR_GL_MD",
]
FX_KEYWORDS = [
    "FxRates",
    "FX Rates",
    "Fx Rates",
]

ALIASES = {
    "company": [
        "Company Code",
        "Company",
        "CoCd",
        "CompanyCode",
        "BKPF-BUKRS",
        "BSIS-BUKRS",
        "BSAS-BUKRS",
        "BUKRS",
        "Empr",
    ],
    "year": [
        "Fiscal Year",
        "FiscalYear",
        "Year",
        "Fisc.Year",
        "BKPF-GJAHR",
        "BSIS-GJAHR",
        "BSAS-GJAHR",
        "GJAHR",
        "Ano",
    ],
    "document": [
        "Document Number",
        "Document No.",
        "DocumentNo",
        "Accounting Document",
        "Journal Entry",
        "BKPF-BELNR",
        "BSIS-BELNR",
        "BSAS-BELNR",
        "BELNR",
        "Nº doc.",
        "Nº doc",
    ],
    "line": [
        "Line Number",
        "Line Item",
        "Item",
        "Document Item",
        "BSIS-BUZEI",
        "BSAS-BUZEI",
        "BUZEI",
        "Line",
        "Itm",
    ],
    "account": [
        "GL Account",
        "G/L Account",
        "G/L Acct",
        "G/L acct",
        "G/L",
        "G/L Acct.",
        "Account",
        "Account Number",
        "BSIS-HKONT",
        "BSAS-HKONT",
        "HKONT",
        "Account Code",
        "Razão",
        "Razao",
        "Cta.Razão",
        "Cta.Razao",
    ],
    "transaction": [
        "Transaction Code",
        "Transaction",
        "Transactn Code",
        "TCode",
        "BKPF-TCODE",
        "TCODE",
        "CódT",
        "CodT",
    ],
    "indicator": [
        "Debit/Credit Indicator",
        "Debit/Credit Ind.",
        "Debit/Credit",
        "D/C Indicator",
        "BSIS-SHKZG",
        "BSAS-SHKZG",
        "SHKZG",
        "D/C",
    ],
    "document_amount": [
        "Document Amount",
        "Amount in Document Currency",
        "Amount in doc. curr.",
        "Amount in doc.curr.",
        "Amount in Doc. Curr.",
        "Amount in DC",
        "Amt in doc. curr.",
        "Amt.in doc.curr.",
        "Amount",
        "BSIS-WRBTR",
        "BSAS-WRBTR",
        "WRBTR",
        "Montante",
    ],
    "currency": [
        "Document Currency",
        "Document Curr.",
        "DocumentCurrency",
        "Doc. Currency",
        "Doc. Curr.",
        "Doc.Curr.",
        "DocCurr",
        "Currency",
        "Curr.",
        "Crcy",
        "BKPF-WAERS",
        "WAERS",
        "Moeda",
    ],
    "posting_date": [
        "Posting Date",
        "Pstng Date",
        "PostingDate",
        "BKPF-BUDAT",
        "BUDAT",
        "Dt.lçto.",
        "Dt.lçto",
        "Dt.lcto.",
        "Dt.lcto",
    ],
    "document_date": [
        "Document Date",
        "Doc. Date",
        "Doc.Date",
        "BKPF-BLDAT",
        "BLDAT",
        "Data doc.",
        "Data doc",
    ],
}

OUTPUT_COLUMNS = [
    "Company Code",
    "Company Name",
    "GL Account",
    "GL Account Description",
    "Transaction Code",
    "Transaction Code Description",
    "Report Currency",
    "Count",
    "Credit",
    "Debit",
    "Amount in Reporting Currency",
]

AMOUNT_COLUMNS = {
    "Credit",
    "Debit",
    "Amount in Reporting Currency",
}

INTEGER_COLUMNS = {
    "Count",
}

COLUMN_WIDTHS = {
    "Company Code": 14,
    "Company Name": 30,
    "GL Account": 18,
    "GL Account Description": 38,
    "Transaction Code": 18,
    "Transaction Code Description": 38,
    "Report Currency": 17,
    "Count": 12,
    "Credit": 20,
    "Debit": 20,
    "Amount in Reporting Currency": 30,
}


def _text(series):
    return (
        series
        .fillna("")
        .astype(str)
        .str.strip()
        .replace(
            {
                "nan": "",
                "None": "",
                "<NA>": "",
            }
        )
    )


def _header(value):
    if value is None:
        text = ""
    else:
        text = (
            str(value)
            .replace("\ufeff", "")
            .replace("\xa0", " ")
            .strip()
            .casefold()
        )

    text = " ".join(text.split())

    while (
        "." in text
        and text.rsplit(".", 1)[-1].isdigit()
    ):
        text = text.rsplit(".", 1)[0]

    return text


def _code(series):
    return _text(series).str.replace(
        r"\.0$",
        "",
        regex=True,
    )


def _company(series):
    values = _code(series)
    numeric = values.str.fullmatch(r"\d+")

    values.loc[numeric] = (
        values.loc[numeric]
        .str.lstrip("0")
        .replace("", "0")
    )

    return values


def _number(series):
    text = _text(series)
    parentheses = text.str.match(r"^\(.*\)$")

    text = text.str.replace(
        r"[()\s]",
        "",
        regex=True,
    )

    both_separators = (
        text.str.contains(",", regex=False)
        & text.str.contains(".", regex=False)
    )

    comma_decimal = text.str.contains(
        r",\d{1,6}$",
        regex=True,
    )

    text = text.where(
        ~both_separators,
        (
            text
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        ),
    )

    text = text.where(
        both_separators | ~comma_decimal,
        (
            text
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        ),
    )

    text = text.where(
        both_separators | comma_decimal,
        text.str.replace(
            ",",
            "",
            regex=False,
        ),
    )

    values = pd.to_numeric(
        text,
        errors="coerce",
    )

    return values.where(
        ~parentheses,
        -values.abs(),
    )


def _date(series):
    text = _text(series)

    parsed = pd.Series(
        pd.NaT,
        index=series.index,
        dtype="datetime64[ns]",
    )

    iso_mask = text.str.match(
        r"^\d{4}-\d{1,2}-\d{1,2}(?:\s|$)"
    )
    compact_mask = text.str.fullmatch(
        r"\d{8}"
    )

    parsed.loc[iso_mask] = pd.to_datetime(
        text.loc[iso_mask].str[:10],
        format="%Y-%m-%d",
        errors="coerce",
    )

    parsed.loc[compact_mask] = pd.to_datetime(
        text.loc[compact_mask],
        format="%Y%m%d",
        errors="coerce",
    )

    remaining_mask = ~(
        iso_mask
        | compact_mask
    )

    parsed.loc[remaining_mask] = pd.to_datetime(
        text.loc[remaining_mask],
        errors="coerce",
        dayfirst=True,
    )

    numeric = pd.to_numeric(
        series,
        errors="coerce",
    )

    excel_dates = pd.to_datetime(
        numeric,
        unit="D",
        origin="1899-12-30",
        errors="coerce",
    )

    return parsed.where(
        parsed.notna(),
        excel_dates,
    )


def _date_value(value):
    return _date(
        pd.Series([value])
    ).iloc[0]


def _company_filter(value):
    if value is None:
        text = ""
    else:
        text = str(value).strip()

    if (
        not text
        or text.upper() in {
            "ALL",
            "TODAS",
            "TODOS",
        }
    ):
        return []

    for separator in [
        ";",
        "|",
        "\n",
        "\r",
        "\t",
        " ",
    ]:
        text = text.replace(
            separator,
            ",",
        )

    normalized = _company(
        pd.Series(
            text.split(",")
        )
    ).tolist()

    return [
        company
        for company in normalized
        if company
    ]


def _candidates(
    dataframe,
    aliases,
):
    wanted = {
        _header(alias)
        for alias in aliases
    }

    return [
        column
        for column in dataframe.columns
        if _header(column) in wanted
    ]


def _compact_header(value):
    return re.sub(
        r"[^a-z0-9]+",
        "",
        _header(value),
    )


def _looks_like_logical_header(
    value,
    logical_name,
):
    """
    Recognize shortened SAP ALV labels without confusing related fields.
    """
    normalized = _header(value)
    compact = _compact_header(value)

    if logical_name == "account":
        if any(
            word in normalized
            for word in (
                "description",
                "text",
                "name",
                "type",
            )
        ):
            return False

        return (
            compact in {
                "gl",
                "glaccount",
                "glacct",
                "account",
                "accountnumber",
                "acct",
            }
            or normalized.startswith("g/l")
        )

    if logical_name == "document_amount":
        if not (
            "amount" in normalized
            or "amt" in normalized
        ):
            return False

        if any(
            word in normalized
            for word in (
                "local",
                "report",
                " usd",
                " lc",
            )
        ):
            return False

        return (
            any(
                word in normalized
                for word in (
                    "document",
                    "doc.",
                    "doc ",
                    " dc",
                )
            )
            or compact.endswith("dc")
        )

    if logical_name == "currency":
        if (
            "amount" in normalized
            or "amt" in normalized
            or "rate" in normalized
        ):
            return False

        return compact in {
            "currency",
            "curr",
            "crcy",
            "documentcurrency",
            "documentcurr",
            "doccurrency",
            "doccurr",
            "waers",
        }

    return False


def _logical_candidates(
    dataframe,
    logical_name,
    aliases,
):
    exact_candidates = _candidates(
        dataframe,
        aliases,
    )

    fuzzy_candidates = [
        column
        for column in dataframe.columns
        if (
            column not in exact_candidates
            and _looks_like_logical_header(
                column,
                logical_name,
            )
        )
    ]

    return (
        exact_candidates
        + fuzzy_candidates
    )


def _candidate_score(
    series,
    logical_name,
):
    clean = _text(series)
    populated = clean.ne("").mean()

    if logical_name == "indicator":
        consistent = (
            clean
            .str.upper()
            .isin(
                {
                    "S",
                    "H",
                    "D",
                    "C",
                    "DEBIT",
                    "CREDIT",
                }
            )
            .mean()
        )
    elif logical_name == "document_amount":
        consistent = (
            _number(series)
            .notna()
            .mean()
        )
    elif logical_name in {
        "posting_date",
        "document_date",
    }:
        consistent = (
            _date(series)
            .notna()
            .mean()
        )
    else:
        consistent = clean.ne("").mean()

    return populated + consistent


def _resolve(
    dataframe,
    source_name,
):
    resolved = {}
    missing = []

    for logical_name, aliases in ALIASES.items():
        candidates = _logical_candidates(
            dataframe,
            logical_name,
            aliases,
        )

        if not candidates:
            if logical_name == "document_date":
                resolved[logical_name] = None
            else:
                missing.append(
                    f"{logical_name}: "
                    f"expected one of {aliases}"
                )
            continue

        resolved[logical_name] = max(
            candidates,
            key=lambda column: (
                _candidate_score(
                    dataframe[column],
                    logical_name,
                ),
                -dataframe.columns.get_loc(
                    column
                ),
            ),
        )

    if missing:
        available_headers = dataframe.attrs.get(
            "source_headers",
            dataframe.columns,
        )

        available = ", ".join(
            repr(str(column))
            for column in available_headers
        )

        raise ValueError(
            f"Missing required columns in "
            f"GL14 {source_name}:\n- "
            + "\n- ".join(missing)
            + "\nAvailable headers read by "
            f"GL14: {available}"
        )

    return resolved


def _read(
    path,
    usecols=None,
    nrows=None,
):
    path = Path(path)
    extension = path.suffix.lower()

    if extension in {
        ".xlsx",
        ".xls",
    }:
        dataframe = pd.read_excel(
            path,
            dtype=object,
            usecols=usecols,
            nrows=nrows,
        )
    elif extension == ".csv":
        dataframe = pd.read_csv(
            path,
            dtype=object,
            usecols=usecols,
            nrows=nrows,
        )
    elif extension == ".txt":
        dataframe = pd.read_csv(
            path,
            sep="\t",
            dtype=object,
            usecols=usecols,
            nrows=nrows,
        )
    else:
        raise ValueError(
            f"Unsupported input extension: "
            f"{path.suffix}"
        )

    dataframe.columns = [
        str(column).strip()
        for column in dataframe.columns
    ]

    return dataframe.dropna(
        how="all"
    )


def _period_suffix(context):
    value = _date_value(
        context["module"].get(
            "to",
            "",
        )
    )

    if pd.isna(value):
        raise ValueError(
            "GL14 requires a valid "
            "module TO date."
        )

    return value.strftime(
        "%Y%m%d"
    )


def _find_file(
    context,
    keywords,
    required=False,
):
    folder = Path(
        context["input_folder"]
    )
    suffix = _period_suffix(
        context
    ).casefold()
    matches = []

    for path in folder.rglob("*"):
        if not path.is_file():
            continue

        if (
            path.suffix.lower()
            not in ALLOWED_EXTENSIONS
        ):
            continue

        normalized_name = (
            path.stem.casefold()
        )

        if (
            suffix in normalized_name
            and any(
                keyword.casefold()
                in normalized_name
                for keyword in keywords
            )
        ):
            matches.append(path)

    matches = sorted(
        set(matches)
    )

    if len(matches) > 1:
        raise ValueError(
            "Multiple GL14 inputs found: "
            + ", ".join(
                str(path)
                for path in matches
            )
        )

    if (
        not matches
        and required
    ):
        raise FileNotFoundError(
            "GL14 input not found for "
            f"{keywords[0]}_{suffix}."
        )

    if not matches:
        return None

    return matches[0]


def _load_source(
    context,
    source_name,
    keywords,
    path=None,
):
    started = perf_counter()

    if path is None:
        path = _find_file(
            context,
            keywords,
        )

    if path is None:
        print(
            f"GL14 {source_name} "
            "input file: not found"
        )
        print(
            f"GL14 timing - read "
            f"{source_name}: "
            f"{perf_counter() - started:.2f} "
            "seconds"
        )
        return pd.DataFrame()

    headers = _read(
        path,
        nrows=0,
    )

    header_candidate_positions = []

    for logical_name, aliases in ALIASES.items():
        candidates = set(
            _logical_candidates(
                headers,
                logical_name,
                aliases,
            )
        )

        header_candidate_positions.extend(
            index
            for index, column
            in enumerate(headers.columns)
            if column in candidates
        )

    # Physical positions are used instead of labels because pandas may
    # recalculate .1/.2 suffixes after a label-based usecols selection.
    usecols = sorted(
        set(
            header_candidate_positions
        )
    )

    dataframe = _read(
        path,
        usecols=usecols,
    )

    dataframe.attrs[
        "source_headers"
    ] = list(
        headers.columns
    )

    print(
        f"GL14 {source_name} "
        f"input file: {path}"
    )
    print(
        f"GL14 {source_name} "
        f"rows read: {len(dataframe)}"
    )
    print(
        f"GL14 timing - read "
        f"{source_name}: "
        f"{perf_counter() - started:.2f} "
        "seconds"
    )

    return dataframe


def _prepare(
    dataframe,
    context,
    source_name,
):
    if dataframe.empty:
        return pd.DataFrame()

    columns = _resolve(
        dataframe,
        source_name,
    )

    companies = _company_filter(
        context["module"].get(
            "companies",
            "",
        )
    )

    company = _company(
        dataframe[
            columns["company"]
        ]
    )

    posting_date = _date(
        dataframe[
            columns["posting_date"]
        ]
    )

    from_date = _date_value(
        context["module"].get(
            "from",
            "",
        )
    )

    to_date = _date_value(
        context["module"].get(
            "to",
            "",
        )
    )

    filter_mask = posting_date.between(
        from_date,
        to_date,
        inclusive="both",
    )

    if companies:
        filter_mask &= company.isin(
            companies
        )

    selected = dataframe.loc[
        filter_mask
    ]

    posting_date = posting_date.loc[
        filter_mask
    ]

    debit_credit = _text(
        selected[
            columns["indicator"]
        ]
    ).str.upper()

    document_amount = _number(
        selected[
            columns["document_amount"]
        ]
    ).abs()

    is_debit = debit_credit.isin(
        {
            "S",
            "D",
            "DEBIT",
        }
    )

    is_credit = debit_credit.isin(
        {
            "H",
            "C",
            "CREDIT",
        }
    )

    invalid_indicator = ~(
        is_debit
        | is_credit
    )

    if invalid_indicator.any():
        invalid_values = sorted(
            debit_credit.loc[
                invalid_indicator
            ]
            .drop_duplicates()
            .tolist()
        )

        raise ValueError(
            f"GL14 {source_name} has "
            "unsupported D/C indicators: "
            f"{invalid_values}"
        )

    result = pd.DataFrame(
        index=selected.index
    )

    result["Company Code"] = (
        company.loc[
            filter_mask
        ]
    )

    result["Fiscal Year"] = _code(
        selected[
            columns["year"]
        ]
    )

    result["Document Number"] = _code(
        selected[
            columns["document"]
        ]
    )

    result["Line Number"] = _code(
        selected[
            columns["line"]
        ]
    )

    result["GL Account"] = _code(
        selected[
            columns["account"]
        ]
    )

    result["Transaction Code"] = _text(
        selected[
            columns["transaction"]
        ]
    ).str.upper()

    result["Document Currency"] = _text(
        selected[
            columns["currency"]
        ]
    ).str.upper()

    if columns["document_date"] is not None:
        result["Conversion Date"] = _date(
            selected[
                columns["document_date"]
            ]
        )
    else:
        result["Conversion Date"] = (
            posting_date
        )

    result["Posting Date"] = (
        posting_date
    )

    result[
        "Document Amount Signed"
    ] = document_amount.where(
        is_debit,
        -document_amount,
    )

    result["Debit Document"] = (
        document_amount.where(
            is_debit,
            0.0,
        )
    )

    result["Credit Document"] = (
        document_amount.where(
            is_credit,
            0.0,
        )
    )

    result["Source"] = (
        source_name
    )

    result["Journal ID"] = (
        result["Company Code"]
        + "|"
        + result["Fiscal Year"]
        + "|"
        + result["Document Number"]
    )

    result = result.reset_index(
        drop=True
    )

    print(
        f"GL14 {source_name} rows "
        "after CONFIG filters: "
        f"{len(result)}"
    )

    return result


def _deduplicate(
    bsis,
    bsas,
):
    combined = pd.concat(
        [
            bsis,
            bsas,
        ],
        ignore_index=True,
    )

    if combined.empty:
        return combined, 0

    line_key = [
        "Company Code",
        "Fiscal Year",
        "Document Number",
        "Line Number",
    ]

    combined["_priority"] = (
        combined["Source"].map(
            {
                "BSIS": 0,
                "BSAS": 1,
            }
        )
    )

    before = len(combined)

    combined = (
        combined
        .sort_values(
            "_priority",
            kind="stable",
        )
        .drop_duplicates(
            line_key,
            keep="last",
        )
    )

    combined = (
        combined
        .drop(
            columns="_priority"
        )
        .reset_index(
            drop=True
        )
    )

    removed = (
        before
        - len(combined)
    )

    return combined, removed


def _load_optional(
    context,
    keywords,
    label,
):
    path = _find_file(
        context,
        keywords,
    )

    if path is None:
        print(
            f"GL14 {label} "
            "input file: not found"
        )
        return pd.DataFrame()

    print(
        f"GL14 {label} "
        f"input file: {path}"
    )

    return _read(path)


def _fx_columns(dataframe):
    aliases = {
        "type": [
            "Exchange Rate Type",
            "ExRt",
            "CgCâ",
            "CgCa",
            "KURST",
            "TCot",
        ],
        "from": [
            "From Currency",
            "From",
            "De",
            "FCURR",
        ],
        "to": [
            "To Currency",
            "To",
            "Para",
            "TCURR",
        ],
        "date": [
            "Valid From",
            "Vál.desde",
            "Val.desde",
            "GDATU",
        ],
        "rate": [
            "Exchange Rate",
            "Exch. Rate",
            "Taxa câmbio",
            "Taxa cambio",
            "UKURS",
            "Rate",
        ],
    }

    result = {}

    for (
        logical_name,
        possible_names,
    ) in aliases.items():
        candidates = _candidates(
            dataframe,
            possible_names,
        )

        if not candidates:
            if logical_name == "type":
                result[logical_name] = None
                continue

            available_headers = ", ".join(
                repr(str(column))
                for column in dataframe.columns
            )

            print(
                "GL14 FxRates cannot be used "
                f"because {logical_name} is missing; "
                f"expected one of {possible_names}. "
                f"Available headers: {available_headers}. "
                "Amount in Reporting Currency will "
                "remain blank when the document "
                "currency is not USD."
            )

            return None

        result[logical_name] = (
            candidates[0]
        )

    result["from_factor"] = (
        _candidates(
            dataframe,
            [
                "From Factor",
                "Factor From",
                "Ratio (from)",
                "Fator (origem)",
                "FFACT",
            ],
        )
        or [None]
    )[0]

    result["to_factor"] = (
        _candidates(
            dataframe,
            [
                "To Factor",
                "Factor To",
                "Ratio (to)",
                "Fator (para)",
                "TFACT",
            ],
        )
        or [None]
    )[0]

    return result


def _add_reporting_currency(
    dataframe,
    fx_dataframe,
):
    started = perf_counter()

    result = dataframe.copy()
    result["FX Factor"] = pd.NA

    if result.empty:
        print(
            "GL14 timing - FX: "
            f"{perf_counter() - started:.2f} "
            "seconds"
        )
        return result

    document_currency = (
        result["Document Currency"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    is_usd = document_currency.isin(
        {
            "USD",
            "$",
        }
    )

    result.loc[
        is_usd,
        "FX Factor",
    ] = 1.0

    fx_usable = False

    if not fx_dataframe.empty:
        columns = _fx_columns(
            fx_dataframe
        )

        if columns is not None:
            fx_usable = True

            if columns["type"] is None:
                rate_type_values = pd.Series(
                    "M",
                    index=fx_dataframe.index,
                    dtype="object",
                )
            else:
                rate_type_values = _text(
                    fx_dataframe[
                        columns["type"]
                    ]
                ).str.upper()

            if columns["from_factor"] is None:
                from_factor = pd.Series(
                    1.0,
                    index=fx_dataframe.index,
                )
            else:
                from_factor = _number(
                    fx_dataframe[
                        columns["from_factor"]
                    ]
                ).fillna(1.0)

            if columns["to_factor"] is None:
                to_factor = pd.Series(
                    1.0,
                    index=fx_dataframe.index,
                )
            else:
                to_factor = _number(
                    fx_dataframe[
                        columns["to_factor"]
                    ]
                ).fillna(1.0)

            normalized_fx = pd.DataFrame(
                {
                    "Rate Type": rate_type_values,
                    "From Currency": _text(
                        fx_dataframe[
                            columns["from"]
                        ]
                    ).str.upper(),
                    "To Currency": _text(
                        fx_dataframe[
                            columns["to"]
                        ]
                    ).str.upper(),
                    "Valid From": _date(
                        fx_dataframe[
                            columns["date"]
                        ]
                    ).dt.normalize(),
                    "Rate": _number(
                        fx_dataframe[
                            columns["rate"]
                        ]
                    ),
                    "From Factor": from_factor,
                    "To Factor": to_factor,
                }
            ).dropna(
                subset=[
                    "Valid From",
                    "Rate",
                ]
            )

            normalized_fx[
                "Adjusted Rate"
            ] = (
                normalized_fx["Rate"]
                * normalized_fx["To Factor"]
                / normalized_fx["From Factor"]
            )
        else:
            normalized_fx = pd.DataFrame(
                columns=[
                    "Rate Type",
                    "From Currency",
                    "To Currency",
                    "Valid From",
                    "Rate",
                    "From Factor",
                    "To Factor",
                    "Adjusted Rate",
                ]
            )

        requested_rates = (
            result.loc[
                ~is_usd,
                [
                    "Document Currency",
                    "Conversion Date",
                ],
            ]
            .drop_duplicates()
            .copy()
        )

        rate_cache = {}

        for (
            currency,
            requested_date,
        ) in requested_rates.itertuples(
            index=False,
            name=None,
        ):
            factor = None

            if (
                currency
                and pd.notna(
                    requested_date
                )
            ):
                requested_date = (
                    requested_date
                    .normalize()
                )

                minimum_date = (
                    requested_date
                    - pd.Timedelta(
                        days=10
                    )
                )

                eligible = normalized_fx[
                    (
                        normalized_fx[
                            "Valid From"
                        ]
                        <= requested_date
                    )
                    & (
                        normalized_fx[
                            "Valid From"
                        ]
                        >= minimum_date
                    )
                ]

                for rate_type in (
                    "EN",
                    "M",
                ):
                    rate_type_rows = eligible[
                        eligible[
                            "Rate Type"
                        ]
                        == rate_type
                    ]

                    direct = (
                        rate_type_rows[
                            (
                                rate_type_rows[
                                    "From Currency"
                                ]
                                == currency
                            )
                            & (
                                rate_type_rows[
                                    "To Currency"
                                ]
                                == REPORT_CURRENCY
                            )
                        ]
                        .sort_values(
                            "Valid From",
                            ascending=False,
                        )
                    )

                    inverse = (
                        rate_type_rows[
                            (
                                rate_type_rows[
                                    "From Currency"
                                ]
                                == REPORT_CURRENCY
                            )
                            & (
                                rate_type_rows[
                                    "To Currency"
                                ]
                                == currency
                            )
                        ]
                        .sort_values(
                            "Valid From",
                            ascending=False,
                        )
                    )

                    if not direct.empty:
                        adjusted_rate = (
                            direct.iloc[0][
                                "Adjusted Rate"
                            ]
                        )

                        if (
                            pd.notna(
                                adjusted_rate
                            )
                            and adjusted_rate != 0
                        ):
                            factor = float(
                                adjusted_rate
                            )
                            break

                    if not inverse.empty:
                        adjusted_rate = (
                            inverse.iloc[0][
                                "Adjusted Rate"
                            ]
                        )

                        if (
                            pd.notna(
                                adjusted_rate
                            )
                            and adjusted_rate != 0
                        ):
                            factor = (
                                1.0
                                / float(
                                    adjusted_rate
                                )
                            )
                            break

            rate_cache[
                (
                    currency,
                    requested_date,
                )
            ] = factor

        if rate_cache:
            key_index = (
                pd.MultiIndex
                .from_frame(
                    result[
                        [
                            "Document Currency",
                            "Conversion Date",
                        ]
                    ]
                )
            )

            factors = pd.Series(
                rate_cache,
                dtype="float64",
            )

            result.loc[
                ~is_usd,
                "FX Factor",
            ] = factors.reindex(
                key_index[
                    ~is_usd
                ]
            ).to_numpy()

    numeric_factor = pd.to_numeric(
        result["FX Factor"],
        errors="coerce",
    )

    result["Debit"] = (
        result["Debit Document"]
        * numeric_factor
    )

    result["Credit"] = (
        result["Credit Document"]
        * numeric_factor
    )

    result[
        "Amount in Reporting Currency"
    ] = (
        result["Document Amount Signed"]
        * numeric_factor
    )

    missing_fx = (
        numeric_factor.isna()
    )

    missing_currencies = ", ".join(
        sorted(
            currency
            for currency in result.loc[
                missing_fx,
                "Document Currency",
            ]
            .dropna()
            .unique()
            if currency != ""
        )
    )

    print(
        "GL14 FX missing rows: "
        f"{int(missing_fx.sum())}; "
        "currencies: "
        f"{missing_currencies or 'none'}"
    )

    print(
        "GL14 FxRates used: "
        f"{'yes' if fx_usable else 'no'}"
    )

    print(
        "GL14 timing - FX: "
        f"{perf_counter() - started:.2f} "
        "seconds"
    )

    return result


def _master_maps(master_dataframe):
    if master_dataframe.empty:
        return {}, {}, {}

    company_candidates = _candidates(
        master_dataframe,
        [
            "Company Code",
            "BUKRS",
            "Empr",
        ],
    )

    company_name_candidates = _candidates(
        master_dataframe,
        [
            "Company Name",
            "BUTXT",
            "Nome da empresa",
            "Empresa",
        ],
    )

    account_candidates = _candidates(
        master_dataframe,
        [
            "GL Account",
            "SAKNR",
            "HKONT",
            "Cta.Razão",
            "Razão",
        ],
    )

    account_name_candidates = _candidates(
        master_dataframe,
        [
            "GL Account Description",
            "TXT50",
            "TXT20",
            "Account Name",
            "Description",
            "TxtDescr",
        ],
    )

    transaction_candidates = _candidates(
        master_dataframe,
        [
            "Transaction Code",
            "TCODE",
            "BKPF-TCODE",
            "CódT",
        ],
    )

    transaction_name_candidates = _candidates(
        master_dataframe,
        [
            "Transaction Code Description",
            "Transaction Description",
            "TCODE Description",
        ],
    )

    company_column = (
        company_candidates[0]
        if company_candidates
        else None
    )

    company_name_column = (
        company_name_candidates[0]
        if company_name_candidates
        else None
    )

    account_column = (
        account_candidates[0]
        if account_candidates
        else None
    )

    account_name_column = (
        account_name_candidates[0]
        if account_name_candidates
        else None
    )

    transaction_column = (
        transaction_candidates[0]
        if transaction_candidates
        else None
    )

    transaction_name_column = (
        transaction_name_candidates[0]
        if transaction_name_candidates
        else None
    )

    company_map = {}
    account_map = {}
    transaction_map = {}

    if (
        company_column is not None
        and company_name_column is not None
    ):
        company_data = pd.DataFrame(
            {
                "key": _company(
                    master_dataframe[
                        company_column
                    ]
                ),
                "value": _text(
                    master_dataframe[
                        company_name_column
                    ]
                ),
            }
        )

        company_data = (
            company_data[
                (
                    company_data["key"]
                    != ""
                )
                & (
                    company_data["value"]
                    != ""
                )
            ]
            .drop_duplicates(
                "key",
                keep="first",
            )
        )

        company_map = (
            company_data
            .set_index(
                "key"
            )["value"]
            .to_dict()
        )

    if (
        account_column is not None
        and account_name_column is not None
    ):
        if company_column is None:
            account_company = pd.Series(
                "",
                index=master_dataframe.index,
            )
        else:
            account_company = _company(
                master_dataframe[
                    company_column
                ]
            )

        account_data = pd.DataFrame(
            {
                "company": account_company,
                "account": _code(
                    master_dataframe[
                        account_column
                    ]
                ),
                "value": _text(
                    master_dataframe[
                        account_name_column
                    ]
                ),
            }
        )

        account_data = (
            account_data[
                (
                    account_data["account"]
                    != ""
                )
                & (
                    account_data["value"]
                    != ""
                )
            ]
            .drop_duplicates(
                [
                    "company",
                    "account",
                ],
                keep="first",
            )
        )

        account_map = (
            account_data
            .set_index(
                [
                    "company",
                    "account",
                ]
            )["value"]
            .to_dict()
        )

    if (
        transaction_column is not None
        and transaction_name_column is not None
    ):
        transaction_data = pd.DataFrame(
            {
                "key": _text(
                    master_dataframe[
                        transaction_column
                    ]
                ).str.upper(),
                "value": _text(
                    master_dataframe[
                        transaction_name_column
                    ]
                ),
            }
        )

        transaction_data = (
            transaction_data[
                (
                    transaction_data["key"]
                    != ""
                )
                & (
                    transaction_data["value"]
                    != ""
                )
            ]
            .drop_duplicates(
                "key",
                keep="first",
            )
        )

        transaction_map = (
            transaction_data
            .set_index(
                "key"
            )["value"]
            .to_dict()
        )

    return (
        company_map,
        account_map,
        transaction_map,
    )


def create_gl14_summary(
    bsis_dataframe,
    bsas_dataframe,
    master_dataframe,
    fx_dataframe,
    context,
):
    bsis = _prepare(
        bsis_dataframe,
        context,
        "BSIS",
    )

    bsas = _prepare(
        bsas_dataframe,
        context,
        "BSAS",
    )

    deduplication_started = (
        perf_counter()
    )

    lines, removed = _deduplicate(
        bsis,
        bsas,
    )

    print(
        "GL14 rows after "
        "deduplication: "
        f"{len(lines)}; "
        f"removed: {removed}"
    )

    print(
        "GL14 timing - "
        "deduplication: "
        f"{perf_counter() - deduplication_started:.2f} "
        "seconds"
    )

    lines = _add_reporting_currency(
        lines,
        fx_dataframe,
    )

    aggregation_started = (
        perf_counter()
    )

    group_key = [
        "Company Code",
        "GL Account",
        "Transaction Code",
    ]

    output = (
        lines
        .groupby(
            group_key,
            sort=False,
            observed=True,
            dropna=False,
        )
        .agg(
            Count=(
                "Journal ID",
                "nunique",
            ),
            Credit=(
                "Credit",
                lambda values: values.sum(
                    min_count=1
                ),
            ),
            Debit=(
                "Debit",
                lambda values: values.sum(
                    min_count=1
                ),
            ),
            **{
                "Amount in Reporting Currency": (
                    "Amount in Reporting Currency",
                    lambda values: values.sum(
                        min_count=1
                    ),
                )
            },
        )
        .reset_index()
    )

    print(
        "GL14 timing - aggregation: "
        f"{perf_counter() - aggregation_started:.2f} "
        "seconds"
    )

    if output.duplicated(
        group_key
    ).any():
        raise AssertionError(
            "GL14 final aggregation "
            "contains duplicate keys."
        )

    if (
        output["Count"]
        .le(0)
        .any()
    ):
        raise AssertionError(
            "GL14 Count validation failed."
        )

    for column_name in (
        "Debit",
        "Credit",
    ):
        if (
            output[column_name]
            .dropna()
            .lt(0)
            .any()
        ):
            raise AssertionError(
                f"GL14 {column_name} "
                "must be positive or zero."
            )

        detail_total = (
            lines[column_name]
            .sum()
        )

        output_total = (
            output[column_name]
            .sum()
        )

        if abs(
            output_total
            - detail_total
        ) > 0.01:
            raise AssertionError(
                f"GL14 {column_name} "
                "reconciliation failed."
            )

    detail_net = lines[
        "Amount in Reporting Currency"
    ].sum()

    output_net = output[
        "Amount in Reporting Currency"
    ].sum()

    if abs(
        output_net
        - detail_net
    ) > 0.01:
        raise AssertionError(
            "GL14 reporting currency "
            "reconciliation failed."
        )

    enrichment_started = (
        perf_counter()
    )

    (
        company_map,
        account_map,
        transaction_map,
    ) = _master_maps(
        master_dataframe
    )

    output["Company Name"] = (
        output["Company Code"]
        .map(company_map)
        .fillna("")
    )

    output[
        "GL Account Description"
    ] = [
        account_map.get(
            (
                company_code,
                gl_account,
            ),
            account_map.get(
                (
                    "",
                    gl_account,
                ),
                "",
            ),
        )
        for (
            company_code,
            gl_account,
        ) in output[
            [
                "Company Code",
                "GL Account",
            ]
        ].itertuples(
            index=False,
            name=None,
        )
    ]

    output[
        "Transaction Code Description"
    ] = (
        output["Transaction Code"]
        .map(transaction_map)
        .fillna("")
    )

    output[
        "Report Currency"
    ] = REPORT_CURRENCY

    output["Count"] = (
        output["Count"]
        .astype("int64")
    )

    output = (
        output[
            OUTPUT_COLUMNS
        ]
        .sort_values(
            group_key,
            kind="stable",
        )
        .reset_index(
            drop=True
        )
    )

    if (
        list(output.columns)
        != OUTPUT_COLUMNS
    ):
        raise AssertionError(
            "GL14 output columns do not "
            "match the required matrix."
        )

    print(
        "GL14 timing - GL Master "
        "Data enrichment: "
        f"{perf_counter() - enrichment_started:.2f} "
        "seconds"
    )

    return output


def _output_file(context):
    return (
        Path(
            context["output_folder"]
        )
        / (
            "LBR_Results_GL_"
            f"{_period_suffix(context)}"
            ".xlsx"
        )
    )


def _write(
    context,
    dataframe,
):
    started = perf_counter()

    output_file = _output_file(
        context
    )

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if (
        not output_file.exists()
        and importlib.util.find_spec(
            "xlsxwriter"
        )
        is not None
    ):
        with pd.ExcelWriter(
            output_file,
            engine="xlsxwriter",
        ) as writer:
            dataframe.to_excel(
                writer,
                sheet_name=SHEET_NAME,
                index=False,
            )

            workbook = writer.book
            worksheet = (
                writer.sheets[
                    SHEET_NAME
                ]
            )

            header_format = (
                workbook.add_format(
                    {
                        "bold": True,
                        "bg_color": (
                            f"#{HEADER_FILL}"
                        ),
                    }
                )
            )

            amount_format = (
                workbook.add_format(
                    {
                        "num_format": (
                            AMOUNT_FORMAT
                        )
                    }
                )
            )

            integer_format = (
                workbook.add_format(
                    {
                        "num_format": (
                            INTEGER_FORMAT
                        )
                    }
                )
            )

            for (
                column_index,
                column_name,
            ) in enumerate(
                dataframe.columns
            ):
                worksheet.write(
                    0,
                    column_index,
                    column_name,
                    header_format,
                )

                if (
                    column_name
                    in AMOUNT_COLUMNS
                ):
                    cell_format = (
                        amount_format
                    )
                elif (
                    column_name
                    in INTEGER_COLUMNS
                ):
                    cell_format = (
                        integer_format
                    )
                else:
                    cell_format = None

                worksheet.set_column(
                    column_index,
                    column_index,
                    COLUMN_WIDTHS.get(
                        column_name,
                        18,
                    ),
                    cell_format,
                )

            worksheet.autofilter(
                0,
                0,
                max(
                    len(dataframe),
                    1,
                ),
                len(dataframe.columns) - 1,
            )

            worksheet.freeze_panes(
                1,
                0,
            )
    else:
        temporary_file = (
            Path(
                tempfile.gettempdir()
            )
            / f"gl14_{output_file.name}"
        )

        try:
            if output_file.exists():
                shutil.copy2(
                    output_file,
                    temporary_file,
                )

                workbook = load_workbook(
                    temporary_file
                )
            else:
                workbook = Workbook()

                workbook.remove(
                    workbook.active
                )

            if (
                SHEET_NAME
                in workbook.sheetnames
            ):
                del workbook[
                    SHEET_NAME
                ]

            worksheet = (
                workbook.create_sheet(
                    SHEET_NAME
                )
            )

            worksheet.append(
                list(
                    dataframe.columns
                )
            )

            for row in dataframe.itertuples(
                index=False,
                name=None,
            ):
                worksheet.append(
                    [
                        (
                            None
                            if pd.isna(value)
                            else value
                        )
                        for value in row
                    ]
                )

            header_fill = PatternFill(
                fill_type="solid",
                fgColor=HEADER_FILL,
            )

            for cell in worksheet[1]:
                cell.font = Font(
                    bold=True
                )
                cell.fill = (
                    header_fill
                )

            worksheet.freeze_panes = (
                "A2"
            )

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            for (
                column_index,
                column_name,
            ) in enumerate(
                dataframe.columns,
                start=1,
            ):
                column_letter = (
                    get_column_letter(
                        column_index
                    )
                )

                worksheet.column_dimensions[
                    column_letter
                ].width = COLUMN_WIDTHS.get(
                    column_name,
                    18,
                )

                if (
                    column_name
                    in AMOUNT_COLUMNS
                ):
                    number_format = (
                        AMOUNT_FORMAT
                    )
                elif (
                    column_name
                    in INTEGER_COLUMNS
                ):
                    number_format = (
                        INTEGER_FORMAT
                    )
                else:
                    number_format = None

                if number_format is not None:
                    for row_index in range(
                        2,
                        worksheet.max_row + 1,
                    ):
                        worksheet.cell(
                            row=row_index,
                            column=column_index,
                        ).number_format = (
                            number_format
                        )

            workbook.save(
                temporary_file
            )

            shutil.copy2(
                temporary_file,
                output_file,
            )

        except PermissionError as error:
            raise PermissionError(
                "Could not save output "
                f"workbook: {output_file}. "
                "Close the workbook and "
                "run again."
            ) from error

        finally:
            if temporary_file.exists():
                temporary_file.unlink()

    print(
        f"GL14 output file: "
        f"{output_file}"
    )

    print(
        f"GL14 output sheet: "
        f"{SHEET_NAME}"
    )

    print(
        f"GL14 output rows: "
        f"{len(dataframe)}"
    )

    print(
        "GL14 timing - write: "
        f"{perf_counter() - started:.2f} "
        "seconds"
    )


def run_gl_014(context):
    started = perf_counter()
    detection_started = (
        perf_counter()
    )

    bsis_path = _find_file(
        context,
        BSIS_KEYWORDS,
    )

    bsas_path = _find_file(
        context,
        BSAS_KEYWORDS,
    )

    print(
        "GL14 timing - input "
        "detection: "
        f"{perf_counter() - detection_started:.2f} "
        "seconds"
    )

    if (
        bsis_path is None
        and bsas_path is None
    ):
        raise FileNotFoundError(
            "GL14 requires at least one "
            "BSIS or BSAS input file."
        )

    bsis_dataframe = _load_source(
        context,
        "BSIS",
        BSIS_KEYWORDS,
        bsis_path,
    )

    bsas_dataframe = _load_source(
        context,
        "BSAS",
        BSAS_KEYWORDS,
        bsas_path,
    )

    optional_started = (
        perf_counter()
    )

    master_dataframe = _load_optional(
        context,
        MASTER_KEYWORDS,
        "GL Master Data",
    )

    fx_dataframe = _load_optional(
        context,
        FX_KEYWORDS,
        "FxRates",
    )

    print(
        "GL14 timing - read GL Master "
        "Data and FxRates: "
        f"{perf_counter() - optional_started:.2f} "
        "seconds"
    )

    output = create_gl14_summary(
        bsis_dataframe,
        bsas_dataframe,
        master_dataframe,
        fx_dataframe,
        context,
    )

    _write(
        context,
        output,
    )

    print(
        "GL14 total seconds: "
        f"{perf_counter() - started:.2f}"
    )


__all__ = [
    "create_gl14_summary",
    "run_gl_014",
]
