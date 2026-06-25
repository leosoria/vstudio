"""
Common CD utilities.

This module contains reusable functions for Cash Disbursements controls.

Design rules:
- The CD input file is a module-level base extract, not a control-specific file.
- The input file name is resolved from the CD module PARAM1 in config.xlsx.
- Every CD control must be independent.
- Every CD control must write only its own sheet.
- No CD control must delete sheets from other controls.
- The final CD output workbook follows:
    output/LBR_Results_CD_YYYYMMDD.xlsx
"""

import warnings
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


ALLOWED_EXTENSIONS = [".xlsx", ".xls", ".csv", ".txt"]

CD_DEFAULT_INPUT_KEYWORD = "LBR Cash"
CD_HEADER_FILL = "D9EAF7"
CD_DATE_FORMAT = "dd/mm/yyyy"
CD_AMOUNT_FORMAT = '#,##0.00'
CD_INTEGER_FORMAT = '#,##0'


def has_allowed_extension(file_path):
    """
    Return True if the file extension is supported.
    """
    file_path = Path(file_path)

    return file_path.suffix.lower() in ALLOWED_EXTENSIONS


def normalize_text(value):
    """
    Normalize text values.
    """
    if value is None:
        return ""

    value_text = str(value).strip()

    if value_text.lower() == "nan":
        return ""

    return value_text


def normalize_header(value):
    """
    Normalize a header for matching.
    """
    return normalize_text(value).lower()


def normalize_company_output(value):
    """
    Normalize company code for output display.

    Examples:
    - 0030 -> 30
    - 0034 -> 34
    - 0052 -> 52
    - 0093 -> 93
    """
    value_text = normalize_text(value)

    if value_text == "":
        return ""

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if value_text.isdigit():
        return str(int(value_text))

    return value_text


def parse_companies_filter(value):
    """
    Parse module COMPANIES filter into normalized values.

    Accepted values:
    - ALL means no filter.
    - Blank means no filter.
    - 0030,0034,0052
    - 0030;0034;0052
    - 0030|0034|0052
    - 0030 0034 0052
    - 30 34 52

    Notes:
    - The input file has one company per row in column Empr.
    - Values may come as 0030 or 30.
    """
    value_text = normalize_text(value)

    if value_text == "":
        return []

    if value_text.upper() in ["ALL", "TODAS", "TODOS"]:
        return []

    for separator in [";", "|", "\n", "\r", "\t"]:
        value_text = value_text.replace(separator, ",")

    # If there are no commas but there are spaces, treat spaces as separators.
    if "," not in value_text and " " in value_text:
        value_text = ",".join(value_text.split())

    companies = []

    for item in value_text.split(","):
        item = normalize_text(item)

        if item == "":
            continue

        if item.upper() in ["ALL", "TODAS", "TODOS"]:
            return []

        companies.append(normalize_company_output(item))

    return companies


def to_datetime_value(value):
    """
    Convert a value to pandas datetime.

    Handles:
    - Excel dates
    - Python datetime values
    - ISO dates like 2026-02-28
    - Local dates like 28/02/2026
    - SAP/Portuguese dates like 17/12/2025
    """
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NaT

    if len(value_text) >= 10 and value_text[4:5] == "-" and value_text[7:8] == "-":
        return pd.to_datetime(value_text, errors="coerce", dayfirst=False)

    return pd.to_datetime(value_text, errors="coerce", dayfirst=True)


def parse_number(value):
    """
    Parse a number from SAP/Excel text.

    Handles examples:
    - 1048.46
    - 1,048.46
    - 1.048,46
    - 1048,46
    - blank
    """
    if pd.isna(value):
        return pd.NA

    if isinstance(value, (int, float)):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NA

    value_text = value_text.replace(" ", "")

    # Parentheses as negative.
    is_negative = value_text.startswith("(") and value_text.endswith(")")

    if is_negative:
        value_text = value_text[1:-1]

    # If both separators exist, infer decimal separator from the last one.
    if "," in value_text and "." in value_text:
        last_comma = value_text.rfind(",")
        last_dot = value_text.rfind(".")

        if last_comma > last_dot:
            # Brazilian/European: 1.048,46
            value_text = value_text.replace(".", "")
            value_text = value_text.replace(",", ".")
        else:
            # US/Excel: 1,048.46
            value_text = value_text.replace(",", "")
    elif "," in value_text and "." not in value_text:
        # Treat comma as decimal separator.
        value_text = value_text.replace(",", ".")

    try:
        parsed_number = float(value_text)
    except ValueError:
        return pd.NA

    if is_negative:
        parsed_number = parsed_number * -1

    return parsed_number


def get_period_suffix(module_config):
    """
    Return YYYYMMDD suffix using module TO date.
    """
    to_date = module_config.get("to", "")
    parsed_date = to_datetime_value(to_date)

    if pd.isna(parsed_date):
        raise ValueError(
            "Could not determine CD output period because module TO date is empty or invalid."
        )

    return parsed_date.strftime("%Y%m%d")


def get_cd_output_file(context):
    """
    Return final CD output workbook path.

    Pattern:
        output/LBR_Results_CD_YYYYMMDD.xlsx
    """
    output_folder = Path(context["output_folder"])
    module_config = context["module"]
    period_suffix = get_period_suffix(module_config)

    return output_folder / f"LBR_Results_CD_{period_suffix}.xlsx"


def find_all_files_containing(base_folder, text_to_find):
    """
    Find all supported files under base_folder whose filename contains text_to_find.
    """
    base_folder = Path(base_folder)
    text_to_find = normalize_text(text_to_find).lower()

    matched_files = []

    for file_path in base_folder.rglob("*"):
        if not file_path.is_file():
            continue

        if not has_allowed_extension(file_path):
            continue

        if text_to_find in file_path.name.lower():
            matched_files.append(file_path)

    return sorted(matched_files)


def find_cd_base_input_file(context):
    """
    Find the module-level CD input file.

    Search keyword priority:
    1. CD module PARAM1 in config.xlsx.
    2. Default keyword: LBR Cash.

    This allows the same input file to be used by CD001, CD002, CD003 and CD004.
    """
    input_folder = Path(context["input_folder"])
    module_config = context["module"]

    keyword = normalize_text(module_config.get("param1", ""))

    if keyword == "":
        keyword = CD_DEFAULT_INPUT_KEYWORD

    matched_files = find_all_files_containing(
        base_folder=input_folder,
        text_to_find=keyword,
    )

    if len(matched_files) == 0:
        raise FileNotFoundError(
            f"CD input file not found. "
            f"Expected one file in input/ containing '{keyword}' in the filename. "
            f"Input folder: {input_folder}"
        )

    if len(matched_files) > 1:
        matched_files_text = "\n".join(
            f"- {file_path}"
            for file_path in matched_files
        )

        raise ValueError(
            f"Multiple CD input files found using keyword '{keyword}'. "
            f"Please make MODULE PARAM1 more specific in config.xlsx.\n"
            f"{matched_files_text}"
        )

    return matched_files[0]


def read_input_file(file_path, sheet_name=0):
    """
    Read input file into a dataframe.

    Supported:
    - xlsx
    - xls
    - csv
    - txt
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    extension = file_path.suffix.lower()

    if extension in [".xlsx", ".xls"]:
        return pd.read_excel(file_path, sheet_name=sheet_name)

    if extension == ".csv":
        return pd.read_csv(file_path)

    if extension == ".txt":
        return pd.read_csv(file_path, sep="\t")

    raise ValueError(f"Unsupported input file extension: {file_path.suffix}")


def clean_dataframe(dataframe):
    """
    Clean dataframe columns and empty rows/columns.
    """
    result = dataframe.copy()

    result = result.dropna(axis=0, how="all")
    result = result.dropna(axis=1, how="all")

    result.columns = [
        normalize_text(column)
        for column in result.columns
    ]

    return result


def find_column(dataframe, possible_names):
    """
    Find a dataframe column using exact case-insensitive matching.
    """
    normalized_lookup = {
        normalize_header(column): column
        for column in dataframe.columns
    }

    for possible_name in possible_names:
        normalized_name = normalize_header(possible_name)

        if normalized_name in normalized_lookup:
            return normalized_lookup[normalized_name]

    return None


def require_columns(dataframe, required_columns):
    """
    Validate required columns.

    required_columns example:
    {
        "company_code": ["Empr"],
        "vendor_code": ["Fornecedor"],
    }

    Returns:
    {
        "company_code": "Empr",
        "vendor_code": "Fornecedor",
    }
    """
    resolved_columns = {}
    missing_columns = []

    for logical_name, possible_names in required_columns.items():
        column_name = find_column(dataframe, possible_names)

        if column_name is None:
            missing_columns.append(
                f"{logical_name}: expected one of {possible_names}"
            )
        else:
            resolved_columns[logical_name] = column_name

    if missing_columns:
        raise ValueError(
            "Missing required CD input columns:\n- "
            + "\n- ".join(missing_columns)
        )

    return resolved_columns


def get_optional_column(dataframe, possible_names):
    """
    Return a column if found, otherwise None.
    """
    return find_column(dataframe, possible_names)


def get_series_or_blank(dataframe, column_name):
    """
    Return dataframe column if available, otherwise blank string.
    """
    if column_name is None:
        return ""

    return dataframe[column_name]


def filter_by_company(dataframe, company_column, companies_filter):
    """
    Filter dataframe by module companies.

    If companies_filter is blank, returns all rows.
    """
    result = dataframe.copy()
    companies = parse_companies_filter(companies_filter)

    if not companies:
        return result

    normalized_company = result[company_column].apply(normalize_company_output)

    return result[normalized_company.isin(companies)].copy()


def apply_signed_amount(amount_value, debit_credit_indicator):
    """
    Apply CD sign logic.

    Rule for this process:
    - H: positive
    - S: negative
    - Other: keep as original
    """
    amount = parse_number(amount_value)
    indicator = normalize_text(debit_credit_indicator).upper()

    if pd.isna(amount):
        return pd.NA

    amount = abs(float(amount))

    if indicator == "S":
        return amount * -1

    return amount


def describe_debit_credit_indicator(value):
    """
    Return a friendly description for SHKZG / D-C.
    """
    indicator = normalize_text(value).upper()

    if indicator == "H":
        return "H - Payment / Positive"

    if indicator == "S":
        return "S - Reversal / Negative"

    return "Other / Review"


def build_key(*values):
    """
    Build a pipe-separated key from values.
    """
    cleaned_values = [
        normalize_text(value)
        for value in values
    ]

    return "|".join(cleaned_values)


def load_cd_base_data(context):
    """
    Load and normalize the common CD base input.

    This function should be reused by:
    - CD001
    - CD002
    - CD003
    - CD004
    """
    input_file = find_cd_base_input_file(context)

    print(f"CD base input file: {input_file}")

    dataframe = read_input_file(input_file)
    dataframe = clean_dataframe(dataframe)

    return dataframe


def open_or_create_cd_output_workbook(output_file):
    """
    Open the final CD output workbook if it exists.
    Otherwise create a new workbook.
    """
    output_file = Path(output_file)

    if output_file.exists():
        return load_workbook(output_file)

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def recreate_cd_sheet(workbook, sheet_name):
    """
    Recreate one CD sheet.

    Rules:
    - Delete only the target sheet if it exists.
    - Do not delete sheets from other controls.
    """
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    return workbook.create_sheet(sheet_name)


def save_cd_output_workbook(workbook, output_file):
    """
    Save the final CD workbook.

    If the output is open in Excel, raise a clear message.
    """
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook.save(output_file)
    except PermissionError as error:
        raise PermissionError(
            f"Could not save output workbook: {output_file}. "
            "The file may be open in Excel or locked by OneDrive. "
            "Close the workbook and run again."
        ) from error


def write_dataframe_to_sheet(worksheet, dataframe):
    """
    Write a pandas dataframe to an openpyxl worksheet.
    """
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None

            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()

            worksheet.cell(row=row_index, column=column_index, value=value)


def apply_standard_cd_formatting(
    worksheet,
    dataframe,
    date_columns=None,
    amount_columns=None,
    integer_columns=None,
):
    """
    Apply standard visual formatting used by CD controls.

    Pattern:
    - Headers in bold.
    - Fill fgColor D9EAF7.
    - Autofilter.
    - Freeze panes A2.
    - Column width based on content.
    - Date and amount formats by column name.
    """
    date_columns = set(date_columns or [])
    amount_columns = set(amount_columns or [])
    integer_columns = set(integer_columns or [])

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=CD_HEADER_FILL,
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    column_name_by_index = {
        index + 1: column_name
        for index, column_name in enumerate(dataframe.columns)
    }

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        column_name = column_name_by_index.get(column_index, "")

        max_length = len(str(column_name))

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

            if cell.row == 1:
                continue

            if column_name in date_columns:
                cell.number_format = CD_DATE_FORMAT

            if column_name in amount_columns:
                cell.number_format = CD_AMOUNT_FORMAT

            if column_name in integer_columns:
                cell.number_format = CD_INTEGER_FORMAT

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)
