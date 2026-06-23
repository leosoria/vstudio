"""
AR_006 - Salesforce credit limit change history.

This control creates/replaces only the AR06 sheet in the shared AR output file.

Logic for AR6:
- Load Account History, Customer, Salesforce current credit limits, and FX Rates input data.
- Exclude intercompany customers.
- Build Salesforce ERP customer code -> SAP customer map using Customer Grupo -> Cliente.
- Read Account History rows where Campo/Compromisso = Limite Aprovado.
- Keep only rows with Data de edição between module FROM and module TO dates.
- Parse old/new/current BRL credit limits using VBA-compatible money parsing.
- Replace current limit with Salesforce current limit when status is not Bloqueado or Expirado.
- Use monthly FX rate from FX Rates for update month and current month.
- Fallback to the common BRL->USD FX rate if a monthly FX rate is not available.
- Calculate deltas, change percentages, and change type.
- Write/recreate only the AR06 sheet in the shared output workbook.
- Apply visual filters
  Change Type is Increase or Decrease.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters

from core.ar_common import (
    build_intercompany_exclusion_set,
    get_ar_output_file,
    load_ar_input_data,
    normalize_company_output,
    normalize_customer_key,
    normalize_erp_code,
    open_or_create_ar_output_workbook,
    recreate_ar_sheet,
    save_ar_output_workbook,
    to_datetime_value,
    to_number,
    validate_required_columns,
)


AR06_SHEET_NAME = "AR06"

AR06_HEADERS = [
    "Company",
    "Customer Code",
    "Customer Name",
    "Log Instance",
    "Update Date",
    "Edited By",
    "Create Date",
    "Main Currency",
    "System Currency",
    "Prev Credit Limit Main",
    "New Credit Limit Main",
    "Current Credit Limit Main",
    "Update Rate Date",
    "Update Rate",
    "Current Rate Date",
    "Current Rate",
    "Prev Credit Limit USD",
    "New Credit Limit USD",
    "Current Credit Limit USD",
    "Delta Main (New-Prev)",
    "Delta Main (Current-New)",
    "Change % Main",
    "Change % USD",
    "Change Type",
]

ACCOUNT_HISTORY_REQUIRED_COLUMNS = [
    "Nome da conta",
    "Campo/Compromisso",
    "Valor antigo",
    "Novo valor",
    "Data de edição",
    "Editado por",
    "Data de criação",
    "Código ERP",
    "Limite Aprovado",
]

CUSTOMER_REQUIRED_COLUMNS = [
    "Cliente",
    "Grupo",
]

SALESFORCE_CURRENT_REQUIRED_COLUMNS = [
    "Código ERP",
    "Limite Aprovado",
    "Status do Limite",
]

FX_RATES_REQUIRED_COLUMNS = [
    "MONTH",
    "SPOT - BS",
]

EXCLUDED_SALESFORCE_LIMIT_STATUSES = [
    "BLOQUEADO",
    "EXPIRADO",
]


def run_ar_006(context):
    """
    Execute AR_006.

    This function:
    - Loads common AR input data.
    - Loads AR06-specific Account History and monthly FX Rates inputs.
    - Builds AR06 output rows.
    - Opens or creates the shared AR output workbook.
    - Recreates only the AR06 sheet.
    - Saves the shared output workbook.
    """
    input_data = context.get("ar_input_data")

    if input_data is None:
        input_data = load_ar_input_data(context)

    input_data = ensure_ar06_specific_input_data(input_data, context)

    output_rows = build_ar06_rows(input_data, context)

    output_file = get_ar_output_file(context)
    workbook = open_or_create_ar_output_workbook(output_file)
    worksheet = recreate_ar_sheet(workbook, AR06_SHEET_NAME)

    write_headers(worksheet)
    write_rows(worksheet, output_rows)
    format_worksheet(worksheet)

    save_ar_output_workbook(workbook, output_file)

    print(f"AR_006 completed. Rows written: {len(output_rows)}")
    print(f"Output file: {output_file}")


def ensure_ar06_specific_input_data(input_data, context):
    """
    Add AR06-specific inputs to the common AR input data dictionary.

    Existing load_ar_input_data(context) loads the common files used by AR_001
    through AR_005, but AR_006 also needs:
    - Account History_YYYYMMDD
    - Exchange Rates FY*_YYYYMMDD or another monthly FX file with MONTH / SPOT - BS

    This function keeps AR_006 independent and avoids requiring changes in
    core/ar_common.py.
    """
    input_data = dict(input_data)

    period_suffix = input_data.get("period_suffix") or get_period_suffix_from_context(
        context
    )

    if "account_history" not in input_data:
        account_history_file = find_first_input_file(
            context["input_folder"],
            ["Account History", period_suffix],
        )

        if account_history_file is None:
            raise FileNotFoundError(
                "AR_006 Account History file was not found. "
                f"Expected file name containing: Account History and {period_suffix}"
            )

        input_data["account_history_file"] = account_history_file
        input_data["account_history"] = read_salesforce_like_file(
            account_history_file,
        )

    if "fx_rate_monthly" not in input_data:
        fx_rate_monthly_file = find_first_input_file(
            context["input_folder"],
            ["Exchange Rates", period_suffix],
        )

        if fx_rate_monthly_file is None:
            fx_rate_monthly_file = find_first_input_file(
                context["input_folder"],
                ["FX Rates", period_suffix],
            )

        if fx_rate_monthly_file is None:
            raise FileNotFoundError(
                "AR_006 monthly FX Rates file was not found. "
                "Expected file name containing Exchange Rates or FX Rates and "
                f"{period_suffix}"
            )

        input_data["fx_rate_monthly_file"] = fx_rate_monthly_file
        input_data["fx_rate_monthly"] = read_monthly_fx_rates_file(
            fx_rate_monthly_file,
        )

    return input_data


def build_ar06_rows(input_data, context):
    """
    Build AR06 output rows.
    """
    account_history_df = input_data["account_history"]
    customer_df = input_data["customer"]
    sf_current_df = input_data["salesforce_current"]
    fx_rate_monthly_df = input_data["fx_rate_monthly"]

    validate_ar006_inputs(
        account_history_df=account_history_df,
        customer_df=customer_df,
        sf_current_df=sf_current_df,
        fx_rate_monthly_df=fx_rate_monthly_df,
    )

    module_config = context.get("module", {})
    company_value = get_company_value(module_config)
    date_from = get_module_date(module_config, "from")
    date_to = get_module_date(module_config, "to")
    fallback_rate = get_fallback_rate(input_data)

    excluded_customers = build_intercompany_exclusion_set()
    customer_to_sap_map = build_customer_code_to_sap_map(customer_df)
    salesforce_current_limit_map = build_ar06_salesforce_current_limit_map(
        sf_current_df,
    )
    salesforce_status_map = build_ar06_salesforce_status_map(sf_current_df)
    fx_rate_map = build_fx_rate_map(fx_rate_monthly_df)

    output_rows = []
    log_counter_map = {}

    for _, row in account_history_df.iterrows():
        field_name = clean_text(row.get("Campo/Compromisso", ""))

        if field_name != "Limite Aprovado":
            continue

        update_datetime = to_datetime_value(row.get("Data de edição", ""))

        if update_datetime is None or str(update_datetime) == "NaT":
            continue

        update_date = update_datetime.normalize()

        if update_date.date() < date_from.date():
            continue

        if update_date.date() > date_to.date():
            continue

        customer_code = normalize_erp_code(row.get("Código ERP", ""))

        if customer_code == "":
            continue

        sap_customer_key = customer_to_sap_map.get(customer_code, "")

        if sap_customer_key in excluded_customers:
            continue

        prev_main = parse_ar06_money(row.get("Valor antigo", ""))
        new_main = parse_ar06_money(row.get("Novo valor", ""))
        current_main = parse_ar06_money(row.get("Limite Aprovado", ""))

        if customer_code in salesforce_current_limit_map:
            current_status = salesforce_status_map.get(customer_code, "")

            if current_status.upper() not in EXCLUDED_SALESFORCE_LIMIT_STATUSES:
                current_main = salesforce_current_limit_map[customer_code]
            else:
                current_main = ""

        update_rate_date = update_date
        current_rate_date = date_to.normalize()

        update_month_key = month_key(update_rate_date)
        current_month_key = month_key(current_rate_date)

        update_rate = fx_rate_map.get(update_month_key, fallback_rate)
        current_rate = fx_rate_map.get(current_month_key, fallback_rate)

        prev_usd = ""
        new_usd = ""
        current_usd = ""

        if is_number(prev_main):
            prev_usd = prev_main * update_rate

        if is_number(new_main):
            new_usd = new_main * update_rate

        if is_number(current_main):
            current_usd = current_main * current_rate

        delta_new_prev = ""
        delta_current_new = ""
        change_main = ""
        change_usd = ""
        change_type = "No Change"

        if is_number(prev_main) and is_number(new_main):
            delta_new_prev = new_main - prev_main

            if new_main > prev_main:
                change_type = "Increase"
            elif new_main < prev_main:
                change_type = "Decrease"

            if prev_main != 0:
                change_main = delta_new_prev / prev_main * 100

        if is_number(current_main) and is_number(new_main):
            delta_current_new = current_main - new_main

        if is_number(prev_usd) and is_number(new_usd):
            if prev_usd != 0:
                change_usd = (new_usd - prev_usd) / prev_usd * 100

        if customer_code not in log_counter_map:
            log_counter_map[customer_code] = 1
        else:
            log_counter_map[customer_code] = log_counter_map[customer_code] + 1

        output_rows.append(
            {
                "Company": normalize_company_output(company_value),
                "Customer Code": customer_code,
                "Customer Name": clean_text(row.get("Nome da conta", "")),
                "Log Instance": log_counter_map[customer_code],
                "Update Date": update_rate_date.to_pydatetime(),
                "Edited By": clean_text(row.get("Editado por", "")),
                "Create Date": date_or_blank(row.get("Data de criação", "")),
                "Main Currency": "BRL",
                "System Currency": "USD",
                "Prev Credit Limit Main": prev_main,
                "New Credit Limit Main": new_main,
                "Current Credit Limit Main": current_main,
                "Update Rate Date": update_rate_date.to_pydatetime(),
                "Update Rate": update_rate,
                "Current Rate Date": current_rate_date.to_pydatetime(),
                "Current Rate": current_rate,
                "Prev Credit Limit USD": prev_usd,
                "New Credit Limit USD": new_usd,
                "Current Credit Limit USD": current_usd,
                "Delta Main (New-Prev)": delta_new_prev,
                "Delta Main (Current-New)": delta_current_new,
                "Change % Main": change_main,
                "Change % USD": change_usd,
                "Change Type": change_type,
            }
        )

    return output_rows


def validate_ar006_inputs(
    account_history_df,
    customer_df,
    sf_current_df,
    fx_rate_monthly_df,
):
    """
    Validate required input columns before processing.
    """
    missing_account_history_columns = validate_required_columns(
        account_history_df,
        ACCOUNT_HISTORY_REQUIRED_COLUMNS,
    )

    if missing_account_history_columns:
        raise ValueError(
            "AR_006 missing required columns in Account History data: "
            f"{missing_account_history_columns}"
        )

    missing_customer_columns = validate_required_columns(
        customer_df,
        CUSTOMER_REQUIRED_COLUMNS,
    )

    if missing_customer_columns:
        raise ValueError(
            "AR_006 missing required columns in Customer data: "
            f"{missing_customer_columns}"
        )

    missing_sf_current_columns = validate_required_columns(
        sf_current_df,
        SALESFORCE_CURRENT_REQUIRED_COLUMNS,
    )

    if missing_sf_current_columns:
        raise ValueError(
            "AR_006 missing required columns in Salesforce current data: "
            f"{missing_sf_current_columns}"
        )

    missing_fx_rate_columns = validate_required_columns(
        fx_rate_monthly_df,
        FX_RATES_REQUIRED_COLUMNS,
    )

    if missing_fx_rate_columns:
        raise ValueError(
            "AR_006 missing required columns in FX Rates data: "
            f"{missing_fx_rate_columns}"
        )


def build_customer_code_to_sap_map(customer_df):
    """
    Build Salesforce ERP customer code -> SAP customer number map.

    - Customer Grupo is the Salesforce / ERP customer code.
    - Customer Cliente is the SAP customer number used for intercompany exclusion.
    - Keep the first mapping found for each Grupo.
    """
    customer_to_sap_map = {}

    for _, row in customer_df.iterrows():
        customer_code_key = normalize_erp_code(row.get("Grupo", ""))
        sap_customer_number_key = normalize_customer_key(row.get("Cliente", ""))

        if customer_code_key == "":
            continue

        if customer_code_key not in customer_to_sap_map:
            customer_to_sap_map[customer_code_key] = sap_customer_number_key

    return customer_to_sap_map


def build_ar06_salesforce_current_limit_map(sf_current_df):
    """
    Build Salesforce current credit-limit map by ERP customer code.

    - Use first Código ERP occurrence only.
    - Parse Limite Aprovado with AR6_ParseMoney-compatible logic.
    - Do not filter status here; status is checked later when applying currentMain.
    """
    limit_map = {}

    for _, row in sf_current_df.iterrows():
        customer_code_key = normalize_erp_code(row.get("Código ERP", ""))

        if customer_code_key == "":
            continue

        if customer_code_key not in limit_map:
            limit_map[customer_code_key] = parse_ar06_money(
                row.get("Limite Aprovado", "")
            )

    return limit_map


def build_ar06_salesforce_status_map(sf_current_df):
    """
    Build Salesforce current status map by ERP customer code.

    - Use first Código ERP occurrence only.
    """
    status_map = {}

    for _, row in sf_current_df.iterrows():
        customer_code_key = normalize_erp_code(row.get("Código ERP", ""))

        if customer_code_key == "":
            continue

        if customer_code_key not in status_map:
            status_map[customer_code_key] = clean_text(
                row.get("Status do Limite", "")
            )

    return status_map


def build_fx_rate_map(fx_rate_monthly_df):
    """
    Build monthly FX map using MONTH -> SPOT - BS.

    Key:
        First day of each month as YYYY-MM-01 string.
    """
    fx_rate_map = {}

    for _, row in fx_rate_monthly_df.iterrows():
        fx_month = to_datetime_value(row.get("MONTH", ""))

        if fx_month is None or str(fx_month) == "NaT":
            continue

        fx_month_key = month_key(fx_month)

        if fx_month_key not in fx_rate_map:
            fx_rate_map[fx_month_key] = to_number(
                row.get("SPOT - BS", 0),
                default=0.0,
            )

    return fx_rate_map


def read_salesforce_like_file(file_path):
    """
    Read Account History / Salesforce-like file.

    Account History has Salesforce-style exports and may contain one or more
    leading rows before the actual header. This function detects the header row
    by searching for required AR06 Account History headers.
    """
    file_path = Path(file_path)

    if file_path.suffix.lower() in [".csv", ".txt"]:
        raw_df = pd.read_csv(
            file_path,
            header=None,
            dtype=object,
            sep=None,
            engine="python",
        )
    else:
        raw_df = pd.read_excel(
            file_path,
            header=None,
            dtype=object,
        )

    header_row_index = detect_header_row(
        raw_df,
        [
            "Nome da conta",
            "Campo/Compromisso",
            "Valor antigo",
            "Novo valor",
            "Data de edição",
            "Código ERP",
            "Limite Aprovado",
        ],
    )

    if header_row_index is None:
        raise ValueError(
            f"Could not detect Account History header row in file: {file_path}"
        )

    headers = raw_df.iloc[header_row_index].fillna("").astype(str).str.strip().tolist()
    data_df = raw_df.iloc[header_row_index + 1 :].copy()
    data_df.columns = headers
    data_df = data_df.dropna(how="all")
    data_df = data_df.loc[:, [column != "" for column in data_df.columns]]

    return data_df.reset_index(drop=True)


def read_monthly_fx_rates_file(file_path):
    """
    Read monthly FX Rates file.

    Supports:
    1. A normalized sheet/table with headers:
       COUNTRY | MONTH | SPOT - BS | Average - P&L
    2. The original Exchange Rates workbook where each relevant sheet is named
       FX Rates - Mon YY and contains a Brazil Real row.

    The output dataframe always contains:
    - COUNTRY
    - MONTH
    - SPOT - BS
    - Average - P&L
    """
    file_path = Path(file_path)

    if file_path.suffix.lower() in [".csv", ".txt"]:
        df = pd.read_csv(
            file_path,
            dtype=object,
            sep=None,
            engine="python",
        )

        if validate_required_columns(df, FX_RATES_REQUIRED_COLUMNS):
            raise ValueError(
                f"Monthly FX Rates file does not contain required columns: {file_path}"
            )

        return df

    excel_file = pd.ExcelFile(file_path)

    normalized_sheets = []

    for sheet_name in excel_file.sheet_names:
        raw_df = pd.read_excel(
            excel_file,
            sheet_name=sheet_name,
            header=None,
            dtype=object,
        )

        header_row_index = detect_header_row(
            raw_df,
            [
                "COUNTRY",
                "MONTH",
                "SPOT - BS",
                "Average - P&L",
            ],
        )

        if header_row_index is not None:
            headers = (
                raw_df.iloc[header_row_index]
                .fillna("")
                .astype(str)
                .str.strip()
                .tolist()
            )
            sheet_df = raw_df.iloc[header_row_index + 1 :].copy()
            sheet_df.columns = headers
            sheet_df = sheet_df.dropna(how="all")
            sheet_df = sheet_df.loc[:, [column != "" for column in sheet_df.columns]]

            if not validate_required_columns(sheet_df, FX_RATES_REQUIRED_COLUMNS):
                normalized_sheets.append(sheet_df)

    if normalized_sheets:
        return pd.concat(normalized_sheets, ignore_index=True)

    output_rows = []

    for sheet_name in excel_file.sheet_names:
        if not sheet_name.startswith("FX Rates -"):
            continue

        if "(PY)" in sheet_name.upper():
            continue

        fx_month = parse_fx_month_from_sheet_name(sheet_name)

        if fx_month is None:
            continue

        raw_df = pd.read_excel(
            excel_file,
            sheet_name=sheet_name,
            header=None,
            dtype=object,
        )

        header_row_index = detect_header_row(
            raw_df,
            [
                "COUNTRY",
                "MONTH",
                "SPOT - BS",
                "Average - P&L",
            ],
        )

        if header_row_index is None:
            continue

        headers = raw_df.iloc[header_row_index].fillna("").astype(str).str.strip()
        sheet_df = raw_df.iloc[header_row_index + 1 :].copy()
        sheet_df.columns = headers
        sheet_df = sheet_df.dropna(how="all")

        country_column = find_column_name(sheet_df, "COUNTRY")
        spot_column = find_column_name(sheet_df, "SPOT - BS")
        average_column = find_column_name(sheet_df, "Average - P&L")

        if country_column is None or spot_column is None:
            continue

        for _, row in sheet_df.iterrows():
            country = clean_text(row.get(country_column, ""))

            if country != "Brazil Real":
                continue

            output_rows.append(
                {
                    "COUNTRY": country,
                    "MONTH": fx_month,
                    "SPOT - BS": row.get(spot_column, ""),
                    "Average - P&L": (
                        row.get(average_column, "") if average_column else ""
                    ),
                }
            )

    return pd.DataFrame(
        output_rows,
        columns=[
            "COUNTRY",
            "MONTH",
            "SPOT - BS",
            "Average - P&L",
        ],
    )


def detect_header_row(raw_df, required_headers):
    """
    Return zero-based index of the row containing all required headers.
    """
    required_normalized = {
        normalize_header_name(header)
        for header in required_headers
    }

    for row_index in range(len(raw_df)):
        current_headers = {
            normalize_header_name(value)
            for value in raw_df.iloc[row_index].tolist()
            if clean_text(value) != ""
        }

        if required_normalized.issubset(current_headers):
            return row_index

    return None


def find_column_name(dataframe, header_name):
    """
    Find actual dataframe column name matching a header case-insensitively.
    """
    expected_header = normalize_header_name(header_name)

    for column in dataframe.columns:
        if normalize_header_name(column) == expected_header:
            return column

    return None


def normalize_header_name(value):
    """
    Normalize header text for comparisons.
    """
    return clean_text(value).upper()


def find_first_input_file(input_folder, required_parts):
    """
    Find first file in input_folder whose name contains all required parts.
    """
    input_folder = Path(input_folder)

    for file_path in sorted(input_folder.iterdir()):
        if not file_path.is_file():
            continue

        file_name_upper = file_path.name.upper()

        if all(str(part).upper() in file_name_upper for part in required_parts):
            if file_path.suffix.lower() in ["", ".xlsx", ".xls", ".csv", ".txt"]:
                return file_path

    return None


def get_period_suffix_from_context(context):
    """
    Return module TO date as YYYYMMDD.
    """
    module_config = context.get("module", {})
    date_to = get_module_date(module_config, "to")

    return date_to.strftime("%Y%m%d")


def get_company_value(module_config):
    """
    Return module company value for output.

    If config has several companies, keep the raw value normalized by
    normalize_company_output at row-write time. For the usual single-company
    execution this converts 0030 -> 30, 0034 -> 34, 0052 -> 52.
    """
    possible_keys = [
        "company",
        "Company",
        "companies",
        "Companies",
        "COMPANY",
        "COMPANIES",
    ]

    for key in possible_keys:
        if key in module_config:
            return module_config[key]

    return ""


def get_module_date(module_config, date_key):
    """
    Return module date from context["module"].

    date_key should be:
    - "from"
    - "to"
    """
    possible_keys = [
        date_key,
        date_key.upper(),
        date_key.capitalize(),
        f"date_{date_key}",
        f"Date {date_key}",
        f"DATE {date_key.upper()}",
    ]

    for key in possible_keys:
        if key in module_config:
            date_value = to_datetime_value(module_config[key])

            if date_value is not None and str(date_value) != "NaT":
                return date_value.normalize()

    raise ValueError(
        f"AR_006 requires module {date_key.upper()} date in config.xlsx."
    )


def get_fallback_rate(input_data):
    """
    Return fallback FX rate.

    Parameters LBR FXRATE.
    Python fallback priority:
    1. input_data['fx_rate_brl_to_usd'] from load_ar_input_data(context)
    2. input_data['fx_rate_details']['final_fx_rate']
    3. 0.0
    """
    fallback_rate = to_number(
        input_data.get("fx_rate_brl_to_usd", 0),
        default=0.0,
    )

    if fallback_rate != 0:
        return fallback_rate

    fx_rate_details = input_data.get("fx_rate_details", {})

    if isinstance(fx_rate_details, dict):
        return to_number(
            fx_rate_details.get("final_fx_rate", 0),
            default=0.0,
        )

    return 0.0


def month_key(date_value):
    """
    Return first day of month key as YYYY-MM-01 string.
    """
    date_value = to_datetime_value(date_value)

    return date_value.replace(day=1).strftime("%Y-%m-%d")


def parse_fx_month_from_sheet_name(sheet_name):
    """
    Parse month from sheet names like:
    - FX Rates - Mar 26
    - FX Rates - Mar-26
    - FX Rates - March 2026
    """
    month_text = sheet_name.replace("FX Rates -", "").strip()
    month_text = month_text.replace("-", " ")
    month_parts = month_text.split()

    if len(month_parts) < 2:
        return None

    month_name = month_parts[0][:3].upper()
    year_text = month_parts[1]

    month_map = {
        "JAN": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "SEP": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
    }

    if month_name not in month_map:
        return None

    try:
        year = int(float(year_text))
    except ValueError:
        return None

    if year < 100:
        year = 2000 + year

    return pd.Timestamp(year=year, month=month_map[month_name], day=1)


def parse_ar06_money(value):
    """
    Parse amount values Returns:
    - float when value can be parsed as numeric
    - "" when source is blank or cannot be parsed

    Handles examples:
    - R$ 100.000,00 -> 100000.0
    - R$5,000,000.00 -> 5000000.0
    - BRL 1.234,56 -> 1234.56
    - blank -> ""
    """
    if is_blank(value):
        return ""

    if isinstance(value, (int, float)):
        return float(value)

    money_text = str(value).strip()
    money_text = money_text.replace("R$", "")
    money_text = money_text.replace("BRL", "")
    money_text = money_text.replace(" ", "")
    money_text = money_text.replace("\xa0", "")

    if money_text == "":
        return ""

    comma_position = money_text.rfind(",")
    dot_position = money_text.rfind(".")

    if comma_position >= 0 and dot_position >= 0:
        if comma_position > dot_position:
            money_text = money_text.replace(".", "")
            money_text = money_text.replace(",", ".")
        else:
            money_text = money_text.replace(",", "")
    elif comma_position >= 0:
        if len(money_text) - comma_position - 1 <= 2:
            money_text = money_text.replace(",", ".")
        else:
            money_text = money_text.replace(",", "")
    elif dot_position >= 0:
        if len(money_text) - dot_position - 1 > 2:
            money_text = money_text.replace(".", "")

    try:
        return float(money_text)
    except ValueError:
        return ""


def date_or_blank(value):
    """
    Convert value to Python datetime without time, or blank.
    """
    date_value = to_datetime_value(value)

    if date_value is None or str(date_value) == "NaT":
        return ""

    return date_value.normalize().to_pydatetime()


def write_headers(worksheet):
    """
    Write AR06 headers.
    """
    for column_index, header in enumerate(AR06_HEADERS, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


def write_rows(worksheet, ar06_rows):
    """
    Write AR06 rows.
    """
    for row_index, row_data in enumerate(ar06_rows, start=2):
        for column_index, header in enumerate(AR06_HEADERS, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=row_data.get(header, ""),
            )


def format_worksheet(worksheet):
    """
    Apply AR06 worksheet formatting.
    """
    date_columns = [
        "Update Date",
        "Create Date",
        "Update Rate Date",
        "Current Rate Date",
    ]

    money_columns = [
        "Prev Credit Limit Main",
        "New Credit Limit Main",
        "Current Credit Limit Main",
        "Prev Credit Limit USD",
        "New Credit Limit USD",
        "Current Credit Limit USD",
        "Delta Main (New-Prev)",
        "Delta Main (Current-New)",
    ]

    rate_columns = [
        "Update Rate",
        "Current Rate",
    ]

    percent_columns = [
        "Change % Main",
        "Change % USD",
    ]

    integer_columns = [
        "Log Instance",
    ]

    text_columns = [
        "Company",
        "Customer Code",
    ]

    for header in date_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "dd/mm/yyyy"

    for header in money_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "#,##0.00"

    for header in rate_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0.000000"

    for header in percent_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0.00"

    for header in integer_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0"

    for header in text_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "@"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    apply_change_type_filter(worksheet)


def apply_change_type_filter(worksheet):
    """
    Apply the AR06 visual filter for Change Type = Increase or Decrease.

    Important OpenPyXL detail:
    - OpenPyXL can write the autofilter criteria to the file.
    - OpenPyXL does not calculate/apply the filter result like Excel does.
    - To make the workbook open already filtered, rows that do not match
      the criteria must also be hidden explicitly.
    """
    if worksheet.max_row <= 1:
        return

    change_type_column_index = get_header_column_index(
        worksheet,
        "Change Type",
    )

    filter_column = FilterColumn(colId=change_type_column_index - 1)
    filter_column.filters = Filters(filter=["Increase", "Decrease"])

    worksheet.auto_filter.filterColumn.append(filter_column)

    for row_index in range(2, worksheet.max_row + 1):
        change_type_value = worksheet.cell(
            row=row_index,
            column=change_type_column_index,
        ).value

        worksheet.row_dimensions[row_index].hidden = (
            clean_text(change_type_value) not in ["Increase", "Decrease"]
        )


def get_header_column_index(worksheet, header_name):
    """
    Return one-based column index for a header in row 1.
    """
    expected_header = str(header_name).strip().upper()

    for column_index in range(1, worksheet.max_column + 1):
        current_header = worksheet.cell(row=1, column=column_index).value
        current_header = (
            ""
            if current_header is None
            else str(current_header).strip().upper()
        )

        if current_header == expected_header:
            return column_index

    raise ValueError(
        f"Column not found in sheet '{worksheet.title}': {header_name}"
    )


def clean_text(value):
    """
    Return stripped text, treating pandas-like blanks as empty string.
    """
    if is_blank(value):
        return ""

    return str(value).strip()


def is_blank(value):
    """
    Return True for None, empty string, or pandas-like NaN text.
    """
    if value is None:
        return True

    value_text = str(value).strip()

    if value_text == "":
        return True

    if value_text.lower() in ["nan", "nat", "none"]:
        return True

    return False


def is_number(value):
    """
    Return True only for numeric values.
    """
    return isinstance(value, (int, float))
