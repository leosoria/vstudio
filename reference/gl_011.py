"""GL_011 - Potential Duplicate General Journals: Same Description And Amount."""

import posixpath
import shutil
import tempfile
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from time import perf_counter
from xml.etree import ElementTree
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.gl_common import (
    build_company_name_map,
    build_gl_account_name_map,
    filter_by_company,
    get_gl_output_file,
    get_optional_column,
    load_gl_bsas_data,
    load_gl_bsis_data,
    load_gl_fx_rates_data,
    load_gl_master_data,
    normalize_company_output,
    normalize_fx_rates,
    open_or_create_gl_output_workbook,
    recreate_gl_sheet,
    require_columns,
    save_gl_output_workbook,
    select_fx_rate_to_usd,
    to_datetime_value,
    write_dataframe_to_sheet,
    write_single_sheet_workbook_fast,
)

SHEET_NAME = "GL11"
REPORT_CURRENCY = "USD"

REQUIRED_COLUMNS = {
    "company_code": ["Empr", "Empr.1", "BUKRS", "BKPF-BUKRS", "BSIS-BUKRS", "BSAS-BUKRS"],
    "fiscal_year": ["Ano", "GJAHR", "BKPF-GJAHR", "BSIS-GJAHR", "BSAS-GJAHR"],
    "document_number": ["Nº doc.", "Nº doc..1", "Nº doc", "BELNR", "BKPF-BELNR", "BSIS-BELNR", "BSAS-BELNR"],
    "line_number": ["Itm", "BUZEI", "BSIS-BUZEI", "BSAS-BUZEI"],
    "gl_account": ["Razão", "Razao", "Cta.Razão", "Cta.Razão.1", "HKONT", "BSIS-HKONT", "BSAS-HKONT"],
    "debit_credit": ["D/C", "SHKZG", "BSIS-SHKZG", "BSAS-SHKZG"],
    "amount_local": ["Montante em MI", "DMBTR", "BSIS-DMBTR", "BSAS-DMBTR"],
    "amount_document": ["Montante", "WRBTR", "BSIS-WRBTR", "BSAS-WRBTR"],
    "document_currency": ["Moeda", "Moeda.1", "WAERS", "BKPF-WAERS"],
    "document_date": ["Data doc.", "Data doc", "BLDAT", "BKPF-BLDAT"],
    "entry_date": ["Dt.entr.", "Dt.entr", "CPUDT", "BKPF-CPUDT"],
    "posting_date": ["Dt.lçto.", "Dt.lçto", "Dt.lcto.", "Dt.lcto", "BUDAT", "BKPF-BUDAT"],
}

OPTIONAL_COLUMNS = {
    "document_type": ["Tp.doc.", "Tp.doc", "BLART", "BKPF-BLART"],
    "document_text": ["Texto cabeçalho documento", "Texto cabecalho documento", "BKTXT", "BKPF-BKTXT"],
    "line_text": ["Texto", "SGTXT", "BSIS-SGTXT", "BSAS-SGTXT"],
    "create_user": ["Pré-edição", "Pre-edição", "Pré-edicao", "Pre-edicao", "PPNAM", "BKPF-PPNAM"],
    "approver_user": ["Nome do usuário", "Nome do usuario", "USNAM", "BKPF-USNAM"],
    "transaction_code": ["CódT", "CodT", "TCODE", "BKPF-TCODE"],
}

OUTPUT_COLUMNS = [
    "Company Code", "Company Name", "Document Text", "Effective Description",
    "GL Account", "GL Account Description", "Document Type", "Document Number",
    "Line Number", "Debit", "Credit", "Amount in Reporting Currency",
    "Report Currency", "Amount in Reporting Currency Rounded",
    "Amount in Document Currency", "Document Currency", "Document Date",
    "Date Entered", "Posting Date", "Fiscal Year", "Fiscal Period",
    "Create User ID", "Create User Name", "Approver User ID",
    "Approver User Name", "Transaction Code", "Transaction Code Description",
    "Duplicate Journal Count", "Duplicate Line Count", "Duplicate Key",
    "FX Method", "FX Rate", "FX Rate Date", "Source",
]

DATE_COLUMNS = ["Document Date", "Date Entered", "Posting Date", "FX Rate Date"]
AMOUNT_COLUMNS = [
    "Debit", "Credit", "Amount in Reporting Currency",
    "Amount in Reporting Currency Rounded", "Amount in Document Currency", "FX Rate",
]
INTEGER_COLUMNS = ["Duplicate Journal Count", "Duplicate Line Count"]

COLUMN_WIDTHS = {
    "Company Code": 14,
    "Company Name": 28,
    "Document Text": 36,
    "Effective Description": 42,
    "GL Account": 18,
    "GL Account Description": 34,
    "Document Type": 15,
    "Document Number": 20,
    "Line Number": 14,
    "Debit": 18,
    "Credit": 18,
    "Amount in Reporting Currency": 28,
    "Report Currency": 16,
    "Amount in Reporting Currency Rounded": 34,
    "Amount in Document Currency": 28,
    "Document Currency": 18,
    "Document Date": 15,
    "Date Entered": 15,
    "Posting Date": 15,
    "Fiscal Year": 14,
    "Fiscal Period": 15,
    "Create User ID": 18,
    "Create User Name": 24,
    "Approver User ID": 18,
    "Approver User Name": 24,
    "Transaction Code": 18,
    "Transaction Code Description": 32,
    "Duplicate Journal Count": 24,
    "Duplicate Line Count": 21,
    "Duplicate Key": 52,
    "FX Method": 34,
    "FX Rate": 18,
    "FX Rate Date": 15,
    "Source": 14,
}

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _blank(index):
    return pd.Series("", index=index, dtype="object")


def _clean_text(series):
    return series.fillna("").astype(str).str.strip().replace({"nan": "", "None": "", "<NA>": ""})


def _normalize_code(series):
    return _clean_text(series).str.replace(r"\.0$", "", regex=True)


def _normalize_company(series):
    return _normalize_code(series).map(normalize_company_output)


def _optional(dataframe, resolved, name):
    column = resolved.get(name)
    return dataframe[column] if column is not None else _blank(dataframe.index)


def _parse_number(series):
    text = _clean_text(series)
    negative = text.str.match(r"^\(.*\)$")
    text = text.str.replace(r"[()]", "", regex=True).str.replace(" ", "", regex=False)
    both = text.str.contains(",", regex=False) & text.str.contains(".", regex=False)
    comma_decimal = text.str.contains(r",\d{1,6}$", regex=True)
    text = text.where(~both, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | ~comma_decimal, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | comma_decimal, text.str.replace(",", "", regex=False))
    values = pd.to_numeric(text, errors="coerce")
    return values.where(~negative, -values.abs())


def _parse_date(series):
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    numeric = pd.to_numeric(series, errors="coerce")
    excel_dates = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    return parsed.where(parsed.notna(), excel_dates)


def _round_half_up_value(value):
    if pd.isna(value):
        return pd.NA

    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return pd.NA

    return float(decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _round_half_up_series(series):
    return pd.to_numeric(series, errors="coerce").map(_round_half_up_value)


def _resolve_optional_columns(dataframe):
    return {name: get_optional_column(dataframe, aliases) for name, aliases in OPTIONAL_COLUMNS.items()}


def _company_currency_map(master_dataframe):
    if master_dataframe.empty:
        return {}

    company = get_optional_column(master_dataframe, ["BUKRS", "Empr", "Company Code"])
    currency = get_optional_column(master_dataframe, ["WAERS", "Moeda", "Currency", "Company Main Currency"])

    if company is None or currency is None:
        return {}

    frame = pd.DataFrame({
        "company": _normalize_company(master_dataframe[company]),
        "currency": _clean_text(master_dataframe[currency]).str.upper(),
    })
    frame = frame[(frame["company"] != "") & (frame["currency"] != "")].drop_duplicates("company")
    return frame.set_index("company")["currency"].to_dict()


def _prepare_source(source_dataframe, context, source_name):
    if source_dataframe.empty:
        return pd.DataFrame()

    required = require_columns(source_dataframe, REQUIRED_COLUMNS, f"GL11 {source_name}")
    optional = _resolve_optional_columns(source_dataframe)
    module = context["module"]

    filtered = filter_by_company(source_dataframe, required["company_code"], module.get("companies", ""))
    entry_date = _parse_date(filtered[required["entry_date"]])
    from_date = to_datetime_value(module.get("from", ""))
    to_date = to_datetime_value(module.get("to", ""))
    mask = entry_date.between(from_date, to_date, inclusive="both")

    filtered = filtered.loc[mask].copy()
    entry_date = entry_date.loc[filtered.index]
    indicator = _clean_text(filtered[required["debit_credit"]]).str.upper()
    local_abs = _parse_number(filtered[required["amount_local"]]).abs()
    document_abs = _parse_number(filtered[required["amount_document"]]).abs()

    result = pd.DataFrame(index=filtered.index)
    result["Company Code"] = _normalize_company(filtered[required["company_code"]])
    result["Fiscal Year"] = _normalize_code(filtered[required["fiscal_year"]])
    result["Document Number"] = _normalize_code(filtered[required["document_number"]])
    result["Line Number"] = _normalize_code(filtered[required["line_number"]])
    result["GL Account"] = _normalize_code(filtered[required["gl_account"]])
    result["Document Type"] = _clean_text(_optional(filtered, optional, "document_type"))
    result["Document Text"] = _clean_text(_optional(filtered, optional, "document_text"))
    result["Line Text"] = _clean_text(_optional(filtered, optional, "line_text"))
    result["Effective Description"] = result["Line Text"].where(result["Line Text"] != "", result["Document Text"])
    result["Debit"] = local_abs.where(indicator == "S", pd.NA)
    result["Credit"] = local_abs.where(indicator == "H", pd.NA)
    result["_LOCAL_SIGNED"] = local_abs.where(indicator != "H", -local_abs)
    result["Amount in Document Currency"] = document_abs.where(indicator != "H", -document_abs)
    result["Document Currency"] = _clean_text(filtered[required["document_currency"]]).str.upper()
    result["Document Date"] = _parse_date(filtered[required["document_date"]])
    result["Date Entered"] = entry_date
    result["Posting Date"] = _parse_date(filtered[required["posting_date"]])
    result["Create User ID"] = _clean_text(_optional(filtered, optional, "create_user"))
    result["Approver User ID"] = _clean_text(_optional(filtered, optional, "approver_user"))
    result["Transaction Code"] = _clean_text(_optional(filtered, optional, "transaction_code"))
    result["Source"] = source_name

    print(f"GL11 {source_name} rows read: {len(source_dataframe)}")
    print(f"GL11 {source_name} rows after CONFIG filters: {len(result)}")
    return result.reset_index(drop=True)


def _deduplicate_sources(bsis, bsas):
    key = ["Company Code", "Fiscal Year", "Document Number", "Line Number"]
    combined = pd.concat([bsis, bsas], ignore_index=True)

    if combined.empty:
        return combined, 0

    combined["_SOURCE_PRIORITY"] = combined["Source"].map({"BSIS": 0, "BSAS": 1}).fillna(0)
    combined["_ORIGINAL_ORDER"] = range(len(combined))
    overlap = combined.duplicated(key, keep=False)
    sources = combined.loc[overlap].groupby(key, dropna=False)["Source"].agg(lambda values: "/".join(sorted(set(values))))
    combined = combined.sort_values(["_SOURCE_PRIORITY", "_ORIGINAL_ORDER"], kind="stable").drop_duplicates(key, keep="last")

    if not sources.empty:
        combined = combined.merge(sources.rename("_MERGED_SOURCE"), on=key, how="left")
        combined["Source"] = combined["_MERGED_SOURCE"].fillna(combined["Source"])
        combined = combined.drop(columns="_MERGED_SOURCE")

    removed = len(bsis) + len(bsas) - len(combined)
    return combined.drop(columns=["_SOURCE_PRIORITY", "_ORIGINAL_ORDER"]), removed


def _add_usd_amount(dataframe, master_dataframe, fx_dataframe):
    result = dataframe.copy()
    result["Amount in Reporting Currency"] = pd.NA
    result["FX Method"] = ""
    result["FX Rate"] = pd.NA
    result["FX Rate Date"] = pd.NaT

    if result.empty:
        return result

    company_currencies = result["Company Code"].map(_company_currency_map(master_dataframe)).fillna("").str.upper()
    document_currencies = result["Document Currency"].fillna("").str.upper()
    company_usd = company_currencies.isin(["USD", "$"])
    document_usd = ~company_usd & document_currencies.isin(["USD", "$"])
    direct = company_usd | document_usd

    result.loc[company_usd, "Amount in Reporting Currency"] = result.loc[company_usd, "_LOCAL_SIGNED"]
    result.loc[company_usd, "FX Method"] = "Local amount (company currency USD)"
    result.loc[document_usd, "Amount in Reporting Currency"] = result.loc[document_usd, "Amount in Document Currency"]
    result.loc[document_usd, "FX Method"] = "Document amount (document currency USD)"
    result.loc[direct, "FX Rate"] = 1.0
    result.loc[direct, "FX Rate Date"] = result.loc[direct, "Document Date"]

    pending = ~direct

    if pending.any() and fx_dataframe.empty:
        print(
            "GL11 FxRates not provided; reporting currency amounts remain blank for non-USD "
            "document/company currencies. GL11 LHA duplicate logic uses reporting currency, "
            "so those rows are excluded from duplicate-key detection."
        )

    if pending.any() and not fx_dataframe.empty:
        normalized_fx = normalize_fx_rates(fx_dataframe)
        conversion_currencies = document_currencies.where(document_currencies != "", company_currencies)
        keys = pd.DataFrame({
            "currency": conversion_currencies.loc[pending],
            "date": result.loc[pending, "Document Date"],
        }).drop_duplicates()
        rates = []

        for row in keys.itertuples(index=False):
            details = select_fx_rate_to_usd(normalized_fx, row.currency, row.date)
            rates.append({
                "currency": row.currency,
                "date": row.date,
                "factor": pd.NA if details is None else details["fx_to_usd"],
                "method": "" if details is None else details["method"],
                "rate": pd.NA if details is None else details["usd_rate"],
                "rate_date": pd.NaT if details is None else details["rate_date"],
            })

        rate_frame = pd.DataFrame(rates)
        pending_frame = pd.DataFrame({
            "_index": result.index[pending],
            "currency": conversion_currencies.loc[pending],
            "date": result.loc[pending, "Document Date"],
        }).merge(rate_frame, on=["currency", "date"], how="left").set_index("_index")
        factor = pd.to_numeric(pending_frame["factor"], errors="coerce")

        result.loc[pending_frame.index, "Amount in Reporting Currency"] = result.loc[pending_frame.index, "Amount in Document Currency"] * factor
        result.loc[pending_frame.index, "FX Method"] = pending_frame["method"]
        result.loc[pending_frame.index, "FX Rate"] = pending_frame["rate"]
        result.loc[pending_frame.index, "FX Rate Date"] = pending_frame["rate_date"]

    result["Amount in Reporting Currency"] = pd.to_numeric(result["Amount in Reporting Currency"], errors="coerce")
    return result


def create_gl11_duplicate_journals(bsis_dataframe, bsas_dataframe, master_dataframe, fx_dataframe, context):
    stage_started = perf_counter()
    bsis = _prepare_source(bsis_dataframe, context, "BSIS")
    bsas = _prepare_source(bsas_dataframe, context, "BSAS")
    print(f"GL11 preparation seconds: {perf_counter() - stage_started:.2f}")

    stage_started = perf_counter()
    combined, removed = _deduplicate_sources(bsis, bsas)
    print(f"GL11 combined rows: {len(bsis) + len(bsas)}")
    print(f"GL11 technical duplicates removed: {removed}")
    print(f"GL11 technical deduplication seconds: {perf_counter() - stage_started:.2f}")

    invalid_company = combined["Company Code"].eq("").sum() if not combined.empty else 0
    invalid_description = combined["Effective Description"].eq("").sum() if not combined.empty else 0
    basic_valid = (
        (combined["Company Code"] != "")
        & (combined["Fiscal Year"] != "")
        & (combined["Document Number"] != "")
        & (combined["Line Number"] != "")
        & (combined["Effective Description"] != "")
    )

    stage_started = perf_counter()
    working = _add_usd_amount(combined.loc[basic_valid].copy(), master_dataframe, fx_dataframe)
    print(f"GL11 reporting currency seconds: {perf_counter() - stage_started:.2f}")

    working["Amount in Reporting Currency Rounded"] = _round_half_up_series(working["Amount in Reporting Currency"])
    working["Amount in Reporting Currency Rounded"] = working["Amount in Reporting Currency Rounded"].mask(
        working["Amount in Reporting Currency Rounded"].eq(0),
        0.0,
    )
    available = working["Amount in Reporting Currency Rounded"].notna()
    candidates = working.loc[available].copy()
    group_key = ["Company Code", "Effective Description", "Amount in Reporting Currency Rounded"]
    journal_key = ["Company Code", "Fiscal Year", "Document Number"]

    stage_started = perf_counter()
    candidates["_JOURNAL_ID"] = pd.factorize(pd.MultiIndex.from_frame(candidates[journal_key]), sort=False)[0]
    group_statistics = candidates.groupby(group_key, sort=False, observed=True, dropna=False).agg(
        **{
            "Duplicate Journal Count": ("_JOURNAL_ID", "nunique"),
            "Duplicate Line Count": ("_JOURNAL_ID", "size"),
        }
    ).reset_index()
    duplicate_groups = group_statistics[group_statistics["Duplicate Journal Count"] > 1].copy()

    if duplicate_groups.empty:
        output = pd.DataFrame(columns=OUTPUT_COLUMNS)
    else:
        output = candidates.merge(duplicate_groups, on=group_key, how="inner")
        output["Company Name"] = output["Company Code"].map(build_company_name_map(master_dataframe)).fillna("")
        output["GL Account Description"] = output["GL Account"].map(build_gl_account_name_map(master_dataframe)).fillna("")
        output["Report Currency"] = REPORT_CURRENCY
        output["Create User Name"] = ""
        output["Approver User Name"] = ""
        output["Fiscal Period"] = output["Posting Date"].dt.strftime("%Y-%m").fillna("")
        output["Transaction Code Description"] = ""
        amount_key = output["Amount in Reporting Currency Rounded"].map(lambda value: f"{value:.2f}")
        output["Duplicate Key"] = output["Company Code"] + " | " + output["Effective Description"] + " | " + amount_key
        output = output.sort_values(group_key + ["Fiscal Year", "Document Number", "Line Number"], kind="stable").reset_index(drop=True)
        output = output[OUTPUT_COLUMNS]

    print(f"GL11 duplicate detection seconds: {perf_counter() - stage_started:.2f}")
    print(f"GL11 rows without Company Code: {invalid_company}")
    print(f"GL11 rows without Effective Description: {invalid_description}")
    print(f"GL11 rows with reporting amount: {available.sum()}")
    print(f"GL11 rows without reporting amount: {(~available).sum()}")
    print(f"GL11 rows with rounded zero amount: {working['Amount in Reporting Currency Rounded'].eq(0).sum()}")
    print(f"GL11 duplicate keys: {len(duplicate_groups)}")
    print(f"GL11 output rows: {len(output)}")
    return output


def _apply_gl11_fast_formatting(worksheet, dataframe):
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    column_positions = {column_name: column_index for column_index, column_name in enumerate(dataframe.columns, start=1)}

    for column_name, width in COLUMN_WIDTHS.items():
        column_index = column_positions.get(column_name)
        if column_index is not None:
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

    format_by_column = {
        **{column_name: "dd/mm/yyyy" for column_name in DATE_COLUMNS},
        **{column_name: "#,##0.00;[Red]-#,##0.00" for column_name in AMOUNT_COLUMNS},
        **{column_name: "0" for column_name in INTEGER_COLUMNS},
    }

    for column_name, number_format in format_by_column.items():
        column_index = column_positions.get(column_name)
        if column_index is None:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format


def _find_sheet_xml_path(workbook_archive, sheet_name):
    workbook_xml = ElementTree.fromstring(workbook_archive.read("xl/workbook.xml"))
    relationship_id = None

    for sheet in workbook_xml.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
        if str(sheet.attrib.get("name", "")).strip().casefold() == sheet_name.casefold():
            relationship_id = sheet.attrib.get(f"{{{DOCUMENT_REL_NS}}}id")
            break

    if relationship_id is None:
        return None

    relationships_xml = ElementTree.fromstring(workbook_archive.read("xl/_rels/workbook.xml.rels"))

    for relationship in relationships_xml.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        if relationship.attrib.get("Id") != relationship_id:
            continue
        target = relationship.attrib.get("Target", "")
        if target.startswith("/"):
            return target.lstrip("/")
        return posixpath.normpath(posixpath.join("xl", target))

    return None


def _column_letter_to_number(column_letter):
    value = 0
    for character in column_letter:
        if character.isalpha():
            value = value * 26 + ord(character.upper()) - 64
    return value


def _extract_gl11_style_ids(workbook_archive, sheet_xml_path):
    style_ids = {"header": None, "date": None, "amount": None, "integer": None}
    if sheet_xml_path is None:
        return style_ids

    try:
        root = ElementTree.fromstring(workbook_archive.read(sheet_xml_path))
    except KeyError:
        return style_ids

    for cell in root.findall(f".//{{{SPREADSHEET_NS}}}c"):
        reference = cell.attrib.get("r", "")
        style_id = cell.attrib.get("s")
        if not reference or style_id is None:
            continue
        column = "".join(character for character in reference if character.isalpha())
        row_text = "".join(character for character in reference if character.isdigit())
        if row_text == "1" and style_ids["header"] is None:
            style_ids["header"] = style_id
        if row_text in ["", "1"]:
            continue
        column_index = _column_letter_to_number(column)
        if not 1 <= column_index <= len(OUTPUT_COLUMNS):
            continue
        column_name = OUTPUT_COLUMNS[column_index - 1]
        if column_name in DATE_COLUMNS and style_ids["date"] is None:
            style_ids["date"] = style_id
        elif column_name in AMOUNT_COLUMNS and style_ids["amount"] is None:
            style_ids["amount"] = style_id
        elif column_name in INTEGER_COLUMNS and style_ids["integer"] is None:
            style_ids["integer"] = style_id
        if all(value is not None for value in style_ids.values()):
            break
    return style_ids


def _excel_serial_date(value):
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return (value - date(1899, 12, 30)).days
    return None


def _xml_cell(reference, value, style_id=None):
    style_attribute = f' s="{style_id}"' if style_id else ""
    if pd.isna(value):
        return f'<c r="{reference}"{style_attribute}/>'
    serial_date = _excel_serial_date(value)
    if serial_date is not None:
        return f'<c r="{reference}"{style_attribute}><v>{serial_date}</v></c>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{reference}"{style_attribute}><v>{value}</v></c>'
    text = escape(str(value), {'"': '&quot;'})
    return f'<c r="{reference}" t="inlineStr"{style_attribute}><is><t>{text}</t></is></c>'


def _build_gl11_sheet_xml(dataframe, style_ids):
    row_count = len(dataframe) + 1
    column_count = len(dataframe.columns)
    last_column = get_column_letter(column_count)
    dimension = f"A1:{last_column}{max(row_count, 1)}"
    column_positions = {column_name: column_index for column_index, column_name in enumerate(dataframe.columns, start=1)}
    column_xml = []

    for column_name, width in COLUMN_WIDTHS.items():
        column_index = column_positions.get(column_name)
        if column_index is not None:
            column_xml.append(f'<col min="{column_index}" max="{column_index}" width="{width}" customWidth="1"/>')

    rows = []
    rows.append("<row r=\"1\">" + "".join(
        _xml_cell(f"{get_column_letter(column_index)}1", column_name, style_ids.get("header"))
        for column_index, column_name in enumerate(dataframe.columns, start=1)
    ) + "</row>")

    for row_number, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        cells = []
        for column_index, value in enumerate(row, start=1):
            column_name = dataframe.columns[column_index - 1]
            style_id = None
            if column_name in DATE_COLUMNS:
                style_id = style_ids.get("date")
            elif column_name in AMOUNT_COLUMNS:
                style_id = style_ids.get("amount")
            elif column_name in INTEGER_COLUMNS:
                style_id = style_ids.get("integer")
            cells.append(_xml_cell(f"{get_column_letter(column_index)}{row_number}", value, style_id))
        rows.append(f'<row r="{row_number}">' + "".join(cells) + "</row>")

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{SPREADSHEET_NS}" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<dimension ref="{dimension}"/>'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        '<selection pane="bottomLeft"/></sheetView></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<cols>{"".join(column_xml)}</cols>'
        f'<sheetData>{"".join(rows)}</sheetData>'
        f'<autoFilter ref="{dimension}"/>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
        '</worksheet>'
    ).encode("utf-8")


def _replace_existing_gl11_sheet_fast(output_file, output_dataframe):
    output_file = Path(output_file)
    if not output_file.exists():
        return False

    with tempfile.TemporaryDirectory(prefix="lbr_gl11_zip_") as temporary_folder:
        local_output_file = Path(temporary_folder) / output_file.name
        with zipfile.ZipFile(output_file, "r") as source_archive:
            sheet_xml_path = _find_sheet_xml_path(source_archive, SHEET_NAME)
            if sheet_xml_path is None:
                return False
            style_ids = _extract_gl11_style_ids(source_archive, sheet_xml_path)
            sheet_xml = _build_gl11_sheet_xml(output_dataframe, style_ids)
            with zipfile.ZipFile(local_output_file, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1, allowZip64=True) as target_archive:
                for member in source_archive.infolist():
                    if member.is_dir():
                        target_archive.writestr(member, b"")
                        continue
                    if member.filename == sheet_xml_path:
                        target_archive.writestr(member, sheet_xml)
                        continue
                    with source_archive.open(member, "r") as source_member:
                        with target_archive.open(member, "w") as target_member:
                            shutil.copyfileobj(source_member, target_member, length=1024 * 1024)
        try:
            shutil.copy2(local_output_file, output_file)
        except PermissionError as error:
            raise PermissionError(f"Could not replace GL output workbook: {output_file}. Close the workbook and run again.") from error
    return True


def _create_lightweight_local_workbook(source_file, local_file):
    source_file = Path(source_file)
    local_file = Path(local_file)
    with zipfile.ZipFile(source_file, "r") as source_archive:
        gl11_xml_path = _find_sheet_xml_path(source_archive, SHEET_NAME)
        empty_xml = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="{SPREADSHEET_NS}"><sheetData/></worksheet>'.encode("utf-8")
        with zipfile.ZipFile(local_file, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1, allowZip64=True) as local_archive:
            for member in source_archive.infolist():
                if member.is_dir():
                    local_archive.writestr(member, b"")
                    continue
                if member.filename == gl11_xml_path:
                    local_archive.writestr(member, empty_xml)
                    continue
                with source_archive.open(member, "r") as source_member:
                    with local_archive.open(member, "w") as local_member:
                        shutil.copyfileobj(source_member, local_member, length=1024 * 1024)


def write_gl11_output(output_dataframe, context):
    write_started = perf_counter()
    output_file = get_gl_output_file(context)

    if not output_file.exists() and write_single_sheet_workbook_fast(
        output_file=output_file,
        sheet_name=SHEET_NAME,
        dataframe=output_dataframe,
        date_columns=DATE_COLUMNS,
        amount_columns=AMOUNT_COLUMNS,
        integer_columns=INTEGER_COLUMNS,
    ):
        print(f"GL11 output workbook: {output_file}")
        print("GL11 sheet written: GL11")
        print(f"GL11 workbook write seconds: {perf_counter() - write_started:.2f}")
        return output_file

    if output_file.exists():
        stage_started = perf_counter()
        if _replace_existing_gl11_sheet_fast(output_file, output_dataframe):
            print(f"GL11 direct sheet XML replacement seconds: {perf_counter() - stage_started:.2f}")
            print(f"GL11 output workbook: {output_file}")
            print("GL11 sheet replaced: GL11")
            print(f"GL11 workbook write seconds: {perf_counter() - write_started:.2f}")
            return output_file
        print("GL11 sheet does not exist in current workbook; falling back to openpyxl sheet creation.")

    with tempfile.TemporaryDirectory(prefix="lbr_gl11_") as temporary_folder:
        local_output_file = Path(temporary_folder) / output_file.name
        if output_file.exists():
            stage_started = perf_counter()
            _create_lightweight_local_workbook(output_file, local_output_file)
            print(f"GL11 local staging seconds: {perf_counter() - stage_started:.2f}")
        else:
            print("GL11 output workbook does not exist; creating it locally with openpyxl.")

        stage_started = perf_counter()
        workbook = open_or_create_gl_output_workbook(local_output_file)
        print(f"GL11 workbook open seconds: {perf_counter() - stage_started:.2f}")
        worksheet = recreate_gl_sheet(workbook, SHEET_NAME)

        stage_started = perf_counter()
        write_dataframe_to_sheet(worksheet=worksheet, dataframe=output_dataframe)
        print(f"GL11 worksheet data write seconds: {perf_counter() - stage_started:.2f}")

        stage_started = perf_counter()
        _apply_gl11_fast_formatting(worksheet, output_dataframe)
        print(f"GL11 worksheet formatting seconds: {perf_counter() - stage_started:.2f}")

        stage_started = perf_counter()
        save_gl_output_workbook(workbook, local_output_file)
        print(f"GL11 local workbook save seconds: {perf_counter() - stage_started:.2f}")
        workbook.close()

        try:
            shutil.copy2(local_output_file, output_file)
        except PermissionError as error:
            raise PermissionError(f"Could not replace GL output workbook: {output_file}. Close the workbook and run again.") from error

    print(f"GL11 output workbook: {output_file}")
    print("GL11 sheet replaced: GL11")
    print(f"GL11 workbook write seconds: {perf_counter() - write_started:.2f}")
    return output_file


def run_gl_011(context):
    """Execute GL_011 and write or replace only the GL11 worksheet."""
    started = perf_counter()
    print("=" * 80)
    print("Running GL_011 - Potential Duplicate General Journals: Same Description And Amount")
    print("LHA logic: Company + Effective Description + signed USD amount rounded to 2 decimals; distinct journals > 1")

    bsis = load_gl_bsis_data(context)
    bsas = load_gl_bsas_data(context)
    master = load_gl_master_data(context)
    fx_rates = load_gl_fx_rates_data(context)

    if bsis.empty and bsas.empty:
        raise FileNotFoundError("GL11 requires BSIS and/or BSAS journal input.")

    result = create_gl11_duplicate_journals(bsis, bsas, master, fx_rates, context)
    write_gl11_output(result, context)
    print(f"GL11 companies reported: {result['Company Code'].nunique() if not result.empty else 0}")
    print(f"GL11 effective descriptions reported: {result['Effective Description'].nunique() if not result.empty else 0}")
    print(f"GL11 elapsed seconds: {perf_counter() - started:.2f}")
    return result


__all__ = ["run_gl_011"]
