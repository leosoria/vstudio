"""
FAM_004 - Fixed Assets Comparison vs Reporting Pack.

Analysis:
- Module: Fixed Asset Management
- Analysis Code: FAM_04
- Analysis Title: Fixed Assets Comparison vs Reporting Pack

Description:
Compares the fixed assets listing generated from SAP AR01 against the Fixed
Assets section of the Reporting Pack.

Rules:
- This control writes/replaces only sheet FAM04.
- It does not read FAM01, FAM02 or FAM03 output sheets.
- It rebuilds the FAM01-like normalized dataframe from SAP AR01 and FX rates.
"""

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from modules.FAM.fam_001 import (
    REQUIRED_COLUMNS,
    build_fam_001_dataframe,
    remove_ar01_summary_rows,
    resolve_optional_columns,
)
from core.fam_common import (
    apply_standard_fam_formatting,
    build_fx_rate_lookup,
    filter_by_company,
    get_fam_output_file,
    load_fam_ar01_data,
    load_fx_rates_data,
    normalize_text,
    open_or_create_fam_output_workbook,
    parse_number,
    recreate_fam_sheet,
    require_columns,
    save_fam_output_workbook,
    write_dataframe_to_sheet,
)


SHEET_NAME = "FAM04"
DEFAULT_REPORTING_PACK_DATE_ROW = 5
DEFAULT_TOLERANCE = 0.01

FAM04_METRICS = [
    "Suma de CAP histórico",
    "Suma de CAP en fecha inicio",
    "Suma de Amortiz.acumul.en fecha de inicio",
    "Suma de VNC en fecha inicio",
    "Suma de Capitalización",
    "Suma de Valoración",
    "Suma de CAP en fecha de fin",
    "Suma de VNC en fecha de fin",
    "Suma de Depreciación en fecha de fin",
    "Suma de Amortización acumulada en fecha fin",
]

FAM01_SOURCE_COLUMNS = {
    "Suma de CAP histórico": "CAP histórico",
    "Suma de CAP en fecha inicio": "CAP en fecha inicio (LC)",
    "Suma de Amortiz.acumul.en fecha de inicio": "Amortiz.acumul.en fecha de inicio",
    "Suma de VNC en fecha inicio": "VNC en fecha inicio",
    "Suma de Capitalización": "Capitalización",
    "Suma de Valoración": "Valoración",
    "Suma de CAP en fecha de fin": "CAP en fecha de fin (LC)",
    "Suma de VNC en fecha de fin": "VNC en fecha de fin (LC)",
    "Suma de Depreciación en fecha de fin": "Depreciación en fecha de fin",
    "Suma de Amortización acumulada en fecha fin": "Amortización acumulada en fecha fin (LC)",
}

REPORTING_PACK_LINES = {
    "Suma de CAP en fecha inicio": (89, 13),
    "Suma de CAP en fecha de fin": (142, 37),
    "Suma de VNC en fecha de fin": (222, 73),
    "Suma de Amortización acumulada en fecha fin": (209, 67),
}

FAM04_BALANCE_FILE_KEYWORD = "BAL"
FAM04_MOVEMENT_FILE_KEYWORD = "MOV"
DEFAULT_DEPRECIATION_AREA = "01"

BALANCE_REQUIRED_COLUMNS = {
    "company_code": "BUKRS",
    "asset_class": "ANLKL",
    "depreciation_area": "AFABE",
    "fiscal_year": "GJAHR",
    "apc_opening": "KANSW",
}

BALANCE_DEPRECIATION_COLUMNS = [
    "KNAFA",
    "KSAFA",
    "KAAFA",
    "KMAFA",
]

MOVEMENT_REQUIRED_COLUMNS = {
    "company_code": "BUKRS",
    "asset_class": "ANLKL",
    "depreciation_area": "AFABE",
    "fiscal_year": "GJAHR",
    "transaction_type": "BWASL",
    "amount": "ANBTR",
}

MOVEMENT_DEPRECIATION_COLUMNS = [
    "NAFAB",
    "SAFAB",
]

SUPPORT_COLUMN_ALIASES = {
    "EMPR": "BUKRS",
    "EMPRESA": "BUKRS",
    "BUKRS": "BUKRS",
    "IMOBILIZADO": "ANLN1",
    "ANLN1": "ANLN1",
    "SBN": "ANLN2",
    "SUBNUMERO": "ANLN2",
    "ANLN2": "ANLN2",
    "CLASSE": "ANLKL",
    "CLASE": "ANLKL",
    "ANLKL": "ANLKL",
    "DENOMINACAO DO IMOBILIZADO": "TXT50",
    "DENOMINACION DEL ACTIVO FIJO": "TXT50",
    "TXT50": "TXT50",
    "AR": "AFABE",
    "AREA": "AFABE",
    "AFABE": "AFABE",
    "ANO": "GJAHR",
    "ANIO": "GJAHR",
    "EJERCICIO": "GJAHR",
    "GJAHR": "GJAHR",
    "VAL AQUIS ACUM": "KANSW",
    "VALOR AQUISICAO ACUM": "KANSW",
    "VALOR ADQUISICION ACUM": "KANSW",
    "KANSW": "KANSW",
    "DEPR NORM ACUM": "KNAFA",
    "DEPRECIACION NORMAL ACUM": "KNAFA",
    "KNAFA": "KNAFA",
    "DEPR ESPE ACUM": "KSAFA",
    "DEPRECIACION ESPECIAL ACUM": "KSAFA",
    "KSAFA": "KSAFA",
    "DEPR EXTR ACUM": "KAAFA",
    "DEPRECIACION EXTRAORD ACUM": "KAAFA",
    "KAAFA": "KAAFA",
    "RESERVA ACUM": "KMAFA",
    "KMAFA": "KMAFA",
    "N DOC": "BELNR",
    "NUM DOC": "BELNR",
    "DOCUMENTO": "BELNR",
    "BELNR": "BELNR",
    "ITM": "BUZEI",
    "ITEM": "BUZEI",
    "BUZEI": "BUZEI",
    "TIPO DE MOVIMENTO": "BWASL",
    "TIPO DE MOVIMIENTO": "BWASL",
    "BWASL": "BWASL",
    "DATA REF": "BZDAT",
    "FECHA REF": "BZDAT",
    "BZDAT": "BZDAT",
    "MONTANTE LANCADO": "ANBTR",
    "IMPORTE CONTABILIZADO": "ANBTR",
    "ANBTR": "ANBTR",
    "DEPR NORM MOV": "NAFAB",
    "DEPRECIACION NORMAL MOV": "NAFAB",
    "NAFAB": "NAFAB",
    "DEPR ESPE MOV": "SAFAB",
    "DEPRECIACION ESPECIAL MOV": "SAFAB",
    "SAFAB": "SAFAB",
}

CAPITALIZATION_TRANSACTION_PREFIXES = ("1",)
VALUATION_TRANSACTION_PREFIXES = ("7", "8")


REPORTING_PACK_COMPANY_ALIASES = {
    # Brazil RP files received for FY26 use TM1 entity names in the filename,
    # for example: "FY26 Reporting Pack - Brazil PLLAL_TM1 - Feb 26_20260228".
    # Keep these aliases in code rather than config.xlsx so FAM04 can resolve
    # the RP file from input/ without requiring file parameters.
    "LAR": ("LAR", "PLLAL", "BRAZIL PLLAL", "LAM PLLAL LLC"),
    "LBR": ("LBR", "PLOG", "BRAZIL PLOG", "PTLS", "BRAZIL PTLS", "WESERVICE", "BRAZIL WESERVICE", "JOURNALS", "BRAZIL JOURNALS"),
    # SAP company codes in the FAM01 dataframe are normalized without leading
    # zeroes (for example 0030 becomes 30), so include the numeric CoCo values
    # used by Brazil to avoid RP lookup misses such as "company 30".
    "30": ("30", "0030", "PLOG", "BRAZIL PLOG"),
    "34": ("34", "0034", "PTLS", "BRAZIL PTLS"),
    "52": ("52", "0052", "WESERVICE", "BRAZIL WESERVICE"),
    "93": ("93", "0093", "PLLAL", "BRAZIL PLLAL", "LAM PLLAL LLC"),
    "LCC": ("LCC",),
    "LCL": ("LCL",),
    "LCN": ("LCN",),
    "LCO": ("LCO",),
    "LCR": ("LCR",),
    "LEC": ("LEC",),
    "LPE": ("LPE",),
    "LPR": ("LPR",),
    "LRO": ("LRO",),
    "LSV": ("LSV",),
    "LUY": ("LUY",),
}

REPORTING_PACK_MONTH_ALIASES = {
    1: ("JAN", "JANUARY", "ENE", "ENERO", "01"),
    2: ("FEB", "FEBRUARY", "FEBRERO", "02"),
    3: ("MAR", "MARCH", "MARZO", "03"),
    4: ("APR", "APRIL", "ABR", "ABRIL", "04"),
    5: ("MAY", "MAYO", "05"),
    6: ("JUN", "JUNE", "JUNIO", "06"),
    7: ("JUL", "JULY", "JULIO", "07"),
    8: ("AUG", "AUGUST", "AGO", "AGOSTO", "08"),
    9: ("SEP", "SEPT", "SEPTEMBER", "SEPTIEMBRE", "09"),
    10: ("OCT", "OCTOBER", "OCTUBRE", "10"),
    11: ("NOV", "NOVEMBER", "NOVIEMBRE", "11"),
    12: ("DEC", "DECEMBER", "DIC", "DICIEMBRE", "12"),
}

AMOUNT_COLUMNS = set(FAM04_METRICS)
DATE_COLUMNS = set()
INTEGER_COLUMNS = set()
POSITIVE_DEPRECIATION_METRICS = {
    "Suma de Amortiz.acumul.en fecha de inicio",
    "Suma de Depreciación en fecha de fin",
    "Suma de Amortización acumulada en fecha fin",
}


def normalize_lookup_text(value):
    """Normalize text for Reporting Pack filename and sheet matching."""
    value_text = normalize_text(value).upper()
    value_text = unicodedata.normalize("NFKD", value_text)
    value_text = "".join(character for character in value_text if not unicodedata.combining(character))

    return re.sub(r"[^A-Z0-9]+", " ", value_text).strip()


def normalize_sheet_lookup_name(value):
    """Normalize sheet names, tolerating case, spaces, underscores and hyphens."""
    return normalize_lookup_text(value).replace(" ", "").lower()


def fiscal_year(year, month):
    """Return LHA fiscal year number where FY runs from March to February."""
    return (year + 1 if month >= 3 else year) % 100


def fiscal_position(month):
    """Return the fiscal month position where March is 1 and February is 12."""
    return month - 2 if month >= 3 else month + 10


def get_company_aliases(company_code):
    """Return configured filename aliases for a company, plus the company code."""
    company = normalize_company_code(company_code)
    aliases = list(REPORTING_PACK_COMPANY_ALIASES.get(company, ()))

    if company not in aliases:
        aliases.append(company)

    return [normalize_lookup_text(alias) for alias in aliases if normalize_lookup_text(alias) != ""]


def parse_reporting_pack_filename(file_path):
    """Parse Reporting Pack filename using the LHA debug_rp selection concepts."""
    normalized_name = normalize_lookup_text(file_path.stem)
    tokens = normalized_name.split()

    if "REPORTING" not in tokens or "PACK" not in tokens:
        return None

    fiscal_year_match = re.search(r"\bFY\s*(\d{2})\b", normalized_name)
    parsed_fiscal_year = int(fiscal_year_match.group(1)) if fiscal_year_match else None

    parsed_month = None
    for month_number, aliases in REPORTING_PACK_MONTH_ALIASES.items():
        normalized_aliases = {normalize_lookup_text(alias) for alias in aliases}
        if any(alias in tokens for alias in normalized_aliases):
            parsed_month = month_number
            break

    if parsed_month is None:
        compact_name = normalized_name.replace(" ", "")
        period_match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", compact_name)
        if period_match:
            parsed_month = int(period_match.group(2))

    return {
        "normalized_name": normalized_name,
        "fy": parsed_fiscal_year,
        "month": parsed_month,
    }


def get_reporting_pack_input_folder(context):
    """Return the coded Reporting Pack input folder.

    FAM04 intentionally does not read file paths or filename patterns from
    config.xlsx parameters. Reporting Pack files are expected in input/.
    """
    return Path(context["input_folder"])


def get_fam04_tolerance(context):
    """Return PARAM2 tolerance, defaulting to 0.01 when blank or invalid."""
    control = context.get("control", {})
    raw_tolerance = control.get("param2", "")
    parsed_tolerance = parse_number(raw_tolerance)

    if pd.isna(parsed_tolerance):
        return DEFAULT_TOLERANCE

    return abs(float(parsed_tolerance))


def get_analysis_date_to(context):
    """Return module TO date as a Python date."""
    raw_date = context["module"].get("to", "")
    parsed_date = pd.to_datetime(raw_date, errors="coerce")

    if pd.isna(parsed_date):
        raise ValueError(f"FAM04 could not parse module TO date: {raw_date}")

    return parsed_date.date()


def get_analysis_date_from(context):
    """Return module FROM date as a Python date when configured."""
    raw_date = context["module"].get("from", "")
    parsed_date = pd.to_datetime(raw_date, errors="coerce")

    if pd.isna(parsed_date):
        return None

    return parsed_date.date()


def normalize_company_code(value):
    """Normalize company code for matching filenames and FAM dataframe values."""
    value_text = normalize_text(value).upper()

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if value_text.isdigit():
        return str(int(value_text))

    return value_text


def parse_active_companies(context, fam_dataframe):
    """Return active company codes from config, falling back to dataframe companies."""
    companies_filter = normalize_text(context["module"].get("companies", ""))

    if companies_filter == "" or companies_filter.upper() == "ALL":
        return sorted(
            normalize_company_code(company)
            for company in fam_dataframe["CoCo"].dropna().unique()
            if normalize_company_code(company) != ""
        )

    normalized = companies_filter.replace(";", ",").replace("|", ",")
    companies = []

    for token in normalized.replace(" ", ",").split(","):
        company = normalize_company_code(token)
        if company != "" and company not in companies:
            companies.append(company)

    return companies


def find_fixed_assets_sheet(workbook):
    """Locate Fixed_Assets sheet using tolerant normalization."""
    for sheet_name in workbook.sheetnames:
        if normalize_sheet_lookup_name(sheet_name) == "fixedassets":
            return sheet_name

    return None


def values_match_reporting_date(value, target_date):
    """Return True when a Reporting Pack header value matches DATE TO."""
    if value is None:
        return False

    parsed_date = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed_date):
        return False

    return parsed_date.date() == target_date


def find_reporting_pack_date_column(worksheet, target_date, date_row=DEFAULT_REPORTING_PACK_DATE_ROW):
    """Find the column in the Reporting Pack date row that matches DATE TO."""
    for cell in worksheet[date_row]:
        if values_match_reporting_date(cell.value, target_date):
            return cell.column

    return None


def numeric_cell_value(worksheet, row_number, column_number):
    """Read a numeric RP cell, returning None when it is blank or non-numeric."""
    value = worksheet.cell(row=row_number, column=column_number).value
    parsed_value = parse_number(value)

    if pd.isna(parsed_value):
        return None

    return float(parsed_value)


def find_reporting_pack_file(context, company_code, date_to):
    """Find the Reporting Pack workbook for company and period.

    This follows the LHA FAM debug_rp.py approach: parse Reporting Pack
    filenames, match by company alias, fiscal year and fiscal month coverage
    instead of relying on a simple company-code substring.
    """
    input_folder = get_reporting_pack_input_folder(context)
    target_fiscal_year = fiscal_year(date_to.year, date_to.month)
    target_fiscal_position = fiscal_position(date_to.month)
    company_aliases = get_company_aliases(company_code)

    candidates = []
    for file_path in input_folder.glob("*.xls*"):
        if file_path.name.startswith("~$"):
            continue
        if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue

        file_info = parse_reporting_pack_filename(file_path)
        if file_info is None:
            continue

        normalized_name = file_info["normalized_name"]
        normalized_tokens = set(normalized_name.split())
        if not any(alias in normalized_tokens or alias in normalized_name for alias in company_aliases):
            continue

        if file_info["fy"] is not None and file_info["fy"] != target_fiscal_year:
            continue

        if file_info["month"] is not None and fiscal_position(file_info["month"]) < target_fiscal_position:
            continue

        candidates.append((file_path, file_info))

    if not candidates:
        return None

    candidates = sorted(
        candidates,
        key=lambda candidate: (
            candidate[1]["month"] is None,
            fiscal_position(candidate[1]["month"]) if candidate[1]["month"] else 99,
            len(candidate[0].name),
            candidate[0].name.lower(),
        ),
    )

    return candidates[0][0]


def read_reporting_pack_values(context, company_code, date_to):
    """Read mapped Reporting Pack line values for one company."""
    reporting_pack_file = find_reporting_pack_file(context, company_code, date_to)

    if reporting_pack_file is None:
        return None, f"RP (no se encontro el Reporting Pack de {company_code} en input/)"

    workbook = load_workbook(reporting_pack_file, data_only=True, read_only=False)

    try:
        sheet_name = find_fixed_assets_sheet(workbook)
        if sheet_name is None:
            return None, f"RP (no se encontro la hoja Fixed_Assets en {reporting_pack_file.name})"

        worksheet = workbook[sheet_name]
        date_column = find_reporting_pack_date_column(worksheet, date_to)
        if date_column is None:
            return None, f"RP (no se encontro la fecha {date_to} en {reporting_pack_file.name})"

        values = {}
        for metric, row_numbers in REPORTING_PACK_LINES.items():
            row_values = [
                numeric_cell_value(worksheet, row_number, date_column)
                for row_number in row_numbers
            ]
            numeric_values = [value for value in row_values if value is not None]
            values[metric] = sum(numeric_values) if numeric_values else None

        return values, None
    finally:
        workbook.close()



def sanitize_excel_value(value):
    """Return a value that is safe to serialize into an Excel workbook."""
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    if isinstance(value, str):
        cleaned_value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if cleaned_value.startswith(("=", "+", "-", "@")):
            return f"'{cleaned_value}"

        return cleaned_value

    if isinstance(value, float) and not pd.notna(value):
        return None

    if isinstance(value, float) and value in (float("inf"), float("-inf")):
        return None

    return value


def sanitize_output_dataframe(dataframe):
    """Clean values before writing to Excel to avoid repair warnings."""
    return dataframe.apply(lambda column: column.map(sanitize_excel_value))


def get_period_suffix(context):
    """Return YYYYMMDD suffix from the module TO date."""
    return get_analysis_date_to(context).strftime("%Y%m%d")


def normalize_depreciation_area(value):
    """Normalize depreciation area values such as 1, 01 or 1.0 to 2 digits."""
    value_text = normalize_text(value)

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if value_text.isdigit():
        return value_text.zfill(2)

    return value_text


def normalize_support_column_name(column):
    """Normalize SQVI column headers to their SAP technical names when needed."""
    column_text = normalize_text(column).upper()
    column_key = normalize_lookup_text(column_text)

    if column_text in SUPPORT_COLUMN_ALIASES:
        return SUPPORT_COLUMN_ALIASES[column_text]

    if column_key in SUPPORT_COLUMN_ALIASES:
        return SUPPORT_COLUMN_ALIASES[column_key]

    return column_text


def normalize_support_columns(dataframe):
    """Rename Portuguese/Spanish SQVI headers to SAP technical field names."""
    normalized_columns = []
    used_columns = set()

    for column in dataframe.columns:
        normalized_column = normalize_support_column_name(column)

        if normalized_column in used_columns:
            suffix = 1
            unique_column = f"{normalized_column}_{suffix}"
            while unique_column in used_columns:
                suffix += 1
                unique_column = f"{normalized_column}_{suffix}"
            normalized_column = unique_column

        normalized_columns.append(normalized_column)
        used_columns.add(normalized_column)

    dataframe = dataframe.copy()
    dataframe.columns = normalized_columns

    return dataframe


def find_fam04_support_file(context, keyword):
    """Find optional SQVI support files saved in input/ for FAM04 enrichment."""
    input_folder = Path(context["input_folder"])
    period_suffix = get_period_suffix(context)
    keyword_normalized = normalize_lookup_text(keyword)

    candidates = []
    for file_path in input_folder.glob("*.xls*"):
        if file_path.name.startswith("~$"):
            continue
        if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue

        normalized_name = normalize_lookup_text(file_path.stem)
        if "FAM" not in normalized_name.split():
            continue
        if keyword_normalized not in normalized_name.split():
            continue
        if period_suffix not in normalized_name.replace(" ", ""):
            continue

        candidates.append(file_path)

    if not candidates:
        return None

    return sorted(candidates, key=lambda path: (len(path.name), path.name.lower()))[0]


def read_optional_support_file(context, keyword):
    """Read an optional FAM04 SQVI support workbook if it exists."""
    support_file = find_fam04_support_file(context, keyword)

    if support_file is None:
        print(f"FAM_004 support file {keyword}: not found in input/ (optional)")
        return pd.DataFrame()

    dataframe = pd.read_excel(support_file)
    dataframe = dataframe.dropna(axis=0, how="all")
    dataframe = dataframe.dropna(axis=1, how="all")
    dataframe = normalize_support_columns(dataframe)

    print(f"FAM_004 support file {keyword}: {support_file}")
    print(f"FAM_004 support file {keyword} rows loaded: {len(dataframe)}")
    print(f"FAM_004 support file {keyword} columns resolved: {list(dataframe.columns)}")

    return dataframe


def filter_support_dataframe(dataframe, context):
    """Filter SQVI support dataframe by depreciation area and fiscal year."""
    if dataframe.empty:
        return dataframe

    result = dataframe.copy()
    fiscal_year_value = get_analysis_date_to(context).year

    if "AFABE" in result.columns:
        result = result[
            result["AFABE"].apply(normalize_depreciation_area) == DEFAULT_DEPRECIATION_AREA
        ].copy()

    if "GJAHR" in result.columns:
        result = result[
            pd.to_numeric(result["GJAHR"], errors="coerce") == fiscal_year_value
        ].copy()

    if "BZDAT" in result.columns:
        date_to = get_analysis_date_to(context)
        date_from = get_analysis_date_from(context)
        posting_dates = pd.to_datetime(result["BZDAT"], errors="coerce").dt.date
        date_mask = posting_dates <= date_to

        if date_from is not None:
            date_mask = date_mask & (posting_dates >= date_from)

        result = result[date_mask.fillna(False)].copy()

    return result


def sum_existing_columns(dataframe, columns):
    """Sum numeric values from columns that exist in a dataframe."""
    existing_columns = [column for column in columns if column in dataframe.columns]

    if not existing_columns:
        return pd.Series(index=dataframe.index, dtype="float64")

    numeric_dataframe = dataframe[existing_columns].apply(
        lambda column: pd.to_numeric(column.apply(parse_number), errors="coerce")
    )

    return numeric_dataframe.sum(axis=1, min_count=1)


def normalize_depreciation_amount(value):
    """Return depreciation amounts as positive values for FAM04 presentation."""
    parsed_value = parse_number(value)

    if pd.isna(parsed_value):
        return None

    return abs(float(parsed_value))


def build_balance_support_summary(context):
    """Build FAM04 metric summary from SQVI balance file ANLA+ANLC."""
    dataframe = read_optional_support_file(context, FAM04_BALANCE_FILE_KEYWORD)
    dataframe = filter_support_dataframe(dataframe, context)

    required_columns = set(BALANCE_REQUIRED_COLUMNS.values())
    if dataframe.empty or not required_columns.issubset(set(dataframe.columns)):
        missing_columns = sorted(required_columns - set(dataframe.columns))
        if missing_columns:
            print(f"FAM_004 BAL support skipped. Missing columns: {missing_columns}")
        return pd.DataFrame(columns=["CoCo", "Clase AF", *FAM04_METRICS])

    working = pd.DataFrame()
    working["CoCo"] = dataframe[BALANCE_REQUIRED_COLUMNS["company_code"]].apply(normalize_company_code)
    working["Clase AF"] = dataframe[BALANCE_REQUIRED_COLUMNS["asset_class"]].apply(normalize_text)
    working["Suma de CAP en fecha inicio"] = pd.to_numeric(
        dataframe[BALANCE_REQUIRED_COLUMNS["apc_opening"]].apply(parse_number),
        errors="coerce",
    )
    working["Suma de Amortiz.acumul.en fecha de inicio"] = sum_existing_columns(
        dataframe,
        BALANCE_DEPRECIATION_COLUMNS,
    ).apply(normalize_depreciation_amount)
    working["Suma de VNC en fecha inicio"] = (
        working["Suma de CAP en fecha inicio"]
        - working["Suma de Amortiz.acumul.en fecha de inicio"]
    )

    summary = (
        working
        .groupby(["CoCo", "Clase AF"], dropna=False)[[
            "Suma de CAP en fecha inicio",
            "Suma de Amortiz.acumul.en fecha de inicio",
            "Suma de VNC en fecha inicio",
        ]]
        .sum(min_count=1)
        .reset_index()
    )
    print(f"FAM_004 BAL support summary rows: {len(summary)}")

    return summary


def classify_movement_amount(row, prefixes):
    """Return ANBTR when BWASL starts with one of the requested prefixes."""
    transaction_type = normalize_text(row.get(MOVEMENT_REQUIRED_COLUMNS["transaction_type"], ""))

    if transaction_type.endswith(".0"):
        transaction_type = transaction_type[:-2]

    if not transaction_type.startswith(prefixes):
        return None

    return parse_number(row.get(MOVEMENT_REQUIRED_COLUMNS["amount"], None))


def build_movement_support_summary(context):
    """Build FAM04 metric summary from SQVI movement file ANLA+ANEP."""
    dataframe = read_optional_support_file(context, FAM04_MOVEMENT_FILE_KEYWORD)
    dataframe = filter_support_dataframe(dataframe, context)

    required_columns = set(MOVEMENT_REQUIRED_COLUMNS.values())
    if dataframe.empty or not required_columns.issubset(set(dataframe.columns)):
        missing_columns = sorted(required_columns - set(dataframe.columns))
        if missing_columns:
            print(f"FAM_004 MOV support skipped. Missing columns: {missing_columns}")
        return pd.DataFrame(columns=["CoCo", "Clase AF", *FAM04_METRICS])

    working = pd.DataFrame()
    working["CoCo"] = dataframe[MOVEMENT_REQUIRED_COLUMNS["company_code"]].apply(normalize_company_code)
    working["Clase AF"] = dataframe[MOVEMENT_REQUIRED_COLUMNS["asset_class"]].apply(normalize_text)
    working["Suma de Capitalización"] = dataframe.apply(
        lambda row: classify_movement_amount(row, CAPITALIZATION_TRANSACTION_PREFIXES),
        axis=1,
    )
    working["Suma de Valoración"] = dataframe.apply(
        lambda row: classify_movement_amount(row, VALUATION_TRANSACTION_PREFIXES),
        axis=1,
    )
    working["Suma de Depreciación en fecha de fin"] = sum_existing_columns(
        dataframe,
        MOVEMENT_DEPRECIATION_COLUMNS,
    ).apply(normalize_depreciation_amount)

    summary = (
        working
        .groupby(["CoCo", "Clase AF"], dropna=False)[[
            "Suma de Capitalización",
            "Suma de Valoración",
            "Suma de Depreciación en fecha de fin",
        ]]
        .sum(min_count=1)
        .reset_index()
    )
    print(f"FAM_004 MOV support summary rows: {len(summary)}")

    return summary


def build_fam04_support_summary(context):
    """Combine optional SQVI balance and movement summaries for FAM04."""
    balance_summary = build_balance_support_summary(context)
    movement_summary = build_movement_support_summary(context)

    summaries = [
        summary
        for summary in [balance_summary, movement_summary]
        if not summary.empty
    ]

    if not summaries:
        print("FAM_004 SQVI support: no BAL/MOV enrichment applied.")
        return pd.DataFrame(columns=["CoCo", "Clase AF", *FAM04_METRICS])

    combined = pd.concat(summaries, ignore_index=True, sort=False)
    metric_columns = [column for column in FAM04_METRICS if column in combined.columns]

    summary = (
        combined
        .groupby(["CoCo", "Clase AF"], dropna=False)[metric_columns]
        .sum(min_count=1)
        .reset_index()
    )
    print(f"FAM_004 SQVI support total summary rows: {len(summary)}")

    return summary


def apply_support_summary(grouped, support_summary, company_code):
    """Overlay optional SQVI support metrics onto a company grouped FAM04 block."""
    if support_summary.empty:
        return grouped

    company = normalize_company_code(company_code)
    company_support = support_summary[
        support_summary["CoCo"].apply(normalize_company_code) == company
    ].copy()

    if company_support.empty:
        return grouped

    company_support = company_support.drop(columns=["CoCo"])
    result = grouped.merge(company_support, on="Clase AF", how="outer", suffixes=("", "_support"))

    for metric in FAM04_METRICS:
        support_column = f"{metric}_support"
        if support_column in result.columns:
            result[metric] = result[support_column].combine_first(result.get(metric))
            result = result.drop(columns=[support_column])

    return result[["Clase AF", *FAM04_METRICS]]


def build_fam04_base_dataframe(source_dataframe, required_columns, optional_columns, fx_lookup):
    """Build the independent FAM01-like dataframe used by FAM04."""
    return build_fam_001_dataframe(
        source_dataframe=source_dataframe,
        required_columns=required_columns,
        optional_columns=optional_columns,
        fx_lookup=fx_lookup,
    )


def build_company_block(fam_dataframe, company_code, rp_values=None, rp_error=None, support_summary=None):
    """Build one output block: company title, pivot, Total general, RP and Diferencia."""
    company = normalize_company_code(company_code)
    sub_dataframe = fam_dataframe[
        fam_dataframe["CoCo"].apply(normalize_company_code) == company
    ].copy()

    empty_row = {"Clase AF": None, **{metric: None for metric in FAM04_METRICS}}
    header_row = {**empty_row, "Clase AF": f"Company: {company_code}"}

    if sub_dataframe.empty:
        return pd.DataFrame([{**header_row}, {**empty_row, "Clase AF": "sin datos"}])

    for source_column in FAM01_SOURCE_COLUMNS.values():
        sub_dataframe[source_column] = pd.to_numeric(
            sub_dataframe[source_column].apply(parse_number),
            errors="coerce",
        )

    for metric in POSITIVE_DEPRECIATION_METRICS:
        source_column = FAM01_SOURCE_COLUMNS[metric]
        sub_dataframe[source_column] = sub_dataframe[source_column].abs()

    grouped = (
        sub_dataframe
        .groupby("Clase AF", dropna=False)[list(FAM01_SOURCE_COLUMNS.values())]
        .sum(min_count=1)
    )
    grouped = grouped.rename(columns={source: metric for metric, source in FAM01_SOURCE_COLUMNS.items()})
    grouped = grouped.reset_index()
    grouped = grouped[["Clase AF", *FAM04_METRICS]]
    grouped = apply_support_summary(
        grouped=grouped,
        support_summary=support_summary if support_summary is not None else pd.DataFrame(),
        company_code=company_code,
    )

    total_values = grouped[FAM04_METRICS].sum(min_count=1)
    total_row = {"Clase AF": "Total general", **total_values.to_dict()}

    rp_row = {**empty_row, "Clase AF": "RP"}
    difference_row = {**empty_row, "Clase AF": "Diferencia"}

    if rp_error:
        rp_row["Clase AF"] = rp_error
    elif rp_values:
        for metric in FAM04_METRICS:
            rp_value = rp_values.get(metric)
            rp_row[metric] = rp_value
            total_value = total_row.get(metric)
            if rp_value is not None and total_value is not None and pd.notna(total_value):
                difference_row[metric] = float(total_value) - float(rp_value)

    block_rows = [header_row]
    block_rows.extend(grouped.to_dict("records"))
    block_rows.extend([total_row, empty_row, rp_row, difference_row, empty_row])

    return pd.DataFrame(block_rows, columns=["Clase AF", *FAM04_METRICS])


def build_fam_004_dataframe(fam_dataframe, context):
    """Build all FAM04 company blocks."""
    date_to = get_analysis_date_to(context)
    companies = parse_active_companies(context, fam_dataframe)
    support_summary = build_fam04_support_summary(context)
    sections = []

    for company_code in companies:
        rp_values, rp_error = read_reporting_pack_values(context, company_code, date_to)
        sections.append(
            build_company_block(
                fam_dataframe=fam_dataframe,
                company_code=company_code,
                rp_values=rp_values,
                rp_error=rp_error,
                support_summary=support_summary,
            )
        )

    if not sections:
        return pd.DataFrame([{"Clase AF": "sin datos"}], columns=["Clase AF", *FAM04_METRICS])

    return pd.concat(sections, ignore_index=True)


def run_fam_004(context):
    """Run FAM_004 and write the FAM04 sheet."""
    source_dataframe = load_fam_ar01_data(context)

    fx_dataframe = load_fx_rates_data(context)
    fx_lookup = build_fx_rate_lookup(fx_dataframe)

    required_columns = require_columns(source_dataframe, REQUIRED_COLUMNS)
    optional_columns = resolve_optional_columns(source_dataframe)

    source_dataframe = remove_ar01_summary_rows(
        source_dataframe=source_dataframe,
        required_columns=required_columns,
    )

    source_dataframe = filter_by_company(
        dataframe=source_dataframe,
        company_column=required_columns["company_code"],
        companies_filter=context["module"].get("companies", ""),
    )

    fam_dataframe = build_fam04_base_dataframe(
        source_dataframe=source_dataframe,
        required_columns=required_columns,
        optional_columns=optional_columns,
        fx_lookup=fx_lookup,
    )

    output_dataframe = build_fam_004_dataframe(
        fam_dataframe=fam_dataframe,
        context=context,
    )
    output_dataframe = sanitize_output_dataframe(output_dataframe)

    output_file = get_fam_output_file(context)
    workbook = open_or_create_fam_output_workbook(output_file)
    worksheet = recreate_fam_sheet(workbook, SHEET_NAME)

    write_dataframe_to_sheet(
        worksheet=worksheet,
        dataframe=output_dataframe,
    )

    apply_standard_fam_formatting(
        worksheet=worksheet,
        dataframe=output_dataframe,
        date_columns=DATE_COLUMNS,
        amount_columns=AMOUNT_COLUMNS,
        integer_columns=INTEGER_COLUMNS,
    )

    save_fam_output_workbook(
        workbook=workbook,
        output_file=output_file,
    )

    tolerance = get_fam04_tolerance(context)
    print(f"FAM_004 output file: {output_file}")
    print(f"FAM_004 sheet: {SHEET_NAME}")
    print(f"FAM_004 tolerance: {tolerance}")
    print(f"FAM_004 rows: {len(output_dataframe)}")
