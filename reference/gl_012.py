"""GL_012 - Number Of General Journals To General Ledger Accounts.

This control creates an account-level summary of general journal activity from
BSIS and BSAS.  It intentionally keeps all GL12-specific logic local so the
control remains independent and does not require changes to core/gl_common.py.

Accounting sign convention used in this report:
- D/C = S is a debit and is signed positive.
- D/C = H is a credit and is signed negative.
- Debit is reported as a positive subtotal for S lines.
- Credit is reported as a negative subtotal for H lines, matching the signed
  accounting net used in Amount in Reporting Currency.
"""

import importlib.util
import shutil
import tempfile
from pathlib import Path
from time import perf_counter

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAME = "GL12"
REPORT_CURRENCY = "USD"
HEADER_FILL = "D9EAF7"
DATE_FORMAT = "dd/mm/yyyy"
AMOUNT_FORMAT = '#,##0.00;[Red]-#,##0.00'
INTEGER_FORMAT = '#,##0'

ALLOWED_EXTENSIONS = [".xlsx", ".xls", ".csv", ".txt"]

BSIS_KEYWORDS = ["LBR GL_JE_BSIS", "LBR_GL_JE_BSIS"]
BSAS_KEYWORDS = ["LBR GL_JE_BSAS", "LBR_GL_JE_BSAS"]
MASTER_KEYWORDS = ["LBR GL_MD", "LBR_GL_MD"]
FX_KEYWORDS = ["FxRates", "FX Rates", "Fx Rates"]

REQUIRED_COLUMNS = {
    "company_code": ["Empr", "Empr.1", "BUKRS", "BKPF-BUKRS", "BSIS-BUKRS", "BSAS-BUKRS"],
    "fiscal_year": ["Ano", "Ano.1", "GJAHR", "BKPF-GJAHR", "BSIS-GJAHR", "BSAS-GJAHR"],
    "document_number": ["Nº doc.", "Nº doc..1", "Nº doc", "BELNR", "BKPF-BELNR", "BSIS-BELNR", "BSAS-BELNR"],
    "line_number": ["Itm", "BUZEI", "BSIS-BUZEI", "BSAS-BUZEI"],
    "gl_account": ["Razão", "Razao", "Cta.Razão", "Cta.Razão.1", "Cta.Razão.2", "Cta.Razao", "Cta.Razao.1", "HKONT", "BSIS-HKONT", "BSAS-HKONT"],
    "debit_credit": ["D/C", "SHKZG", "BSIS-SHKZG", "BSAS-SHKZG"],
    "amount_document": ["Montante", "WRBTR", "BSIS-WRBTR", "BSAS-WRBTR"],
    "amount_local": ["Montante em MI", "DMBTR", "BSIS-DMBTR", "BSAS-DMBTR"],
    "document_currency": ["Moeda", "Moeda.1", "WAERS", "BKPF-WAERS"],
    "document_date": ["Data doc.", "Data doc", "BLDAT", "BKPF-BLDAT"],
    "entry_date": ["Dt.entr.", "Dt.entr", "CPUDT", "BKPF-CPUDT"],
    "posting_date": ["Dt.lçto.", "Dt.lçto", "Dt.lcto.", "Dt.lcto", "BUDAT", "BKPF-BUDAT"],
}

OUTPUT_COLUMNS = [
    "Company Code",
    "Company Name",
    "GL Account",
    "GL Account Description",
    "Count",
    "Line Count",
    "Sum Absolute Amount",
    "Report Currency",
    "Debit",
    "Credit",
    "Amount in Reporting Currency",
    "Amount in Document Currency",
    "Document Currency",
    "Fiscal Year",
    "Period From",
    "Period To",
    "First Posting Date",
    "Last Posting Date",
    "First Date Entered",
    "Last Date Entered",
    "Source",
    "Source Count BSIS",
    "Source Count BSAS",
    "FX Method",
    "FX Rate",
    "FX Rate Date",
]

DATE_COLUMNS = [
    "Period From",
    "Period To",
    "First Posting Date",
    "Last Posting Date",
    "First Date Entered",
    "Last Date Entered",
    "FX Rate Date",
]
AMOUNT_COLUMNS = [
    "Sum Absolute Amount",
    "Debit",
    "Credit",
    "Amount in Reporting Currency",
    "Amount in Document Currency",
    "FX Rate",
]
INTEGER_COLUMNS = ["Count", "Line Count", "Source Count BSIS", "Source Count BSAS"]
COLUMN_WIDTHS = {
    "Company Code": 14,
    "Company Name": 30,
    "GL Account": 18,
    "GL Account Description": 36,
    "Count": 12,
    "Line Count": 14,
    "Sum Absolute Amount": 24,
    "Report Currency": 16,
    "Debit": 18,
    "Credit": 18,
    "Amount in Reporting Currency": 28,
    "Amount in Document Currency": 28,
    "Document Currency": 18,
    "Fiscal Year": 14,
    "Period From": 15,
    "Period To": 15,
    "First Posting Date": 18,
    "Last Posting Date": 18,
    "First Date Entered": 18,
    "Last Date Entered": 18,
    "Source": 18,
    "Source Count BSIS": 18,
    "Source Count BSAS": 18,
    "FX Method": 36,
    "FX Rate": 18,
    "FX Rate Date": 15,
}


def _elapsed(label, started):
    print(f"GL12 {label} seconds: {perf_counter() - started:.2f}")


def _normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in ["nan", "none", "<na>"]:
        return ""
    return text


def _normalize_header(value):
    return _normalize_text(value).lower()


def _clean_text(series):
    return series.fillna("").astype(str).str.strip().replace({"nan": "", "None": "", "<NA>": ""})


def _normalize_code(series):
    return _clean_text(series).str.replace(r"\.0$", "", regex=True)


def _normalize_company_output(value):
    text = _normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit():
        return str(int(text))
    return text


def _normalize_company(series):
    return _normalize_code(series).map(_normalize_company_output)


def _parse_companies_filter(value):
    text = _normalize_text(value)
    if text == "" or text.upper() in ["ALL", "TODAS", "TODOS"]:
        return []
    for separator in [";", "|", "\n", "\r", "\t"]:
        text = text.replace(separator, ",")
    if "," not in text and " " in text:
        text = ",".join(text.split())
    companies = []
    for item in text.split(","):
        item = _normalize_company_output(item.strip())
        if item == "":
            continue
        if item.upper() in ["ALL", "TODAS", "TODOS"]:
            return []
        companies.append(item)
    return companies


def _parse_date_value(value):
    if pd.isna(value):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value
    text = _normalize_text(value)
    if text == "":
        return pd.NaT
    if text.endswith(".0"):
        text = text[:-2]
    if len(text) == 8 and text.isdigit():
        return pd.to_datetime(text, format="%Y%m%d", errors="coerce")
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        return pd.to_datetime(text[:10], format="%Y-%m-%d", errors="coerce")
    return pd.to_datetime(text, errors="coerce", dayfirst=True)


def _parse_date(series):
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    numeric = pd.to_numeric(series, errors="coerce")
    excel_dates = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    return parsed.where(parsed.notna(), excel_dates)


def _parse_number(series):
    text = _clean_text(series)
    negative_parentheses = text.str.match(r"^\(.*\)$")
    text = text.str.replace(r"[()]", "", regex=True).str.replace(" ", "", regex=False)
    both = text.str.contains(",", regex=False) & text.str.contains(".", regex=False)
    comma_decimal = text.str.contains(r",\d{1,6}$", regex=True)
    text = text.where(~both, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | ~comma_decimal, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | comma_decimal, text.str.replace(",", "", regex=False))
    values = pd.to_numeric(text, errors="coerce")
    return values.where(~negative_parentheses, -values.abs())


def _find_column(dataframe, possible_names):
    normalized_lookup = {_normalize_header(column): column for column in dataframe.columns}
    for possible_name in possible_names:
        column = normalized_lookup.get(_normalize_header(possible_name))
        if column is not None:
            return column
    return None


def _require_columns(dataframe, required_columns, source_name):
    resolved = {}
    missing = []
    for logical_name, aliases in required_columns.items():
        column = _find_column(dataframe, aliases)
        if column is None:
            missing.append(f"{logical_name}: expected one of {aliases}")
        else:
            resolved[logical_name] = column
    if missing:
        raise ValueError(f"Missing required columns in {source_name}:\n- " + "\n- ".join(missing))
    return resolved


def _period_suffix(context):
    parsed = _parse_date_value(context["module"].get("to", ""))
    if pd.isna(parsed):
        raise ValueError("Could not determine GL12 output period because module TO date is empty or invalid.")
    return parsed.strftime("%Y%m%d")


def _find_files_containing(base_folder, text_to_find):
    text = _normalize_text(text_to_find).lower()
    result = []
    for file_path in Path(base_folder).rglob("*"):
        if file_path.is_file() and file_path.suffix.lower() in ALLOWED_EXTENSIONS and text in file_path.name.lower():
            result.append(file_path)
    return sorted(result)


def _find_period_file(context, keyword_prefixes, required=False, label="input"):
    input_folder = Path(context["input_folder"])
    suffix = _period_suffix(context)
    for keyword_prefix in keyword_prefixes:
        keyword = f"{keyword_prefix}_{suffix}"
        matches = _find_files_containing(input_folder, keyword)
        if len(matches) > 1:
            raise ValueError(f"Multiple GL12 {label} files found using keyword '{keyword}': " + ", ".join(str(path) for path in matches))
        if len(matches) == 1:
            return matches[0]
    if required:
        raise FileNotFoundError(f"GL12 required {label} file was not found for period suffix {suffix}.")
    return None


def _read_input_file(file_path, usecols=None):
    file_path = Path(file_path)
    if file_path.suffix.lower() in [".xlsx", ".xls"]:
        dataframe = pd.read_excel(file_path, dtype=object, usecols=usecols)
    elif file_path.suffix.lower() == ".csv":
        dataframe = pd.read_csv(file_path, dtype=object, usecols=usecols)
    elif file_path.suffix.lower() == ".txt":
        dataframe = pd.read_csv(file_path, sep="\t", dtype=object, usecols=usecols)
    else:
        raise ValueError(f"Unsupported input file extension: {file_path.suffix}")
    dataframe = dataframe.dropna(axis=0, how="all").dropna(axis=1, how="all")
    dataframe.columns = [_normalize_text(column) for column in dataframe.columns]
    return dataframe


def _read_input_headers(file_path):
    file_path = Path(file_path)
    if file_path.suffix.lower() in [".xlsx", ".xls"]:
        dataframe = pd.read_excel(file_path, dtype=object, nrows=0)
    elif file_path.suffix.lower() == ".csv":
        dataframe = pd.read_csv(file_path, dtype=object, nrows=0)
    elif file_path.suffix.lower() == ".txt":
        dataframe = pd.read_csv(file_path, sep="\t", dtype=object, nrows=0)
    else:
        raise ValueError(f"Unsupported input file extension: {file_path.suffix}")
    dataframe.columns = [_normalize_text(column) for column in dataframe.columns]
    return list(dataframe.columns)


def _resolve_source_usecols(input_file, source_name):
    header_dataframe = pd.DataFrame(columns=_read_input_headers(input_file))
    resolved = _require_columns(header_dataframe, REQUIRED_COLUMNS, f"GL12 {source_name}")
    return sorted(set(resolved.values()))


def _load_source(context, source_name, keywords):
    started = perf_counter()
    input_file = _find_period_file(context, keywords, required=False, label=source_name)
    if input_file is None:
        print(f"GL12 {source_name} input file: not found.")
        _elapsed(f"load {source_name}", started)
        return pd.DataFrame()
    print(f"GL12 {source_name} input file: {input_file}")
    usecols = _resolve_source_usecols(input_file, source_name)
    dataframe = _read_input_file(input_file, usecols=usecols)
    dataframe["Source"] = source_name
    print(f"GL12 {source_name} rows read: {len(dataframe)}")
    _elapsed(f"load {source_name}", started)
    return dataframe


def _load_optional(context, keywords, label):
    input_file = _find_period_file(context, keywords, required=False, label=label)
    if input_file is None:
        print(f"GL12 {label} input file: not found.")
        return pd.DataFrame()
    print(f"GL12 {label} input file: {input_file}")
    return _read_input_file(input_file)


def _prepare_source(source_dataframe, context, source_name):
    if source_dataframe.empty:
        return pd.DataFrame()
    started = perf_counter()
    required = _require_columns(source_dataframe, REQUIRED_COLUMNS, f"GL12 {source_name}")
    module = context["module"]
    filtered = source_dataframe
    companies = _parse_companies_filter(module.get("companies", ""))
    if companies:
        company_values = _normalize_company(filtered[required["company_code"]])
        filtered = filtered.loc[company_values.isin(companies)]
    posting_date = _parse_date(filtered[required["posting_date"]])
    from_date = _parse_date_value(module.get("from", ""))
    to_date = _parse_date_value(module.get("to", ""))
    filtered = filtered.loc[posting_date.between(from_date, to_date, inclusive="both")].copy()
    posting_date = posting_date.loc[filtered.index]
    _elapsed(f"filter config {source_name}", started)

    started = perf_counter()
    indicator = _clean_text(filtered[required["debit_credit"]]).str.upper()
    document_abs = _parse_number(filtered[required["amount_document"]]).abs()
    local_abs = _parse_number(filtered[required["amount_local"]]).abs()
    document_signed = document_abs.where(indicator != "H", -document_abs)
    local_signed = local_abs.where(indicator != "H", -local_abs)
    debit_document = document_abs.where(indicator == "S", 0.0)
    credit_document = (-document_abs).where(indicator == "H", 0.0)

    result = pd.DataFrame(index=filtered.index)
    result["Company Code"] = _normalize_company(filtered[required["company_code"]])
    result["Fiscal Year"] = _normalize_code(filtered[required["fiscal_year"]])
    result["Document Number"] = _normalize_code(filtered[required["document_number"]])
    result["Line Number"] = _normalize_code(filtered[required["line_number"]])
    result["GL Account"] = _normalize_code(filtered[required["gl_account"]])
    result["Debit Document"] = debit_document
    result["Credit Document"] = credit_document
    result["Amount in Document Currency"] = document_signed
    result["_LOCAL_SIGNED"] = local_signed
    result["Document Currency"] = _clean_text(filtered[required["document_currency"]]).str.upper()
    result["Document Date"] = _parse_date(filtered[required["document_date"]])
    result["Date Entered"] = _parse_date(filtered[required["entry_date"]])
    result["Posting Date"] = posting_date
    result["Source"] = source_name
    result["Journal ID"] = result["Company Code"] + "|" + result["Fiscal Year"] + "|" + result["Document Number"]
    _elapsed(f"minimum normalization and signed amounts {source_name}", started)
    print(f"GL12 {source_name} rows after CONFIG filters: {len(result)}")
    return result.reset_index(drop=True)


def _deduplicate_sources(bsis, bsas):
    combined = pd.concat([bsis, bsas], ignore_index=True)
    if combined.empty:
        return combined, 0
    key = ["Company Code", "Fiscal Year", "Document Number", "Line Number"]
    combined["_SOURCE_PRIORITY"] = combined["Source"].map({"BSIS": 0, "BSAS": 1}).fillna(0)
    combined["_ORIGINAL_ORDER"] = range(len(combined))
    overlap = combined.duplicated(key, keep=False)
    sources = combined.loc[overlap].groupby(key, dropna=False)["Source"].agg(lambda values: "/".join(sorted(set(values))))
    combined = combined.sort_values(["_SOURCE_PRIORITY", "_ORIGINAL_ORDER"], kind="stable").drop_duplicates(key, keep="last")
    if not sources.empty:
        combined = combined.merge(sources.rename("_MERGED_SOURCE"), on=key, how="left")
        combined["Source"] = combined["_MERGED_SOURCE"].fillna(combined["Source"])
        combined = combined.drop(columns="_MERGED_SOURCE")
    removed = len(bsis) + len(bsas) - len(combined)
    return combined.drop(columns=["_SOURCE_PRIORITY", "_ORIGINAL_ORDER"]), removed


def _build_company_name_map(master_dataframe):
    if master_dataframe.empty:
        return {}
    company = _find_column(master_dataframe, ["BUKRS", "Empr", "Company Code"])
    name = _find_column(master_dataframe, ["BUTXT", "Nome da firma", "Nome da empresa", "Empresa", "Company", "Company Name"])
    if company is None or name is None:
        return {}
    frame = pd.DataFrame({"Company Code": _normalize_company(master_dataframe[company]), "Company Name": _clean_text(master_dataframe[name])})
    frame = frame[(frame["Company Code"] != "") & (frame["Company Name"] != "")].drop_duplicates("Company Code")
    return frame.set_index("Company Code")["Company Name"].to_dict()


def _build_gl_account_name_map(master_dataframe):
    if master_dataframe.empty:
        return {}
    account = _find_column(master_dataframe, ["SAKNR", "HKONT", "Cta.Razão", "Cta.Razao", "Cta.Razão.1", "Razão", "Razao", "Conta", "Account", "GL Account"])
    text = _find_column(master_dataframe, ["TXT50", "TxtDescr", "TXT20", "Texto breve", "Texto", "Account Name", "Description", "GL Account Description"])
    language = _find_column(master_dataframe, ["SPRAS", "Idioma", "Language"])
    if account is None or text is None:
        return {}
    frame = pd.DataFrame({"GL Account": _normalize_code(master_dataframe[account]), "Text": _clean_text(master_dataframe[text])})
    if language is not None:
        priority = _clean_text(master_dataframe[language]).str.upper().map({"PT": 1, "P": 1, "ES": 2, "S": 2, "EN": 3, "E": 3}).fillna(9)
        frame["Priority"] = priority
        frame = frame.sort_values("Priority", kind="stable")
    frame = frame[(frame["GL Account"] != "") & (frame["Text"] != "")].drop_duplicates("GL Account")
    return frame.set_index("GL Account")["Text"].to_dict()


def _normalize_fx_rates(fx_dataframe):
    if fx_dataframe.empty:
        return pd.DataFrame(columns=["Rate Type", "From Currency", "To Currency", "Valid From", "Rate", "Factor From", "Factor To"])
    columns = _require_columns(
        fx_dataframe,
        {
            "rate_type": ["CgCâ", "CgCa", "KURST", "TCot", "Tipo cotización", "Tipo cotizacion"],
            "from_currency": ["De", "FCURR", "Moeda de procedência", "Moeda de procedencia", "From Currency"],
            "to_currency": ["Para", "TCURR", "Moeda destino", "To Currency"],
            "valid_from": ["Vál.desde", "Val.desde", "GDATU", "Data válida desde", "Data valida desde", "Valid From"],
            "rate": ["Taxa câmbio", "Taxa cambio", "UKURS", "Cotação", "Cotacao", "Exchange Rate", "Rate"],
        },
        "GL12 FxRates",
    )
    factor_from = _find_column(fx_dataframe, ["Fator (origem)", "Fator origem", "FFACT", "From Factor", "Factor From"])
    factor_to = _find_column(fx_dataframe, ["Fator (para)", "Fator para", "TFACT", "To Factor", "Factor To"])
    result = pd.DataFrame()
    result["Rate Type"] = _clean_text(fx_dataframe[columns["rate_type"]]).str.upper()
    result["From Currency"] = _clean_text(fx_dataframe[columns["from_currency"]]).str.upper()
    result["To Currency"] = _clean_text(fx_dataframe[columns["to_currency"]]).str.upper()
    result["Valid From"] = _parse_date(fx_dataframe[columns["valid_from"]]).dt.normalize()
    result["Rate"] = _parse_number(fx_dataframe[columns["rate"]])
    result["Factor From"] = 1 if factor_from is None else _parse_number(fx_dataframe[factor_from]).fillna(1)
    result["Factor To"] = 1 if factor_to is None else _parse_number(fx_dataframe[factor_to]).fillna(1)
    return result.dropna(subset=["Valid From"])


def _select_fx_rate_to_usd(normalized_fx, currency, requested_date):
    currency = _normalize_text(currency).upper()
    requested_date = _parse_date_value(requested_date)
    if currency == "" or pd.isna(requested_date):
        return None
    if currency in ["USD", "$"]:
        return {"factor": 1.0, "rate": 1.0, "rate_date": requested_date.normalize(), "method": "Document currency USD, rate 1.000000"}
    if normalized_fx.empty:
        return None
    requested_date = requested_date.normalize()
    for rate_type in ["EN", "M"]:
        for days_back in range(0, 11):
            date = requested_date - pd.Timedelta(days=days_back)
            candidates = normalized_fx[(normalized_fx["Rate Type"] == rate_type) & (normalized_fx["Valid From"] == date)]
            direct = candidates[(candidates["From Currency"] == currency) & (candidates["To Currency"] == "USD")]
            if not direct.empty:
                row = direct.iloc[0]
                adjusted = float(row["Rate"]) * float(row["Factor To"] or 1) / float(row["Factor From"] or 1)
                if adjusted != 0:
                    return {"factor": adjusted, "rate": 1 / adjusted, "rate_date": date, "method": f"FxRates {rate_type}: {currency}->USD; USD = amount * FX"}
            inverse = candidates[(candidates["From Currency"] == "USD") & (candidates["To Currency"] == currency)]
            if not inverse.empty:
                row = inverse.iloc[0]
                adjusted = float(row["Rate"]) * float(row["Factor To"] or 1) / float(row["Factor From"] or 1)
                if adjusted != 0:
                    return {"factor": 1 / adjusted, "rate": adjusted, "rate_date": date, "method": f"FxRates {rate_type}: USD->{currency}; USD = amount / USD Rate"}
    return None


def _add_reporting_currency(dataframe, fx_dataframe):
    started = perf_counter()
    result = dataframe.copy()
    result["Amount in Reporting Currency"] = pd.NA
    result["Debit"] = pd.NA
    result["Credit"] = pd.NA
    result["FX Method"] = ""
    result["FX Rate"] = pd.NA
    result["FX Rate Date"] = pd.NaT
    if result.empty:
        _elapsed("FX / reporting currency", started)
        return result

    normalized_fx = _normalize_fx_rates(fx_dataframe) if not fx_dataframe.empty else pd.DataFrame()
    currencies = result["Document Currency"].fillna("").str.upper()
    dates = result["Document Date"]
    keys = pd.DataFrame({"currency": currencies, "date": dates}).drop_duplicates()
    rates = []
    for row in keys.itertuples(index=False):
        details = _select_fx_rate_to_usd(normalized_fx, row.currency, row.date)
        rates.append(
            {
                "currency": row.currency,
                "date": row.date,
                "factor": pd.NA if details is None else details["factor"],
                "rate": pd.NA if details is None else details["rate"],
                "rate_date": pd.NaT if details is None else details["rate_date"],
                "method": "" if details is None else details["method"],
            }
        )
    rate_frame = pd.DataFrame(rates)
    result = result.merge(rate_frame, left_on=["Document Currency", "Document Date"], right_on=["currency", "date"], how="left")
    factor = pd.to_numeric(result["factor"], errors="coerce")
    result["Amount in Reporting Currency"] = result["Amount in Document Currency"] * factor
    result["Debit"] = result["Debit Document"] * factor
    result["Credit"] = result["Credit Document"] * factor
    result["FX Method"] = result["method"].fillna("")
    result["FX Rate"] = result["rate"]
    result["FX Rate Date"] = result["rate_date"]
    result = result.drop(columns=["currency", "date", "factor", "rate", "rate_date", "method"])
    _elapsed("FX / reporting currency", started)
    return result


def _aggregate(working, context):
    started = perf_counter()
    if working.empty:
        _elapsed("grouping", started)
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    valid = working[(working["Company Code"] != "") & (working["GL Account"] != "") & (working["Journal ID"] != "||")].copy()
    valid["Report Currency"] = REPORT_CURRENCY
    valid["_ABS_REPORTING"] = valid["Amount in Reporting Currency"].abs()
    valid["_SOURCE_BSIS"] = valid["Source"].str.contains("BSIS", na=False).astype("int64")
    valid["_SOURCE_BSAS"] = valid["Source"].str.contains("BSAS", na=False).astype("int64")
    group_key = ["Company Code", "GL Account", "Report Currency"]
    summary = valid.groupby(group_key, sort=False, observed=True, dropna=False).agg(
        **{
            "Count": ("Journal ID", "nunique"),
            "Line Count": ("Journal ID", "size"),
            "Sum Absolute Amount": ("_ABS_REPORTING", "sum"),
            "Debit": ("Debit", "sum"),
            "Credit": ("Credit", "sum"),
            "Amount in Reporting Currency": ("Amount in Reporting Currency", "sum"),
            "Amount in Document Currency": ("Amount in Document Currency", "sum"),
            "Document Currency": ("Document Currency", lambda values: ", ".join(sorted(value for value in values.dropna().unique() if value != ""))[:32767]),
            "Fiscal Year": ("Fiscal Year", lambda values: ", ".join(sorted(value for value in values.dropna().unique() if value != ""))[:32767]),
            "First Posting Date": ("Posting Date", "min"),
            "Last Posting Date": ("Posting Date", "max"),
            "First Date Entered": ("Date Entered", "min"),
            "Last Date Entered": ("Date Entered", "max"),
            "Source": ("Source", lambda values: ", ".join(sorted(value for value in values.dropna().unique() if value != ""))[:32767]),
            "Source Count BSIS": ("_SOURCE_BSIS", "sum"),
            "Source Count BSAS": ("_SOURCE_BSAS", "sum"),
            "FX Method": ("FX Method", lambda values: ", ".join(sorted(value for value in values.dropna().unique() if value != ""))[:32767]),
            "FX Rate": ("FX Rate", "mean"),
            "FX Rate Date": ("FX Rate Date", "max"),
        }
    ).reset_index()
    period_from = _parse_date_value(context["module"].get("from", ""))
    period_to = _parse_date_value(context["module"].get("to", ""))
    summary["Period From"] = period_from
    summary["Period To"] = period_to
    _elapsed("grouping", started)
    return summary


def _enrich(summary, master_dataframe):
    started = perf_counter()
    if summary.empty:
        _elapsed("enrichment", started)
        return summary.reindex(columns=OUTPUT_COLUMNS)
    company_names = _build_company_name_map(master_dataframe)
    account_names = _build_gl_account_name_map(master_dataframe)
    summary["Company Name"] = summary["Company Code"].map(company_names).fillna("")
    summary["GL Account Description"] = summary["GL Account"].map(account_names).fillna("")
    summary = summary.sort_values(["Company Code", "Count", "Sum Absolute Amount", "GL Account"], ascending=[True, False, False, True], kind="stable")
    summary = summary.reindex(columns=OUTPUT_COLUMNS)
    _elapsed("enrichment", started)
    return summary


def _get_output_file(context):
    return Path(context["output_folder"]) / f"LBR_Results_GL_{_period_suffix(context)}.xlsx"


def _write_fast(output_file, dataframe):
    if output_file.exists():
        return False

    if importlib.util.find_spec("xlsxwriter") is None:
        print("GL12 xlsxwriter package not found. Using openpyxl writer.")
        return False
    with pd.ExcelWriter(output_file, engine="xlsxwriter", datetime_format=DATE_FORMAT, date_format=DATE_FORMAT) as writer:
        dataframe.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        workbook = writer.book
        worksheet = writer.sheets[SHEET_NAME]
        header_format = workbook.add_format({"bold": True, "bg_color": f"#{HEADER_FILL}"})
        date_format = workbook.add_format({"num_format": DATE_FORMAT})
        amount_format = workbook.add_format({"num_format": AMOUNT_FORMAT})
        integer_format = workbook.add_format({"num_format": INTEGER_FORMAT})
        for column_index, column_name in enumerate(dataframe.columns):
            worksheet.write(0, column_index, column_name, header_format)
            width = COLUMN_WIDTHS.get(column_name, min(max(len(str(column_name)) + 2, 12), 45))
            fmt = None
            if column_name in DATE_COLUMNS:
                fmt = date_format
            elif column_name in AMOUNT_COLUMNS:
                fmt = amount_format
            elif column_name in INTEGER_COLUMNS:
                fmt = integer_format
            worksheet.set_column(column_index, column_index, width, fmt)
        worksheet.autofilter(0, 0, max(len(dataframe), 1), max(len(dataframe.columns) - 1, 0))
        worksheet.freeze_panes(1, 0)
    return True


def _write_with_openpyxl(output_file, dataframe):
    temp_folder = Path(tempfile.gettempdir()) / "lbr_gl12"
    temp_folder.mkdir(parents=True, exist_ok=True)
    temp_file = temp_folder / output_file.name

    if output_file.exists():
        shutil.copy2(output_file, temp_file)
        workbook = load_workbook(temp_file)
    else:
        temp_file = temp_folder / output_file.name
        workbook = Workbook()
        workbook.remove(workbook.active)
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    worksheet = workbook.create_sheet(SHEET_NAME)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            worksheet.cell(row=row_index, column=column_index, value=value)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    column_positions = {column_name: index for index, column_name in enumerate(dataframe.columns, start=1)}
    for column_name, width in COLUMN_WIDTHS.items():
        column_index = column_positions.get(column_name)
        if column_index is not None:
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
    for column_name in DATE_COLUMNS:
        column_index = column_positions.get(column_name)
        if column_index is not None:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = DATE_FORMAT
    for column_name in AMOUNT_COLUMNS:
        column_index = column_positions.get(column_name)
        if column_index is not None:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = AMOUNT_FORMAT
    for column_name in INTEGER_COLUMNS:
        column_index = column_positions.get(column_name)
        if column_index is not None:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = INTEGER_FORMAT
    try:
        workbook.save(temp_file)
        shutil.copy2(temp_file, output_file)
    except PermissionError as error:
        raise PermissionError(f"Could not save output workbook: {output_file}. The file may be open in Excel or locked by OneDrive. Close the workbook and run again.") from error
    finally:
        try:
            if temp_file.exists():
                temp_file.unlink()
        except OSError:
            pass


def _write_output(context, dataframe):
    started = perf_counter()
    output_file = _get_output_file(context)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wrote_fast = _write_fast(output_file, dataframe)
    if not wrote_fast:
        _write_with_openpyxl(output_file, dataframe)
    print(f"GL12 output file: {output_file}")
    print(f"GL12 output sheet: {SHEET_NAME}")
    print(f"GL12 output rows: {len(dataframe)}")
    _elapsed("write Excel", started)


def create_gl12_summary(bsis_dataframe, bsas_dataframe, master_dataframe, fx_dataframe, context):
    bsis = _prepare_source(bsis_dataframe, context, "BSIS")
    bsas = _prepare_source(bsas_dataframe, context, "BSAS")

    started = perf_counter()
    combined, removed = _deduplicate_sources(bsis, bsas)
    print(f"GL12 combined rows: {len(bsis) + len(bsas)}")
    print(f"GL12 technical duplicates removed: {removed}")
    _elapsed("technical deduplication", started)

    working = _add_reporting_currency(combined, fx_dataframe)
    summary = _aggregate(working, context)
    output = _enrich(summary, master_dataframe)
    return output


def run_gl_012(context):
    started = perf_counter()
    bsis_dataframe = _load_source(context, "BSIS", BSIS_KEYWORDS)
    bsas_dataframe = _load_source(context, "BSAS", BSAS_KEYWORDS)
    if bsis_dataframe.empty and bsas_dataframe.empty:
        raise FileNotFoundError("GL12 requires at least one BSIS or BSAS input file for the active period.")

    stage_started = perf_counter()
    master_dataframe = _load_optional(context, MASTER_KEYWORDS, "master")
    fx_dataframe = _load_optional(context, FX_KEYWORDS, "FxRates")
    _elapsed("load master/FxRates", stage_started)

    output = create_gl12_summary(bsis_dataframe, bsas_dataframe, master_dataframe, fx_dataframe, context)
    _write_output(context, output)
    print(f"GL12 total seconds: {perf_counter() - started:.2f}")
