"""
CD_004 - Vendors With Large Numbers Of Cash Disbursements.

Analysis:
- Module: Cash Disbursements
- Analysis Code: CD_04
- Analysis Title: Vendors With Large Numbers Of Cash Disbursements

Description:
Highlights vendors with unusually high counts of disbursements.

Procedure:
Rank vendors by count and review for split transactions.

Analytic logic:
Count disbursements per vendor and compare each vendor to its company peers.
Vendors are selected when their count of unique payment documents is at or
above both:
- a minimum absolute threshold; and
- the configured company-level percentile threshold.

Context:
Focus on vendors with many small payments or potential policy breaches.

Configuration:
CD_004 reads dynamic parameters from the CD sheet control row.

Recommended simple setup in sheet CD:
    PARAM1 = 10
    PARAM2 = 0.90

Meaning:
    PARAM1 = MIN_QTY_PAYMENTS
    PARAM2 = PEER_PERCENTILE

Also supported:
    PARAM1 = MIN_QTY_PAYMENTS=10
    PARAM2 = PEER_PERCENTILE=0.90

    PARAM1 = MIN_QTY_PAYMENTS=10
    PARAM2 = PEER_PERCENTILE=0.90;OUTPUT_LEVEL=DETAIL

    PARAM1 = MIN_QTY_PAYMENTS=10
    PARAM2 = PEER_PERCENTILE=0.90;OUTPUT_LEVEL=PAYMENT

Output levels:
- DETAIL:
    Keeps all CD_001 detail rows for selected vendors.
    This preserves detail, but the number of rows can be higher than Qty Payments
    because one KEY_DOC can have multiple detail lines.
- PAYMENT:
    Keeps one row per Company + Vendor + KEY_DOC.
    This makes CD04 closer to the unique payment count used by Qty Payments.

Auditability:
CD_004 writes a small parameter box on the right side of the CD04 sheet, in the
first rows after the last data column. It includes:
- Minimum Qty Payments
- Peer Percentile
- Output Level
- Effective company thresholds

Design notes:
- CD_004 uses the same payment universe as CD_001 by reusing the CD_001
  detail dataframe builder.
- It does not read the CD01 sheet from the output workbook.
- It writes/replaces only the CD04 sheet.
- It does not delete sheets from other controls.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.cd_common import (
    apply_standard_cd_formatting,
    get_cd_output_file,
    load_cd_base_data,
    normalize_text,
    open_or_create_cd_output_workbook,
    recreate_cd_sheet,
    save_cd_output_workbook,
    write_dataframe_to_sheet,
)
from modules.CD.cd_001 import build_cd_001_dataframe


SHEET_NAME = "CD04"

DEFAULT_MIN_QTY_PAYMENTS = 10
DEFAULT_PEER_PERCENTILE = 0.90
DEFAULT_OUTPUT_LEVEL = "DETAIL"

VALID_OUTPUT_LEVELS = {
    "DETAIL",
    "PAYMENT",
}

PARAMETER_BOX_FILL = "D9EAF7"

OUTPUT_COLUMNS = [
    "CoCo",
    "Company",
    "KEY_DOC",
    "Payment DocEntry",
    "Payment DocNum",
    "Journal TransId",
    "Vendor Code",
    "Vendor Name",
    "KEY_VENDOR",
    "Payment Date",
    "Tax Date",
    "Payment Currency",
    "Company Main Currency",
    "Payment Amount",
    "Payment Amount USD",
    "Counter Reference",
    "Canceled",
    "Transfer Amount",
    "Cash Amount",
    "Check Amount",
    "Credit Card Amount",
    "Qty Payments",
]

DATE_COLUMNS = {
    "Payment Date",
    "Tax Date",
}

AMOUNT_COLUMNS = {
    "Payment Amount",
    "Payment Amount USD",
    "Transfer Amount",
    "Cash Amount",
    "Check Amount",
    "Credit Card Amount",
}

INTEGER_COLUMNS = {
    "Payment DocEntry",
    "Payment DocNum",
    "Journal TransId",
    "Qty Payments",
}


def get_control_config_value(control_config, key, default_value=""):
    """
    Return a control config value using case-insensitive key matching.

    The config reader normally exposes lowercase keys such as param1/param2,
    but this helper also supports PARAM1/PARAM2 style keys defensively.
    """
    key_normalized = normalize_text(key).lower()

    for config_key, config_value in control_config.items():
        if normalize_text(config_key).lower() == key_normalized:
            return config_value

    return default_value


def parse_integer_parameter(value, parameter_name):
    """
    Parse a positive integer parameter.
    """
    value_text = normalize_text(value)

    if value_text == "":
        raise ValueError(f"{parameter_name} is blank.")

    try:
        parsed_value = int(float(value_text))
    except ValueError as error:
        raise ValueError(
            f"{parameter_name} must be a positive integer. "
            f"Received: {value_text}"
        ) from error

    if parsed_value < 1:
        raise ValueError(
            f"{parameter_name} must be greater than or equal to 1. "
            f"Received: {parsed_value}"
        )

    return parsed_value


def parse_percentile_parameter(value, parameter_name):
    """
    Parse percentile parameter.

    Accepted values:
    - 0.90
    - 90%
    - 90

    Values greater than 1 and less than or equal to 100 are treated as
    percentages and divided by 100.
    """
    value_text = normalize_text(value)

    if value_text == "":
        raise ValueError(f"{parameter_name} is blank.")

    is_percentage = value_text.endswith("%")

    if is_percentage:
        value_text = value_text[:-1]

    value_text = value_text.replace(",", ".")

    try:
        parsed_value = float(value_text)
    except ValueError as error:
        raise ValueError(
            f"{parameter_name} must be a decimal between 0 and 1 "
            f"or a percentage between 0% and 100%. Received: {value}"
        ) from error

    if is_percentage or parsed_value > 1:
        parsed_value = parsed_value / 100

    if parsed_value <= 0 or parsed_value > 1:
        raise ValueError(
            f"{parameter_name} must be greater than 0 and less than or equal to 1. "
            f"Received: {parsed_value}"
        )

    return parsed_value


def parse_output_level_parameter(value):
    """
    Parse output level parameter.

    Accepted values:
    - DETAIL
    - PAYMENT
    """
    value_text = normalize_text(value).upper()

    if value_text == "":
        return DEFAULT_OUTPUT_LEVEL

    if value_text not in VALID_OUTPUT_LEVELS:
        raise ValueError(
            "OUTPUT_LEVEL must be DETAIL or PAYMENT. "
            f"Received: {value}"
        )

    return value_text


def split_parameter_items(parameter_text):
    """
    Split a config parameter into items.

    Semicolon, pipe and line breaks are supported as separators.

    Comma is intentionally not used as a main separator because some users may
    write decimal comma values such as 0,90.
    """
    value_text = normalize_text(parameter_text)

    if value_text == "":
        return []

    for separator in ["|", "\n", "\r", "\t"]:
        value_text = value_text.replace(separator, ";")

    return [
        normalize_text(item)
        for item in value_text.split(";")
        if normalize_text(item) != ""
    ]


def normalize_parameter_key(value):
    """
    Normalize a parameter key for matching.
    """
    value_text = normalize_text(value).upper()
    value_text = value_text.replace(" ", "")
    value_text = value_text.replace("-", "_")

    return value_text


def apply_named_cd_004_parameter(
    parameter_key,
    parameter_value,
    current_parameters,
):
    """
    Apply one named CD_004 parameter.
    """
    normalized_key = normalize_parameter_key(parameter_key)

    if normalized_key in [
        "MIN_QTY_PAYMENTS",
        "MINQTYPAYMENTS",
        "MIN_QTY",
        "MINQTY",
        "MIN",
    ]:
        current_parameters["min_qty_payments"] = parse_integer_parameter(
            value=parameter_value,
            parameter_name="MIN_QTY_PAYMENTS",
        )
        return current_parameters

    if normalized_key in [
        "PEER_PERCENTILE",
        "PEERPERCENTILE",
        "PERCENTILE",
        "PCTL",
    ]:
        current_parameters["peer_percentile"] = parse_percentile_parameter(
            value=parameter_value,
            parameter_name="PEER_PERCENTILE",
        )
        return current_parameters

    if normalized_key in [
        "OUTPUT_LEVEL",
        "OUTPUTLEVEL",
        "LEVEL",
    ]:
        current_parameters["output_level"] = parse_output_level_parameter(
            parameter_value
        )
        return current_parameters

    raise ValueError(
        f"Unknown CD_004 parameter: {parameter_key}. "
        "Allowed parameters are MIN_QTY_PAYMENTS, PEER_PERCENTILE and OUTPUT_LEVEL."
    )


def parse_named_items(items, current_parameters):
    """
    Parse named key=value items from PARAM1 or PARAM2.
    """
    positional_items = []

    for item in items:
        if "=" not in item:
            positional_items.append(item)
            continue

        key, value = item.split("=", 1)

        current_parameters = apply_named_cd_004_parameter(
            parameter_key=key,
            parameter_value=value,
            current_parameters=current_parameters,
        )

    return positional_items, current_parameters


def parse_cd_004_parameters(context):
    """
    Parse CD_004 parameters from the CD sheet control row.

    Preferred CD sheet setup:
        ID Control | Execute | PARAM1 | PARAM2
        CD004      | T       | 10     | 0.90

    Meaning:
        PARAM1 = MIN_QTY_PAYMENTS
        PARAM2 = PEER_PERCENTILE

    Additional supported examples:
        PARAM1 = MIN_QTY_PAYMENTS=10
        PARAM2 = PEER_PERCENTILE=0.90

        PARAM1 = MIN_QTY_PAYMENTS=10
        PARAM2 = PEER_PERCENTILE=0.90;OUTPUT_LEVEL=PAYMENT

        PARAM1 = MIN_QTY_PAYMENTS=10;PEER_PERCENTILE=0.90;OUTPUT_LEVEL=DETAIL
        PARAM2 =

    Defaults are used when PARAM1/PARAM2 are blank.
    """
    control_config = context.get("control", {})

    param1_text = normalize_text(
        get_control_config_value(
            control_config=control_config,
            key="param1",
            default_value="",
        )
    )

    param2_text = normalize_text(
        get_control_config_value(
            control_config=control_config,
            key="param2",
            default_value="",
        )
    )

    parameters = {
        "min_qty_payments": DEFAULT_MIN_QTY_PAYMENTS,
        "peer_percentile": DEFAULT_PEER_PERCENTILE,
        "output_level": DEFAULT_OUTPUT_LEVEL,
    }

    param1_items = split_parameter_items(param1_text)
    param2_items = split_parameter_items(param2_text)

    param1_positional_items, parameters = parse_named_items(
        items=param1_items,
        current_parameters=parameters,
    )

    param2_positional_items, parameters = parse_named_items(
        items=param2_items,
        current_parameters=parameters,
    )

    if len(param1_positional_items) > 1:
        raise ValueError(
            "CD_004 PARAM1 accepts one positional value only: MIN_QTY_PAYMENTS. "
            "For multiple values use named syntax, for example: "
            "MIN_QTY_PAYMENTS=10;PEER_PERCENTILE=0.90"
        )

    if len(param2_positional_items) > 2:
        raise ValueError(
            "CD_004 PARAM2 accepts at most two positional values: "
            "PEER_PERCENTILE;OUTPUT_LEVEL. "
            "For more values use named syntax."
        )

    if len(param1_positional_items) == 1:
        parameters["min_qty_payments"] = parse_integer_parameter(
            value=param1_positional_items[0],
            parameter_name="MIN_QTY_PAYMENTS",
        )

    if len(param2_positional_items) >= 1:
        first_param2_value = normalize_text(param2_positional_items[0]).upper()

        if first_param2_value in VALID_OUTPUT_LEVELS:
            parameters["output_level"] = parse_output_level_parameter(
                first_param2_value
            )
        else:
            parameters["peer_percentile"] = parse_percentile_parameter(
                value=param2_positional_items[0],
                parameter_name="PEER_PERCENTILE",
            )

    if len(param2_positional_items) == 2:
        parameters["output_level"] = parse_output_level_parameter(
            param2_positional_items[1]
        )

    return parameters


def calculate_company_peer_thresholds(
    vendor_summary_dataframe,
    min_qty_payments,
    peer_percentile,
):
    """
    Calculate peer thresholds by company.

    Each company receives a threshold based on the configured percentile of
    vendor payment counts. The effective threshold is the higher of:
    - min_qty_payments
    - the company percentile threshold

    This keeps the control focused on vendors with high payment frequency while
    still adapting to different company sizes and activity levels.
    """
    if vendor_summary_dataframe.empty:
        return {}

    thresholds = {}

    for company, group in vendor_summary_dataframe.groupby("Company", dropna=False):
        percentile_value = group["Qty Payments"].quantile(peer_percentile)

        if pd.isna(percentile_value):
            threshold = min_qty_payments
        else:
            threshold = max(min_qty_payments, int(round(float(percentile_value))))

        thresholds[company] = threshold

    return thresholds


def build_vendor_summary_dataframe(detail_dataframe):
    """
    Build vendor-level payment count summary used by CD_004.

    Aggregation rules:
    - Group by Company, Vendor Code and Vendor Name.
    - Qty Payments counts unique KEY_DOC values to avoid overcounting payments
      that have more than one line.
    - Total Amount USD sums Payment Amount USD.
    - First Payment is the earliest Payment Date.
    - Last Payment is the latest Payment Date.
    """
    if detail_dataframe.empty:
        return pd.DataFrame(
            columns=[
                "Company",
                "Vendor Code",
                "Vendor Name",
                "Qty Payments",
                "Total Amount USD",
                "First Payment",
                "Last Payment",
            ]
        )

    working_dataframe = detail_dataframe.copy()

    working_dataframe["Payment Date"] = pd.to_datetime(
        working_dataframe["Payment Date"],
        errors="coerce",
    )

    working_dataframe["Payment Amount USD"] = pd.to_numeric(
        working_dataframe["Payment Amount USD"],
        errors="coerce",
    )

    working_dataframe["Canceled"] = (
        working_dataframe["Canceled"]
        .fillna("N")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    working_dataframe = working_dataframe[working_dataframe["Canceled"] != "Y"].copy()

    summary_dataframe = (
        working_dataframe
        .groupby(
            [
                "Company",
                "Vendor Code",
                "Vendor Name",
            ],
            dropna=False,
        )
        .agg(
            **{
                "Qty Payments": ("KEY_DOC", "nunique"),
                "Total Amount USD": ("Payment Amount USD", "sum"),
                "First Payment": ("Payment Date", "min"),
                "Last Payment": ("Payment Date", "max"),
            }
        )
        .reset_index()
    )

    return summary_dataframe


def apply_output_level(output_dataframe, output_level):
    """
    Apply CD_004 output level.

    DETAIL:
        Keep all CD_001 detail rows for selected vendors.

    PAYMENT:
        Keep one row per Company + Vendor + KEY_DOC. This makes the output
        closer to the unique payment count used by Qty Payments.
    """
    if output_dataframe.empty:
        return output_dataframe

    if output_level == "DETAIL":
        return output_dataframe

    if output_level == "PAYMENT":
        return (
            output_dataframe
            .sort_values(
                by=[
                    "Company",
                    "Vendor Code",
                    "Vendor Name",
                    "KEY_DOC",
                    "Payment Date",
                ],
                ascending=[True, True, True, True, True],
                na_position="last",
            )
            .drop_duplicates(
                subset=[
                    "Company",
                    "Vendor Code",
                    "Vendor Name",
                    "KEY_DOC",
                ],
                keep="first",
            )
            .copy()
        )

    raise ValueError(
        "OUTPUT_LEVEL must be DETAIL or PAYMENT. "
        f"Received: {output_level}"
    )


def build_cd_004_result(
    detail_dataframe,
    min_qty_payments=DEFAULT_MIN_QTY_PAYMENTS,
    peer_percentile=DEFAULT_PEER_PERCENTILE,
    output_level=DEFAULT_OUTPUT_LEVEL,
):
    """
    Build CD_004 output dataframe and metadata.

    Returns:
        output_dataframe, company_thresholds
    """
    output_level = parse_output_level_parameter(output_level)

    if detail_dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), {}

    working_dataframe = detail_dataframe.copy()

    working_dataframe["Payment Date"] = pd.to_datetime(
        working_dataframe["Payment Date"],
        errors="coerce",
    )

    working_dataframe["Payment Amount USD"] = pd.to_numeric(
        working_dataframe["Payment Amount USD"],
        errors="coerce",
    )

    working_dataframe["Canceled"] = (
        working_dataframe["Canceled"]
        .fillna("N")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    working_dataframe = working_dataframe[working_dataframe["Canceled"] != "Y"].copy()

    if working_dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), {}

    vendor_summary_dataframe = build_vendor_summary_dataframe(
        detail_dataframe=working_dataframe,
    )

    if vendor_summary_dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), {}

    company_thresholds = calculate_company_peer_thresholds(
        vendor_summary_dataframe=vendor_summary_dataframe,
        min_qty_payments=min_qty_payments,
        peer_percentile=peer_percentile,
    )

    vendor_summary_dataframe["Company Qty Threshold"] = vendor_summary_dataframe[
        "Company"
    ].map(company_thresholds)

    selected_vendor_summary_dataframe = vendor_summary_dataframe[
        vendor_summary_dataframe["Qty Payments"]
        >= vendor_summary_dataframe["Company Qty Threshold"]
    ].copy()

    if selected_vendor_summary_dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), company_thresholds

    qty_payments_by_vendor = (
        selected_vendor_summary_dataframe
        .set_index(["Company", "Vendor Code", "Vendor Name"])["Qty Payments"]
        .to_dict()
    )

    selected_vendor_keys = set(qty_payments_by_vendor.keys())

    working_dataframe["Vendor Count Key"] = list(
        zip(
            working_dataframe["Company"],
            working_dataframe["Vendor Code"],
            working_dataframe["Vendor Name"],
        )
    )

    output_dataframe = working_dataframe[
        working_dataframe["Vendor Count Key"].isin(selected_vendor_keys)
    ].copy()

    if output_dataframe.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS), company_thresholds

    output_dataframe["Qty Payments"] = output_dataframe["Vendor Count Key"].map(
        qty_payments_by_vendor
    )

    output_dataframe = apply_output_level(
        output_dataframe=output_dataframe,
        output_level=output_level,
    )

    output_dataframe = output_dataframe[OUTPUT_COLUMNS]

    output_dataframe = output_dataframe.sort_values(
        by=[
            "Company",
            "Qty Payments",
            "Vendor Code",
            "Vendor Name",
            "Payment Date",
            "KEY_DOC",
        ],
        ascending=[True, False, True, True, True, True],
        na_position="last",
    ).reset_index(drop=True)

    return output_dataframe, company_thresholds


def build_cd_004_dataframe(
    detail_dataframe,
    min_qty_payments=DEFAULT_MIN_QTY_PAYMENTS,
    peer_percentile=DEFAULT_PEER_PERCENTILE,
    output_level=DEFAULT_OUTPUT_LEVEL,
):
    """
    Build CD_004 output dataframe from CD_001 detail rows.

    This wrapper is kept for consistency with other CD controls and for easier
    testing. It returns only the dataframe.
    """
    output_dataframe, _ = build_cd_004_result(
        detail_dataframe=detail_dataframe,
        min_qty_payments=min_qty_payments,
        peer_percentile=peer_percentile,
        output_level=output_level,
    )

    return output_dataframe


def write_cd_004_parameter_box(
    worksheet,
    dataframe,
    min_qty_payments,
    peer_percentile,
    output_level,
    company_thresholds,
):
    """
    Write a small CD_004 parameter box to the right of the output table.

    The box starts in the first rows, two columns after the last CD04 data
    column. It is written after the standard dataframe formatting is applied so
    the worksheet autofilter remains limited to the main data table.
    """
    start_row = 1
    start_col = len(dataframe.columns) + 3

    label_col = start_col
    value_col = start_col + 1

    label_col_letter = get_column_letter(label_col)
    value_col_letter = get_column_letter(value_col)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=PARAMETER_BOX_FILL,
    )

    title_cell = worksheet.cell(
        row=start_row,
        column=label_col,
        value="CD04 Parameters Used",
    )
    title_cell.font = Font(bold=True)
    title_cell.fill = header_fill

    worksheet.merge_cells(
        start_row=start_row,
        start_column=label_col,
        end_row=start_row,
        end_column=value_col,
    )

    parameter_header_row = start_row + 2

    worksheet.cell(
        row=parameter_header_row,
        column=label_col,
        value="Parameter",
    )
    worksheet.cell(
        row=parameter_header_row,
        column=value_col,
        value="Value",
    )

    for column in [label_col, value_col]:
        cell = worksheet.cell(row=parameter_header_row, column=column)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    parameter_rows = [
        ("Minimum Qty Payments", min_qty_payments),
        ("Peer Percentile", peer_percentile),
        ("Output Level", output_level),
    ]

    for row_offset, (parameter_name, parameter_value) in enumerate(
        parameter_rows,
        start=1,
    ):
        row_number = parameter_header_row + row_offset

        worksheet.cell(
            row=row_number,
            column=label_col,
            value=parameter_name,
        )

        value_cell = worksheet.cell(
            row=row_number,
            column=value_col,
            value=parameter_value,
        )

        if parameter_name == "Peer Percentile":
            value_cell.number_format = "0%"

    threshold_title_row = parameter_header_row + len(parameter_rows) + 3

    threshold_title_cell = worksheet.cell(
        row=threshold_title_row,
        column=label_col,
        value="CD04 Company Thresholds",
    )
    threshold_title_cell.font = Font(bold=True)
    threshold_title_cell.fill = header_fill

    worksheet.merge_cells(
        start_row=threshold_title_row,
        start_column=label_col,
        end_row=threshold_title_row,
        end_column=value_col,
    )

    threshold_header_row = threshold_title_row + 2

    worksheet.cell(
        row=threshold_header_row,
        column=label_col,
        value="Company",
    )
    worksheet.cell(
        row=threshold_header_row,
        column=value_col,
        value="Effective Qty Threshold",
    )

    for column in [label_col, value_col]:
        cell = worksheet.cell(row=threshold_header_row, column=column)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    if not company_thresholds:
        worksheet.cell(
            row=threshold_header_row + 1,
            column=label_col,
            value="No company thresholds calculated",
        )
    else:
        for row_offset, (company, threshold) in enumerate(
            sorted(
                company_thresholds.items(),
                key=lambda item: normalize_text(item[0]),
            ),
            start=1,
        ):
            row_number = threshold_header_row + row_offset

            worksheet.cell(
                row=row_number,
                column=label_col,
                value=company,
            )
            worksheet.cell(
                row=row_number,
                column=value_col,
                value=threshold,
            )

    worksheet.column_dimensions[label_col_letter].width = 35
    worksheet.column_dimensions[value_col_letter].width = 28


def print_cd_004_summary(
    detail_dataframe,
    output_dataframe,
    output_file,
    min_qty_payments,
    peer_percentile,
    output_level,
    company_thresholds,
):
    """
    Print CD_004 validation summary.
    """
    if output_dataframe.empty:
        selected_vendor_count = 0
        selected_key_doc_count = 0
        max_qty_payments = 0
    else:
        selected_vendor_count = output_dataframe["KEY_VENDOR"].nunique()
        selected_key_doc_count = output_dataframe["KEY_DOC"].nunique()
        max_qty_payments = output_dataframe["Qty Payments"].max()

    if company_thresholds:
        min_company_threshold = min(company_thresholds.values())
        max_company_threshold = max(company_thresholds.values())
    else:
        min_company_threshold = 0
        max_company_threshold = 0

    print("CD_004 validation summary")
    print("-------------------------")
    print(f"CD_004 source/detail rows: {len(detail_dataframe)}")
    print(f"CD_004 candidate rows written: {len(output_dataframe)}")
    print(f"CD_004 selected vendors: {selected_vendor_count}")
    print(f"CD_004 selected unique KEY_DOC: {selected_key_doc_count}")
    print(f"Minimum Qty Payments threshold: {min_qty_payments}")
    print(f"Peer percentile threshold: {peer_percentile:.0%}")
    print(f"Output level: {output_level}")
    print(f"Minimum company effective threshold: {min_company_threshold}")
    print(f"Maximum company effective threshold: {max_company_threshold}")
    print(f"Maximum Qty Payments: {max_qty_payments}")
    print(f"Output file: {output_file}")
    print(f"Output sheet: {SHEET_NAME}")
    print()


def run_cd_004(context):
    """
    Run CD_004.
    """
    cd_004_parameters = parse_cd_004_parameters(context)

    min_qty_payments = cd_004_parameters["min_qty_payments"]
    peer_percentile = cd_004_parameters["peer_percentile"]
    output_level = cd_004_parameters["output_level"]

    source_dataframe = load_cd_base_data(context)

    detail_dataframe = build_cd_001_dataframe(
        source_dataframe=source_dataframe,
        context=context,
    )

    output_dataframe, company_thresholds = build_cd_004_result(
        detail_dataframe=detail_dataframe,
        min_qty_payments=min_qty_payments,
        peer_percentile=peer_percentile,
        output_level=output_level,
    )

    if len(output_dataframe) == 0:
        print()
        print("WARNING: CD_004 generated 0 rows.")
        print("Possible causes:")
        print("- No vendors met the high payment count threshold.")
        print("- No rows with D/C = S were found.")
        print("- COMPANIES filter does not match Empr values.")
        print("- The input file is empty or the wrong sheet was read.")
        print("- The module PARAM1 matched the wrong input file.")
        print("- CD_004 PARAM1/PARAM2 thresholds are too restrictive.")
        print()

    output_file = get_cd_output_file(context)

    workbook = open_or_create_cd_output_workbook(output_file)

    worksheet = recreate_cd_sheet(
        workbook=workbook,
        sheet_name=SHEET_NAME,
    )

    write_dataframe_to_sheet(
        worksheet=worksheet,
        dataframe=output_dataframe,
    )

    apply_standard_cd_formatting(
        worksheet=worksheet,
        dataframe=output_dataframe,
        date_columns=DATE_COLUMNS,
        amount_columns=AMOUNT_COLUMNS,
        integer_columns=INTEGER_COLUMNS,
    )

    write_cd_004_parameter_box(
        worksheet=worksheet,
        dataframe=output_dataframe,
        min_qty_payments=min_qty_payments,
        peer_percentile=peer_percentile,
        output_level=output_level,
        company_thresholds=company_thresholds,
    )

    save_cd_output_workbook(
        workbook=workbook,
        output_file=output_file,
    )

    print_cd_004_summary(
        detail_dataframe=detail_dataframe,
        output_dataframe=output_dataframe,
        output_file=output_file,
        min_qty_payments=min_qty_payments,
        peer_percentile=peer_percentile,
        output_level=output_level,
        company_thresholds=company_thresholds,
    )
