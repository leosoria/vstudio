"""Common infrastructure for independent LBR Purchase Order controls.

This module contains input discovery, SAP Excel normalization, configuration
filters and output-workbook handling. Control-specific analytics do not belong
here. In particular, PO01 split detection and creator-coverage rules belong in
``modules/PO/po_001.py``.
"""

import os
import re
import tempfile
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


PO_INPUT_PREFIX = "LBR PO Lines"
PO_INPUT_EXTENSION = ".xlsx"
PO_INPUT_SHEET = "Sheet1"
PO_HEADER_ROW = 1
PO_OUTPUT_PREFIX = "LBR_Results_PO"
ALLOWED_INPUT_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
    ".xls",
}
INVALID_TEXT_VALUES = {
    "",
    "nan",
    "none",
    "<na>",
}
HEADER_FILL = "D9EAF7"
DATE_NUMBER_FORMAT = "dd/mm/yyyy"
AMOUNT_NUMBER_FORMAT = "#,##0.00"


PO_LINE_ALIASES = {
    "Company": (
        "CoCd",
        "Company",
        "Company Code",
        "BUKRS",
        "EKKO-BUKRS",
    ),
    "PO Number": (
        "Purch.Doc.",
        "Purchasing Document",
        "PO Number",
        "EBELN",
        "EKKO-EBELN",
    ),
    "PO Line": (
        "Item",
        "PO Line",
        "EBELP",
        "EKPO-EBELP",
    ),
    "Vendor Code": (
        "Vendor",
        "Vendor Code",
        "LIFNR",
        "EKKO-LIFNR",
    ),
    "PO Doc Date": (
        "Doc. Date",
        "Document Date",
        "PO Document Date",
        "BEDAT",
        "EKKO-BEDAT",
    ),
    "PO Creator ID": (
        "Created by",
        "Created By",
        "PO Creator ID",
        "ERNAM",
        "EKKO-ERNAM",
    ),
    "Item Code": (
        "Material",
        "Item Code",
        "MATNR",
        "EKPO-MATNR",
    ),
    "PO Quantity": (
        "PO Quantity",
        "Order Quantity",
        "MENGE",
        "EKPO-MENGE",
    ),
    "PO UOM": (
        "OUn.1",
        "Order Unit",
        "Unit of Measure",
        "MEINS",
        "EKPO-MEINS",
    ),
    "PO Unit Price": (
        "Net Price",
        "PO Unit Price",
        "NETPR",
        "EKPO-NETPR",
    ),
    "PO Price Unit": (
        "Per",
        "Price Unit",
        "PEINH",
        "EKPO-PEINH",
    ),
    "PO Doc Currency": (
        "Crcy.1",
        "Currency",
        "WAERS",
        "EKKO-WAERS",
    ),
    "PO Line Total": (
        "Net Value",
        "Net Order Value",
        "NETWR",
        "EKPO-NETWR",
    ),
    "PO Material Description": (
        "Short Text",
        "Description",
        "TXZ01",
        "EKPO-TXZ01",
    ),
    "PO Document Type": (
        "Type",
        "Document Type",
        "BSART",
        "EKKO-BSART",
    ),
    "Purchasing Organization": (
        "POrg",
        "Purchasing Organization",
        "EKORG",
        "EKKO-EKORG",
    ),
    "Purchasing Group": (
        "PGr",
        "Purchasing Group",
        "EKGRP",
        "EKKO-EKGRP",
    ),
    "Plant": (
        "Plnt",
        "Plant",
        "WERKS",
        "EKPO-WERKS",
    ),
    "PO Line Deleted": (
        "DCI",
        "Deletion Indicator",
        "LOEKZ",
        "EKPO-LOEKZ",
    ),
    "PO Delivery Completed": (
        "FIn",
        "Delivery Completed",
        "ELIKZ",
        "EKPO-ELIKZ",
    ),
    "PR Number": (
        "Purch.Req.",
        "Purchase Requisition",
        "PR Number",
        "BANFN",
        "EKPO-BANFN",
    ),
    "PR Line": (
        "Item.1",
        "Requisition Item",
        "PR Line",
        "BNFPO",
        "EKPO-BNFPO",
    ),
}


PO01_REQUIRED_FIELDS = (
    "Company",
    "PO Number",
    "PO Line",
    "Vendor Code",
    "PO Doc Date",
    "PO Creator ID",
    "Item Code",
)


def normalize_text(value):
    """Return trimmed text without pandas missing-value placeholders."""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()

    if text.casefold() in INVALID_TEXT_VALUES:
        return ""

    return text


def normalize_lookup(value):
    """Return an accent-, punctuation- and case-insensitive lookup key."""
    text = unicodedata.normalize(
        "NFKD",
        normalize_text(value),
    )

    text = "".join(
        char
        for char in text
        if not unicodedata.combining(char)
    )

    return re.sub(
        r"[^a-z0-9]+",
        "",
        text.casefold(),
    )


def normalize_identifier(value):
    """Normalize an SAP identifier while preserving leading zeroes in text."""
    text = normalize_text(value)

    if re.fullmatch(
        r"\d+\.0",
        text,
    ):
        return text[:-2]

    return text


def normalize_company(value):
    """Normalize an SAP company code to four digits when it is numeric."""
    text = normalize_identifier(
        value
    )

    if text.isdigit():
        return text.zfill(4)

    return text.upper()


def parse_config_date(
    value,
    field_name,
):
    """Parse one required configuration date."""
    text = normalize_text(
        value
    )

    if re.fullmatch(
        r"\d{4}-\d{2}-\d{2}",
        text,
    ):
        parsed = pd.to_datetime(
            text,
            format="%Y-%m-%d",
            errors="coerce",
        )
    else:
        parsed = pd.to_datetime(
            value,
            errors="coerce",
            dayfirst=True,
        )

    if pd.isna(parsed):
        raise ValueError(
            f"PO CONFIG {field_name} is "
            f"empty or invalid: {value!r}."
        )

    return pd.Timestamp(
        parsed
    ).normalize()


def get_period_suffix(module_config):
    """Return YYYYMMDD from the PO module TO date."""
    parsed_to = parse_config_date(
        module_config.get(
            "to",
            "",
        ),
        "TO",
    )

    return parsed_to.strftime(
        "%Y%m%d"
    )


def parse_config_companies(value):
    """Return normalized company codes or None when COMPANIES is ALL."""
    is_collection = isinstance(
        value,
        (
            list,
            tuple,
            set,
        ),
    )

    if (
        not is_collection
        and normalize_text(value).upper()
        == "ALL"
    ):
        return None

    if is_collection:
        raw_values = list(value)
    else:
        raw_values = re.split(
            r"[,;|\n]+",
            normalize_text(value),
        )

    companies = {
        normalize_company(item)
        for item in raw_values
        if normalize_company(item)
    }

    if not companies:
        raise ValueError(
            "PO CONFIG COMPANIES is empty; "
            "use ALL or provide at least "
            "one company."
        )

    return companies


def discover_po_input_files(
    input_folder,
    module_config,
):
    """Return period-matching PO Lines files below input_folder."""
    input_folder = Path(
        input_folder
    )

    if not input_folder.is_dir():
        raise FileNotFoundError(
            "PO input folder was not found: "
            f"{input_folder}"
        )

    period_suffix = get_period_suffix(
        module_config
    )

    expected_stem = normalize_lookup(
        f"{PO_INPUT_PREFIX} "
        f"{period_suffix}"
    )

    return sorted(
        path
        for path in input_folder.rglob("*")
        if path.is_file()
        and path.suffix.casefold()
        in ALLOWED_INPUT_EXTENSIONS
        and not path.name.startswith("~$")
        and normalize_lookup(path.stem)
        == expected_stem
    )


def find_po_input_file(context):
    """Require exactly one PO Lines workbook for CONFIG TO."""
    input_folder = Path(
        context["input_folder"]
    )

    module_config = context[
        "module"
    ]

    matches = discover_po_input_files(
        input_folder,
        module_config,
    )

    period_suffix = get_period_suffix(
        module_config
    )

    expected_name = (
        f"{PO_INPUT_PREFIX}_"
        f"{period_suffix}"
        f"{PO_INPUT_EXTENSION}"
    )

    if not matches:
        raise FileNotFoundError(
            "PO Lines input was not found. "
            "Expected exactly one file "
            f"equivalent to '{expected_name}' "
            f"below {input_folder}."
        )

    if len(matches) > 1:
        details = "\n".join(
            f"- {path}"
            for path in matches
        )

        raise ValueError(
            "Multiple PO Lines inputs were "
            f"found for {period_suffix}; "
            "expected exactly one:"
            f"\n{details}"
        )

    return matches[0]


def resolve_sheet_name(
    file_path,
    expected_sheet=PO_INPUT_SHEET,
):
    """Resolve the explicitly configured input sheet case-insensitively."""
    try:
        excel_file = pd.ExcelFile(
            file_path
        )
    except PermissionError as error:
        raise PermissionError(
            "PO input cannot be opened "
            "because it is locked: "
            f"{file_path}. Close Excel "
            "and retry."
        ) from error

    expected = normalize_lookup(
        expected_sheet
    )

    matches = [
        sheet
        for sheet in excel_file.sheet_names
        if normalize_lookup(sheet)
        == expected
    ]

    if len(matches) != 1:
        raise ValueError(
            "PO input must contain exactly "
            f"one sheet named "
            f"'{expected_sheet}'. "
            f"Available sheets: "
            f"{excel_file.sheet_names}"
        )

    return matches[0]


def inspect_input_workbook(
    file_path,
    sheet_name,
):
    """Fail on formulas in the selected data sheet and return dimensions."""
    try:
        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=False,
        )
    except PermissionError as error:
        raise PermissionError(
            "PO input cannot be read "
            "because it is locked: "
            f"{file_path}. Close Excel "
            "and retry."
        ) from error
    except Exception as error:
        raise ValueError(
            "PO input workbook is not "
            f"readable: {file_path}: "
            f"{error}"
        ) from error

    worksheet = workbook[
        sheet_name
    ]

    formula_count = sum(
        cell.data_type == "f"
        for row in worksheet.iter_rows()
        for cell in row
    )

    dimensions = {
        "physical_rows_including_header": (
            worksheet.max_row
        ),
        "physical_columns": (
            worksheet.max_column
        ),
        "formula_cells": (
            formula_count
        ),
    }

    workbook.close()

    if formula_count:
        raise ValueError(
            f"PO input sheet '{sheet_name}' "
            f"contains {formula_count} "
            "formulas. Export stored SAP "
            "values instead of formulas."
        )

    return dimensions


def resolve_po_columns(
    dataframe,
    required_fields=PO01_REQUIRED_FIELDS,
):
    """Resolve real SAP headers and return logical-to-physical mapping."""
    normalized_headers = {}

    for column in dataframe.columns:
        normalized_headers.setdefault(
            normalize_lookup(column),
            [],
        ).append(
            column
        )

    mapping = {}
    missing = []
    ambiguous = []

    for (
        logical_name,
        aliases,
    ) in PO_LINE_ALIASES.items():
        matches = []

        for alias in aliases:
            matches.extend(
                normalized_headers.get(
                    normalize_lookup(alias),
                    [],
                )
            )

        matches = list(
            dict.fromkeys(matches)
        )

        if len(matches) == 1:
            mapping[logical_name] = (
                matches[0]
            )
        elif len(matches) > 1:
            ambiguous.append(
                f"{logical_name}: "
                f"{matches}"
            )
        elif logical_name in required_fields:
            missing.append(
                logical_name
            )

    if missing or ambiguous:
        messages = []

        if missing:
            messages.append(
                "missing required fields: "
                f"{missing}"
            )

        if ambiguous:
            messages.append(
                "ambiguous fields: "
                f"{ambiguous}"
            )

        raise ValueError(
            "PO Lines header validation "
            f"failed ({'; '.join(messages)}). "
            f"Available headers: "
            f"{list(dataframe.columns)}"
        )

    return mapping


def remove_residual_rows(
    dataframe,
    columns,
):
    """Remove SAP export footer rows and fail on partially populated keys."""
    identity = [
        columns["Company"],
        columns["PO Number"],
        columns["PO Line"],
    ]

    blank = pd.DataFrame(
        {
            column: (
                dataframe[column]
                .map(normalize_text)
                .eq("")
            )
            for column in identity
        },
        index=dataframe.index,
    )

    residual_mask = blank.all(
        axis=1
    )

    partial_mask = (
        blank.any(axis=1)
        & ~residual_mask
    )

    if partial_mask.any():
        raise ValueError(
            "PO Lines contains "
            f"{int(partial_mask.sum())} "
            "rows with a partially blank "
            "Company + PO Number + "
            "PO Line key."
        )

    filtered = (
        dataframe
        .loc[~residual_mask]
        .copy()
    )

    return (
        filtered,
        int(
            residual_mask.sum()
        ),
    )


def filter_companies(
    dataframe,
    company_column,
    companies,
):
    """Filter rows to CONFIG COMPANIES or retain all when configured ALL."""
    normalized = (
        dataframe[
            company_column
        ]
        .map(normalize_company)
    )

    if companies is None:
        filtered = dataframe.copy()

        filtered[
            company_column
        ] = normalized

        return (
            filtered,
            0,
        )

    mask = normalized.isin(
        companies
    )

    filtered = (
        dataframe
        .loc[mask]
        .copy()
    )

    filtered[
        company_column
    ] = normalized.loc[
        mask
    ]

    return (
        filtered,
        int(
            (~mask).sum()
        ),
    )


def filter_dates(
    dataframe,
    date_column,
    date_from,
    date_to,
):
    """Filter PO document dates inclusively and fail on invalid dates."""
    raw = dataframe[
        date_column
    ]

    parsed = pd.to_datetime(
        raw,
        errors="coerce",
        dayfirst=True,
    ).dt.normalize()

    invalid = (
        raw.map(normalize_text).ne("")
        & parsed.isna()
    )

    if invalid.any():
        raise ValueError(
            "PO Lines contains "
            f"{int(invalid.sum())} "
            "nonblank invalid PO "
            "document dates."
        )

    mask = parsed.between(
        date_from,
        date_to,
        inclusive="both",
    )

    filtered = (
        dataframe
        .loc[mask]
        .copy()
    )

    filtered[
        date_column
    ] = parsed.loc[
        mask
    ]

    return (
        filtered,
        int(
            (~mask).sum()
        ),
    )


def normalize_po_dataframe(
    dataframe,
    columns,
):
    """Rename resolved columns and normalize identifiers."""
    renamed = dataframe.rename(
        columns={
            physical: logical
            for logical, physical
            in columns.items()
        }
    )

    for column in (
        "Company",
        "PO Number",
        "PO Line",
        "Vendor Code",
        "PO Creator ID",
        "Item Code",
        "PR Number",
        "PR Line",
    ):
        if column in renamed.columns:
            renamed[column] = (
                renamed[column]
                .map(normalize_identifier)
            )

    if "Company" in renamed.columns:
        renamed["Company"] = (
            renamed["Company"]
            .map(normalize_company)
        )

    return renamed


def load_po_lines(
    context,
    required_fields=PO01_REQUIRED_FIELDS,
):
    """Load, validate and CONFIG-filter PO Lines with audit metrics."""
    input_file = find_po_input_file(
        context
    )

    sheet_name = resolve_sheet_name(
        input_file
    )

    workbook_metrics = (
        inspect_input_workbook(
            input_file,
            sheet_name,
        )
    )

    try:
        raw = pd.read_excel(
            input_file,
            sheet_name=sheet_name,
            header=PO_HEADER_ROW - 1,
            dtype=object,
        )
    except PermissionError as error:
        raise PermissionError(
            "PO input cannot be read "
            "because it is locked: "
            f"{input_file}. Close Excel "
            "and retry."
        ) from error

    raw = (
        raw
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
    )

    if raw.empty:
        raise ValueError(
            f"PO input sheet "
            f"'{sheet_name}' contains "
            "no data rows."
        )

    columns = resolve_po_columns(
        raw,
        required_fields=required_fields,
    )

    (
        population,
        residual_rows,
    ) = remove_residual_rows(
        raw,
        columns,
    )

    if population.empty:
        raise ValueError(
            "PO input contains no valid "
            "rows after residual rows "
            "were excluded."
        )

    companies = parse_config_companies(
        context["module"].get(
            "companies",
            "",
        )
    )

    date_from = parse_config_date(
        context["module"].get(
            "from",
            "",
        ),
        "FROM",
    )

    date_to = parse_config_date(
        context["module"].get(
            "to",
            "",
        ),
        "TO",
    )

    if date_from > date_to:
        raise ValueError(
            f"PO CONFIG FROM "
            f"{date_from.date()} is "
            f"after TO {date_to.date()}."
        )

    (
        by_company,
        excluded_company,
    ) = filter_companies(
        population,
        columns["Company"],
        companies,
    )

    (
        by_date,
        excluded_date,
    ) = filter_dates(
        by_company,
        columns["PO Doc Date"],
        date_from,
        date_to,
    )

    normalized = normalize_po_dataframe(
        by_date,
        columns,
    )

    if companies is None:
        companies_metric = [
            "ALL",
        ]
    else:
        companies_metric = sorted(
            companies
        )

    metrics = {
        **workbook_metrics,
        "input_file": str(
            input_file
        ),
        "input_sheet": (
            sheet_name
        ),
        "header_row": (
            PO_HEADER_ROW
        ),
        "rows_read": (
            len(raw)
        ),
        "residual_rows": (
            residual_rows
        ),
        "rows_after_residuals": (
            len(population)
        ),
        "excluded_by_company": (
            excluded_company
        ),
        "excluded_by_date": (
            excluded_date
        ),
        "rows_after_config_filters": (
            len(normalized)
        ),
        "companies": (
            companies_metric
        ),
        "date_from": (
            date_from
        ),
        "date_to": (
            date_to
        ),
    }

    if normalized.empty:
        raise ValueError(
            "PO input has zero rows "
            "after CONFIG filters. "
            f"Rows read={len(raw)}, "
            f"residual={residual_rows}, "
            "excluded company="
            f"{excluded_company}, "
            "excluded date="
            f"{excluded_date}."
        )

    return normalized, metrics


def get_po_output_file(context):
    """Return output/LBR_Results_PO_YYYYMMDD.xlsx."""
    output_folder = Path(
        context["output_folder"]
    )

    period_suffix = get_period_suffix(
        context["module"]
    )

    return (
        output_folder
        / (
            f"{PO_OUTPUT_PREFIX}_"
            f"{period_suffix}.xlsx"
        )
    )


def excel_safe_value(value):
    """Convert pandas/numpy values to safe openpyxl scalars."""
    if value is None or pd.isna(value):
        return None

    if isinstance(
        value,
        pd.Timestamp,
    ):
        return value.to_pydatetime()

    if isinstance(
        value,
        (
            datetime,
            date,
            str,
            int,
            float,
            bool,
        ),
    ):
        return value

    return str(value)


def write_control_sheet(
    context,
    sheet_name,
    dataframe,
    date_columns=None,
    amount_columns=None,
):
    """Replace only one PO sheet while preserving all other sheets."""
    output_file = get_po_output_file(
        context
    )

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if output_file.exists():
        try:
            workbook = load_workbook(
                output_file
            )
        except PermissionError as error:
            raise PermissionError(
                "PO output is open or "
                f"locked: {output_file}. "
                "Close Excel and retry."
            ) from error
    else:
        workbook = Workbook()

        workbook.remove(
            workbook.active
        )

    if sheet_name in workbook.sheetnames:
        workbook.remove(
            workbook[sheet_name]
        )

    worksheet = workbook.create_sheet(
        sheet_name
    )

    safe = dataframe.copy()

    for column in safe.columns:
        safe[column] = (
            safe[column]
            .map(excel_safe_value)
        )

    for row in dataframe_to_rows(
        safe,
        index=False,
        header=True,
    ):
        worksheet.append(
            row
        )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL,
    )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

    date_columns = set(
        date_columns or []
    )

    amount_columns = set(
        amount_columns or []
    )

    header_positions = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    for column in date_columns:
        if column in header_positions:
            column_letter = (
                get_column_letter(
                    header_positions[column]
                )
            )

            for cell in worksheet[
                column_letter
            ][1:]:
                cell.number_format = (
                    DATE_NUMBER_FORMAT
                )

    for column in amount_columns:
        if column in header_positions:
            column_letter = (
                get_column_letter(
                    header_positions[column]
                )
            )

            for cell in worksheet[
                column_letter
            ][1:]:
                cell.number_format = (
                    AMOUNT_NUMBER_FORMAT
                )

    for index, column in enumerate(
        safe.columns,
        start=1,
    ):
        values = [
            str(column),
        ] + [
            normalize_text(value)
            for value in safe[column]
        ]

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = min(
            max(
                len(value)
                for value in values
            ) + 2,
            45,
        )

    temporary_path = None

    try:
        with tempfile.NamedTemporaryFile(
            prefix=(
                f".{output_file.stem}_"
            ),
            suffix=output_file.suffix,
            dir=output_file.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(
                temporary_file.name
            )

        workbook.save(
            temporary_path
        )

        os.replace(
            temporary_path,
            output_file,
        )

    except PermissionError as error:
        raise PermissionError(
            "PO output cannot be saved "
            "because it is open or locked: "
            f"{output_file}."
        ) from error

    finally:
        workbook.close()

        if (
            temporary_path is not None
            and temporary_path.exists()
        ):
            temporary_path.unlink()

    return output_file
