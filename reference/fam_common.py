"""
Common FAM utilities.

This module contains reusable functions for Fixed Asset Management controls.

Design rules:
- The FAM AR01 input file is resolved by code using:
    input/LBR FAM AR_YYYYMMDD.xlsx
- The FX rates input file is resolved by code using:
    input/FxRates_YYYYMMDD.xlsx
- YYYYMMDD comes from the FAM module TO date in config.xlsx.
- The FAM input file name is not configured in PARAM1/PARAM2.
- Every FAM control must be independent.
- Every FAM control must write only its own sheet.
- No FAM control must delete sheets from other controls.
- The final FAM output workbook follows:
    output/LBR_Results_FAM_YYYYMMDD.xlsx
"""

import warnings
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.ar_common import (
    read_sap_fx_rates_file,
    select_fx_rate_row,
)


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


FAM_INPUT_PREFIX = "LBR FAM AR_"
FAM_INPUT_EXTENSION = ".xlsx"

FX_INPUT_PREFIX = "FxRates_"
FX_INPUT_EXTENSION = ".xlsx"

FAM_HEADER_FILL = "D9EAF7"
FAM_DATE_FORMAT = "dd/mm/yyyy"
FAM_AMOUNT_FORMAT = '#,##0.00'
FAM_INTEGER_FORMAT = '#,##0'


def normalize_text(value):
    """
    Normalize text values.
    """
    if value is None:
        return ""

    value_text = str(value).strip()

    if value_text.lower() == "nan":
        return ""

    return value_text


def normalize_header(value):
    """
    Normalize a header for matching.
    """
    return normalize_text(value).lower()


def normalize_currency(value):
    """
    Normalize currency values.
    """
    return normalize_text(value).upper()


def normalize_company_output(value):
    """
    Normalize company code for output display.

    For FAM, keep leading zeroes because SAP company codes usually require them.
    """
    value_text = normalize_text(value)

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    return value_text


def normalize_company_for_filter(value):
    """
    Normalize company code for comparison.

    This allows config COMPANIES to contain either:
    - 0034
    - 34
    """
    value_text = normalize_text(value)

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if value_text.isdigit():
        return str(int(value_text))

    return value_text.upper()


def parse_companies_filter(value):
    """
    Parse module COMPANIES filter into normalized values.

    Accepted values:
    - ALL means no filter.
    - Blank means no filter.
    - 0034,0052
    - 0034;0052
    - 0034|0052
    - 0034 0052
    """
    value_text = normalize_text(value)

    if value_text == "":
        return []

    if value_text.upper() in ["ALL", "TODAS", "TODOS"]:
        return []

    for separator in [";", "|", "\n", "\r", "\t"]:
        value_text = value_text.replace(separator, ",")

    if "," not in value_text and " " in value_text:
        value_text = ",".join(value_text.split())

    companies = []

    for item in value_text.split(","):
        item = normalize_text(item)

        if item == "":
            continue

        if item.upper() in ["ALL", "TODAS", "TODOS"]:
            return []

        companies.append(normalize_company_for_filter(item))

    return companies


def to_datetime_value(value):
    """
    Convert a value to pandas datetime.

    Handles:
    - Excel dates
    - Python datetime values
    - ISO dates like 2026-02-28
    - SAP dates like 28.02.2026
    - Local dates like 28/02/2026
    """
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NaT

    if len(value_text) >= 10 and value_text[4:5] == "-" and value_text[7:8] == "-":
        return pd.to_datetime(value_text, errors="coerce", dayfirst=False)

    return pd.to_datetime(value_text, errors="coerce", dayfirst=True)


def parse_number(value):
    """
    Parse a number from SAP/Excel text.

    Handles:
    - 326,987.93
    - -326,987.93
    - 1.00
    - 0.48
    - 1.048,46
    - blank
    """
    if pd.isna(value):
        return pd.NA

    if isinstance(value, (int, float)):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NA

    value_text = value_text.replace(" ", "")

    is_negative = value_text.startswith("(") and value_text.endswith(")")

    if is_negative:
        value_text = value_text[1:-1]

    if "," in value_text and "." in value_text:
        last_comma = value_text.rfind(",")
        last_dot = value_text.rfind(".")

        if last_comma > last_dot:
            value_text = value_text.replace(".", "")
            value_text = value_text.replace(",", ".")
        else:
            value_text = value_text.replace(",", "")
    elif "," in value_text and "." not in value_text:
        value_text = value_text.replace(",", ".")

    try:
        parsed_number = float(value_text)
    except ValueError:
        return pd.NA

    if is_negative:
        parsed_number = parsed_number * -1

    return parsed_number


def get_period_suffix(module_config):
    """
    Return YYYYMMDD suffix using module TO date.
    """
    to_date = module_config.get("to", "")
    parsed_date = to_datetime_value(to_date)

    if pd.isna(parsed_date):
        raise ValueError(
            "Could not determine FAM period because module TO date is empty or invalid."
        )

    return parsed_date.strftime("%Y%m%d")


def get_fam_input_file(context):
    """
    Return expected FAM AR01 input file.

    Pattern:
        input/LBR FAM AR_YYYYMMDD.xlsx
    """
    input_folder = Path(context["input_folder"])
    module_config = context["module"]
    period_suffix = get_period_suffix(module_config)

    input_file = input_folder / f"{FAM_INPUT_PREFIX}{period_suffix}{FAM_INPUT_EXTENSION}"

    if not input_file.exists():
        raise FileNotFoundError(
            f"FAM input file not found.\n"
            f"Expected file: {input_file}\n"
            f"Please export AR01 and save it using this exact naming convention:\n"
            f"{FAM_INPUT_PREFIX}{period_suffix}{FAM_INPUT_EXTENSION}"
        )

    return input_file


def get_fx_rates_file(context):
    """
    Return expected FX rates input file.

    Pattern:
        input/FxRates_YYYYMMDD.xlsx
    """
    input_folder = Path(context["input_folder"])
    module_config = context["module"]
    period_suffix = get_period_suffix(module_config)

    fx_rates_file = input_folder / f"{FX_INPUT_PREFIX}{period_suffix}{FX_INPUT_EXTENSION}"

    if not fx_rates_file.exists():
        raise FileNotFoundError(
            f"FX rates file not found.\n"
            f"Expected file: {fx_rates_file}\n"
            f"Please save FX rates using this exact naming convention:\n"
            f"{FX_INPUT_PREFIX}{period_suffix}{FX_INPUT_EXTENSION}"
        )

    return fx_rates_file


def get_fam_output_file(context):
    """
    Return final FAM output workbook path.

    Pattern:
        output/LBR_Results_FAM_YYYYMMDD.xlsx
    """
    output_folder = Path(context["output_folder"])
    module_config = context["module"]
    period_suffix = get_period_suffix(module_config)

    return output_folder / f"LBR_Results_FAM_{period_suffix}.xlsx"


def read_input_file(file_path):
    """
    Read input Excel file into a dataframe.
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    return pd.read_excel(file_path)


def clean_dataframe(dataframe):
    """
    Clean dataframe columns and empty rows/columns.
    """
    result = dataframe.copy()

    result = result.dropna(axis=0, how="all")
    result = result.dropna(axis=1, how="all")

    result.columns = [
        normalize_text(column)
        for column in result.columns
    ]

    return result


def find_column(dataframe, possible_names):
    """
    Find a dataframe column using exact case-insensitive matching.
    """
    normalized_lookup = {
        normalize_header(column): column
        for column in dataframe.columns
    }

    for possible_name in possible_names:
        normalized_name = normalize_header(possible_name)

        if normalized_name in normalized_lookup:
            return normalized_lookup[normalized_name]

    return None


def require_columns(dataframe, required_columns):
    """
    Validate required columns.

    Returns a dictionary:
        logical_name -> actual dataframe column name
    """
    resolved_columns = {}
    missing_columns = []

    for logical_name, possible_names in required_columns.items():
        column_name = find_column(dataframe, possible_names)

        if column_name is None:
            missing_columns.append(
                f"{logical_name}: expected one of {possible_names}"
            )
        else:
            resolved_columns[logical_name] = column_name

    if missing_columns:
        raise ValueError(
            "Missing required FAM input columns:\n- "
            + "\n- ".join(missing_columns)
        )

    return resolved_columns


def get_optional_column(dataframe, possible_names):
    """
    Return a column if found, otherwise None.
    """
    return find_column(dataframe, possible_names)


def filter_by_company(dataframe, company_column, companies_filter):
    """
    Filter dataframe by module companies.

    If COMPANIES is blank or ALL, returns all rows.
    """
    result = dataframe.copy()
    companies = parse_companies_filter(companies_filter)

    if not companies:
        return result

    normalized_company = result[company_column].apply(normalize_company_for_filter)

    return result[normalized_company.isin(companies)].copy()


def load_fam_ar01_data(context):
    """
    Load and normalize the FAM AR01 input.
    """
    input_file = get_fam_input_file(context)

    dataframe = read_input_file(input_file)
    dataframe = clean_dataframe(dataframe)

    return dataframe


def load_fx_rates_data(context):
    """
    Load FX rates.

    Supported formats:

    1. Simple manual format:
        Currency
        Rate USD (To)
        Fecha rate USD

    2. SAP SQVI / TCURR export format:
        CgCâ / TCot
        De
        Para / A
        Vál.desde / Válido de
        Taxa câmbio / Tipo cambio
        Fator (origem) / Factor (de)
        Fator (para) / Factor (a)
    """
    fx_rates_file = get_fx_rates_file(context)

    raw_dataframe = read_input_file(fx_rates_file)
    raw_dataframe = clean_dataframe(raw_dataframe)

    currency_column = find_column(
        raw_dataframe,
        [
            "Currency",
            "Moeda",
            "Currency From",
            "From Currency",
            "Moneda",
        ],
    )
    rate_column = find_column(
        raw_dataframe,
        [
            "Rate USD (To)",
            "Rate USD",
            "FX Rate",
            "Rate",
            "Tipo de cambio",
            "Taxa",
        ],
    )
    rate_date_column = find_column(
        raw_dataframe,
        [
            "Fecha rate USD",
            "Rate Date",
            "Date",
            "Data",
            "Fecha",
        ],
    )

    if currency_column is not None and rate_column is not None and rate_date_column is not None:
        fx_dataframe = pd.DataFrame()
        fx_dataframe["Currency"] = raw_dataframe[currency_column].apply(normalize_currency)
        fx_dataframe["Rate USD (To)"] = raw_dataframe[rate_column].apply(parse_number)
        fx_dataframe["Fecha rate USD"] = raw_dataframe[rate_date_column].apply(to_datetime_value)

        fx_dataframe = fx_dataframe.dropna(
            subset=[
                "Currency",
                "Rate USD (To)",
            ],
            how="any",
        )

        fx_dataframe = fx_dataframe[fx_dataframe["Currency"] != ""].copy()

        if fx_dataframe.empty:
            raise ValueError(
                f"FX rates file has no valid rates: {fx_rates_file}"
            )

        fx_dataframe = fx_dataframe.sort_values(
            by=[
                "Currency",
                "Fecha rate USD",
            ],
            ascending=[
                True,
                False,
            ],
        )

        fx_dataframe = fx_dataframe.drop_duplicates(
            subset=[
                "Currency",
            ],
            keep="first",
        )

        fx_dataframe.attrs["source_format"] = "simple"
        fx_dataframe.attrs["source_file"] = str(fx_rates_file)

        return fx_dataframe

    sap_fx_dataframe, raw_sap_dataframe = read_sap_fx_rates_file(fx_rates_file)

    sap_fx_dataframe.attrs["source_format"] = "sap_tcurr"
    sap_fx_dataframe.attrs["source_file"] = str(fx_rates_file)
    sap_fx_dataframe.attrs["raw_rows"] = len(raw_sap_dataframe)
    sap_fx_dataframe.attrs["requested_date"] = context["module"].get("to", "")

    return sap_fx_dataframe


def build_fx_rate_lookup(fx_dataframe):
    """
    Build FX lookup by currency.

    Output lookup rate always means:
        local currency per 1 USD

    Therefore:
        USD = LC / Rate USD (To)

    SAP helper select_fx_rate_row returns final_fx_rate as LC -> USD multiplier.
    Therefore for SAP:
        Rate USD (To) = 1 / final_fx_rate
    """
    source_format = fx_dataframe.attrs.get("source_format", "simple")

    lookup = {}

    if source_format == "simple":
        for _, row in fx_dataframe.iterrows():
            currency = normalize_currency(row.get("Currency", ""))
            rate = row.get("Rate USD (To)", pd.NA)
            rate_date = row.get("Fecha rate USD", pd.NaT)

            if currency == "":
                continue

            if pd.isna(rate):
                continue

            lookup[currency] = {
                "rate": float(rate),
                "date": rate_date,
                "status": "OK",
                "source": "simple",
                "notes": "",
            }

        if "USD" not in lookup:
            lookup["USD"] = {
                "rate": 1.0,
                "date": pd.NaT,
                "status": "OK",
                "source": "default",
                "notes": "Default USD rate",
            }

        return lookup

    if source_format == "sap_tcurr":
        requested_date = fx_dataframe.attrs.get("requested_date", "")

        currencies = set()

        if "from_currency" in fx_dataframe.columns:
            currencies.update(
                fx_dataframe["from_currency"]
                .dropna()
                .astype(str)
                .str.strip()
                .str.upper()
                .tolist()
            )

        if "to_currency" in fx_dataframe.columns:
            currencies.update(
                fx_dataframe["to_currency"]
                .dropna()
                .astype(str)
                .str.strip()
                .str.upper()
                .tolist()
            )

        currencies = sorted(
            currency
            for currency in currencies
            if currency not in ["", "USD", "NAN", "NONE"]
        )

        for currency in currencies:
            result = select_fx_rate_row(
                fx_df=fx_dataframe,
                currency=currency,
                requested_date=requested_date,
                requested_rate_type="spot_bs",
                allow_previous_date=True,
                max_previous_days=10,
                allow_tcot_fallback=True,
                allow_future_date=False,
            )

            if result["status"] != "Found":
                lookup[currency] = {
                    "rate": pd.NA,
                    "date": pd.NaT,
                    "status": "Review",
                    "source": "sap_tcurr",
                    "notes": result.get("notes", ""),
                }
                continue

            final_fx_rate = result.get("final_fx_rate", None)

            if final_fx_rate is None or pd.isna(final_fx_rate) or float(final_fx_rate) == 0:
                lookup[currency] = {
                    "rate": pd.NA,
                    "date": result.get("sap_valid_date", pd.NaT),
                    "status": "Review",
                    "source": "sap_tcurr",
                    "notes": "Final FX rate is empty or zero.",
                }
                continue

            rate_usd_to = 1 / float(final_fx_rate)

            lookup[currency] = {
                "rate": rate_usd_to,
                "date": result.get("sap_valid_date", pd.NaT),
                "status": "OK",
                "source": "sap_tcurr",
                "notes": result.get("notes", ""),
                "selected_tcot": result.get("selected_tcot", ""),
                "direction": result.get("direction", ""),
                "sap_raw_rate": result.get("sap_raw_rate", None),
                "adjusted_rate": result.get("adjusted_rate", None),
                "factor_from": result.get("factor_from", None),
                "factor_to": result.get("factor_to", None),
            }

        lookup["USD"] = {
            "rate": 1.0,
            "date": pd.NaT,
            "status": "OK",
            "source": "default",
            "notes": "Default USD rate",
        }

        return lookup

    raise ValueError(
        f"Unsupported FX source format: {source_format}"
    )


def open_or_create_fam_output_workbook(output_file):
    """
    Open an existing FAM output workbook or create a new one.
    """
    output_file = Path(output_file)

    if output_file.exists():
        return load_workbook(output_file)

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def recreate_fam_sheet(workbook, sheet_name):
    """
    Recreate only one sheet.

    This deletes/replaces the requested control sheet only.
    It does not delete sheets from other controls.
    """
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    return workbook.create_sheet(sheet_name)


def write_dataframe_to_sheet(worksheet, dataframe):
    """
    Write a dataframe to an openpyxl worksheet.
    """
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(
            row=1,
            column=column_index,
            value=column_name,
        )

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )


def apply_standard_fam_formatting(
    worksheet,
    dataframe,
    date_columns=None,
    amount_columns=None,
    integer_columns=None,
):
    """
    Apply standard formatting to a FAM output sheet.
    """
    date_columns = date_columns or set()
    amount_columns = amount_columns or set()
    integer_columns = integer_columns or set()

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=FAM_HEADER_FILL,
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    column_name_by_index = {
        column_index: column_name
        for column_index, column_name in enumerate(dataframe.columns, start=1)
    }

    for column_index, column_name in column_name_by_index.items():
        column_letter = get_column_letter(column_index)

        if column_name in date_columns:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = FAM_DATE_FORMAT

        if column_name in amount_columns:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = FAM_AMOUNT_FORMAT

        if column_name in integer_columns:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = FAM_INTEGER_FORMAT

        max_length = len(normalize_text(column_name))

        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            value_length = len(normalize_text(value))

            if value_length > max_length:
                max_length = value_length

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def save_fam_output_workbook(workbook, output_file):
    """
    Save the FAM output workbook.

    If the workbook is open in Excel, show a clear message.
    """
    output_file = Path(output_file)

    try:
        workbook.save(output_file)
    except PermissionError as error:
        raise PermissionError(
            f"Could not save output workbook: {output_file}. "
            f"Close the workbook and run again."
        ) from error
