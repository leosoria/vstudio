"""
Common AR utilities.

This module contains reusable functions for AR controls.

What belongs here:
- File discovery in the input folder.
- Period-based file search using the module TO date.
- Reading Excel/CSV files.
- Salesforce dynamic header detection.
- SAP FX Rates reading from SQVI / TCURR export.
- DataFrame cleanup.
- Company filtering.
- Customer/ERP normalization.
- Common dictionaries used by AR controls.
- Loading common AR input data.

Current FX logic:
- FX source file is expected in input/ as:
    FxRates_YYYYMMDD.xlsx

- Example:
    FxRates_20260228.xlsx

- The file comes from SAP ECC SQVI based on TCURR.

- Expected SAP columns may be:
    CgCâ
    De
    Para
    Vál.desde
    Taxa câmbio
    Fator (origem)
    Fator (para)

- AR_001 currently needs BRL -> USD using Spot / Balance Sheet logic:
    requested_rate_type = "spot_bs"
    preferred TCot = "EN"


--------------------
------Fx Rates------
--------------------

La exportación desde SAP debe realizarse desde SQVI utilizando la tabla TCURR y exportando los siguientes campos:
- TCot
- De (Moneda de procedencia)
- A (Moneda de destino)
- Válido de
- Tipo cambio
- Factor (de)
- Factor (a)

Querie creada llamada IA_FXRATES

En SQVI, el único criterio obligatorio de selección debe ser el rango de fechas. Siempre que sea posible, los filtros de TCot y monedas deben dejarse en blanco para evitar excluir tipos de cambio válidos.

Lógica de negocio:
 - La tasa final de FX siempre debe representar:
       Moneda -> USD

El script debe determinar automáticamente el tipo de cambio correcto luego de cargar la exportación de SAP, considerando que:
- Distintas monedas pueden utilizar diferentes valores de TCot.
- Algunas monedas están registradas como Moneda → USD.
- Otras están registradas como USD → Moneda.
- Por lo tanto, en algunos casos será necesario invertir el tipo de cambio.
- Los factores deben conservarse para fines de validación, aunque no formen parte del resultado final.

Jerarquía recomendada de TCot:
- Tasas diarias / operativas: M
- Spot / Balance Sheet: EN
- Average / P&L: ZGB

- Jerarquía de búsqueda:
1. Buscar la fecha solicitada usando el TCot preferido.
2. Buscar la dirección directa:
       Moneda -> USD
3. Si no se encuentra la dirección directa, buscar la dirección inversa:
       USD -> Moneda
   y calcular:
       final_fx_rate = 1 / adjusted_rate
4. Si no se encuentra la fecha exacta y allow_previous_date=True,
   buscar la última fecha anterior disponible dentro de max_previous_days.
5. Si el fallback de TCot está habilitado, buscar valores alternativos de TCot.
6. Si aún no se encuentra la tasa, devolver Not found o generar error,
   según cómo lo maneje la función que llama.


Lógica de factores SAP:
- Si Factor (origem) y Factor (para) están en blanco, 0/0 o 1/1:
    adjusted_rate = tasa SAP original
- Si los factores son distintos de 1/1 o 0/0:
    adjusted_rate = raw_rate * factor_to / factor_from
- Si factor_from es cero y factor_to no es cero:
    marcar como factor inválido.

Tratamiento de tipos de cambio faltantes:
- Primer intento: fecha exacta y TCot preferido.
- Segundo intento: fecha exacta utilizando la dirección inversa de la moneda.
- Tercer intento (opcional): utilizar la fecha anterior disponible dentro de la tolerancia permitida.
- Cuarto intento (opcional): utilizar un TCot alternativo aprobado.
- Si no se encuentra un tipo de cambio válido, el script debe devolver "Not found" y nunca estimar o inventar un valor.

Examples:

Cierre mensual
TCot: EN
Moneda de procedencia: USD
Moneda de destino: BRL
Fecha: 28.02.2026
=1/UKURS


Históricos
TCot: M
Moneda de procedencia: USD
Moneda de destino: BRL
Fecha: 01.02.2026 - 28.02.2026

--------------------
--------------------
--------------------

    
"""

import warnings
from datetime import timedelta

import pandas as pd

from core.intercompanies import INTERCOMPANIES


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


ALLOWED_EXTENSIONS = ["", ".xlsx", ".xls", ".csv", ".txt"]

SALESFORCE_HEADER_KEYWORDS = [
    "Proprietário da conta",
    "Nome da conta",
    "Razão Social",
    "CNPJ",
    "Código ERP",
    "Limite Aprovado",
    "Status do Limite",
    "Validade",
]

RATE_TYPE_TO_PRIMARY_TCOT = {
    "daily": "M",
    "spot_bs": "EN",
    "average_pl": "ZGB",
}

RATE_TYPE_TO_TCOT_FALLBACKS = {
    "daily": ["M"],
    "spot_bs": ["EN", "M"],
    "average_pl": ["ZGB", "M"],
}

SAP_FX_COLUMN_ALIASES = {
    "tcot": [
        "tcot",
        "cgcâ",
        "cgc",
        "tipo cotação",
        "tipo cotacao",
        "tipo cotización",
        "tipo cotizacion",
    ],
    "from_currency": [
        "de",
        "from",
        "from currency",
        "moeda origem",
        "moneda origen",
    ],
    "to_currency": [
        "para",
        "a",
        "to",
        "to currency",
        "moeda destino",
        "moneda destino",
    ],
    "valid_from": [
        "vál.desde",
        "val.desde",
        "válido de",
        "valido de",
        "valid from",
        "fecha válida",
        "fecha valida",
    ],
    "raw_rate": [
        "taxa câmbio",
        "taxa cambio",
        "tipo cambio",
        "exchange rate",
        "rate",
    ],
    "factor_from": [
        "fator (origem)",
        "factor (de)",
        "factor de",
        "factor from",
    ],
    "factor_to": [
        "fator (para)",
        "factor (a)",
        "factor a",
        "factor to",
    ],
}


def has_allowed_extension(file_path):
    """
    Return True if the file extension is allowed.
    """
    return file_path.suffix.lower() in ALLOWED_EXTENSIONS


def to_datetime_value(value):
    """
    Convert a value to pandas datetime.

    Handles:
    - Excel dates
    - Python datetime values
    - ISO dates like 2026-02-28
    - SAP dates like 28.02.2026
    - Local dates like 28/02/2026
    """
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value

    value_text = str(value).strip()

    if value_text == "":
        return pd.NaT

    if len(value_text) >= 10 and value_text[4:5] == "-" and value_text[7:8] == "-":
        return pd.to_datetime(value_text, errors="coerce", dayfirst=False)

    return pd.to_datetime(value_text, errors="coerce", dayfirst=True)



def get_period_suffix(module_config):
    """
    Return YYYYMMDD based on the module TO date.

    Example:
    module_config["to"] = 2026-02-28
    returns:
    20260228
    """
    to_date = to_datetime_value(module_config.get("to", ""))

    if pd.isna(to_date):
        raise ValueError("Invalid module TO date. Could not build period suffix.")

    return to_date.strftime("%Y%m%d")


def get_rate_sheet_name(module_config):
    """
    Build old Exchange Rates sheet name based on module TO date.

    This function is kept for compatibility, but the current process uses
    SAP FX export files named FxRates_YYYYMMDD.xlsx.
    """
    to_date = to_datetime_value(module_config.get("to", ""))

    if pd.isna(to_date):
        raise ValueError("Invalid module TO date. Could not build rate sheet name.")

    return f"FX Rates - {to_date.strftime('%b %y')}"


def find_all_files_containing(base_folder, text_to_find):
    """
    Find all files where the file name contains a given text.
    """
    matching_files = []

    text_to_find = text_to_find.lower()

    for file_path in base_folder.iterdir():
        if not file_path.is_file():
            continue

        if file_path.name.startswith("~$"):
            continue

        file_name = file_path.name.lower()

        if text_to_find in file_name and has_allowed_extension(file_path):
            matching_files.append(file_path)

    return sorted(matching_files)


def find_first_file_containing(base_folder, text_to_find):
    """
    Find the first file where the file name contains a given text.
    """
    files = find_all_files_containing(base_folder, text_to_find)

    if len(files) == 0:
        return None

    return files[0]


def find_all_files_containing_all(base_folder, required_texts):
    """
    Find all files where the file name contains all required texts.

    Example:
    required_texts = ["ZTFI098", "20260228"]
    """
    matching_files = []

    required_texts_normalized = [
        str(text).strip().lower()
        for text in required_texts
        if str(text).strip() != ""
    ]

    for file_path in base_folder.iterdir():
        if not file_path.is_file():
            continue

        if file_path.name.startswith("~$"):
            continue

        if not has_allowed_extension(file_path):
            continue

        file_name = file_path.name.lower()

        match = True

        for required_text in required_texts_normalized:
            if required_text not in file_name:
                match = False
                break

        if match:
            matching_files.append(file_path)

    return sorted(matching_files)


def find_first_file_containing_all(base_folder, required_texts):
    """
    Find the first file where the file name contains all required texts.
    """
    files = find_all_files_containing_all(base_folder, required_texts)

    if len(files) == 0:
        return None

    return files[0]


def clean_dataframe(dataframe):
    """
    Remove empty columns and strip column names.
    """
    df = dataframe.copy()

    df = df.dropna(axis=1, how="all")
    df.columns = [str(column).strip() for column in df.columns]

    unnamed_columns = [
        column
        for column in df.columns
        if str(column).strip().lower().startswith("unnamed")
    ]

    for column in unnamed_columns:
        if df[column].isna().all():
            df = df.drop(columns=[column])

    return df


def read_table_file(file_path):
    """
    Read an Excel or CSV file using the first row as headers.
    """
    suffix = file_path.suffix.lower()

    try:
        if suffix == ".csv":
            df = pd.read_csv(file_path)
        elif suffix in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file extension: {file_path.suffix}")
    except PermissionError:
        raise PermissionError(
            "Python cannot open this file because access was denied. "
            "Close the file in Excel, make sure it is available locally in OneDrive, "
            f"and try again. File: {file_path}"
        )

    return clean_dataframe(df)


def read_preview_without_headers(file_path, number_of_rows=30):
    """
    Read the first rows without assuming headers.
    """
    suffix = file_path.suffix.lower()

    try:
        if suffix == ".csv":
            return pd.read_csv(file_path, header=None, nrows=number_of_rows)

        if suffix in [".xlsx", ".xls"]:
            return pd.read_excel(file_path, header=None, nrows=number_of_rows)
    except PermissionError:
        raise PermissionError(
            "Python cannot open this file because access was denied. "
            "Close the file in Excel, make sure it is available locally in OneDrive, "
            f"and try again. File: {file_path}"
        )

    raise ValueError(f"Unsupported file extension: {file_path.suffix}")


def detect_header_row(file_path, expected_keywords, max_rows_to_scan=30, minimum_matches=3):
    """
    Detect a real header row by scanning the first rows.

    Returns zero-based row index for pandas.
    Example:
    - Excel row 10 returns 9.
    """
    preview_df = read_preview_without_headers(
        file_path,
        number_of_rows=max_rows_to_scan,
    )

    expected_keywords_normalized = [
        keyword.strip().lower()
        for keyword in expected_keywords
    ]

    for row_index, row in preview_df.iterrows():
        row_values = []

        for value in row.tolist():
            value_text = str(value).strip().lower()

            if value_text in ["", "nan", "none"]:
                continue

            row_values.append(value_text)

        matches = 0

        for keyword in expected_keywords_normalized:
            if keyword in row_values:
                matches += 1

        if matches >= minimum_matches:
            return row_index

    return None


def read_file_with_detected_header(file_path, expected_keywords):
    """
    Detect a header row and read the file using that row.
    """
    header_row = detect_header_row(file_path, expected_keywords)

    if header_row is None:
        raise ValueError("Could not detect header row")

    suffix = file_path.suffix.lower()

    try:
        if suffix == ".csv":
            df = pd.read_csv(file_path, header=header_row)
        elif suffix in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path, header=header_row)
        else:
            raise ValueError(f"Unsupported file extension: {file_path.suffix}")
    except PermissionError:
        raise PermissionError(
            "Python cannot open this file because access was denied. "
            "Close the file in Excel, make sure it is available locally in OneDrive, "
            f"and try again. File: {file_path}"
        )

    df = clean_dataframe(df)

    return df, header_row


def combine_files(file_paths):
    """
    Read and combine multiple standard files into one DataFrame.
    """
    dataframes = []

    for file_path in file_paths:
        df = read_table_file(file_path)
        df["Source File"] = file_path.name
        dataframes.append(df)

    if len(dataframes) == 0:
        return pd.DataFrame()

    return pd.concat(dataframes, ignore_index=True)


def normalize_column_name(column_name):
    """
    Normalize a column name for comparisons.
    """
    return str(column_name).strip().lower()


def normalize_customer_key(value):
    """
    Normalize customer key.

    Replicates VBA-like behavior:
    - Numeric values become number-like text without leading zeros.
    - Non-numeric values become uppercase text.
    """
    if pd.isna(value):
        return ""

    value_text = str(value).strip()

    if value_text == "":
        return ""

    try:
        numeric_value = float(value_text)

        if numeric_value.is_integer():
            return str(int(numeric_value))

        return str(numeric_value)
    except ValueError:
        return value_text.upper()


def normalize_erp_code(value):
    """
    Normalize ERP customer code.
    """
    return normalize_customer_key(value)


def normalize_company_code(value):
    """
    Normalize company code to 4 digits.

    Examples:
    30 -> 0030
    0030 -> 0030
    52 -> 0052
    """
    if pd.isna(value):
        return ""

    value_text = str(value).strip()

    if value_text == "":
        return ""

    try:
        numeric_value = int(float(value_text))
        return str(numeric_value).zfill(4)
    except ValueError:
        return value_text.zfill(4)


def normalize_company_output(value):
    """
    Normalize company code for AR output without leading zeros.

    This is used in AR output sheets.

    Examples:
    - 0030 -> 30
    - 30 -> 30
    - 0034 -> 34
    - 34 -> 34
    - 0052 -> 52
    """

    if pd.isna(value):
        return ""

    value_text = str(value).strip()

    if value_text == "":
        return ""

    try:
        numeric_value = int(float(value_text))
        return str(numeric_value)
    except ValueError:
        normalized_value = value_text.lstrip("0")

        if normalized_value == "":
            return value_text

        return normalized_value



def parse_companies(companies_value):
    """
    Parse COMPANIES from Config.xlsx.

    Examples:
    ALL -> []
    0030,0034,0052 -> ["0030", "0034", "0052"]
    """
    if pd.isna(companies_value):
        return []

    companies_text = str(companies_value).strip()

    if companies_text.upper() == "ALL":
        return []

    parts = companies_text.replace(";", ",").split(",")

    companies = []

    for part in parts:
        company = normalize_company_code(part)

        if company != "":
            companies.append(company)

    return companies


def filter_dataframe_by_companies(dataframe, company_column, companies_value):
    """
    Filter a DataFrame by company codes.

    If COMPANIES = ALL, the original DataFrame is returned.
    """
    companies = parse_companies(companies_value)

    if len(companies) == 0:
        return dataframe.copy()

    df = dataframe.copy()

    if company_column not in df.columns:
        raise ValueError(f"Company column '{company_column}' was not found.")

    df["_NormalizedCompany"] = df[company_column].apply(normalize_company_code)

    df = df[df["_NormalizedCompany"].isin(companies)].copy()

    df = df.drop(columns=["_NormalizedCompany"])

    return df


def get_companies_summary(dataframe, company_column="Empresa"):
    """
    Return unique company codes found in a DataFrame.
    """
    if dataframe.empty:
        return []

    if company_column not in dataframe.columns:
        return []

    companies = dataframe[company_column].apply(normalize_company_code).dropna().unique()

    return sorted([company for company in companies if company != ""])


def to_number(value, default=0.0):
    """
    Convert a value to number.

    Handles:
    - numeric values
    - blanks
    - comma decimal separator
    - thousands separators
    """
    if pd.isna(value):
        return default

    if isinstance(value, (int, float)):
        return float(value)

    value_text = str(value).strip()

    if value_text == "":
        return default

    if "," in value_text and "." not in value_text:
        value_text = value_text.replace(",", ".")
    elif "," in value_text and "." in value_text:
        value_text = value_text.replace(",", "")

    try:
        return float(value_text)
    except ValueError:
        return default


def validate_required_columns(dataframe, required_columns):
    """
    Validate whether a DataFrame contains required columns.
    """
    actual_columns_normalized = {
        normalize_column_name(column)
        for column in dataframe.columns
    }

    missing_columns = []

    for required_column in required_columns:
        required_column_normalized = normalize_column_name(required_column)

        if required_column_normalized not in actual_columns_normalized:
            missing_columns.append(required_column)

    return missing_columns


def read_salesforce_file(file_path):
    """
    Read a Salesforce file using dynamic header detection.
    """
    df, header_row = read_file_with_detected_header(
        file_path,
        SALESFORCE_HEADER_KEYWORDS,
    )

    return {
        "file_path": file_path,
        "file_name": file_path.name,
        "header_row_pandas": header_row,
        "header_row_excel": header_row + 1,
        "dataframe": df,
    }


# ============================================================
# SAP FX rates from SQVI / TCURR export
# ============================================================

def find_fx_rates_file(input_folder, module_config):
    """
    Find SAP FX rates file using period suffix.

    Expected file name:
        FxRates_YYYYMMDD.xlsx

    Example:
        FxRates_20260228.xlsx
    """
    period_suffix = get_period_suffix(module_config)

    return find_first_file_containing_all(
        input_folder,
        ["FxRates", period_suffix],
    )


def build_sap_fx_rename_map(columns):
    """
    Build rename map for SAP FX export columns.
    """
    rename_map = {}

    normalized_columns = {
        normalize_column_name(column): column
        for column in columns
    }

    for internal_name, aliases in SAP_FX_COLUMN_ALIASES.items():
        for alias in aliases:
            alias_normalized = normalize_column_name(alias)

            if alias_normalized in normalized_columns:
                original_column = normalized_columns[alias_normalized]
                rename_map[original_column] = internal_name
                break

    return rename_map


def read_sap_fx_rates_file(fx_rates_file):
    """
    Read SAP SQVI / TCURR FX export and normalize columns.
    """
    raw_df = read_table_file(fx_rates_file)

    rename_map = build_sap_fx_rename_map(raw_df.columns)
    fx_df = raw_df.rename(columns=rename_map).copy()

    required_columns = [
        "tcot",
        "from_currency",
        "to_currency",
        "valid_from",
        "raw_rate",
        "factor_from",
        "factor_to",
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in fx_df.columns
    ]

    if missing_columns:
        raise ValueError(
            "Missing SAP FX required columns after normalization. "
            f"Missing: {missing_columns}. "
            f"Available columns: {list(raw_df.columns)}"
        )

    fx_df["tcot"] = fx_df["tcot"].astype(str).str.strip().str.upper()
    fx_df["from_currency"] = fx_df["from_currency"].astype(str).str.strip().str.upper()
    fx_df["to_currency"] = fx_df["to_currency"].astype(str).str.strip().str.upper()
    fx_df["valid_from"] = fx_df["valid_from"].apply(to_datetime_value)
    fx_df["raw_rate"] = fx_df["raw_rate"].apply(lambda value: to_number(value, default=0.0))
    fx_df["factor_from"] = fx_df["factor_from"].apply(lambda value: to_number(value, default=0.0))
    fx_df["factor_to"] = fx_df["factor_to"].apply(lambda value: to_number(value, default=0.0))

    fx_df = fx_df.dropna(subset=["valid_from"]).copy()

    return fx_df, raw_df


def get_requested_tcots(requested_rate_type, allow_tcot_fallback):
    """
    Return TCot search sequence for requested rate type.
    """
    requested_rate_type = str(requested_rate_type).strip().lower()

    if requested_rate_type not in RATE_TYPE_TO_PRIMARY_TCOT:
        raise ValueError(
            f"Invalid requested_rate_type: {requested_rate_type}. "
            f"Valid values: {list(RATE_TYPE_TO_PRIMARY_TCOT.keys())}"
        )

    if allow_tcot_fallback:
        return RATE_TYPE_TO_TCOT_FALLBACKS[requested_rate_type]

    return [RATE_TYPE_TO_PRIMARY_TCOT[requested_rate_type]]


def calculate_adjusted_rate(raw_rate, factor_from, factor_to):
    """
    Calculate SAP adjusted rate using factors.

    Rules:
    - blank/blank, 0/0 or 1/1: use raw rate as-is.
    - otherwise:
        adjusted_rate = raw_rate * factor_to / factor_from
    """
    raw_rate = to_number(raw_rate, default=0.0)
    factor_from = to_number(factor_from, default=0.0)
    factor_to = to_number(factor_to, default=0.0)

    if factor_from == 0 and factor_to == 0:
        return raw_rate, ""

    if factor_from == 1 and factor_to == 1:
        return raw_rate, ""

    if factor_from == 0 and factor_to != 0:
        return raw_rate, "Invalid factor: origin factor is zero"

    adjusted_rate = raw_rate * factor_to / factor_from

    return adjusted_rate, "Review factor"


def build_candidate_dates(
    requested_date,
    allow_previous_date=True,
    max_previous_days=10,
    allow_future_date=False,
    max_future_days=0,
):
    """
    Build date candidates for FX search.

    Exact date is always first.
    Previous dates are optional.
    Future dates are disabled by default.
    """
    requested_date = to_datetime_value(requested_date)

    if pd.isna(requested_date):
        raise ValueError("Invalid requested FX date.")

    requested_date = requested_date.normalize()

    candidates = [(requested_date, "")]

    if allow_previous_date:
        for days_back in range(1, max_previous_days + 1):
            candidate_date = requested_date - timedelta(days=days_back)
            candidates.append(
                (
                    candidate_date,
                    f"Used previous available date: {candidate_date.strftime('%Y-%m-%d')}",
                )
            )

    if allow_future_date:
        for days_forward in range(1, max_future_days + 1):
            candidate_date = requested_date + timedelta(days=days_forward)
            candidates.append(
                (
                    candidate_date,
                    f"Used future date: {candidate_date.strftime('%Y-%m-%d')}",
                )
            )

    return candidates


def select_fx_rate_row(
    fx_df,
    currency,
    requested_date,
    requested_rate_type="spot_bs",
    allow_previous_date=True,
    max_previous_days=10,
    allow_tcot_fallback=True,
    allow_future_date=False,
):
    """
    Select one FX rate from SAP export.

    Search hierarchy:
    1. Preferred TCot.
    2. Exact date.
    3. Currency -> USD.
    4. USD -> Currency.
    5. Previous dates if allowed.
    6. TCot fallback if allowed.

    Result always represents:
        Currency -> USD
    """
    currency = str(currency).strip().upper()
    requested_date = to_datetime_value(requested_date)

    if pd.isna(requested_date):
        raise ValueError("Invalid requested FX date.")

    requested_date = requested_date.normalize()

    requested_rate_type_normalized = str(requested_rate_type).strip().lower()

    tcots = get_requested_tcots(
        requested_rate_type_normalized,
        allow_tcot_fallback,
    )

    primary_tcot = RATE_TYPE_TO_PRIMARY_TCOT[requested_rate_type_normalized]

    date_candidates = build_candidate_dates(
        requested_date=requested_date,
        allow_previous_date=allow_previous_date,
        max_previous_days=max_previous_days,
        allow_future_date=allow_future_date,
        max_future_days=0,
    )

    for tcot in tcots:
        tcot_note = ""

        if tcot != primary_tcot:
            tcot_note = f"Used TCot fallback: {tcot}"

        for candidate_date, date_note in date_candidates:
            filtered_df = fx_df[
                (fx_df["tcot"] == tcot)
                & (fx_df["valid_from"] == candidate_date)
            ]

            if filtered_df.empty:
                continue

            direct_rows = filtered_df[
                (filtered_df["from_currency"] == currency)
                & (filtered_df["to_currency"] == "USD")
            ]

            if not direct_rows.empty:
                row = direct_rows.iloc[0]

                adjusted_rate, factor_note = calculate_adjusted_rate(
                    row["raw_rate"],
                    row["factor_from"],
                    row["factor_to"],
                )

                notes = [
                    note
                    for note in [date_note, tcot_note, factor_note]
                    if str(note).strip() != ""
                ]

                return {
                    "currency": currency,
                    "requested_rate_type": requested_rate_type_normalized,
                    "selected_tcot": tcot,
                    "direction": f"{currency}->USD",
                    "sap_valid_date": row["valid_from"],
                    "sap_raw_rate": row["raw_rate"],
                    "adjusted_rate": adjusted_rate,
                    "final_fx_rate": adjusted_rate,
                    "factor_from": row["factor_from"],
                    "factor_to": row["factor_to"],
                    "status": "Found",
                    "notes": "; ".join(notes),
                }

            inverse_rows = filtered_df[
                (filtered_df["from_currency"] == "USD")
                & (filtered_df["to_currency"] == currency)
            ]

            if not inverse_rows.empty:
                row = inverse_rows.iloc[0]

                adjusted_rate, factor_note = calculate_adjusted_rate(
                    row["raw_rate"],
                    row["factor_from"],
                    row["factor_to"],
                )

                if adjusted_rate == 0:
                    final_fx_rate = None
                    status = "Not found"
                    inverse_note = "Cannot invert zero adjusted rate"
                else:
                    final_fx_rate = 1 / adjusted_rate
                    status = "Found"
                    inverse_note = "Used inverse direction"

                notes = [
                    note
                    for note in [date_note, tcot_note, factor_note, inverse_note]
                    if str(note).strip() != ""
                ]

                return {
                    "currency": currency,
                    "requested_rate_type": requested_rate_type_normalized,
                    "selected_tcot": tcot,
                    "direction": f"USD->{currency}",
                    "sap_valid_date": row["valid_from"],
                    "sap_raw_rate": row["raw_rate"],
                    "adjusted_rate": adjusted_rate,
                    "final_fx_rate": final_fx_rate,
                    "factor_from": row["factor_from"],
                    "factor_to": row["factor_to"],
                    "status": status,
                    "notes": "; ".join(notes),
                }

    return {
        "currency": currency,
        "requested_rate_type": requested_rate_type_normalized,
        "selected_tcot": "",
        "direction": "",
        "sap_valid_date": pd.NaT,
        "sap_raw_rate": None,
        "adjusted_rate": None,
        "final_fx_rate": None,
        "factor_from": None,
        "factor_to": None,
        "status": "Not found",
        "notes": "Rate not available in SAP export for requested currency/date/type.",
    }


def get_fx_rate_from_sap_export(
    fx_rates_file,
    module_config,
    currency="BRL",
    requested_rate_type="spot_bs",
    allow_previous_date=True,
    max_previous_days=10,
    allow_tcot_fallback=True,
    allow_future_date=False,
):
    """
    Read FX rate from SAP SQVI / TCURR export.

    Current AR_001 usage:
    - currency = BRL
    - requested_rate_type = spot_bs
    - TCot priority = EN, then M if fallback is allowed
    - output rate = BRL -> USD
    """
    fx_df, raw_df = read_sap_fx_rates_file(fx_rates_file)

    result = select_fx_rate_row(
        fx_df=fx_df,
        currency=currency,
        requested_date=module_config.get("to", ""),
        requested_rate_type=requested_rate_type,
        allow_previous_date=allow_previous_date,
        max_previous_days=max_previous_days,
        allow_tcot_fallback=allow_tcot_fallback,
        allow_future_date=allow_future_date,
    )

    result["source_file"] = fx_rates_file
    result["raw_rows"] = len(raw_df)
    result["normalized_rows"] = len(fx_df)

    if result["status"] != "Found":
        raise ValueError(
            f"FX rate not found. Currency: {currency}. "
            f"Requested rate type: {requested_rate_type}. "
            f"File: {fx_rates_file}. "
            f"Notes: {result['notes']}"
        )

    if result["final_fx_rate"] is None or result["final_fx_rate"] == 0:
        raise ValueError(
            f"FX rate is empty or zero. Currency: {currency}. "
            f"File: {fx_rates_file}. "
            f"Details: {result}"
        )

    return result


def get_fx_rate_from_exchange_rates(
    exchange_rates_file,
    module_config,
    currency="BRL",
    rate_column="SPOT - BS",
):
    """
    Legacy function.

    Kept only for compatibility with older code.

    The current process should use:
        get_fx_rate_from_sap_export

    This function still reads the old Exchange Rates workbook if ever needed.
    """
    sheet_name = get_rate_sheet_name(module_config)

    try:
        rates_df = pd.read_excel(exchange_rates_file, sheet_name=sheet_name)
    except PermissionError:
        raise PermissionError(
            "Python cannot open this Exchange Rates file because access was denied. "
            "Close the file in Excel, make sure it is available locally in OneDrive, "
            f"and try again. File: {exchange_rates_file}"
        )

    rates_df = clean_dataframe(rates_df)

    if rates_df.empty:
        raise ValueError(f"Exchange rates sheet '{sheet_name}' is empty.")

    first_column = rates_df.columns[0]

    currency_normalized = currency.strip().upper()

    rate_rows = rates_df[
        rates_df[first_column].astype(str).str.strip().str.upper() == currency_normalized
    ]

    if rate_rows.empty:
        raise ValueError(
            f"Currency {currency} was not found in exchange rates sheet {sheet_name}."
        )

    if rate_column not in rates_df.columns:
        raise ValueError(
            f"Rate column '{rate_column}' was not found in exchange rates sheet {sheet_name}."
        )

    rate_value = rate_rows.iloc[0][rate_column]

    fx_rate = to_number(rate_value, default=0.0)

    if fx_rate == 0:
        raise ValueError(
            f"FX rate for {currency} in column '{rate_column}' is zero or invalid."
        )

    return fx_rate


def build_customer_erp_map(customer_df):
    """
    Build a dictionary mapping SAP customer number to ERP customer code.
    """
    customer_erp_map = {}

    if customer_df.empty:
        return customer_erp_map

    for _, row in customer_df.iterrows():
        sap_customer_number = row.get("Cliente", "")
        erp_customer_code = row.get("Grupo", "")

        sap_customer_key = normalize_customer_key(sap_customer_number)

        if sap_customer_key != "" and sap_customer_key not in customer_erp_map:
            customer_erp_map[sap_customer_key] = str(erp_customer_code).strip()

    return customer_erp_map


def build_intercompany_exclusion_set(intercompany_df=None):
    """
    Build a set of customers to exclude.

    Intercompany exclusions are maintained in:
        core/intercompanies.py

    The optional intercompany_df parameter is kept for future compatibility.
    """
    excluded_customers = set()

    for item in INTERCOMPANIES:
        customer = item.get("customer", "")
        key = normalize_customer_key(customer)

        if key != "":
            excluded_customers.add(key)

    return excluded_customers


def build_intercompany_detail_map():
    """
    Build a dictionary with intercompany exclusion details.

    Returns:
        {
            "534": "PROMONLOGICALIS TECNOLOGIA S/A",
            ...
        }
    """
    detail_map = {}

    for item in INTERCOMPANIES:
        customer = item.get("customer", "")
        excluded_intercompany = item.get("excluded_intercompany", "")

        key = normalize_customer_key(customer)

        if key != "":
            detail_map[key] = excluded_intercompany

    return detail_map


def build_salesforce_status_map(sf_current_df):
    """
    Build dictionary of ERP customer code to Salesforce limit status.
    """
    status_map = {}

    if sf_current_df.empty:
        return status_map

    for _, row in sf_current_df.iterrows():
        erp_customer_code = normalize_erp_code(row.get("Código ERP", ""))
        limit_status = str(row.get("Status do Limite", "")).strip()

        if erp_customer_code != "" and erp_customer_code not in status_map:
            status_map[erp_customer_code] = limit_status

    return status_map


def build_salesforce_limit_map(sf_current_df):
    """
    Build dictionary of valid current credit limits by ERP customer code.

    Logic:
    - ERP > 700000
    - ERP < 999999
    - Status is not Bloqueado
    - Status is not Expirado
    - Sum Limite Aprovado by ERP customer code
    """
    limit_map = {}

    if sf_current_df.empty:
        return limit_map

    for _, row in sf_current_df.iterrows():
        erp_customer_code = normalize_erp_code(row.get("Código ERP", ""))
        credit_limit = to_number(row.get("Limite Aprovado", 0), default=0.0)
        limit_status = str(row.get("Status do Limite", "")).strip()

        if erp_customer_code == "":
            continue

        erp_numeric = to_number(erp_customer_code, default=0.0)

        if erp_numeric <= 700000 or erp_numeric >= 999999:
            continue

        if limit_status.upper() in ["BLOQUEADO", "EXPIRADO"]:
            continue

        if erp_customer_code not in limit_map:
            limit_map[erp_customer_code] = credit_limit
        else:
            limit_map[erp_customer_code] += credit_limit

    return limit_map


def build_due120_set(sf_due120_df):
    """
    Build set of customers with credit limits due within 120 days.
    """
    due120_set = set()

    if sf_due120_df.empty:
        return due120_set

    for _, row in sf_due120_df.iterrows():
        erp_customer_code = normalize_erp_code(row.get("Código ERP", ""))

        if erp_customer_code != "":
            due120_set.add(erp_customer_code)

    return due120_set


def validate_ztfi_files_by_companies(ztfi_files, companies_value):
    """
    Validate that at least one ZTFI file exists for each requested company.

    This validation is only applied when COMPANIES is not ALL.
    """
    companies = parse_companies(companies_value)

    if len(companies) == 0:
        return {
            "status": "OK",
            "missing_companies": [],
        }

    found_companies = set()

    for file_path in ztfi_files:
        file_name = file_path.name

        for company in companies:
            short_company = str(int(company))

            patterns = [
                f"_{company}_",
                f"_{short_company}_",
            ]

            for pattern in patterns:
                if pattern in file_name:
                    found_companies.add(company)

    missing_companies = [
        company
        for company in companies
        if company not in found_companies
    ]

    return {
        "status": "OK" if len(missing_companies) == 0 else "REVIEW REQUIRED",
        "missing_companies": missing_companies,
    }


def load_ar_input_data(context):
    """
    Load common AR input data.

    This function reads:
    - ZTFI098 files
    - Customer file
    - Salesforce current credit limits
    - Salesforce due 120 file
    - SAP FX Rates file from SQVI / TCURR

    Intercompany exclusions are loaded from:
    - core/intercompanies.py

    File discovery uses the module TO date as YYYYMMDD suffix.

    Example:
    TO = 2026-02-28
    period_suffix = 20260228
    """
    input_folder = context["input_folder"]
    module_config = context["module"]

    period_suffix = get_period_suffix(module_config)

    ztfi_files = find_all_files_containing_all(
        input_folder,
        ["ZTFI098", period_suffix],
    )

    customer_file = find_first_file_containing_all(
        input_folder,
        ["LBR Customer", period_suffix],
    )

    sf_current_file = find_first_file_containing_all(
        input_folder,
        ["Customer_credit_limits_curr", period_suffix],
    )

    sf_due120_file = find_first_file_containing_all(
        input_folder,
        ["Credit_limits_due_120", period_suffix],
    )

    fx_rates_file = find_fx_rates_file(
        input_folder,
        module_config,
    )

    intercompany_file = None

    if len(ztfi_files) == 0:
        raise FileNotFoundError(f"No ZTFI098 files were found for period {period_suffix}.")

    if customer_file is None:
        raise FileNotFoundError(f"Customer file was not found for period {period_suffix}.")

    if sf_current_file is None:
        raise FileNotFoundError(
            f"Salesforce current credit limits file was not found for period {period_suffix}."
        )

    if sf_due120_file is None:
        raise FileNotFoundError(
            f"Salesforce due 120 file was not found for period {period_suffix}."
        )

    if fx_rates_file is None:
        raise FileNotFoundError(
            f"SAP FX Rates file was not found for period {period_suffix}. "
            f"Expected file name like: FxRates_{period_suffix}.xlsx"
        )

    ztfi_file_company_validation = validate_ztfi_files_by_companies(
        ztfi_files,
        module_config.get("companies", "ALL"),
    )

    ztfi_df_raw = combine_files(ztfi_files)

    ztfi_df = filter_dataframe_by_companies(
        ztfi_df_raw,
        "Empresa",
        module_config.get("companies", "ALL"),
    )

    customer_df = read_table_file(customer_file)

    intercompany_df = pd.DataFrame()

    sf_current = read_salesforce_file(sf_current_file)
    sf_due120 = read_salesforce_file(sf_due120_file)

    fx_rate_details = get_fx_rate_from_sap_export(
        fx_rates_file,
        module_config,
        currency="BRL",
        requested_rate_type="spot_bs",
        allow_previous_date=True,
        max_previous_days=10,
        allow_tcot_fallback=True,
        allow_future_date=False,
    )

    fx_rate_brl_to_usd = fx_rate_details["final_fx_rate"]

    return {
        "period_suffix": period_suffix,
        "ztfi_files": ztfi_files,
        "customer_file": customer_file,
        "intercompany_file": intercompany_file,
        "intercompany_source": "core/intercompanies.py",
        "sf_current_file": sf_current_file,
        "sf_due120_file": sf_due120_file,

        # New SAP FX source.
        "fx_rates_file": fx_rates_file,
        "fx_rate_details": fx_rate_details,

        # Compatibility key in case AR_001 prints or references old name.
        "exchange_rates_file": fx_rates_file,

        "ztfi_file_company_validation": ztfi_file_company_validation,
        "ztfi_companies": get_companies_summary(ztfi_df, "Empresa"),
        "ztfi_raw": ztfi_df_raw,
        "ztfi": ztfi_df,
        "customer": customer_df,
        "intercompany": intercompany_df,
        "salesforce_current": sf_current["dataframe"],
        "salesforce_due120": sf_due120["dataframe"],
        "salesforce_current_header_row": sf_current["header_row_excel"],
        "salesforce_due120_header_row": sf_due120["header_row_excel"],
        "fx_rate_brl_to_usd": fx_rate_brl_to_usd,
    }


# ============================================================
# AR output workbook helpers
# ============================================================

from openpyxl import Workbook, load_workbook


def get_ar_output_file(context):
    """
    Return final AR output file path.

    All AR controls write to the same workbook:
        output/LBR_Results_AR_YYYYMMDD.xlsx

    The YYYYMMDD suffix is based on the module TO date.
    """

    output_folder = context["output_folder"]
    module_config = context["module"]

    period_suffix = get_period_suffix(module_config)

    return output_folder / f"LBR_Results_AR_{period_suffix}.xlsx"


def open_or_create_ar_output_workbook(output_file):
    """
    Open the final AR output workbook if it exists.
    Otherwise create a new workbook.

    This allows every AR control to run independently.
    """

    if output_file.exists():
        return load_workbook(output_file)

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def recreate_ar_sheet(workbook, sheet_name):
    """
    Recreate a single AR sheet in the final workbook.

    Rules:
    - If the target sheet exists, delete only that sheet.
    - Do not delete sheets from other AR controls.
    - Create the requested sheet at the end of the workbook.
    """

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    return workbook.create_sheet(sheet_name)


def save_ar_output_workbook(workbook, output_file):
    """
    Save the final AR workbook.

    Ensures output folder exists before saving.
    """

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
