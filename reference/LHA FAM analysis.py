"""
modules/FAM/analysis.py  --  Analizador del modulo FAM (tests del Cuadro AF).

Lo invoca run_analysis.py (orquestador). La funcion publica es analyze(row, test, params)
y devuelve un DataFrame (que sera UNA hoja del Results).

Tests:
  FAM001  Cuadro AF (listado completo)
  FAM002  Fixed Assets With No Depreciation  (Amort. acum. fin = 0)
  FAM003  Fixed Assets With Zero Book Value   (VNC fin <= 0)
  FAM004  Comparacion vs Reporting Pack       (pivot por Clase AF + RP + Diferencia, por cia)
"""
import os
import pandas as pd
from core.analysis_base import load_import_df, col, INPUT_DIR
from core.rp_lookup import resolve_rp

# columna del pivot a comparar  ->  filas del Fixed_Assets (RP) que se SUMAN entre si.
# Cada metrica del Cuadro AF se arma sumando DOS lineas del RP (misma pestaña, misma fecha).
RP_LINES = {
    "Suma de CAP en fecha inicio":                 (89, 13),    # CAP inicio = L89 + L13
    "Suma de CAP en fecha de fin":                 (142, 37),   # CAP fin    = L142 + L37
    "Suma de VNC en fecha de fin":                 (222, 73),   # VNC fin    = L222 + L73
    "Suma de Amortización acumulada en fecha fin": (209, 67),   # Amort acum = L209 + L67
}
PIVOT_BASES = ["CAP histórico", "CAP en fecha inicio", "Amortiz.acumul.en fecha de inicio",
               "VNC en fecha inicio", "Capitalización", "Valoración", "CAP en fecha de fin",
               "VNC en fecha de fin", "Depreciación en fecha de fin",
               "Amortización acumulada en fecha fin"]


def analyze(row, test, params) -> pd.DataFrame:
    df = load_import_df(row)
    t = test.upper()
    if t in ("FAM001", "FA001"):                      # Cuadro AF: listado completo
        return df
    if t in ("FAM002", "FA002"):                      # Fixed Assets With No Depreciation
        c = col(df, "Amortización acumulada en fecha fin")
        out = df[df[c].fillna(0).abs() < 0.005].copy()
        out.attrs["highlight"] = c                    # columna analizada (la resalta el formato)
        return out
    if t == "FAM003":                                 # Fixed Assets With Zero Book Value
        c = col(df, "VNC en fecha de fin")
        out = df[df[c].fillna(0) <= 0.005].copy()
        out.attrs["highlight"] = c
        return out
    if t == "FAM004":                                 # comparacion vs Reporting Pack
        return _fam004(row, df)
    return pd.DataFrame([{"Test": test, "Resultado": "sin logica definida"}])


def _find_rp_file(code, date_to):
    # v2: selecciona por compañia (alias) + FYxx (ejercicio marzo-febrero) + mes,
    #     no por "code in filename". Logica en core/rp_lookup.resolve_rp.
    import logging
    path = resolve_rp(INPUT_DIR, code, date_to.year, date_to.month)
    if path:
        logging.info("    RP %-5s -> %s", code, os.path.basename(path))
    else:
        fy = (date_to.year + 1 if date_to.month >= 3 else date_to.year) % 100
        logging.warning("    RP %-5s -> NO ENCONTRADO (FY%02d, mes %d) en %s",
                        code, fy, date_to.month, INPUT_DIR)
    return path


def _rp_values(path, date_to):
    """Lee solo lo necesario de Fixed_Assets (streaming): fecha del To en fila 5 y las lineas."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if s.lower().replace(" ", "_") == "fixed_assets"), None)
    if sheet is None:
        wb.close(); return None
    ws = wb[sheet]
    all_rows = {r for rows in RP_LINES.values() for r in rows}   # todas las filas a leer
    needed = {5} | all_rows
    last = max(needed)
    captured = {}
    for i, rowcells in enumerate(ws.iter_rows(values_only=True), start=1):
        if i in needed:
            captured[i] = rowcells
        if i >= last:
            break
    wb.close()

    fechas = captured.get(5, ())
    target = None
    for idx, v in enumerate(fechas):            # idx 0-based -> columna idx+1
        if hasattr(v, "date") and v.date() == date_to:
            target = idx
            break
    if target is None:
        return None

    def _cell(rownum):
        rowvals = captured.get(rownum, ())
        v = rowvals[target] if target < len(rowvals) else None
        return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

    out = {}
    for pivcol, rows in RP_LINES.items():
        vals = [x for x in (_cell(r) for r in rows) if x is not None]
        out[pivcol] = sum(vals) if vals else None   # suma las 2 lineas; None si ninguna trae numero
    return out


def _pivot_one(sub_df):
    clase = col(sub_df, "Clase AF")
    cols = {}
    for b in PIVOT_BASES:
        try:
            cols[f"Suma de {b}"] = col(sub_df, b)
        except KeyError:
            pass
    piv = sub_df.groupby(clase)[list(cols.values())].sum()
    piv.columns = list(cols.keys())
    piv = piv.reset_index().rename(columns={clase: "Clase AF"})
    total = piv.drop(columns=["Clase AF"]).sum(numeric_only=True)
    total["Clase AF"] = "Total general"
    block = pd.concat([piv, pd.DataFrame([total])[piv.columns]], ignore_index=True)
    return block, list(piv.columns)


def _fam004(row, df: pd.DataFrame) -> pd.DataFrame:
    """Por compania: titulo + pivot por Clase AF + Total general + RP + Diferencia, separados."""
    coco = "CoCo" if "CoCo" in df.columns else None
    codes = [c.code for c in row.companies] if row.companies else (
        sorted(df[coco].dropna().unique()) if coco else [row.scope])
    multi = len(codes) > 1

    sections = []
    for code in codes:
        sub = df[df[coco] == code] if coco else df
        if sub.empty:
            continue
        block, cols = _pivot_one(sub)
        empty = {c: None for c in cols}
        rp_row = dict(empty); rp_row["Clase AF"] = "RP"
        dif_row = dict(empty); dif_row["Clase AF"] = "Diferencia"

        rpfile = _find_rp_file(code, row.date_to)
        if rpfile is None:
            rp_row["Clase AF"] = f"RP (no se encontro el Reporting Pack de {code} en input/)"
        else:
            vals = _rp_values(rpfile, row.date_to)
            if vals is None:
                rp_row["Clase AF"] = f"RP (no se encontro la fecha {row.date_to} en {os.path.basename(rpfile)})"
            else:
                total_row = block[block["Clase AF"] == "Total general"].iloc[0]
                for pivcol, v in vals.items():
                    if pivcol in rp_row:
                        rp_row[pivcol] = v
                        try:
                            dif_row[pivcol] = float(total_row[pivcol]) - float(v)
                        except (TypeError, ValueError):
                            dif_row[pivcol] = None

        # titulo de compania (barra) -- siempre, para que cada cuadro quede claro
        hdr = dict(empty); hdr["Clase AF"] = f"=== {code} ==="
        sections.append(pd.DataFrame([hdr])[cols])
        sections.append(block)
        sections.append(pd.DataFrame([empty, rp_row, dif_row])[cols])
        # margen entre companias
        if multi:
            sections.append(pd.DataFrame([dict(empty), dict(empty), dict(empty)])[cols])

    if not sections:
        return pd.DataFrame([{"Clase AF": "sin datos"}])
    return pd.concat(sections, ignore_index=True)
