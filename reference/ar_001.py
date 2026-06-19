"""
AR_001 - Debtors Age Analysis.

This module generates AR01.

What it does:
1. Loads common AR input data using ar_common.py.
2. Uses Config.xlsx module TO date to identify period files.
3. Uses Config.xlsx COMPANIES to filter ZTFI098 data.
4. Reads BRL to USD FX rate from SAP FX Rates file.
5. Combines all ZTFI098 files for the period.
6. Creates AR01 using ZTFI098 and Customer mapping.
7. Excludes Intercompany customers.
8. Calculates aging buckets.
9. Enriches AR01 with Salesforce credit limit information.
10. Exports AR01 to the shared AR output workbook.

Shared output rule:
- All AR controls write to:
    output/LBR_Results_AR_YYYYMMDD.xlsx
- AR_001 creates the workbook if it does not exist.
- AR_001 replaces only AR01.
- AR_001 does not delete AR02, AR03, AR04, AR05, or AR06.

This module is called from run.py through run_ar_001(context).
"""

import pandas as pd

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.ar_common import (
    build_customer_erp_map,
    build_due120_set,
    build_intercompany_exclusion_set,
    build_salesforce_limit_map,
    build_salesforce_status_map,
    get_ar_output_file,
    load_ar_input_data,
    normalize_company_output,
    normalize_customer_key,
    normalize_erp_code,
    open_or_create_ar_output_workbook,
    recreate_ar_sheet,
    save_ar_output_workbook,
    to_datetime_value,
    to_number,
    validate_required_columns,
)


AR01_SHEET_NAME = "AR01"

AR01_COLUMNS = [
    "Company",
    "Clave de operación",
    "SAP Customer Number",
    "ERP Customer Code",
    "Customer Name",
    "Company Main Currency",
    "Company System Currency",
    "Credit Limit Currency",
    "Credit Limit Main Currency",
    "Credit Limit USD",
    "Posting Date",
    "Due Date",
    "Document Currency",
    "FX Rate",
    "FC Debit",
    "FC Credit",
    "FC Balance",
    "Document Number",
    "Document Type",
    "Balance",
    "Outstanding Balance",
    "DaysPastDue",
    "Future",
    "0-30",
    "31-60",
    "61-90",
    "91-120",
    "121+",
    "Credit Limit Due 120 Flag",
    "Credit Limit Diagnostic",
]

REQUIRED_COLUMNS = {
    "ZTFI098": [
        "Empresa",
        "Nº doc.",
        "Cliente",
        "Grp. Emp.",
        "Nome Cli/For",
        "Data doc.",
        "Dt.base prazo pgto",
        "Moeda transação",
        "Val. da Transação",
        "Moeda BRL",
        "Moeda USD",
        "Referência",
    ],
    "Customer": [
        "Cliente",
        "Grupo",
    ],
    "Salesforce Current": [
        "Código ERP",
        "Limite Aprovado",
        "Status do Limite",
    ],
    "Salesforce Due 120": [
        "Código ERP",
    ],
}


def print_header(title):
    """
    Print a section header.
    """

    print(title)
    print("-" * len(title))


def validate_inputs(input_data):
    """
    Validate required columns for AR_001 input data.
    """

    validation_results = []

    datasets = [
        ("ZTFI098", input_data["ztfi"], REQUIRED_COLUMNS["ZTFI098"]),
        ("Customer", input_data["customer"], REQUIRED_COLUMNS["Customer"]),
        (
            "Salesforce Current",
            input_data["salesforce_current"],
            REQUIRED_COLUMNS["Salesforce Current"],
        ),
        (
            "Salesforce Due 120",
            input_data["salesforce_due120"],
            REQUIRED_COLUMNS["Salesforce Due 120"],
        ),
    ]

    has_errors = False

    for dataset_name, dataframe, required_columns in datasets:
        missing_columns = validate_required_columns(dataframe, required_columns)

        if len(missing_columns) > 0:
            has_errors = True
            status = "REVIEW REQUIRED"
        else:
            status = "OK"

        validation_results.append(
            {
                "dataset": dataset_name,
                "status": status,
                "missing_columns": missing_columns,
            }
        )

    company_validation = input_data.get("ztfi_file_company_validation", {})

    if company_validation.get("status") == "REVIEW REQUIRED":
        has_errors = True

        validation_results.append(
            {
                "dataset": "ZTFI098 files by company",
                "status": "REVIEW REQUIRED",
                "missing_columns": [
                    f"Missing ZTFI file for company {company}"
                    for company in company_validation.get("missing_companies", [])
                ],
            }
        )

    return validation_results, has_errors


def create_ar01_from_ztfi(ztfi_df, customer_df, intercompany_df, input_data):
    """
    Create AR01 from ZTFI098.

    This replicates the VBA Populate_AR01_From_ZTFI098 logic.
    """

    customer_erp_map = build_customer_erp_map(customer_df)
    excluded_customers = build_intercompany_exclusion_set(intercompany_df)

    fx_rate_brl_to_usd = input_data["fx_rate_brl_to_usd"]

    ar_rows = []

    for _, row in ztfi_df.iterrows():
        sap_customer_number = row.get("Cliente", "")
        sap_customer_key = normalize_customer_key(sap_customer_number)

        if sap_customer_key in excluded_customers:
            continue

        erp_customer_code_from_zt = str(row.get("Grp. Emp.", "")).strip()

        if sap_customer_key in customer_erp_map:
            erp_customer_code = str(customer_erp_map[sap_customer_key]).strip()
        else:
            erp_customer_code = erp_customer_code_from_zt

        erp_customer_code = normalize_erp_code(erp_customer_code)

        document_currency = str(row.get("Moeda transação", "")).strip().upper()
        transaction_amount = to_number(row.get("Val. da Transação", 0), default=0.0)

        if document_currency == "BRL":
            applied_fx_rate = fx_rate_brl_to_usd
        else:
            applied_fx_rate = 1.0

        open_amount_usd = transaction_amount * applied_fx_rate

        if open_amount_usd >= 0:
            fc_debit = open_amount_usd
            fc_credit = 0.0
        else:
            fc_debit = 0.0
            fc_credit = open_amount_usd

        ar_rows.append(
            {
                "Company": normalize_company_output(row.get("Empresa", "")),
                "Clave de operación": row.get("Referência", ""),
                "SAP Customer Number": sap_customer_number,
                "ERP Customer Code": erp_customer_code,
                "Customer Name": row.get("Nome Cli/For", ""),
                "Company Main Currency": row.get("Moeda BRL", ""),
                "Company System Currency": row.get("Moeda USD", ""),
                "Credit Limit Currency": "",
                "Credit Limit Main Currency": "",
                "Credit Limit USD": "",
                "Posting Date": to_datetime_value(row.get("Data doc.", "")),
                "Due Date": to_datetime_value(row.get("Dt.base prazo pgto", "")),
                "Document Currency": document_currency,
                "FX Rate": applied_fx_rate,
                "FC Debit": fc_debit,
                "FC Credit": fc_credit,
                "FC Balance": open_amount_usd,
                "Document Number": row.get("Nº doc.", ""),
                "Document Type": "Invoice",
                "Balance": open_amount_usd,
                "Outstanding Balance": open_amount_usd,
                "DaysPastDue": pd.NA,
                "Future": 0.0,
                "0-30": 0.0,
                "31-60": 0.0,
                "61-90": 0.0,
                "91-120": 0.0,
                "121+": 0.0,
                "Credit Limit Due 120 Flag": "",
                "Credit Limit Diagnostic": "",
            }
        )

    return pd.DataFrame(ar_rows, columns=AR01_COLUMNS)


def calculate_ar01_aging(ar01_df, module_config):
    """
    Calculate aging buckets for AR01.

    Cutoff date comes from module CONFIG column TO.
    """

    df = ar01_df.copy()

    cutoff_date = to_datetime_value(module_config.get("to", ""))

    if pd.isna(cutoff_date):
        raise ValueError("Invalid module TO date. Could not calculate aging.")

    aging_columns = ["Future", "0-30", "31-60", "61-90", "91-120", "121+"]

    for column in aging_columns:
        df[column] = 0.0

    df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce")
    df["Outstanding Balance"] = pd.to_numeric(
        df["Outstanding Balance"],
        errors="coerce",
    ).fillna(0)

    valid_due_date = df["Due Date"].notna()

    df["DaysPastDue"] = pd.NA

    df.loc[valid_due_date, "DaysPastDue"] = (
        cutoff_date - df.loc[valid_due_date, "Due Date"]
    ).dt.days

    df["DaysPastDue"] = df["DaysPastDue"].astype("Int64")

    df.loc[
        valid_due_date & (df["DaysPastDue"] <= 0),
        "Future",
    ] = df["Outstanding Balance"]

    df.loc[
        valid_due_date & (df["DaysPastDue"] > 0) & (df["DaysPastDue"] <= 30),
        "0-30",
    ] = df["Outstanding Balance"]

    df.loc[
        valid_due_date & (df["DaysPastDue"] > 30) & (df["DaysPastDue"] <= 60),
        "31-60",
    ] = df["Outstanding Balance"]

    df.loc[
        valid_due_date & (df["DaysPastDue"] > 60) & (df["DaysPastDue"] <= 90),
        "61-90",
    ] = df["Outstanding Balance"]

    df.loc[
        valid_due_date & (df["DaysPastDue"] > 90) & (df["DaysPastDue"] <= 120),
        "91-120",
    ] = df["Outstanding Balance"]

    df.loc[
        valid_due_date & (df["DaysPastDue"] > 120),
        "121+",
    ] = df["Outstanding Balance"]

    return df


def enrich_ar01_with_salesforce(ar01_df, input_data):
    """
    Enrich AR01 with Salesforce credit limit information.

    This replicates the VBA Enrich_AR01_With_Salesforce logic.
    """

    df = ar01_df.copy()

    sf_current_df = input_data["salesforce_current"]
    sf_due120_df = input_data["salesforce_due120"]

    fx_rate_brl_to_usd = input_data["fx_rate_brl_to_usd"]

    status_map = build_salesforce_status_map(sf_current_df)
    limit_map = build_salesforce_limit_map(sf_current_df)
    due120_set = build_due120_set(sf_due120_df)

    credit_limit_currency_values = []
    credit_limit_main_values = []
    credit_limit_usd_values = []
    due120_flag_values = []
    diagnostic_values = []

    for _, row in df.iterrows():
        erp_customer_code = normalize_erp_code(row.get("ERP Customer Code", ""))

        credit_limit_currency = ""
        credit_limit_main = ""
        credit_limit_usd = ""
        diagnostic = ""

        if erp_customer_code == "":
            diagnostic = "Missing ERP Customer Code"

        else:
            erp_numeric = to_number(erp_customer_code, default=0.0)

            if erp_numeric <= 700000 or erp_numeric >= 999999:
                diagnostic = "Invalid ERP range"

            elif erp_customer_code in limit_map:
                credit_limit_amount = float(limit_map[erp_customer_code])

                credit_limit_currency = "BRL"
                credit_limit_main = credit_limit_amount
                credit_limit_usd = credit_limit_amount * fx_rate_brl_to_usd

                diagnostic = "Valid credit limit"

            elif erp_customer_code in status_map:
                credit_limit_currency = "BRL"
                limit_status = str(status_map[erp_customer_code]).strip()

                if limit_status.upper() == "BLOQUEADO":
                    diagnostic = "Blocked credit limit"
                elif limit_status.upper() == "EXPIRADO":
                    diagnostic = "Expired credit limit"
                else:
                    diagnostic = "Salesforce record without valid limit"

            else:
                diagnostic = "ERP not found in Salesforce current"

        if erp_customer_code in due120_set:
            due120_flag = "Yes"
        else:
            due120_flag = "No"

        credit_limit_currency_values.append(credit_limit_currency)
        credit_limit_main_values.append(credit_limit_main)
        credit_limit_usd_values.append(credit_limit_usd)
        due120_flag_values.append(due120_flag)
        diagnostic_values.append(diagnostic)

    df["Credit Limit Currency"] = credit_limit_currency_values
    df["Credit Limit Main Currency"] = credit_limit_main_values
    df["Credit Limit USD"] = credit_limit_usd_values
    df["Credit Limit Due 120 Flag"] = due120_flag_values
    df["Credit Limit Diagnostic"] = diagnostic_values

    return df


def build_aging_summary(ar01_df):
    """
    Build aging summary by bucket.
    """

    bucket_columns = ["Future", "0-30", "31-60", "61-90", "91-120", "121+"]

    rows = []

    for bucket in bucket_columns:
        amount = pd.to_numeric(ar01_df[bucket], errors="coerce").fillna(0).sum()
        count = (pd.to_numeric(ar01_df[bucket], errors="coerce").fillna(0) != 0).sum()

        rows.append(
            {
                "Bucket": bucket,
                "Rows": int(count),
                "Amount": amount,
            }
        )

    return pd.DataFrame(rows)


def write_ar01_output(ar01_df, context):
    """
    Export AR01 to the shared AR output workbook.

    Shared output rule:
    - If final workbook does not exist, create it.
    - If final workbook exists, open it.
    - Replace only AR01.
    - Preserve any other AR sheets.
    """

    output_path = get_ar_output_file(context)

    workbook = open_or_create_ar_output_workbook(output_path)
    worksheet = recreate_ar_sheet(workbook, AR01_SHEET_NAME)

    write_ar01_headers(worksheet)
    write_ar01_rows(worksheet, ar01_df)
    format_ar01_worksheet(worksheet)

    save_ar_output_workbook(workbook, output_path)

    return output_path


def write_ar01_headers(worksheet):
    """
    Write AR01 headers.
    """

    for column_index, header in enumerate(AR01_COLUMNS, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


def write_ar01_rows(worksheet, ar01_df):
    """
    Write AR01 rows to worksheet.
    """

    for row_index, (_, row) in enumerate(ar01_df.iterrows(), start=2):
        for column_index, column_name in enumerate(AR01_COLUMNS, start=1):
            value = clean_excel_value(row.get(column_name, ""))

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )


def format_ar01_worksheet(worksheet):
    """
    Apply AR01 worksheet formatting.
    """

    date_columns = [
        "Posting Date",
        "Due Date",
    ]

    money_columns = [
        "Credit Limit Main Currency",
        "Credit Limit USD",
        "FX Rate",
        "FC Debit",
        "FC Credit",
        "FC Balance",
        "Balance",
        "Outstanding Balance",
        "Future",
        "0-30",
        "31-60",
        "61-90",
        "91-120",
        "121+",
    ]

    integer_columns = [
        "DaysPastDue",
    ]

    text_columns = [
        "Company",
        "SAP Customer Number",
        "ERP Customer Code",
        "Document Number",
    ]

    for header in date_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "yyyy-mm-dd"

    for header in money_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "#,##0.00"

    for header in integer_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0"

    for header in text_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "@"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def get_header_column_index(worksheet, header_name):
    """
    Return one-based column index for a header in row 1.
    """

    expected_header = str(header_name).strip().upper()

    for column_index in range(1, worksheet.max_column + 1):
        current_header = worksheet.cell(row=1, column=column_index).value
        current_header = "" if current_header is None else str(current_header).strip().upper()

        if current_header == expected_header:
            return column_index

    raise ValueError(
        f"Column not found in sheet '{worksheet.title}': {header_name}"
    )


def clean_excel_value(value):
    """
    Convert pandas/numpy values to Excel-safe values.
    """

    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()

    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            return value

    return value


def print_input_summary(input_data):
    """
    Print input summary.
    """

    print_header("AR_001 input data")

    print(f"Period suffix: {input_data['period_suffix']}")
    print(f"FX rate BRL to USD: {input_data['fx_rate_brl_to_usd']}")
    print(f"FX Rates file: {input_data['fx_rates_file'].name}")
    print(f"ZTFI098 files: {len(input_data['ztfi_files'])}")
    print(f"ZTFI098 raw rows: {len(input_data['ztfi_raw'])}")
    print(f"ZTFI098 filtered rows: {len(input_data['ztfi'])}")
    print(f"ZTFI098 companies: {', '.join(input_data['ztfi_companies'])}")
    print(f"Customer rows: {len(input_data['customer'])}")
    print(f"Intercompany rows: {len(input_data['intercompany'])}")
    print(
        "Salesforce current: "
        f"rows={len(input_data['salesforce_current'])} "
        f"header_row={input_data['salesforce_current_header_row']}"
    )
    print(
        "Salesforce due120: "
        f"rows={len(input_data['salesforce_due120'])} "
        f"header_row={input_data['salesforce_due120_header_row']}"
    )

    company_validation = input_data.get("ztfi_file_company_validation", {})

    if company_validation.get("status") == "REVIEW REQUIRED":
        print("ZTFI file company validation: REVIEW REQUIRED")
        print(
            "Missing companies: "
            + ", ".join(company_validation.get("missing_companies", []))
        )
    else:
        print("ZTFI file company validation: OK")

    print()


def print_validation_results(validation_results):
    """
    Print required column validation results.
    """

    print_header("Required columns")

    for result in validation_results:
        print(f"{result['dataset']}: {result['status']}")

        if len(result["missing_columns"]) > 0:
            print("  Missing columns:")

            for column in result["missing_columns"]:
                print(f"  - {column}")

    print()


def print_ar01_summary(ar01_df):
    """
    Print AR01 output summary.
    """

    print_header("AR01 summary")

    print(f"Rows: {len(ar01_df)}")
    print(f"Columns: {len(ar01_df.columns)}")

    print()
    print("Credit limit diagnostics:")

    diagnostic_counts = (
        ar01_df["Credit Limit Diagnostic"]
        .fillna("")
        .replace("", "Blank")
        .value_counts()
    )

    for diagnostic, count in diagnostic_counts.items():
        print(f"- {diagnostic}: {count}")

    print()
    print("Aging summary:")

    aging_summary = build_aging_summary(ar01_df)

    for _, row in aging_summary.iterrows():
        print(f"- {row['Bucket']}: rows={row['Rows']} amount={row['Amount']:,.2f}")

    print()


def run_ar_001(context):
    """
    Main runner for AR_001.
    """

    print()

    module_config = context["module"]

    print_header("AR_001 - Debtors Age Analysis")

    input_data = load_ar_input_data(context)

    print_input_summary(input_data)

    validation_results, has_validation_errors = validate_inputs(input_data)

    print_validation_results(validation_results)

    if has_validation_errors:
        print_header("AR_001 status")
        print("Status: REVIEW REQUIRED")
        print("AR_001 was not generated because required columns or company files are missing.")
        print()
        return

    ar01_df = create_ar01_from_ztfi(
        input_data["ztfi"],
        input_data["customer"],
        input_data["intercompany"],
        input_data,
    )

    ar01_df = calculate_ar01_aging(
        ar01_df,
        module_config,
    )

    ar01_df = enrich_ar01_with_salesforce(
        ar01_df,
        input_data,
    )

    output_path = write_ar01_output(
        ar01_df,
        context,
    )

    print_ar01_summary(ar01_df)

    print_header("AR_001 status")
    print("Status: OK")
    print(f"Output file: {output_path}")
    print()
