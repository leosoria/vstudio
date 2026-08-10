"""GL_015 - Analysis of GL accounts per general journal poster.

The control aggregates the active-period BSIS/BSAS population by company, GL
account and Create User ID (BKPF-PPNAM).  It is intentionally independent and
writes or replaces only the GL15 worksheet.
"""

from time import perf_counter

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.gl_common import (
    filter_by_company,
    get_gl_output_file,
    load_gl_bsas_data,
    load_gl_bsis_data,
    load_gl_fx_rates_data,
    load_gl_master_data,
    normalize_company_output,
    normalize_fx_rates,
    normalize_text,
    open_or_create_gl_output_workbook,
    recreate_gl_sheet,
    save_gl_output_workbook,
    select_fx_rate_to_usd,
    to_datetime_value,
    write_dataframe_to_sheet,
    write_single_sheet_workbook_fast,
)


SHEET_NAME = "GL15"
REPORT_CURRENCY = "USD"
MATERIAL_CREATE_USER_BLANK_PCT = 50.0

OUTPUT_COLUMNS = [
    "Company Code",
    "Company Name",
    "GL Account",
    "GL Account Description",
    "Amount in Reporting Currency",
    "Report Currency",
    "Create User ID",
    "Create User Name",
    "Count",
    "Credit",
    "Debit",
]

# English business headers are deliberately first. Technical SAP names and the
# historical Portuguese extracts remain fallbacks. Duplicate suffixes (.1, .2,
# ...) are evaluated by _resolve_column instead of being tied to column order.
COLUMN_ALIASES = {
    "company_code": ["Company Code", "BKPF-BUKRS", "BUKRS", "BSIS-BUKRS", "BSAS-BUKRS", "Empr"],
    "fiscal_year": ["Fiscal Year", "BKPF-GJAHR", "GJAHR", "BSIS-GJAHR", "BSAS-GJAHR", "Ano"],
    "document_number": ["Document Number", "Accounting Document Number", "BKPF-BELNR", "BELNR", "BSIS-BELNR", "BSAS-BELNR", "Nº doc.", "Nº doc"],
    "line_number": ["Line Number", "Line Item", "BSIS-BUZEI", "BSAS-BUZEI", "BUZEI", "Itm"],
    "gl_account": ["GL Account", "G/L Account", "General Ledger Account", "BSIS-HKONT", "BSAS-HKONT", "HKONT", "Razão", "Razao", "Cta.Razão", "Cta.Razao"],
    "debit_credit": ["Debit/Credit Indicator", "Debit Credit Indicator", "BSIS-SHKZG", "BSAS-SHKZG", "SHKZG", "D/C"],
    "amount_local": ["Amount in Local Currency", "Local Amount", "BSIS-DMBTR", "BSAS-DMBTR", "DMBTR", "Montante em MI"],
    "amount_document": ["Amount in Document Currency", "Document Amount", "BSIS-WRBTR", "BSAS-WRBTR", "WRBTR", "Montante"],
    "document_currency": ["Document Currency", "Currency", "BKPF-WAERS", "WAERS", "Moeda"],
    "posting_date": ["Posting Date", "BKPF-BUDAT", "BUDAT", "Dt.lçto.", "Dt.lçto", "Dt.lcto.", "Dt.lcto"],
    "entry_date": ["Entry Date", "Date Entered", "BKPF-CPUDT", "CPUDT", "Dt.entr.", "Dt.entr"],
    "create_user": ["Create User ID", "Creator ID", "Created By", "BKPF-PPNAM", "PPNAM", "Pré-edição", "Pre-edição", "Pré-edicao", "Pre-edicao"],
    "create_user_name": ["Create User Name", "Creator Name"],
    "approver_user": ["Approver User ID", "BKPF-USNAM", "USNAM", "Nome do usuário", "Nome do usuario"],
}

REQUIRED_FIELDS = [
    "company_code", "fiscal_year", "document_number", "line_number",
    "gl_account", "debit_credit", "amount_document", "document_currency",
    "posting_date", "create_user",
]

COLUMN_WIDTHS = {
    "Company Code": 14,
    "Company Name": 30,
    "GL Account": 18,
    "GL Account Description": 36,
    "Amount in Reporting Currency": 28,
    "Report Currency": 16,
    "Create User ID": 20,
    "Create User Name": 28,
    "Count": 12,
    "Credit": 18,
    "Debit": 18,
}


def _log_timing(stage, started):
    elapsed = perf_counter() - started
    print(f"GL15 timing - {stage}: {elapsed:.2f} seconds")
    return perf_counter()


def _clean_text(series):
    return series.fillna("").astype(str).str.strip().replace(
        {"nan": "", "None": "", "<NA>": ""}
    )


def _normalize_code(series):
    return _clean_text(series).str.replace(r"\.0$", "", regex=True)


def _normalize_company(series):
    return _normalize_code(series).map(normalize_company_output)


def _base_header(value):
    return normalize_text(value).lower()


def _header_without_duplicate_suffix(value):
    return _base_header(value).rsplit(".", 1)[0] if _base_header(value).rsplit(".", 1)[-1].isdigit() else _base_header(value)


def _parse_number(series):
    text = _clean_text(series)
    negative_parentheses = text.str.match(r"^\(.*\)$")
    text = text.str.replace(r"[()]", "", regex=True).str.replace(" ", "", regex=False)
    both = text.str.contains(",", regex=False) & text.str.contains(".", regex=False)
    comma_decimal = text.str.contains(r",\d{1,6}$", regex=True)
    text = text.where(~both, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | ~comma_decimal, text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    text = text.where(both | comma_decimal, text.str.replace(",", "", regex=False))
    values = pd.to_numeric(text, errors="coerce")
    return values.where(~negative_parentheses, -values.abs())


def _parse_date(series):
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    numeric = pd.to_numeric(series, errors="coerce")
    excel_dates = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    return parsed.where(parsed.notna(), excel_dates)


def _candidate_score(dataframe, column, logical_name):
    series = dataframe[column]
    non_null = float(series.notna().mean())
    clean = _clean_text(series)
    non_blank = float(clean.ne("").mean())
    score = non_null + non_blank

    if logical_name in {"amount_local", "amount_document"}:
        score += float(_parse_number(series).notna().mean())
    elif logical_name in {"posting_date", "entry_date"}:
        score += float(_parse_date(series).notna().mean())
    elif logical_name == "debit_credit":
        score += float(clean.str.upper().isin(["S", "H", "D", "C", "DEBIT", "CREDIT"]).mean())
    elif logical_name in {"company_code", "fiscal_year", "document_number", "line_number", "gl_account"}:
        score += float(clean.str.replace(r"\.0$", "", regex=True).str.match(r"^[A-Za-z0-9_-]+$").mean())

    return score


def _resolve_column(dataframe, logical_name, required=False):
    aliases = COLUMN_ALIASES[logical_name]
    columns = list(dataframe.columns)
    for alias in aliases:
        normalized_alias = normalize_text(alias).lower()
        candidates = [
            column for column in columns
            if _header_without_duplicate_suffix(column) == normalized_alias
        ]
        if candidates:
            # Avoid scanning a large source column merely to score it when the
            # header is unambiguous. Content scoring is only needed for .1/.2
            # alternatives of the same semantic header.
            if len(candidates) == 1:
                return candidates[0]
            return max(candidates, key=lambda column: _candidate_score(dataframe, column, logical_name))

    if required:
        raise ValueError(
            f"GL15 missing required field '{logical_name}'. Expected one of: {aliases}"
        )
    return None


def _resolve_columns(dataframe, source_name):
    resolved = {
        field: _resolve_column(dataframe, field, required=field in REQUIRED_FIELDS)
        for field in COLUMN_ALIASES
    }
    print(f"GL15 {source_name} resolved headers:")
    for field in COLUMN_ALIASES:
        column = resolved[field]
        if column is not None:
            null_pct = (dataframe[column].isna() | _clean_text(dataframe[column]).eq("")).mean() * 100
            print(f"  {field}: {column} (blank/null {null_pct:.2f}%)")
    return resolved


def _prepare_source(dataframe, source_name):
    if dataframe.empty:
        return pd.DataFrame()
    resolved = _resolve_columns(dataframe, source_name)
    result = pd.DataFrame(index=dataframe.index)
    result["Company Code"] = _normalize_company(dataframe[resolved["company_code"]])
    result["Fiscal Year"] = _normalize_code(dataframe[resolved["fiscal_year"]])
    result["Document Number"] = _normalize_code(dataframe[resolved["document_number"]])
    result["Line Number"] = _normalize_code(dataframe[resolved["line_number"]])
    result["GL Account"] = _normalize_code(dataframe[resolved["gl_account"]])
    result["Debit/Credit"] = _clean_text(dataframe[resolved["debit_credit"]]).str.upper()
    result["Amount Document Raw"] = _parse_number(dataframe[resolved["amount_document"]])
    local_column = resolved["amount_local"]
    result["Amount Local Raw"] = _parse_number(dataframe[local_column]) if local_column else pd.NA
    result["Document Currency"] = _clean_text(dataframe[resolved["document_currency"]]).str.upper()
    result["Posting Date"] = _parse_date(dataframe[resolved["posting_date"]])
    entry_column = resolved["entry_date"]
    result["Entry Date"] = _parse_date(dataframe[entry_column]) if entry_column else pd.NaT
    result["Create User ID"] = _clean_text(dataframe[resolved["create_user"]])
    name_column = resolved["create_user_name"]
    result["Create User Name"] = _clean_text(dataframe[name_column]) if name_column else ""
    approver_column = resolved["approver_user"]
    result["Approver User ID Evidence"] = (
        _clean_text(dataframe[approver_column]) if approver_column else ""
    )
    result["Source"] = source_name
    return result


def _period_bounds(context):
    period_from = to_datetime_value(context["module"].get("from", ""))
    period_to = to_datetime_value(context["module"].get("to", ""))
    if pd.isna(period_from) or pd.isna(period_to):
        raise ValueError("GL15 requires valid FROM and TO dates in config.xlsx.")
    return period_from.normalize(), period_to.normalize()


def _apply_filters(dataframe, context, period_from, period_to):
    result = filter_by_company(
        dataframe, "Company Code", context["module"].get("companies", "")
    )
    mask = result["Posting Date"].notna() & result["Posting Date"].dt.normalize().between(period_from, period_to)
    return result.loc[mask].copy()


def _deduplicate_sources(dataframe):
    key = ["Company Code", "Fiscal Year", "Document Number", "Line Number"]
    if dataframe[key].eq("").any(axis=None):
        missing = int(dataframe[key].eq("").any(axis=1).sum())
        raise ValueError(f"GL15 cannot build a robust line key for {missing} rows with blank key fields.")
    result = dataframe.copy()
    result["_source_priority"] = result["Source"].map({"BSIS": 1, "BSAS": 2}).fillna(0)
    result["_order"] = range(len(result))
    result = result.sort_values(["_source_priority", "_order"], kind="stable")
    duplicated = result.duplicated(key, keep="last")
    removed = int(duplicated.sum())
    return result.loc[~duplicated].drop(columns=["_source_priority", "_order"]).reset_index(drop=True), removed


def _calculate_amounts(dataframe):
    indicator = dataframe["Debit/Credit"].str.upper()
    amount = dataframe["Amount Document Raw"].abs()
    debit_mask = indicator.isin(["S", "D", "DEBIT"])
    credit_mask = indicator.isin(["H", "C", "CREDIT"])
    invalid = ~(debit_mask | credit_mask)
    if invalid.any():
        values = sorted(indicator.loc[invalid].dropna().unique().tolist())
        raise ValueError(f"GL15 found unsupported Debit/Credit indicators: {values}")
    dataframe["Debit"] = amount.where(debit_mask, 0.0).fillna(0.0)
    dataframe["Credit"] = amount.where(credit_mask, 0.0).fillna(0.0)
    dataframe["Amount Document Signed"] = dataframe["Debit"] - dataframe["Credit"]
    return dataframe


def _report_create_user_sufficiency(dataframe):
    """Report material PPNAM gaps without silently changing poster semantics."""
    line_count = len(dataframe)
    blank_mask = dataframe["Create User ID"].eq("")
    blank_lines = int(blank_mask.sum())
    blank_pct = (blank_lines / line_count * 100.0) if line_count else 0.0
    approver_available = int(
        dataframe["Approver User ID Evidence"].ne("").sum()
    )

    print("GL15 Create User ID input sufficiency:")
    print(f"  BKPF-PPNAM blank lines: {blank_lines} of {line_count} ({blank_pct:.2f}%)")
    print(f"  BKPF-USNAM nonblank lines (comparison only): {approver_available}")

    if blank_pct >= MATERIAL_CREATE_USER_BLANK_PCT:
        print(
            "GL15 INPUT LIMITATION: BKPF-PPNAM has material nullity. The control "
            "keeps blank-poster journals so the population reconciles and does not "
            "silently substitute BKPF-USNAM, because USNAM has different semantics."
        )
        print(
            "GL15 SAP extraction recommendation: obtain BKPF-PPNAM (Create User ID / "
            "Pre-editing User; English header 'Create User ID') joined by BUKRS + "
            "BELNR + GJAHR. The current input is sufficient for amounts and journal "
            "counts, but not for complete poster attribution."
        )

    return blank_lines, blank_pct


def _convert_reporting_currency(dataframe, normalized_fx):
    dataframe["Amount in Reporting Currency"] = pd.Series(pd.NA, index=dataframe.index, dtype="Float64")
    if dataframe.empty:
        return dataframe
    pairs = dataframe[["Document Currency", "Posting Date"]].drop_duplicates()
    lookup_rows = []
    for currency, requested_date in pairs.itertuples(index=False, name=None):
        rate = select_fx_rate_to_usd(normalized_fx, currency, requested_date)
        lookup_rows.append({
            "Document Currency": currency,
            "Posting Date": requested_date,
            "_fx": rate["fx_to_usd"] if rate else pd.NA,
        })
    lookup = pd.DataFrame(lookup_rows)
    before = len(dataframe)
    dataframe = dataframe.merge(lookup, on=["Document Currency", "Posting Date"], how="left", validate="many_to_one")
    if len(dataframe) != before:
        raise AssertionError("GL15 FX lookup multiplied journal lines.")
    dataframe["Amount in Reporting Currency"] = dataframe["Amount Document Signed"] * pd.to_numeric(dataframe["_fx"], errors="coerce")
    return dataframe.drop(columns="_fx")


def _stable_user_names(dataframe):
    named = dataframe.loc[dataframe["Create User Name"].ne(""), ["Create User ID", "Create User Name"]].copy()
    if named.empty:
        return {}
    named["_count"] = named.groupby(["Create User ID", "Create User Name"])["Create User Name"].transform("size")
    named = named.sort_values(["Create User ID", "_count", "Create User Name"], ascending=[True, False, True], kind="stable")
    return named.drop_duplicates("Create User ID").set_index("Create User ID")["Create User Name"].to_dict()


def _master_maps(master):
    if master.empty:
        return {}, {}
    company_col = _first_master_column(master, ["Company Code", "BKPF-BUKRS", "BUKRS", "Empr"])
    account_col = _first_master_column(master, ["GL Account", "G/L Account", "SAKNR", "HKONT", "Cta.Razão", "Cta.Razao", "Razão", "Razao", "Account"])
    company_name_col = _first_master_column(master, ["Company Name", "BUTXT", "Nome da firma", "Nome da empresa", "Empresa", "Company"])
    description_col = _first_master_column(master, ["GL Account Description", "Account Description", "TXT50", "TxtDescr", "TXT20", "Account Name", "Description", "Texto breve"])
    company_map = {}
    account_map = {}
    if company_col and company_name_col:
        frame = pd.DataFrame({"company": _normalize_company(master[company_col]), "name": _clean_text(master[company_name_col])})
        frame = frame.loc[frame["company"].ne("") & frame["name"].ne("")].sort_values(["company", "name"], kind="stable").drop_duplicates("company")
        company_map = frame.set_index("company")["name"].to_dict()
    if company_col and account_col and description_col:
        frame = pd.DataFrame({"company": _normalize_company(master[company_col]), "account": _normalize_code(master[account_col]), "description": _clean_text(master[description_col])})
        frame = frame.loc[frame["company"].ne("") & frame["account"].ne("") & frame["description"].ne("")].sort_values(["company", "account", "description"], kind="stable").drop_duplicates(["company", "account"])
        account_map = frame.set_index(["company", "account"])["description"].to_dict()
    return company_map, account_map


def _first_master_column(dataframe, aliases):
    for alias in aliases:
        matches = [column for column in dataframe.columns if _header_without_duplicate_suffix(column) == alias.lower()]
        if matches:
            return max(matches, key=lambda column: float(dataframe[column].notna().mean()))
    return None


def _aggregate(dataframe, company_map, account_map, user_name_map):
    dataframe["Journal Key"] = dataframe["Company Code"] + "|" + dataframe["Fiscal Year"] + "|" + dataframe["Document Number"]
    keys = ["Company Code", "GL Account", "Create User ID"]
    summary = dataframe.groupby(keys, dropna=False, as_index=False).agg(
        **{
            "Amount in Reporting Currency": ("Amount in Reporting Currency", lambda values: values.sum(min_count=1)),
            "Count": ("Journal Key", "nunique"),
            "Credit": ("Credit", "sum"),
            "Debit": ("Debit", "sum"),
        }
    )
    summary["Company Name"] = summary["Company Code"].map(company_map).fillna("")
    summary["GL Account Description"] = [account_map.get((company, account), "") for company, account in summary[["Company Code", "GL Account"]].itertuples(index=False, name=None)]
    summary["Create User Name"] = summary["Create User ID"].map(user_name_map).fillna("")
    summary["Report Currency"] = REPORT_CURRENCY
    summary["Count"] = summary["Count"].astype("int64")
    return summary[OUTPUT_COLUMNS].sort_values(keys, kind="stable").reset_index(drop=True)


def _validate(population, output):
    keys = ["Company Code", "GL Account", "Create User ID"]
    if output.duplicated(keys).any():
        raise AssertionError("GL15 output contains duplicate aggregation keys.")
    if not output["Count"].gt(0).all():
        raise AssertionError("GL15 Count must be a positive integer.")
    if output[["Debit", "Credit"]].lt(0).any(axis=None):
        raise AssertionError("GL15 Debit and Credit must be positive or zero.")
    for column in ["Debit", "Credit"]:
        if not pd.Series([population[column].sum(), output[column].sum()]).fillna(0).pipe(lambda values: abs(values.iloc[0] - values.iloc[1]) <= 0.01):
            raise AssertionError(f"GL15 {column} did not reconcile after aggregation.")
    before = population["Amount in Reporting Currency"].sum(min_count=1)
    after = output["Amount in Reporting Currency"].sum(min_count=1)
    if not (pd.isna(before) and pd.isna(after)) and abs(float(before) - float(after)) > 0.01:
        raise AssertionError("GL15 reporting amount did not reconcile after aggregation.")
    if list(output.columns) != OUTPUT_COLUMNS:
        raise AssertionError("GL15 output columns do not match the LHA layout.")


def _format_sheet(worksheet, dataframe):
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    positions = {name: index for index, name in enumerate(dataframe.columns, 1)}
    for name, index in positions.items():
        worksheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(name, 18)
        number_format = None
        if name in {"Debit", "Credit", "Amount in Reporting Currency"}:
            number_format = "#,##0.00;[Red]-#,##0.00"
        elif name == "Count":
            number_format = "#,##0"
        if number_format:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, index).number_format = number_format


def _write_output(context, output):
    output_file = get_gl_output_file(context)
    if write_single_sheet_workbook_fast(output_file, SHEET_NAME, output, amount_columns=["Amount in Reporting Currency", "Credit", "Debit"], integer_columns=["Count"]):
        return output_file
    workbook = open_or_create_gl_output_workbook(output_file)
    worksheet = recreate_gl_sheet(workbook, SHEET_NAME)
    write_dataframe_to_sheet(worksheet, output)
    _format_sheet(worksheet, output)
    save_gl_output_workbook(workbook, output_file)
    return output_file


def run_gl_015(context):
    """Run GL_015 and return the standard runner result."""
    total_started = perf_counter()
    stage_started = total_started
    print("Running GL15 - Analysis Of General Ledger Accounts Per General Journal Poster")
    print("GL15 poster decision: Create User ID = BKPF-PPNAM; BKPF-USNAM is not substituted.")
    period_from, period_to = _period_bounds(context)

    bsis = load_gl_bsis_data(context)
    print(f"GL15 BSIS rows read: {len(bsis)}")
    bsas = load_gl_bsas_data(context)
    print(f"GL15 BSAS rows read: {len(bsas)}")
    stage_started = _log_timing("input reading", stage_started)
    if bsis.empty and bsas.empty:
        raise ValueError("GL15 requires at least one BSIS or BSAS input file.")

    frames = []
    for source, raw in [("BSIS", bsis), ("BSAS", bsas)]:
        if raw.empty:
            continue
        prepared = _prepare_source(raw, source)
        filtered = _apply_filters(prepared, context, period_from, period_to)
        print(f"GL15 {source} rows after CONFIG filters: {len(filtered)}")
        frames.append(filtered)
    population = pd.concat(frames, ignore_index=True)
    stage_started = _log_timing("normalization and CONFIG filters", stage_started)

    population, duplicates_removed = _deduplicate_sources(population)
    print(f"GL15 BSIS/BSAS duplicates removed: {duplicates_removed}")
    stage_started = _log_timing("deduplication", stage_started)
    population = _calculate_amounts(population)
    blank_user_lines, blank_user_pct = _report_create_user_sufficiency(population)
    population["Journal Key"] = population["Company Code"] + "|" + population["Fiscal Year"] + "|" + population["Document Number"]
    blank_user_journals = int(population.loc[population["Create User ID"].eq(""), "Journal Key"].nunique())
    print(f"GL15 blank Create User ID lines: {blank_user_lines}; journals: {blank_user_journals}")

    fx = load_gl_fx_rates_data(context)
    normalized_fx = normalize_fx_rates(fx) if not fx.empty else normalize_fx_rates(pd.DataFrame())
    population = _convert_reporting_currency(population, normalized_fx)
    print(f"GL15 FxRates used: {'yes' if not fx.empty else 'no'}")
    stage_started = _log_timing("FX", stage_started)

    user_names = _stable_user_names(population)
    master = load_gl_master_data(context)
    company_map, account_map = _master_maps(master)
    stage_started = _log_timing("GL Master Data", stage_started)
    output = _aggregate(population, company_map, account_map, user_names)
    _validate(population, output)
    stage_started = _log_timing("aggregation and validation", stage_started)

    unique_journals = int(population["Journal Key"].nunique())
    distinct_users = int(population.loc[population["Create User ID"].ne(""), "Create User ID"].nunique())
    print("GL15 reconciliation:")
    print(f"  unique journals: {unique_journals}")
    print(f"  Debit: {population['Debit'].sum():.2f}")
    print(f"  Credit: {population['Credit'].sum():.2f}")
    print(f"  Amount in Reporting Currency: {population['Amount in Reporting Currency'].sum(min_count=1)}")
    print(f"  final groups: {len(output)}")
    print(f"  distinct nonblank users: {distinct_users}")
    print(
        f"  blank-user lines: {blank_user_lines} ({blank_user_pct:.2f}%); "
        f"journals: {blank_user_journals}"
    )

    output_file = _write_output(context, output)
    _log_timing("Excel writing", stage_started)
    total_elapsed = perf_counter() - total_started
    print(f"GL15 rows written: {len(output)}")
    print(f"GL15 output sheet written: {output_file} [{SHEET_NAME}]")
    print(f"GL15 total elapsed: {total_elapsed:.2f} seconds")
    return {"status": "OK", "output_file": output_file, "sheet_name": SHEET_NAME, "rows": len(output)}


__all__ = ["run_gl_015"]
