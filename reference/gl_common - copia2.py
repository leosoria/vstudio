"""
Common GL utilities.
"""

import warnings
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


ALLOWED_EXTENSIONS = [".xlsx", ".xls", ".csv", ".txt"]

GL_BSIS_INPUT_KEYWORDS = [
    "LBR GL_JE_BSIS",
    "LBR_GL_JE_BSIS",
]
GL_BSAS_INPUT_KEYWORDS = [
    "LBR GL_JE_BSAS",
    "LBR_GL_JE_BSAS",
]
GL_MASTER_INPUT_KEYWORDS = [
    "LBR GL_MD",
    "LBR_GL_MD",
]

GL_HEADER_FILL = "D9EAF7"
GL_DATE_FORMAT = "dd/mm/yyyy"
GL_AMOUNT_FORMAT = '#,##0.00'
GL_INTEGER_FORMAT = '#,##0'


def has_allowed_extension(file_path):
    file_path = Path(file_path)

    return file_path.suffix.lower() in ALLOWED_EXTENSIONS


def normalize_text(value):
    if value is None:
        return ""

    value_text = str(value).strip()

    if value_text.lower() == "nan":
        return ""

    return value_text


def normalize_header(value):
    return normalize_text(value).lower()


def normalize_company_output(value):
    value_text = normalize_text(value)

    if value_text == "":
        return ""

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if value_text.isdigit():
        return str(int(value_text))

    return value_text


def normalize_company_match(value):
    return normalize_company_output(value)


def normalize_code_keep_leading_zeroes(value):
    value_text = normalize_text(value)

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    return value_text


def parse_companies_filter(value):
    value_text = normalize_text(value)

    if value_text == "":
        return []

    if value_text.upper() in ["ALL", "TODAS", "TODOS"]:
        return []

    for separator in [";", "|", "\n", "\r", "\t"]:
        value_text = value_text.replace(separator, ",")

    if "," not in value_text and " " in value_text:
        value_text = ",".join(value_text.split())

    companies = []

    for item in value_text.split(","):
        item = normalize_text(item)

        if item == "":
            continue

        if item.upper() in ["ALL", "TODAS", "TODOS"]:
            return []

        companies.append(normalize_company_match(item))

    return companies


def to_datetime_value(value):
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NaT

    if value_text.endswith(".0"):
        value_text = value_text[:-2]

    if len(value_text) == 8 and value_text.isdigit():
        return pd.to_datetime(value_text, format="%Y%m%d", errors="coerce")

    if len(value_text) >= 10 and value_text[4:5] == "-" and value_text[7:8] == "-":
        return pd.to_datetime(value_text, errors="coerce", dayfirst=False)

    return pd.to_datetime(value_text, errors="coerce", dayfirst=True)


def parse_number(value):
    if pd.isna(value):
        return pd.NA

    if isinstance(value, (int, float)):
        return value

    value_text = normalize_text(value)

    if value_text == "":
        return pd.NA

    value_text = value_text.replace(" ", "")

    is_negative = value_text.startswith("(") and value_text.endswith(")")

    if is_negative:
        value_text = value_text[1:-1]

    if "," in value_text and "." in value_text:
        last_comma = value_text.rfind(",")
        last_dot = value_text.rfind(".")

        if last_comma > last_dot:
            value_text = value_text.replace(".", "")
            value_text = value_text.replace(",", ".")
        else:
            value_text = value_text.replace(",", "")
    elif "," in value_text and "." not in value_text:
        value_text = value_text.replace(",", ".")

    try:
        parsed_number = float(value_text)
    except ValueError:
        return pd.NA

    if is_negative:
        parsed_number = parsed_number * -1

    return parsed_number


def get_period_suffix(module_config):
    to_date = module_config.get("to", "")
    parsed_date = to_datetime_value(to_date)

    if pd.isna(parsed_date):
        raise ValueError(
            "Could not determine GL output period because module TO date is empty or invalid."
        )

    return parsed_date.strftime("%Y%m%d")


def get_gl_output_file(context):
    output_folder = Path(context["output_folder"])
    module_config = context["module"]
    period_suffix = get_period_suffix(module_config)

    return output_folder / f"LBR_Results_GL_{period_suffix}.xlsx"


def find_all_files_containing(base_folder, text_to_find):
    base_folder = Path(base_folder)
    text_to_find = normalize_text(text_to_find).lower()

    matched_files = []

    for file_path in base_folder.rglob("*"):
        if not file_path.is_file():
            continue

        if not has_allowed_extension(file_path):
            continue

        if text_to_find in file_path.name.lower():
            matched_files.append(file_path)

    return sorted(matched_files)


def find_optional_period_file(context, keyword_prefix):
    input_folder = Path(context["input_folder"])
    period_suffix = get_period_suffix(context["module"])
    keyword = f"{keyword_prefix}_{period_suffix}"

    matched_files = find_all_files_containing(
        base_folder=input_folder,
        text_to_find=keyword,
    )

    if len(matched_files) == 0:
        return None

    if len(matched_files) > 1:
        matched_files_text = "\n".join(
            f"- {file_path}"
            for file_path in matched_files
        )

        raise ValueError(
            f"Multiple GL input files found using keyword '{keyword}'. "
            f"Expected zero or one file.\n"
            f"{matched_files_text}"
        )

    return matched_files[0]


def find_period_file_using_keywords(context, keyword_prefixes):
    for keyword_prefix in keyword_prefixes:
        input_file = find_optional_period_file(
            context=context,
            keyword_prefix=keyword_prefix,
        )

        if input_file is not None:
            return input_file

    return None


def find_gl_bsis_input_file(context):
    return find_period_file_using_keywords(
        context=context,
        keyword_prefixes=GL_BSIS_INPUT_KEYWORDS,
    )


def find_gl_bsas_input_file(context):
    return find_period_file_using_keywords(
        context=context,
        keyword_prefixes=GL_BSAS_INPUT_KEYWORDS,
    )


def find_gl_master_input_file(context):
    return find_period_file_using_keywords(
        context=context,
        keyword_prefixes=GL_MASTER_INPUT_KEYWORDS,
    )


def read_input_file(file_path, sheet_name=0):
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    extension = file_path.suffix.lower()

    if extension in [".xlsx", ".xls"]:
        return pd.read_excel(file_path, sheet_name=sheet_name, dtype=object)

    if extension == ".csv":
        return pd.read_csv(file_path, dtype=object)

    if extension == ".txt":
        return pd.read_csv(file_path, sep="\t", dtype=object)

    raise ValueError(f"Unsupported input file extension: {file_path.suffix}")


def clean_dataframe(dataframe):
    result = dataframe.copy()

    result = result.dropna(axis=0, how="all")
    result = result.dropna(axis=1, how="all")

    result.columns = [
        normalize_text(column)
        for column in result.columns
    ]

    return result


def find_column(dataframe, possible_names):
    normalized_lookup = {
        normalize_header(column): column
        for column in dataframe.columns
    }

    for possible_name in possible_names:
        normalized_name = normalize_header(possible_name)

        if normalized_name in normalized_lookup:
            return normalized_lookup[normalized_name]

    return None


def require_columns(dataframe, required_columns, source_name="GL input"):
    resolved_columns = {}
    missing_columns = []

    for logical_name, possible_names in required_columns.items():
        column_name = find_column(dataframe, possible_names)

        if column_name is None:
            missing_columns.append(
                f"{logical_name}: expected one of {possible_names}"
            )
        else:
            resolved_columns[logical_name] = column_name

    if missing_columns:
        raise ValueError(
            f"Missing required columns in {source_name}:\n- "
            + "\n- ".join(missing_columns)
        )

    return resolved_columns


def get_optional_column(dataframe, possible_names):
    return find_column(dataframe, possible_names)


def get_value(row, column_name, default=""):
    if column_name is None:
        return default

    return row.get(column_name, default)


def filter_by_company(dataframe, company_column, companies_filter):
    result = dataframe.copy()
    companies = parse_companies_filter(companies_filter)

    if not companies:
        return result

    normalized_company = result[company_column].apply(normalize_company_match)

    return result[normalized_company.isin(companies)].copy()


def apply_sap_debit_credit_sign(amount_value, debit_credit_indicator):
    amount = parse_number(amount_value)
    indicator = normalize_text(debit_credit_indicator).upper()

    if pd.isna(amount):
        return pd.NA

    amount = abs(float(amount))

    if indicator == "H":
        return amount * -1

    return amount


def build_document_key(company_code, fiscal_year, document_number, line_item):
    return "|".join(
        [
            normalize_company_output(company_code),
            normalize_code_keep_leading_zeroes(fiscal_year),
            normalize_code_keep_leading_zeroes(document_number),
            normalize_code_keep_leading_zeroes(line_item),
        ]
    )


def load_gl_source_data(context, source_name, input_file):
    if input_file is None:
        print(f"GL {source_name} input file: not found.")
        return pd.DataFrame()

    print(f"GL {source_name} input file: {input_file}")

    dataframe = read_input_file(input_file)
    dataframe = clean_dataframe(dataframe)
    dataframe["Source"] = source_name

    return dataframe


def load_gl_bsis_data(context):
    input_file = find_gl_bsis_input_file(context)

    return load_gl_source_data(
        context=context,
        source_name="BSIS",
        input_file=input_file,
    )


def load_gl_bsas_data(context):
    input_file = find_gl_bsas_input_file(context)

    return load_gl_source_data(
        context=context,
        source_name="BSAS",
        input_file=input_file,
    )


def load_gl_master_data(context):
    input_file = find_gl_master_input_file(context)

    if input_file is None:
        print("GL master input file: not found. Company and Account Name will be blank.")
        return pd.DataFrame()

    print(f"GL master input file: {input_file}")

    dataframe = read_input_file(input_file)
    dataframe = clean_dataframe(dataframe)

    return dataframe


def build_company_name_map(master_dataframe):
    if master_dataframe.empty:
        return {}

    bukrs_column = get_optional_column(
        master_dataframe,
        [
            "BUKRS",
            "Empr",
        ],
    )
    company_name_column = get_optional_column(
        master_dataframe,
        [
            "BUTXT",
            "Nome da firma",
            "Nome da empresa",
            "Empresa",
            "Company",
        ],
    )

    if bukrs_column is None or company_name_column is None:
        return {}

    result = {}

    for _, row in master_dataframe.iterrows():
        company_code = normalize_company_output(row.get(bukrs_column, ""))
        company_name = normalize_text(row.get(company_name_column, ""))

        if company_code == "":
            continue

        if company_code not in result:
            result[company_code] = company_name

    return result


def build_gl_account_name_map(master_dataframe):
    if master_dataframe.empty:
        return {}

    account_column = get_optional_column(
        master_dataframe,
        [
            "SAKNR",
            "HKONT",
            "Cta.Razão",
            "Cta.Razao",
            "Cta.Razão.1",
            "Cta.Razao.1",
            "Razão",
            "Razao",
            "Conta",
            "Account",
        ],
    )
    language_column = get_optional_column(
        master_dataframe,
        [
            "SPRAS",
            "Idioma",
            "Language",
        ],
    )
    text_column = get_optional_column(
        master_dataframe,
        [
            "TXT50",
            "TxtDescr",
            "TXT20",
            "Texto breve",
            "Texto",
            "Account Name",
            "Description",
        ],
    )

    if account_column is None or text_column is None:
        return {}

    language_priority = {
        "PT": 1,
        "P": 1,
        "ES": 2,
        "S": 2,
        "EN": 3,
        "E": 3,
    }

    candidate_by_account = {}

    for _, row in master_dataframe.iterrows():
        account = normalize_code_keep_leading_zeroes(row.get(account_column, ""))
        text = normalize_text(row.get(text_column, ""))

        if account == "" or text == "":
            continue

        language = normalize_text(row.get(language_column, "")).upper() if language_column else ""
        priority = language_priority.get(language, 9)

        current = candidate_by_account.get(account)

        if current is None or priority < current["priority"]:
            candidate_by_account[account] = {
                "priority": priority,
                "text": text,
            }

    return {
        account: value["text"]
        for account, value in candidate_by_account.items()
    }


def open_or_create_gl_output_workbook(output_file):
    output_file = Path(output_file)

    if output_file.exists():
        return load_workbook(output_file)

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def recreate_gl_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    return workbook.create_sheet(sheet_name)


def save_gl_output_workbook(workbook, output_file):
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook.save(output_file)
    except PermissionError as error:
        raise PermissionError(
            f"Could not save output workbook: {output_file}. "
            "The file may be open in Excel or locked by OneDrive. "
            "Close the workbook and run again."
        ) from error


def write_dataframe_to_sheet(worksheet, dataframe):
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None

            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()

            worksheet.cell(row=row_index, column=column_index, value=value)


def apply_standard_gl_formatting(
    worksheet,
    dataframe,
    date_columns=None,
    amount_columns=None,
    integer_columns=None,
):
    date_columns = set(date_columns or [])
    amount_columns = set(amount_columns or [])
    integer_columns = set(integer_columns or [])

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=GL_HEADER_FILL,
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    column_name_by_index = {
        index + 1: column_name
        for index, column_name in enumerate(dataframe.columns)
    }

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        column_name = column_name_by_index.get(column_index, "")

        max_length = len(str(column_name))

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

            if cell.row == 1:
                continue

            if column_name in date_columns:
                cell.number_format = GL_DATE_FORMAT

            if column_name in amount_columns:
                cell.number_format = GL_AMOUNT_FORMAT

            if column_name in integer_columns:
                cell.number_format = GL_INTEGER_FORMAT

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def write_single_sheet_workbook_fast(
    output_file,
    sheet_name,
    dataframe,
    date_columns=None,
    amount_columns=None,
    integer_columns=None,
):
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    date_columns = set(date_columns or [])
    amount_columns = set(amount_columns or [])
    integer_columns = set(integer_columns or [])

    if output_file.exists():
        return False

    try:
        with pd.ExcelWriter(
            output_file,
            engine="xlsxwriter",
            datetime_format="dd/mm/yyyy",
            date_format="dd/mm/yyyy",
        ) as writer:
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            header_format = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#D9EAF7",
                    "border": 0,
                }
            )

            date_format = workbook.add_format(
                {
                    "num_format": "dd/mm/yyyy",
                }
            )

            amount_format = workbook.add_format(
                {
                    "num_format": "#,##0.00",
                }
            )

            integer_format = workbook.add_format(
                {
                    "num_format": "#,##0",
                }
            )

            for column_index, column_name in enumerate(dataframe.columns):
                worksheet.write(0, column_index, column_name, header_format)

                column_width = min(
                    max(
                        len(str(column_name)) + 2,
                        12,
                    ),
                    45,
                )

                if column_name in date_columns:
                    worksheet.set_column(
                        column_index,
                        column_index,
                        column_width,
                        date_format,
                    )
                elif column_name in amount_columns:
                    worksheet.set_column(
                        column_index,
                        column_index,
                        column_width,
                        amount_format,
                    )
                elif column_name in integer_columns:
                    worksheet.set_column(
                        column_index,
                        column_index,
                        column_width,
                        integer_format,
                    )
                else:
                    worksheet.set_column(
                        column_index,
                        column_index,
                        column_width,
                    )

            last_row = max(len(dataframe), 1)
            last_column = max(len(dataframe.columns) - 1, 0)

            worksheet.autofilter(
                0,
                0,
                last_row,
                last_column,
            )
            worksheet.freeze_panes(1, 0)

    except ModuleNotFoundError:
        print(
            "Fast writer is not available because XlsxWriter is not installed. "
            "Falling back to openpyxl writer."
        )
        return False
    except ImportError:
        print(
            "Fast writer is not available because XlsxWriter is not installed. "
            "Falling back to openpyxl writer."
        )
        return False
    except PermissionError as error:
        raise PermissionError(
            f"Could not save output workbook: {output_file}. "
            "The file may be open in Excel or locked by OneDrive. "
            "Close the workbook and run again."
        ) from error

    return True
