"""
AR_003 automation.

This control creates/replaces only the AR03 sheet in the shared AR output file.

Logic replicated from VBA AR3:
- Load ZTFI098 and Customer input data.
- Exclude intercompany customers.
- Aggregate balances by Company + ERP Customer Code.
- Calculate max days past due and oldest due date.
- Add AR3 diagnostic based on net balance.
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.ar_common import (
    build_customer_erp_map,
    build_intercompany_exclusion_set,
    get_ar_output_file,
    load_ar_input_data,
    normalize_company_output,
    normalize_customer_key,
    normalize_erp_code,
    open_or_create_ar_output_workbook,
    recreate_ar_sheet,
    save_ar_output_workbook,
    to_datetime_value,
    to_number,
    validate_required_columns,
)


AR03_SHEET_NAME = "AR03"

AR03_HEADERS = [
    "Company",
    "ERP Customer Code",
    "SAP Customer Number",
    "Customer Name",
    "Outstanding Balance",
    "Max DaysPastDue",
    "Company Main Currency",
    "Documents Count",
    "Oldest Due Date",
    "AR3 Diagnostic",
]

ZTFI_REQUIRED_COLUMNS = [
    "Empresa",
    "Nº doc.",
    "Cliente",
    "Grp. Emp.",
    "Nome Cli/For",
    "Dt.base prazo pgto",
    "Val. da Transação",
]

CUSTOMER_REQUIRED_COLUMNS = [
    "Cliente",
    "Grupo",
]


def run_ar_003(context):
    """
    Execute AR_003.

    This function:
    - Loads common AR input data.
    - Builds AR03 output rows.
    - Opens or creates the shared AR output workbook.
    - Recreates only the AR03 sheet.
    - Saves the shared output workbook.
    """
    input_data = load_ar_input_data(context)
    output_rows = build_ar03_rows(input_data, context)

    output_file = get_ar_output_file(context)
    workbook = open_or_create_ar_output_workbook(output_file)
    worksheet = recreate_ar_sheet(workbook, AR03_SHEET_NAME)

    write_headers(worksheet)
    write_rows(worksheet, output_rows)
    format_worksheet(worksheet)

    save_ar_output_workbook(workbook, output_file)

    print(f"AR_003 completed. Rows written: {len(output_rows)}")
    print(f"Output file: {output_file}")


def build_ar03_rows(input_data, context):
    """
    Build AR03 output rows.

    Aggregation key:
        Company + ERP Customer Code

    Intercompany customers are excluded using build_intercompany_exclusion_set().
    """
    ztfi_df = input_data["ztfi"]
    customer_df = input_data["customer"]

    validate_required_columns(ztfi_df, ZTFI_REQUIRED_COLUMNS)
    validate_required_columns(customer_df, CUSTOMER_REQUIRED_COLUMNS)

    cutoff_date = get_cutoff_date(input_data, context)

    customer_erp_map = build_customer_erp_map(customer_df)
    intercompany_exclusion_set = build_intercompany_exclusion_set()

    ar03_by_key = {}

    for _, row in ztfi_df.iterrows():
        sap_customer_number = row["Cliente"]
        sap_customer_key = normalize_customer_key(sap_customer_number)

        if sap_customer_key in intercompany_exclusion_set:
            continue

        company_code_raw = row["Empresa"]
        company_output = normalize_company_output(company_code_raw)

        erp_customer_code_from_ztfi = normalize_erp_code(row["Grp. Emp."])

        if sap_customer_key in customer_erp_map:
            erp_customer_code = normalize_erp_code(customer_erp_map[sap_customer_key])
        else:
            erp_customer_code = erp_customer_code_from_ztfi

        if erp_customer_code == "":
            erp_customer_code = sap_customer_key

        ar03_key = f"{company_output}|{erp_customer_code}"

        credit_balance = to_number(row["Val. da Transação"])
        if credit_balance is None:
            credit_balance = 0

        due_date = to_datetime_value(row["Dt.base prazo pgto"])

        if due_date is not None:
            days_past_due = (cutoff_date.date() - due_date.date()).days
        else:
            days_past_due = ""

        if ar03_key not in ar03_by_key:
            ar03_by_key[ar03_key] = {
                "Company": company_output,
                "ERP Customer Code": erp_customer_code,
                "SAP Customer Number": sap_customer_key,
                "Customer Name": row["Nome Cli/For"],
                "Outstanding Balance": 0,
                "Max DaysPastDue": days_past_due,
                "Company Main Currency": "BRL",
                "Documents Count": 0,
                "Oldest Due Date": due_date,
                "AR3 Diagnostic": "",
            }

        ar03_row = ar03_by_key[ar03_key]

        ar03_row["Outstanding Balance"] = (
            ar03_row["Outstanding Balance"] + credit_balance
        )
        ar03_row["Documents Count"] = ar03_row["Documents Count"] + 1

        if isinstance(days_past_due, int):
            current_max_days = ar03_row["Max DaysPastDue"]

            if current_max_days == "":
                ar03_row["Max DaysPastDue"] = days_past_due
            elif days_past_due > int(current_max_days):
                ar03_row["Max DaysPastDue"] = days_past_due

        if due_date is not None:
            current_oldest_due_date = ar03_row["Oldest Due Date"]

            if current_oldest_due_date is None or current_oldest_due_date == "":
                ar03_row["Oldest Due Date"] = due_date
            elif due_date < current_oldest_due_date:
                ar03_row["Oldest Due Date"] = due_date

    output_rows = list(ar03_by_key.values())

    for output_row in output_rows:
        outstanding_balance = output_row["Outstanding Balance"]

        if outstanding_balance < 0:
            output_row["AR3 Diagnostic"] = "Net credit balance identified"
        elif outstanding_balance == 0:
            output_row["AR3 Diagnostic"] = "Zero balance"
        else:
            output_row["AR3 Diagnostic"] = "Net debit balance"

    return output_rows


def get_cutoff_date(input_data, context):
    """
    Return cutoff date used to calculate days past due.

    VBA source used:
        ThisWorkbook.Sheets("Parameters LBR").Range("B7")

    Python source priority:
    1. Date exposed by load_ar_input_data(context), if available.
    2. Module period end date from config.xlsx: context["module"]["to"].
    """
    possible_input_data_keys = [
        "cutoff_date",
        "cut_off_date",
        "period_end_date",
        "to_date",
        "date_to",
    ]

    for key in possible_input_data_keys:
        if key in input_data:
            cutoff_date = to_datetime_value(input_data[key])
            if cutoff_date is not None:
                return cutoff_date

    module_config = context.get("module", {})

    possible_module_keys = [
        "to",
        "To",
        "date_to",
        "period_end_date",
        "cutoff_date",
        "cut_off_date",
    ]

    for key in possible_module_keys:
        if key in module_config:
            cutoff_date = to_datetime_value(module_config[key])
            if cutoff_date is not None:
                return cutoff_date

    raise ValueError(
        "AR_003 requires a cutoff date to calculate Max DaysPastDue. "
        "No valid cutoff date was found in load_ar_input_data(context) or "
        "context['module']. Expected the module end date from config.xlsx, "
        "usually context['module']['to']."
    )


def write_headers(worksheet):
    """
    Write AR03 headers.
    """
    for column_index, header in enumerate(AR03_HEADERS, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


def write_rows(worksheet, ar03_rows):
    """
    Write AR03 detail rows.
    """
    for row_index, row_data in enumerate(ar03_rows, start=2):
        for column_index, header in enumerate(AR03_HEADERS, start=1):
            value = row_data.get(header, "")

            if header == "Oldest Due Date" and value is not None and value != "":
                value = value.date()

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )


def format_worksheet(worksheet):
    """
    Apply AR03 worksheet formatting.
    """
    money_columns = [
        "Outstanding Balance",
    ]

    integer_columns = [
        "Max DaysPastDue",
        "Documents Count",
    ]

    date_columns = [
        "Oldest Due Date",
    ]

    text_columns = [
        "Company",
        "ERP Customer Code",
        "SAP Customer Number",
    ]

    for header in money_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "#,##0.00"

    for header in integer_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0"

    for header in date_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "dd/mm/yyyy"

    for header in text_columns:
        column_index = get_header_column_index(worksheet, header)

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "@"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def get_header_column_index(worksheet, header_name):
    """
    Return one-based column index for a header in row 1.
    """
    expected_header = str(header_name).strip().upper()

    for column_index in range(1, worksheet.max_column + 1):
        current_header = worksheet.cell(row=1, column=column_index).value
        current_header = "" if current_header is None else str(current_header).strip().upper()

        if current_header == expected_header:
            return column_index

    raise ValueError(
        f"Column not found in sheet '{worksheet.title}': {header_name}"
    )


def is_blank(value):
    """
    Return True for None, empty string, or pandas-like NaN text.
    """
    if value is None:
        return True

    value_text = str(value).strip()

    if value_text == "":
        return True

    if value_text.lower() in ["nan", "nat", "none"]:
        return True

    return False
