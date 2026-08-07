"""GL_013 - General journals summarized by GL account per month and company.

This control summarizes BSIS and BSAS journal lines for the active GL period.  It
keeps the control independent from other GL controls and writes/replaces only the
GL13 sheet in the GL output workbook.
"""

from time import perf_counter

import pandas as pd

from core.gl_common import (
    build_company_name_map,
    build_gl_account_name_map,
    filter_by_company,
    get_gl_output_file,
    get_optional_column,
    load_gl_bsas_data,
    load_gl_bsis_data,
    load_gl_fx_rates_data,
    load_gl_master_data,
    normalize_company_output,
    normalize_fx_rates,
    normalize_text,
    open_or_create_gl_output_workbook,
    recreate_gl_sheet,
    require_columns,
    save_gl_output_workbook,
    select_fx_rate_to_usd,
    to_datetime_value,
    write_dataframe_to_sheet,
    write_single_sheet_workbook_fast,
)


SHEET_NAME = "GL13"
REPORT_CURRENCY = "USD"

REQUIRED_COLUMNS = {
    "company_code": ["Empr", "Empr.1", "BUKRS", "BKPF-BUKRS", "BSIS-BUKRS", "BSAS-BUKRS"],
    "fiscal_year": ["Ano", "GJAHR", "BKPF-GJAHR", "BSIS-GJAHR", "BSAS-GJAHR"],
    "document_number": ["Nº doc.", "Nº doc..1", "Nº doc", "BELNR", "BKPF-BELNR", "BSIS-BELNR", "BSAS-BELNR"],
    "gl_account": ["Razão", "Razao", "Cta.Razão", "Cta.Razao", "Cta.Razão.1", "Cta.Razao.1", "HKONT", "BSIS-HKONT", "BSAS-HKONT"],
    "debit_credit": ["D/C", "SHKZG", "BSIS-SHKZG", "BSAS-SHKZG"],
    "amount_document": ["Montante", "WRBTR", "BSIS-WRBTR", "BSAS-WRBTR"],
    "amount_local": ["Montante em MI", "DMBTR", "BSIS-DMBTR", "BSAS-DMBTR"],
    "document_currency": ["Moeda", "Moeda.1", "WAERS", "BKPF-WAERS"],
    "posting_date": ["Dt.lçto.", "Dt.lçto", "Dt.lcto.", "Dt.lcto", "BUDAT", "BKPF-BUDAT"],
}

OPTIONAL_COLUMNS = {
    "line_number": ["Itm", "BUZEI", "BSIS-BUZEI", "BSAS-BUZEI"],
    "entry_date": ["Dt.entr.", "Dt.entr", "CPUDT", "BKPF-CPUDT"],
}

BASE_OUTPUT_COLUMNS = [
    "Company Code",
    "Company Name",
    "GL Account",
    "GL Account Description",
    "Report Currency",
]

TRAILING_OUTPUT_COLUMNS = [
    "Total Amount in Reporting Currency",
    "Total Amount in Local Currency",
    "Total Amount in Document Currency",
    "Journal Count",
    "Line Count",
    "First Posting Date",
    "Last Posting Date",
    "Period From",
    "Period To",
    "Source",
]

DATE_COLUMNS = ["First Posting Date", "Last Posting Date", "Period From", "Period To"]
INTEGER_COLUMNS = ["Journal Count", "Line Count"]

GL13_COLUMN_WIDTHS = {
    "Company Code": 14,
    "Company Name": 28,
    "GL Account": 18,
    "GL Account Description": 34,
    "Report Currency": 16,
    "Total Amount in Reporting Currency": 34,
    "Total Amount in Local Currency": 30,
    "Total Amount in Document Currency": 34,
    "Journal Count": 16,
    "Line Count": 14,
    "First Posting Date": 18,
    "Last Posting Date": 18,
    "Period From": 14,
    "Period To": 14,
    "Source": 16,
}


def _log_stage(stage_name, start_time):
    elapsed = perf_counter() - start_time
    print(f"GL13 timing - {stage_name}: {elapsed:.2f} seconds")
    return perf_counter()


def _blank(index):
    return pd.Series("", index=index, dtype="object")


def _clean_text(series):
    return series.fillna("").astype(str).str.strip().replace(
        {"nan": "", "None": "", "<NA>": ""}
    )


def _normalize_code(series):
    return _clean_text(series).str.replace(r"\.0$", "", regex=True)


def _normalize_company(series):
    return _normalize_code(series).map(normalize_company_output)


def _parse_number(series):
    text = _clean_text(series)
    negative_parentheses = text.str.match(r"^\(.*\)$")
    text = text.str.replace(r"[()]", "", regex=True).str.replace(" ", "", regex=False)
    both_separators = text.str.contains(",", regex=False) & text.str.contains(".", regex=False)
    comma_decimal = text.str.contains(r",\d{1,6}$", regex=True)
    text = text.where(
        ~both_separators,
        text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
    )
    text = text.where(
        both_separators | ~comma_decimal,
        text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
    )
    text = text.where(
        both_separators | comma_decimal,
        text.str.replace(",", "", regex=False),
    )
    values = pd.to_numeric(text, errors="coerce")
    return values.where(~negative_parentheses, -values.abs())


def _parse_date(series):
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    numeric = pd.to_numeric(series, errors="coerce")
    excel_dates = pd.to_datetime(
        numeric,
        unit="D",
        origin="1899-12-30",
        errors="coerce",
    )
    return parsed.where(parsed.notna(), excel_dates)


def _resolve_optional_columns(dataframe):
    return {
        name: get_optional_column(dataframe, aliases)
        for name, aliases in OPTIONAL_COLUMNS.items()
    }


def _prepare_minimal_source(dataframe, source_name):
    if dataframe.empty:
        return pd.DataFrame()

    required = require_columns(
        dataframe=dataframe,
        required_columns=REQUIRED_COLUMNS,
        source_name=f"GL {source_name}",
    )
    optional = _resolve_optional_columns(dataframe)

    prepared = pd.DataFrame(index=dataframe.index)
    prepared["Company Code"] = _normalize_company(dataframe[required["company_code"]])
    prepared["Fiscal Year"] = _normalize_code(dataframe[required["fiscal_year"]])
    prepared["Document Number"] = _normalize_code(dataframe[required["document_number"]])
    prepared["GL Account"] = _normalize_code(dataframe[required["gl_account"]])
    prepared["Debit/Credit"] = _clean_text(dataframe[required["debit_credit"]]).str.upper()
    prepared["Amount in Document Currency Raw"] = _parse_number(dataframe[required["amount_document"]])
    prepared["Amount in Local Currency Raw"] = _parse_number(dataframe[required["amount_local"]])
    prepared["Document Currency"] = _clean_text(dataframe[required["document_currency"]]).str.upper()
    prepared["Posting Date"] = _parse_date(dataframe[required["posting_date"]])
    prepared["Source"] = source_name

    line_column = optional.get("line_number")
    if line_column is None:
        prepared["Line Number"] = ""
    else:
        prepared["Line Number"] = _normalize_code(dataframe[line_column])

    entry_date_column = optional.get("entry_date")
    if entry_date_column is None:
        prepared["Entry Date"] = pd.NaT
    else:
        prepared["Entry Date"] = _parse_date(dataframe[entry_date_column])

    return prepared


def _get_period_bounds(context):
    module_config = context["module"]
    period_from = to_datetime_value(module_config.get("from", ""))
    period_to = to_datetime_value(module_config.get("to", ""))

    if pd.isna(period_from) or pd.isna(period_to):
        raise ValueError("GL13 requires valid FROM and TO dates in config.xlsx.")

    return period_from.normalize(), period_to.normalize()


def _filter_by_period(dataframe, period_from, period_to):
    date_filter = (
        dataframe["Posting Date"].notna()
        & (dataframe["Posting Date"].dt.normalize() >= period_from)
        & (dataframe["Posting Date"].dt.normalize() <= period_to)
    )
    return dataframe.loc[date_filter].copy()


def _build_month_columns(period_from, period_to):
    month_range = pd.period_range(
        start=period_from.to_period("M"),
        end=period_to.to_period("M"),
        freq="M",
    )
    return [str(month) for month in month_range]


def _calculate_signed_amounts(dataframe):
    debit_credit = dataframe["Debit/Credit"].fillna("").astype(str).str.upper()
    document_amount = dataframe["Amount in Document Currency Raw"].abs()
    local_amount = dataframe["Amount in Local Currency Raw"].abs()

    dataframe["Amount in Document Currency Signed"] = document_amount.where(
        debit_credit != "H",
        -document_amount,
    )
    dataframe["Amount in Local Currency Signed"] = local_amount.where(
        debit_credit != "H",
        -local_amount,
    )

    return dataframe


def _build_journal_id(dataframe):
    return (
        dataframe["Company Code"].astype(str)
        + "|"
        + dataframe["Fiscal Year"].astype(str)
        + "|"
        + dataframe["Document Number"].astype(str)
    )


def _build_line_id(dataframe):
    return _build_journal_id(dataframe) + "|" + dataframe["Line Number"].astype(str)


def _convert_to_reporting_currency(dataframe, normalized_fx):
    dataframe["Report Currency"] = REPORT_CURRENCY
    dataframe["FX Method"] = ""
    dataframe["FX Rate"] = pd.NA
    dataframe["FX Rate Date"] = pd.NaT
    dataframe["Amount in Reporting Currency"] = pd.NA

    if dataframe.empty:
        return dataframe

    keys = dataframe[["Document Currency", "Posting Date"]].drop_duplicates().copy()
    fx_lookup = {}

    for row in keys.itertuples(index=False):
        currency = normalize_text(row[0]).upper()
        requested_date = row[1]
        lookup_key = (currency, requested_date.normalize() if not pd.isna(requested_date) else pd.NaT)
        fx_lookup[lookup_key] = select_fx_rate_to_usd(
            normalized_fx_dataframe=normalized_fx,
            currency=currency,
            requested_date=requested_date,
        )

    for (currency, requested_date), fx_rate in fx_lookup.items():
        if fx_rate is None:
            continue

        mask = (
            (dataframe["Document Currency"] == currency)
            & (dataframe["Posting Date"].dt.normalize() == requested_date)
        )
        dataframe.loc[mask, "Amount in Reporting Currency"] = (
            dataframe.loc[mask, "Amount in Document Currency Signed"] * fx_rate["fx_to_usd"]
        )
        dataframe.loc[mask, "FX Method"] = fx_rate["method"]
        dataframe.loc[mask, "FX Rate"] = fx_rate["fx_to_usd"]
        dataframe.loc[mask, "FX Rate Date"] = fx_rate["rate_date"]

    return dataframe


def _build_summary(dataframe, month_values, period_from, period_to):
    dataframe["Journal ID"] = _build_journal_id(dataframe)
    dataframe["Line ID"] = _build_line_id(dataframe)
    dataframe["Posting Month"] = dataframe["Posting Date"].dt.to_period("M").astype(str)

    monthly = dataframe.groupby(
        ["Company Code", "GL Account", "Report Currency", "Posting Month"],
        dropna=False,
        as_index=False,
    ).agg(
        **{
            "Amount in Reporting Currency": (
                "Amount in Reporting Currency",
                lambda values: values.sum(min_count=1),
            ),
        }
    )

    pivot = monthly.pivot_table(
        index=["Company Code", "GL Account", "Report Currency"],
        columns="Posting Month",
        values="Amount in Reporting Currency",
        aggfunc="sum",
        fill_value=0,
        dropna=False,
    )

    for month_value in month_values:
        if month_value not in pivot.columns:
            pivot[month_value] = 0.0

    pivot = pivot[month_values].reset_index()
    month_column_map = {
        month_value: f"Report Amount Month {month_index} ({month_value})"
        for month_index, month_value in enumerate(month_values, start=1)
    }
    pivot = pivot.rename(columns=month_column_map)
    monthly_output_columns = [month_column_map[month_value] for month_value in month_values]

    totals = dataframe.groupby(
        ["Company Code", "GL Account", "Report Currency"],
        dropna=False,
        as_index=False,
    ).agg(
        **{
            "Total Amount in Local Currency": ("Amount in Local Currency Signed", "sum"),
            "Total Amount in Document Currency": ("Amount in Document Currency Signed", "sum"),
            "Journal Count": ("Journal ID", "nunique"),
            "Line Count": ("Line ID", "nunique"),
            "First Posting Date": ("Posting Date", "min"),
            "Last Posting Date": ("Posting Date", "max"),
            "Source": ("Source", lambda values: "+".join(sorted(set(values.dropna().astype(str))))),
        }
    )

    summary = pivot.merge(
        totals,
        on=["Company Code", "GL Account", "Report Currency"],
        how="left",
    )
    summary["Total Amount in Reporting Currency"] = summary[monthly_output_columns].sum(axis=1)
    summary["Period From"] = period_from
    summary["Period To"] = period_to

    return summary, monthly_output_columns


def _enrich_summary(summary, master_dataframe):
    company_name_map = build_company_name_map(master_dataframe)
    account_name_map = build_gl_account_name_map(master_dataframe)

    summary["Company Name"] = summary["Company Code"].map(company_name_map).fillna("")
    summary["GL Account Description"] = summary["GL Account"].map(account_name_map).fillna("")

    return summary


def _apply_gl13_formatting(worksheet, dataframe, amount_columns, date_columns, integer_columns):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    column_positions = {
        column_name: column_index
        for column_index, column_name in enumerate(dataframe.columns, start=1)
    }

    for column_name in dataframe.columns:
        column_index = column_positions[column_name]
        width = GL13_COLUMN_WIDTHS.get(column_name)

        if width is None:
            if column_name.startswith("Report Amount Month"):
                width = 28
            else:
                width = min(max(len(str(column_name)) + 2, 12), 45)

        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    format_by_column = {
        **{column_name: "dd/mm/yyyy" for column_name in date_columns},
        **{column_name: "#,##0.00;[Red]-#,##0.00" for column_name in amount_columns},
        **{column_name: "#,##0" for column_name in integer_columns},
    }

    for column_name, number_format in format_by_column.items():
        column_index = column_positions.get(column_name)

        if column_index is None:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format


def _write_output(context, output_dataframe, amount_columns):
    output_file = get_gl_output_file(context)
    date_columns = DATE_COLUMNS
    integer_columns = INTEGER_COLUMNS

    if write_single_sheet_workbook_fast(
        output_file=output_file,
        sheet_name=SHEET_NAME,
        dataframe=output_dataframe,
        date_columns=date_columns,
        amount_columns=amount_columns,
        integer_columns=integer_columns,
    ):
        print(f"GL13 output file created: {output_file}")
        return output_file

    workbook = open_or_create_gl_output_workbook(output_file)
    worksheet = recreate_gl_sheet(workbook, SHEET_NAME)
    write_dataframe_to_sheet(worksheet, output_dataframe)
    _apply_gl13_formatting(
        worksheet=worksheet,
        dataframe=output_dataframe,
        amount_columns=amount_columns,
        date_columns=date_columns,
        integer_columns=integer_columns,
    )
    save_gl_output_workbook(workbook, output_file)
    print(f"GL13 output sheet written: {output_file} [{SHEET_NAME}]")

    return output_file


def run_gl_013(context):
    """Run GL_013."""
    total_start = perf_counter()
    stage_start = perf_counter()

    print("Running GL13 - General Journals Summarized By General Ledger Account Per Month And Per Company Code")
    print("GL13 input sufficiency: current BSIS, BSAS, GL_MD and optional FxRates inputs are sufficient; no new SAP SQVI download is required.")

    period_from, period_to = _get_period_bounds(context)
    month_values = _build_month_columns(period_from, period_to)
    print(f"GL13 period: {period_from.date()} to {period_to.date()}")
    print(f"GL13 months included: {len(month_values)} ({', '.join(month_values)})")

    bsis = load_gl_bsis_data(context)
    stage_start = _log_stage("carga BSIS", stage_start)

    bsas = load_gl_bsas_data(context)
    stage_start = _log_stage("carga BSAS", stage_start)

    if bsis.empty and bsas.empty:
        raise ValueError("GL13 requires at least one BSIS or BSAS input file.")

    prepared_frames = []

    if not bsis.empty:
        prepared_frames.append(_prepare_minimal_source(bsis, "BSIS"))

    if not bsas.empty:
        prepared_frames.append(_prepare_minimal_source(bsas, "BSAS"))

    journals = pd.concat(prepared_frames, ignore_index=True)
    del prepared_frames
    stage_start = _log_stage("normalización mínima", stage_start)

    journals = filter_by_company(
        dataframe=journals,
        company_column="Company Code",
        companies_filter=context["module"].get("companies", ""),
    )
    journals = _filter_by_period(journals, period_from, period_to)
    stage_start = _log_stage("filtro config", stage_start)

    journals = _calculate_signed_amounts(journals)
    stage_start = _log_stage("cálculo de importes firmados", stage_start)

    journals["Posting Month"] = journals["Posting Date"].dt.to_period("M").astype(str)
    stage_start = _log_stage("cálculo de Posting Month", stage_start)

    fx_rates = load_gl_fx_rates_data(context)
    normalized_fx = normalize_fx_rates(fx_rates) if not fx_rates.empty else normalize_fx_rates(pd.DataFrame())
    journals = _convert_to_reporting_currency(journals, normalized_fx)
    stage_start = _log_stage("FX / reporting currency", stage_start)

    summary, monthly_output_columns = _build_summary(
        dataframe=journals,
        month_values=month_values,
        period_from=period_from,
        period_to=period_to,
    )
    stage_start = _log_stage("agrupamiento mensual", stage_start)
    stage_start = _log_stage("pivot mensual", stage_start)

    master = load_gl_master_data(context)
    summary = _enrich_summary(summary, master)
    stage_start = _log_stage("enriquecimiento", stage_start)

    output_columns = BASE_OUTPUT_COLUMNS + monthly_output_columns + TRAILING_OUTPUT_COLUMNS
    summary = summary[output_columns].sort_values(
        by=["Company Code", "GL Account", "Report Currency"],
        kind="mergesort",
    )

    amount_columns = monthly_output_columns + [
        "Total Amount in Reporting Currency",
        "Total Amount in Local Currency",
        "Total Amount in Document Currency",
    ]

    output_file = _write_output(context, summary, amount_columns)
    stage_start = _log_stage("escritura Excel", stage_start)

    print(f"GL13 rows written: {len(summary)}")
    print(f"GL13 total elapsed: {perf_counter() - total_start:.2f} seconds")

    return {
        "status": "OK",
        "output_file": output_file,
        "sheet_name": SHEET_NAME,
        "rows": len(summary),
    }
