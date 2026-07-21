import sys

sys.dont_write_bytecode = True

import shutil
import tempfile
import traceback
from datetime import datetime
from pathlib import Path
from time import perf_counter

from openpyxl import load_workbook


def print_header(title):
    """
    Print a section header.
    """
    print(title)
    print("-" * len(title))


def format_elapsed_time(seconds):
    """
    Format elapsed seconds as HH:MM:SS.
    """
    total_seconds = int(round(seconds))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


class TeeStream:
    """
    Write console output to the original stream and to a log file.
    """

    def __init__(self, *streams):
        self.streams = streams

    def write(self, message):
        for stream in self.streams:
            stream.write(message)
            stream.flush()

    def flush(self):
        for stream in self.streams:
            stream.flush()


def create_log_file(project_folder):
    """
    Create a timestamped log file path under project-level logs folder.
    """
    logs_folder = Path(project_folder) / "logs"
    logs_folder.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    return logs_folder / f"LBR_Run_{timestamp}.log"


def setup_console_logging(project_folder):
    """
    Mirror stdout and stderr to a run log file.
    """
    log_file = create_log_file(project_folder)
    log_handle = log_file.open("w", encoding="utf-8")

    original_stdout = sys.stdout
    original_stderr = sys.stderr

    sys.stdout = TeeStream(original_stdout, log_handle)
    sys.stderr = TeeStream(original_stderr, log_handle)

    return {
        "log_file": log_file,
        "log_handle": log_handle,
        "original_stdout": original_stdout,
        "original_stderr": original_stderr,
    }


def restore_console_logging(logging_context):
    """
    Restore stdout/stderr and close the log file.
    """
    if logging_context is None:
        return

    sys.stdout = logging_context["original_stdout"]
    sys.stderr = logging_context["original_stderr"]

    try:
        logging_context["log_handle"].close()
    except OSError:
        pass


def normalize_excel_header(value):
    """
    Normalize Excel headers for case-insensitive matching.
    """
    if value is None:
        return ""

    return str(value).strip().lower()


def normalize_excel_control_id(value):
    """
    Normalize control IDs for config status updates.
    """
    import re

    control_id = str(value).strip().upper()

    if control_id == "" or control_id.lower() == "nan":
        return ""

    match = re.fullmatch(r"([A-Z]+)_?0*([0-9]+)", control_id)

    if not match:
        return control_id

    return f"{match.group(1)}_{int(match.group(2)):03d}"


def find_excel_sheet_name(workbook, requested_sheet_name):
    """
    Resolve an Excel sheet name using case-insensitive matching.
    """
    requested_sheet_name = str(requested_sheet_name).strip().lower()

    for sheet_name in workbook.sheetnames:
        if str(sheet_name).strip().lower() == requested_sheet_name:
            return sheet_name

    return None


def find_or_create_header_column(worksheet, header_name):
    """
    Find a header column in row 1 or create it at the end.
    """
    target_header = normalize_excel_header(header_name)

    for cell in worksheet[1]:
        if normalize_excel_header(cell.value) == target_header:
            return cell.column

    column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=column_index, value=header_name)

    return column_index


def find_header_column(worksheet, possible_names):
    """
    Find a header column in row 1.
    """
    possible_names = {
        normalize_excel_header(possible_name)
        for possible_name in possible_names
    }

    for cell in worksheet[1]:
        if normalize_excel_header(cell.value) in possible_names:
            return cell.column

    return None


def write_control_test_status_to_workbook(workbook, control, status):
    """
    Write TEST_STATUS for one control into an already-open config workbook.
    """
    sheet_name = find_excel_sheet_name(workbook, control["module"])

    if sheet_name is None:
        raise ValueError(
            f"Could not update TEST_STATUS because module sheet "
            f"'{control['module']}' was not found in config.xlsx."
        )

    worksheet = workbook[sheet_name]
    id_control_column = find_header_column(worksheet, ["ID Control", "TEST"])

    if id_control_column is None:
        raise ValueError(
            f"Could not update TEST_STATUS because sheet '{sheet_name}' does not "
            "contain ID Control or TEST column."
        )

    test_status_column = find_or_create_header_column(worksheet, "TEST_STATUS")
    expected_control_id = normalize_excel_control_id(control["id_control"])

    for row_index in range(2, worksheet.max_row + 1):
        raw_control_id = worksheet.cell(
            row=row_index,
            column=id_control_column,
        ).value
        control_id = normalize_excel_control_id(raw_control_id)

        if control_id != expected_control_id:
            continue

        worksheet.cell(
            row=row_index,
            column=test_status_column,
            value=str(status)[:32767],
        )
        return

    raise ValueError(
        f"Could not update TEST_STATUS because control "
        f"'{control['id_control']}' was not found in sheet '{sheet_name}'."
    )


def write_module_config_test_status_to_workbook(workbook, module_name, status):
    """
    Write TEST_STATUS for one module row into CONFIG sheet in an already-open workbook.
    """
    sheet_name = find_excel_sheet_name(workbook, "CONFIG")

    if sheet_name is None:
        raise ValueError(
            "Could not update CONFIG TEST_STATUS because CONFIG sheet was not found."
        )

    worksheet = workbook[sheet_name]
    module_column = 1
    test_status_column = find_or_create_header_column(worksheet, "TEST_STATUS")
    expected_module_name = str(module_name).strip().lower()

    for row_index in range(2, worksheet.max_row + 1):
        current_module_name = worksheet.cell(
            row=row_index,
            column=module_column,
        ).value
        current_module_name = str(current_module_name).strip().lower()

        if current_module_name != expected_module_name:
            continue

        worksheet.cell(
            row=row_index,
            column=test_status_column,
            value=str(status)[:32767],
        )
        return

    raise ValueError(
        f"Could not update CONFIG TEST_STATUS because module "
        f"'{module_name}' was not found in CONFIG sheet."
    )


def open_config_workbook_for_status(config_path):
    """
    Open config.xlsx to update TEST_STATUS values.
    """
    config_path = Path(config_path)

    try:
        return load_workbook(config_path)
    except PermissionError as error:
        raise PermissionError(
            f"Could not update TEST_STATUS in config file: {config_path}. "
            "The file may be open in Excel or locked by OneDrive. "
            "Close the workbook and run again."
        ) from error


def save_config_workbook_for_status(workbook, config_path):
    """
    Save config.xlsx after TEST_STATUS updates.
    """
    config_path = Path(config_path)

    try:
        workbook.save(config_path)
    except PermissionError as error:
        raise PermissionError(
            f"Could not update TEST_STATUS in config file: {config_path}. "
            "The file may be open in Excel or locked by OneDrive. "
            "Close the workbook and run again."
        ) from error


def get_error_status_message(error):
    """
    Build a clear TEST_STATUS error message.
    """
    if isinstance(error, PermissionError):
        return (
            f"ERROR: PermissionError: {error}. "
            "Close the workbook and run again."
        )

    return f"ERROR: {type(error).__name__}: {error}"


def update_test_statuses(config_path, module_statuses=None, control_statuses=None):
    """
    Update multiple TEST_STATUS cells with one open/save cycle.

    This avoids repeatedly saving config.xlsx to OneDrive, which is slow.
    """
    workbook = open_config_workbook_for_status(config_path)

    for module_name, status in module_statuses or []:
        write_module_config_test_status_to_workbook(
            workbook=workbook,
            module_name=module_name,
            status=status,
        )

    for control, status in control_statuses or []:
        write_control_test_status_to_workbook(
            workbook=workbook,
            control=control,
            status=status,
        )

    save_config_workbook_for_status(
        workbook=workbook,
        config_path=config_path,
    )


def update_control_test_status(config_path, control, status):
    """
    Update TEST_STATUS for one control in its module sheet.

    This writes to the original config.xlsx, not the temporary reader copy.
    """
    update_test_statuses(
        config_path=config_path,
        control_statuses=[(control, status)],
    )


def update_module_config_test_status(config_path, module_name, status):
    """
    Update TEST_STATUS for one module row in CONFIG/config/Config sheet.
    """
    update_test_statuses(
        config_path=config_path,
        module_statuses=[(module_name, status)],
    )


def initialize_module_config_statuses(config_path, active_modules):
    """
    Mark active module rows in CONFIG sheet as PENDING.
    """
    update_test_statuses(
        config_path=config_path,
        module_statuses=[
            (module["module"], "PENDING")
            for module in active_modules
        ],
    )


def update_successful_module_config_statuses(config_path, active_modules):
    """
    Mark active module rows in CONFIG sheet as OK after all controls finish.
    """
    update_test_statuses(
        config_path=config_path,
        module_statuses=[
            (module["module"], "OK")
            for module in active_modules
        ],
    )


def initialize_active_control_statuses(config_path, active_controls):
    """
    Mark active controls in module sheets as PENDING before execution starts.
    """
    update_test_statuses(
        config_path=config_path,
        control_statuses=[
            (control, "PENDING")
            for control in active_controls
        ],
    )


def initialize_run_test_statuses(config_path, active_modules, active_controls):
    """
    Initialize CONFIG and control TEST_STATUS values with one save.
    """
    update_test_statuses(
        config_path=config_path,
        module_statuses=[
            (module["module"], "PENDING")
            for module in active_modules
        ],
        control_statuses=[
            (control, "PENDING")
            for control in active_controls
        ],
    )


def cleanup_python_cache(project_folder):
    """
    Remove all __pycache__ folders and .pyc files inside the project,
    excluding virtual environments and other non-project folders.

    This should run before importing project modules.
    """
    project_folder = Path(project_folder)

    excluded_folder_names = {
        ".venv",
        "venv",
        "env",
        ".git",
        ".mypy_cache",
        ".pytest_cache",
    }

    removed_files = 0
    removed_folders = 0
    failed_items = []

    def is_excluded_path(path):
        path_parts = set(path.parts)

        return any(
            excluded_folder_name in path_parts
            for excluded_folder_name in excluded_folder_names
        )

    # 1. Remove .pyc files outside excluded folders.
    for pyc_file in project_folder.rglob("*.pyc"):
        if not pyc_file.is_file():
            continue

        if is_excluded_path(pyc_file):
            continue

        try:
            pyc_file.unlink()
            removed_files += 1
        except OSError as error:
            failed_items.append(
                {
                    "path": pyc_file,
                    "error": str(error),
                }
            )

    # 2. Remove __pycache__ folders outside excluded folders.
    pycache_folders = [
        folder
        for folder in project_folder.rglob("__pycache__")
        if folder.is_dir() and not is_excluded_path(folder)
    ]

    pycache_folders = sorted(
        pycache_folders,
        key=lambda folder: len(folder.parts),
        reverse=True,
    )

    for pycache_folder in pycache_folders:
        try:
            shutil.rmtree(pycache_folder)
            removed_folders += 1
        except OSError as error:
            failed_items.append(
                {
                    "path": pycache_folder,
                    "error": str(error),
                }
            )

    return {
        "removed_files": removed_files,
        "removed_folders": removed_folders,
        "failed_items": failed_items,
    }


def print_cache_cleanup_result(cache_cleanup_result, title="Python cache cleanup"):
    """
    Print Python cache cleanup result.

    Kept as a utility function, but main() intentionally does not call it
    so regular run logs stay focused on modules, controls, outputs and errors.
    """
    print(
        f"{title}: "
        f"{cache_cleanup_result['removed_files']} .pyc file(s) removed, "
        f"{cache_cleanup_result['removed_folders']} __pycache__ folder(s) removed."
    )

    if cache_cleanup_result["failed_items"]:
        print(f"{title} warnings:")

        for failed_item in cache_cleanup_result["failed_items"]:
            print(f"  Could not remove: {failed_item['path']}")
            print(f"  Reason: {failed_item['error']}")


def load_dependencies():
    """
    Import project dependencies after cache cleanup.

    This avoids trying to delete __pycache__ folders while imported modules
    are still being initialized or locked by the current Python process.
    """
    from core.config_reader import read_active_configuration

    from modules.AR.ar_001 import run_ar_001
    from modules.AR.ar_002 import run_ar_002
    from modules.AR.ar_003 import run_ar_003
    from modules.AR.ar_004 import run_ar_004
    from modules.AR.ar_005 import run_ar_005
    from modules.AR.ar_006 import run_ar_006

    from modules.CD.cd_001 import run_cd_001
    from modules.CD.cd_002 import run_cd_002
    from modules.CD.cd_003 import run_cd_003
    from modules.CD.cd_004 import run_cd_004

    from modules.FAM.fam_001 import run_fam_001
    from modules.FAM.fam_002 import run_fam_002
    from modules.FAM.fam_003 import run_fam_003
    from modules.FAM.fam_004 import run_fam_004

    from modules.GL.gl_001 import run_gl_001

    control_runners = {
        "AR_001": run_ar_001,
        "AR_002": run_ar_002,
        "AR_003": run_ar_003,
        "AR_004": run_ar_004,
        "AR_005": run_ar_005,
        "AR_006": run_ar_006,
        "CD_001": run_cd_001,
        "CD_002": run_cd_002,
        "CD_003": run_cd_003,
        "CD_004": run_cd_004,
        "FAM_001": run_fam_001,
        "FAM_002": run_fam_002,
        "FAM_003": run_fam_003,
        "FAM_004": run_fam_004,
        "GL_001": run_gl_001,
    }

    return read_active_configuration, control_runners


def create_temp_copy(file_path):
    """
    Create a temporary copy of an input workbook and return its path.

    This is useful for config.xlsx because the original workbook may be open
    in Excel. Reading from a temporary copy reduces lock-related issues.
    """
    file_path = Path(file_path)

    temp_folder = Path(tempfile.gettempdir()) / "lbr_analysis"
    temp_folder.mkdir(parents=True, exist_ok=True)

    temp_file = temp_folder / file_path.name

    shutil.copy2(file_path, temp_file)

    return temp_file


def cleanup_temp_config_copy(temp_config_path, original_config_path):
    """
    Delete temporary config copy when it is different from the original file.
    """
    if temp_config_path is None:
        return

    temp_config_path = Path(temp_config_path)
    original_config_path = Path(original_config_path)

    try:
        if temp_config_path.exists() and temp_config_path.resolve() != original_config_path.resolve():
            temp_config_path.unlink()
    except OSError:
        pass


def is_file_locked(file_path):
    """
    Return True if a file appears to be locked by another process.

    This is mainly useful on Windows when Excel has a workbook open.
    """
    file_path = Path(file_path)

    if not file_path.exists():
        return False

    try:
        with file_path.open("a+b"):
            return False
    except OSError:
        return True


def print_project_paths(project_folder, config_path, input_folder, output_folder):
    """
    Print main project paths.
    """
    print()
    print_header("LBR Automation Runner")
    print(f"Project folder: {project_folder}")
    print(f"Config file: {config_path}")
    print(f"Input folder: {input_folder}")
    print(f"Output folder: {output_folder}")
    print()


def validate_project_paths(config_path, input_folder, output_folder):
    """
    Validate required project paths before processing.
    """
    if not config_path.exists():
        print("ERROR: config.xlsx was not found.")
        print(f"Expected file: {config_path}")
        print()
        return False

    if not input_folder.exists():
        print("ERROR: input folder was not found.")
        print(f"Expected folder: {input_folder}")
        print()
        return False

    output_folder.mkdir(parents=True, exist_ok=True)

    return True


def print_active_modules(active_modules):
    """
    Print active modules from config.xlsx.
    """
    print_header("Active modules")

    if len(active_modules) == 0:
        print("No active modules found.")
        print()
        return

    for module in active_modules:
        print(
            f"{module['module']} | "
            f"From: {module['from']} | "
            f"To: {module['to']} | "
            f"Companies: {module['companies']}"
        )

    print()


def print_active_controls(active_controls):
    """
    Print active controls from config.xlsx.
    """
    print_header("Active controls")

    if len(active_controls) == 0:
        print("No active controls found.")
        print()
        return

    for control in active_controls:
        print(f"{control['module']} | {control['id_control']} | {control['name']}")

    print()


def create_execution_summary(active_controls):
    """
    Create an execution summary grouped by module.
    """
    summary_by_module = {}

    for control in active_controls:
        module_name = control["module"]

        if module_name not in summary_by_module:
            summary_by_module[module_name] = {
                "total": 0,
                "ok": 0,
                "error": 0,
                "pending": 0,
                "elapsed_seconds": 0.0,
                "controls": [],
                "status": "PENDING",
            }

        summary_by_module[module_name]["total"] += 1
        summary_by_module[module_name]["pending"] += 1
        summary_by_module[module_name]["controls"].append(
            {
                "id_control": control["id_control"],
                "name": control["name"],
                "status": "PENDING",
                "elapsed_seconds": None,
                "message": "",
            }
        )

    return summary_by_module


def update_execution_summary_control(
    summary_by_module,
    module_name,
    control_id,
    status,
    elapsed_seconds=None,
    message="",
):
    """
    Update one control entry in the execution summary.
    """
    module_summary = summary_by_module[module_name]

    for control_summary in module_summary["controls"]:
        if control_summary["id_control"] != control_id:
            continue

        previous_status = control_summary["status"]

        if previous_status == "OK":
            module_summary["ok"] -= 1
        elif previous_status == "ERROR":
            module_summary["error"] -= 1
        elif previous_status == "PENDING":
            module_summary["pending"] -= 1

        control_summary["status"] = status
        control_summary["elapsed_seconds"] = elapsed_seconds
        control_summary["message"] = message

        if status == "OK":
            module_summary["ok"] += 1
        elif status == "ERROR":
            module_summary["error"] += 1
        else:
            module_summary["pending"] += 1

        if elapsed_seconds is not None:
            module_summary["elapsed_seconds"] += elapsed_seconds

        if module_summary["error"] > 0:
            module_summary["status"] = "ERROR"
        elif module_summary["pending"] == 0:
            module_summary["status"] = "OK"
        else:
            module_summary["status"] = "PENDING"

        return


def print_execution_summary(summary_by_module):
    """
    Print final execution summary grouped by module and control.
    """
    print()
    print_header("Execution summary")

    if len(summary_by_module) == 0:
        print("No controls were executed.")
        print()
        return

    for module_name, module_summary in summary_by_module.items():
        elapsed_seconds = module_summary["elapsed_seconds"]
        print(
            f"Module {module_name}: {module_summary['status']} | "
            f"Controls: {module_summary['ok']} OK, "
            f"{module_summary['error']} ERROR, "
            f"{module_summary['pending']} PENDING | "
            f"Elapsed: {format_elapsed_time(elapsed_seconds)} "
            f"({elapsed_seconds:.2f} seconds)"
        )

        for control_summary in module_summary["controls"]:
            elapsed_seconds = control_summary["elapsed_seconds"]

            if elapsed_seconds is None:
                elapsed_text = "not completed"
            else:
                elapsed_text = (
                    f"{format_elapsed_time(elapsed_seconds)} "
                    f"({elapsed_seconds:.2f} seconds)"
                )

            line = (
                f"  - {control_summary['id_control']}: "
                f"{control_summary['status']} | {elapsed_text}"
            )

            if control_summary["message"]:
                line = f"{line} | {control_summary['message']}"

            print(line)

    print()


def run_active_controls(
    active_controls,
    modules_by_name,
    project_folder,
    config_path,
    input_folder,
    output_folder,
    control_runners,
):
    """
    Execute all active controls found in config.xlsx.
    """
    print_header("Execution")
    summary_by_module = create_execution_summary(active_controls)
    active_control_counts_by_module = {}
    completed_control_counts_by_module = {}
    current_module_name = None

    for control in active_controls:
        module_name = control["module"]
        active_control_counts_by_module[module_name] = (
            active_control_counts_by_module.get(module_name, 0) + 1
        )
        completed_control_counts_by_module[module_name] = 0

    for control in active_controls:
        control_id = control["id_control"]
        module_name = control["module"]

        if module_name != current_module_name:
            if current_module_name is not None:
                print()

            print_header(f"Module {module_name}")
            current_module_name = module_name

        if control_id not in control_runners:
            message = f"ERROR: No runner implemented for {control_id}"
            print(message)
            update_execution_summary_control(
                summary_by_module=summary_by_module,
                module_name=module_name,
                control_id=control_id,
                status="ERROR",
                message=message,
            )
            print_execution_summary(summary_by_module)
            update_test_statuses(
                config_path=config_path,
                module_statuses=[(module_name, "ERROR")],
                control_statuses=[(control, message)],
            )
            raise ValueError(message)

        context = {
            "project_folder": project_folder,
            "config_path": config_path,
            "input_folder": input_folder,
            "output_folder": output_folder,
            "module": modules_by_name[module_name],
            "control": control,
        }

        print(f"Running {control_id} - {control['name']}")

        control_start_time = perf_counter()

        try:
            control_runners[control_id](context)
            control_elapsed_time = perf_counter() - control_start_time
            completed_control_counts_by_module[module_name] += 1
            update_execution_summary_control(
                summary_by_module=summary_by_module,
                module_name=module_name,
                control_id=control_id,
                status="OK",
                elapsed_seconds=control_elapsed_time,
            )

            if (
                completed_control_counts_by_module[module_name]
                == active_control_counts_by_module[module_name]
            ):
                update_test_statuses(
                    config_path=config_path,
                    module_statuses=[(module_name, "OK")],
                    control_statuses=[(control, "OK")],
                )
            else:
                update_test_statuses(
                    config_path=config_path,
                    control_statuses=[(control, "OK")],
                )
        except PermissionError as error:
            control_elapsed_time = perf_counter() - control_start_time
            status_message = get_error_status_message(error)
            update_execution_summary_control(
                summary_by_module=summary_by_module,
                module_name=module_name,
                control_id=control_id,
                status="ERROR",
                elapsed_seconds=control_elapsed_time,
                message=status_message,
            )
            update_test_statuses(
                config_path=config_path,
                module_statuses=[(module_name, "ERROR")],
                control_statuses=[(control, status_message)],
            )
            print()
            print(f"ERROR while running {control_id}.")
            print("A file may be open in Excel or locked by another process.")
            print(f"Details: {error}")
            print()
            print_execution_summary(summary_by_module)
            raise
        except BaseException as error:
            control_elapsed_time = perf_counter() - control_start_time
            status_message = f"ERROR: {type(error).__name__}: {error}"
            update_execution_summary_control(
                summary_by_module=summary_by_module,
                module_name=module_name,
                control_id=control_id,
                status="ERROR",
                elapsed_seconds=control_elapsed_time,
                message=status_message,
            )
            update_test_statuses(
                config_path=config_path,
                module_statuses=[(module_name, "ERROR")],
                control_statuses=[(control, status_message)],
            )
            print()
            print(f"ERROR while running {control_id}.")
            print(f"Details: {type(error).__name__}: {error}")
            print()
            print_execution_summary(summary_by_module)
            raise

        print(
            f"{control_id} elapsed time: "
            f"{format_elapsed_time(control_elapsed_time)} "
            f"({control_elapsed_time:.2f} seconds)"
        )
        print()

    print_execution_summary(summary_by_module)


def main():
    """
    Main automation runner.

    This script:
    1. Cleans old Python cache before importing project modules.
    2. Imports project dependencies after cleanup.
    3. Reads config.xlsx from a temporary copy.
    4. Detects active modules.
    5. Detects active controls.
    6. Executes registered Python runners.
    7. Prints elapsed time by control and total run.

    Note:
    - sys.dont_write_bytecode = True prevents new .pyc files from being written.
    - .venv is excluded from cache cleanup.
    """
    run_start_time = perf_counter()
    temp_config_path = None
    logging_context = None

    project_folder = Path(__file__).parent
    config_path = project_folder / "config.xlsx"
    input_folder = project_folder / "input"
    output_folder = project_folder / "output"

    output_folder.mkdir(parents=True, exist_ok=True)
    logging_context = setup_console_logging(project_folder)

    cleanup_python_cache(project_folder)

    read_active_configuration, control_runners = load_dependencies()

    print_project_paths(
        project_folder=project_folder,
        config_path=config_path,
        input_folder=input_folder,
        output_folder=output_folder,
    )
    print(f"Log file: {logging_context['log_file']}")
    print()

    if not validate_project_paths(
        config_path=config_path,
        input_folder=input_folder,
        output_folder=output_folder,
    ):
        restore_console_logging(logging_context)
        return

    try:
        try:
            temp_config_path = create_temp_copy(config_path)
        except OSError as error:
            print("ERROR: config.xlsx could not be copied for reading.")
            print("Please close it if it is open in Excel and try again.")
            print(f"Config file: {config_path}")
            print(f"Details: {error}")
            print()
            return

        active_configuration = read_active_configuration(temp_config_path)

        active_modules = active_configuration["active_modules"]
        active_controls = active_configuration["active_controls"]
        modules_by_name = active_configuration["modules_by_name"]

        print_active_modules(active_modules)

        if len(active_modules) == 0:
            return

        print_active_controls(active_controls)

        if len(active_controls) == 0:
            return

        print("Initializing TEST_STATUS for active modules and controls...")
        initialize_run_test_statuses(
            config_path=config_path,
            active_modules=active_modules,
            active_controls=active_controls,
        )
        print("TEST_STATUS initialized.")
        print()

        run_active_controls(
            active_controls=active_controls,
            modules_by_name=modules_by_name,
            project_folder=project_folder,
            config_path=config_path,
            input_folder=input_folder,
            output_folder=output_folder,
            control_runners=control_runners,
        )

    finally:
        exception_info = sys.exc_info()

        cleanup_temp_config_copy(
            temp_config_path=temp_config_path,
            original_config_path=config_path,
        )

        run_elapsed_time = perf_counter() - run_start_time

        print()
        print_header("Runner status")
        if exception_info[0] is None:
            print("Finished.")
        else:
            print("Finished with errors.")
            print()
            print_header("Unhandled exception")
            traceback.print_exception(*exception_info)
            print()
        print(
            f"Total elapsed time: "
            f"{format_elapsed_time(run_elapsed_time)} "
            f"({run_elapsed_time:.2f} seconds)"
        )
        print()

        print(f"Log file: {logging_context['log_file']}")

        restore_console_logging(logging_context)


if __name__ == "__main__":
    main()
